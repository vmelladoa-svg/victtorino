"""
Diagnostico de margen completo cruzando:
  - matching_defontana_ml_<fecha>.xlsx  (508 items con costo)
  - candidatos_pausar_<fecha>.xlsx      (estado buy box por listing)

Asume ML_COST_FACTOR = 0.40 (comision ML + IVA + envio + operacion).
  ingreso_neto = price_ml * (1 - 0.40)
  margen_clp   = ingreso_neto - costo_vigente
  margen_pct   = margen_clp / price_ml * 100      (sobre venta)
  markup_pct   = margen_clp / costo_vigente * 100 (sobre costo)
  precio_break_even = costo / 0.60  (precio minimo sin perder)

Output: margen_diagnostico_<fecha>.xlsx con 6 hojas:
  RESUMEN
  TODOS_LOS_LISTINGS
  ALERTA_A_PERDIDA
  ALERTA_MARGEN_BAJO        (<10%)
  OPORTUNIDAD_MARGEN_ALTO   (>40%)
  PIERDE_VS_EXT_CON_MARGEN  (los que pueden competir bajando precio)
  PIERDE_VS_EXT_SIN_CHANCE  (externo te gana incluso al break-even)
"""
from __future__ import annotations
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
MATCHING_XLSX = ROOT / f"matching_defontana_ml_{time.strftime('%Y-%m-%d')}.xlsx"
BUYBOX_XLSX = ROOT / f"candidatos_pausar_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"margen_diagnostico_{time.strftime('%Y-%m-%d')}.xlsx"

# Modelo segun reporte "Margen por canal" Victor 2026-05-26:
#   venta_neta = price_ml / 1.19  (saca IVA)
#   cobro_canal_ml = venta_neta * 0.42  (cargo ML sobre venta neta)
#   ingreso_efectivo = venta_neta - cobro_canal_ml = venta_neta * 0.58
#                    = price_ml / 1.19 * 0.58 = price_ml * 0.4874
#   margen = ingreso_efectivo - costo
#   margen_pct = margen / venta_neta  (asi calcula Victor)
IVA = 0.19
ML_CHANNEL_FEE = 0.32  # sobre venta neta (ajustado 2026-05-26 por Victor)
NETO_FACTOR = (1 / (1 + IVA)) * (1 - ML_CHANNEL_FEE)  # 0.4874


def cargar_hoja(path: Path, nombre: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    if nombre not in wb.sheetnames:
        return []
    ws = wb[nombre]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def calcular_margen(price: float, costo: float) -> dict:
    if price is None or costo is None or price <= 0 or costo <= 0:
        return {"ingreso_neto": None, "margen_clp": None, "margen_pct": None,
                "markup_pct": None, "precio_breakeven": None, "estado_margen": "sin_datos"}
    venta_neta = price / (1 + IVA)
    ingreso = venta_neta * (1 - ML_CHANNEL_FEE)  # = price * NETO_FACTOR
    margen = ingreso - costo
    margen_pct = margen / venta_neta * 100 if venta_neta else 0  # como Victor lo mide
    markup_pct = margen / costo * 100
    breakeven = costo / NETO_FACTOR  # precio_publicado min sin perder
    if margen < 0:
        estado = "a_perdida"
    elif margen_pct < 5:
        estado = "margen_critico"
    elif margen_pct < 15:
        estado = "margen_bajo"
    elif margen_pct < 30:
        estado = "margen_sano"
    elif margen_pct < 45:
        estado = "margen_alto"
    else:
        estado = "margen_muy_alto"
    return {
        "ingreso_neto": round(ingreso),
        "margen_clp": round(margen),
        "margen_pct": round(margen_pct, 1),
        "markup_pct": round(markup_pct, 1),
        "precio_breakeven": round(breakeven),
        "estado_margen": estado,
    }


def fmt(v) -> str:
    try:
        return f"${int(v):,}".replace(",", ".")
    except Exception:
        return "-"


def main() -> int:
    if not MATCHING_XLSX.exists() or not BUYBOX_XLSX.exists():
        print(f"ERROR: faltan {MATCHING_XLSX.name} o {BUYBOX_XLSX.name}")
        return 1

    matched = cargar_hoja(MATCHING_XLSX, "MATCHED")
    print(f"Items con costo conocido: {len(matched)}")

    # info buy box: indexar por item_id
    bb_pausar = cargar_hoja(BUYBOX_XLSX, "PAUSAR_SEGURO")
    bb_revisar = cargar_hoja(BUYBOX_XLSX, "REVISAR_MANUAL")
    estado_bb: dict[str, dict] = {}
    for r in bb_pausar:
        estado_bb[r.get("item_id")] = {"estado_bb": "PAUSAR_SEGURO", **r}
    for r in bb_revisar:
        estado_bb[r.get("item_id")] = {"estado_bb": "PIERDE_VS_EXTERNO", **r}

    # construir filas enriquecidas
    rows: list[dict] = []
    for it in matched:
        price = it.get("price_ml")
        costo = it.get("costo_vigente")
        m = calcular_margen(price, costo)
        bb = estado_bb.get(it.get("item_id"), {})
        ganador_precio = bb.get("ganador_precio")
        ganador_alias = bb.get("ganador_cuenta")
        ganador_seller_nick = bb.get("ext_seller_nickname")

        # ¿puedo competir vs externo bajando precio?
        puede_competir = None
        if bb.get("estado_bb") == "PIERDE_VS_EXTERNO" and ganador_precio and m["precio_breakeven"]:
            puede_competir = m["precio_breakeven"] < ganador_precio
        elif bb.get("estado_bb") == "PIERDE_VS_EXTERNO":
            puede_competir = None

        rows.append({
            **it,
            **m,
            "estado_bb": bb.get("estado_bb") or "no_overlap_o_gana",
            "ganador_cuenta": ganador_alias,
            "ganador_precio": ganador_precio,
            "ext_seller_nick": ganador_seller_nick,
            "diff_pct_vs_ganador": bb.get("diff_pct_vs_ganador"),
            "puede_competir_vs_ext": puede_competir,
            "margen_si_iguala_ext": (
                round(ganador_precio * NETO_FACTOR - costo)
                if (ganador_precio and costo) else None
            ),
            "margen_pct_si_iguala_ext": (
                round((ganador_precio * NETO_FACTOR - costo) / (ganador_precio / (1 + IVA)) * 100, 1)
                if (ganador_precio and costo and ganador_precio > 0) else None
            ),
        })

    # ---------- distribuciones ----------
    estados = Counter(r["estado_margen"] for r in rows)
    print("\n========== ESTADO DE MARGEN (508 items con costo) ==========")
    for k in ["a_perdida", "margen_critico", "margen_bajo", "margen_sano",
              "margen_alto", "margen_muy_alto", "sin_datos"]:
        n = estados.get(k, 0)
        pct = n / len(rows) * 100 if rows else 0
        print(f"  {k:<20}: {n:>4}  ({pct:>5.1f}%)")

    # por cuenta
    print("\n========== POR CUENTA ==========")
    por_cuenta = defaultdict(Counter)
    for r in rows:
        por_cuenta[r["cuenta"]][r["estado_margen"]] += 1
    for c in ["C1", "C2", "C3"]:
        ct = por_cuenta[c]
        total = sum(ct.values())
        print(f"  {c} ({total} items):")
        for k in ["a_perdida", "margen_critico", "margen_bajo", "margen_sano",
                  "margen_alto", "margen_muy_alto"]:
            if ct.get(k):
                print(f"    {k:<18}: {ct[k]:>4} ({ct[k]/total*100:.1f}%)")

    # cruce con buy box
    print("\n========== CRUCE BUY BOX x MARGEN ==========")
    cross = defaultdict(Counter)
    for r in rows:
        cross[r["estado_bb"]][r["estado_margen"]] += 1
    for bb_estado, counts in cross.items():
        tot = sum(counts.values())
        print(f"  {bb_estado} ({tot}):")
        for k, n in sorted(counts.items(), key=lambda x: -x[1]):
            print(f"    {k:<18}: {n} ({n/tot*100:.1f}%)")

    # ---------- alertas ----------
    a_perdida = [r for r in rows if r["estado_margen"] == "a_perdida"]
    margen_bajo = [r for r in rows if r["estado_margen"] in ("margen_critico", "margen_bajo")]
    margen_alto = [r for r in rows if r["estado_margen"] in ("margen_alto", "margen_muy_alto")]
    pierde_con_chance = [r for r in rows
                          if r["estado_bb"] == "PIERDE_VS_EXTERNO" and r["puede_competir_vs_ext"] is True]
    pierde_sin_chance = [r for r in rows
                          if r["estado_bb"] == "PIERDE_VS_EXTERNO" and r["puede_competir_vs_ext"] is False]

    # margen evitable: items con margen alto Y perdiendo buy box (¿bajar precio?)
    margen_alto_y_pierde = [r for r in rows
                            if r["estado_margen"] in ("margen_alto", "margen_muy_alto")
                            and r["estado_bb"] in ("PIERDE_VS_EXTERNO", "PAUSAR_SEGURO")]

    print(f"\n========== ALERTAS ==========")
    print(f"  A perdida                            : {len(a_perdida)}")
    print(f"  Margen <15%                          : {len(margen_bajo)}")
    print(f"  Margen alto >30%                     : {len(margen_alto)}")
    print(f"  PIERDE vs ext PERO puede competir    : {len(pierde_con_chance)}")
    print(f"  PIERDE vs ext SIN chance (a perdida) : {len(pierde_sin_chance)}")
    print(f"  Margen alto Y perdiendo buy box      : {len(margen_alto_y_pierde)}")

    # ---------- xlsx ----------
    wb = Workbook()
    wb.remove(wb.active)

    # RESUMEN
    res = wb.create_sheet("RESUMEN")
    res.append(["==== ESTADO DE MARGEN (508 items) ===="])
    for k in ["a_perdida", "margen_critico", "margen_bajo", "margen_sano",
              "margen_alto", "margen_muy_alto", "sin_datos"]:
        n = estados.get(k, 0)
        res.append([k, n, f"{n/len(rows)*100:.1f}%" if rows else ""])
    res.append([])
    res.append(["==== POR CUENTA ===="])
    for c in ["C1", "C2", "C3"]:
        ct = por_cuenta[c]
        res.append([c, sum(ct.values())])
        for k in ["a_perdida", "margen_critico", "margen_bajo", "margen_sano",
                  "margen_alto", "margen_muy_alto"]:
            if ct.get(k):
                res.append([f"  {k}", ct[k]])
    res.append([])
    res.append(["==== ALERTAS ===="])
    res.append(["A perdida", len(a_perdida)])
    res.append(["Margen <15%", len(margen_bajo)])
    res.append(["Margen alto >30%", len(margen_alto)])
    res.append(["PIERDE vs ext PERO puede competir", len(pierde_con_chance)])
    res.append(["PIERDE vs ext SIN chance (a perdida)", len(pierde_sin_chance)])
    res.append(["Margen alto Y perdiendo buy box", len(margen_alto_y_pierde)])
    res.column_dimensions["A"].width = 42
    res.column_dimensions["B"].width = 10
    res.column_dimensions["C"].width = 10

    cols = [
        "cuenta", "item_id", "title_ml", "seller_sku", "codigo_defontana",
        "costo_vigente", "price_ml", "ingreso_neto", "margen_clp",
        "margen_pct", "markup_pct", "precio_breakeven", "estado_margen",
        "listing_type_id", "available_quantity_ml", "stock_defontana",
        "estado_bb", "ganador_cuenta", "ext_seller_nick", "ganador_precio",
        "diff_pct_vs_ganador", "puede_competir_vs_ext",
        "margen_si_iguala_ext", "margen_pct_si_iguala_ext",
        "catalog_product_id", "permalink",
    ]

    fills_margen = {
        "a_perdida":       PatternFill("solid", fgColor="EF5350"),  # rojo fuerte
        "margen_critico":  PatternFill("solid", fgColor="FF8A65"),
        "margen_bajo":     PatternFill("solid", fgColor="FFCC80"),
        "margen_sano":     PatternFill("solid", fgColor="A5D6A7"),
        "margen_alto":     PatternFill("solid", fgColor="64B5F6"),
        "margen_muy_alto": PatternFill("solid", fgColor="9575CD"),
        "sin_datos":       PatternFill("solid", fgColor="ECEFF1"),
    }

    def dump(nombre: str, data: list[dict], header_color: str = "455A64") -> None:
        ws = wb.create_sheet(nombre)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=header_color)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for r in sorted(data, key=lambda x: (x.get("margen_pct") or 0)):
            ws.append([r.get(c) for c in cols])
            last = ws.max_row
            ws.cell(row=last, column=cols.index("estado_margen") + 1).fill = fills_margen.get(r["estado_margen"], None) or PatternFill()
        ws.freeze_panes = "C2"
        widths = {
            "title_ml": 55, "permalink": 45, "ext_seller_nick": 18,
            "catalog_product_id": 16, "item_id": 16, "codigo_defontana": 14,
            "seller_sku": 16, "puede_competir_vs_ext": 10,
        }
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 13)

    dump("TODOS_LOS_LISTINGS", rows, "455A64")
    dump("ALERTA_A_PERDIDA", a_perdida, "C62828")
    dump("ALERTA_MARGEN_BAJO", margen_bajo, "EF6C00")
    dump("OPORTUNIDAD_MARGEN_ALTO", margen_alto, "1565C0")
    dump("PIERDE_VS_EXT_CON_CHANCE", pierde_con_chance, "2E7D32")
    dump("PIERDE_VS_EXT_SIN_CHANCE", pierde_sin_chance, "AD1457")
    dump("MARGEN_ALTO_Y_PIERDE_BB", margen_alto_y_pierde, "6A1B9A")

    wb.save(DST)
    print(f"\nXlsx: {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
