"""
Para cada catalog_product_id con >=2 cuentas nuestras, consulta el endpoint
/products/{cpid} de ML para saber quien gana el buy box AHORA mismo.

Despues clasifica cada listing nuestro:
  - GANA_BUY_BOX        : tu listing tiene el buy box
  - PIERDE_VS_NUESTRO   : pierde contra otra cuenta tuya (canibalizacion real)
  - PIERDE_VS_EXTERNO   : pierde contra un seller externo
  - SIN_STOCK           : stock 0, no compite
  - SIN_BUY_BOX         : la ficha no tiene buy box (raro)

Output:
  - Consola: distribucion por categoria
  - candidatos_pausar_<fecha>.xlsx con 2 hojas:
      PAUSAR_SEGURO  : C1/C2 que PIERDE_VS_NUESTRO contra C3
      REVISAR_MANUAL : pierden vs externo, o casos ambiguos
"""
from __future__ import annotations
import sys
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from ml_catalogo_publicaciones import (
    CUENTAS, Cuenta, _request,
)

ROOT = Path(__file__).parent
SRC = ROOT / f"catalogo_publicaciones_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"candidatos_pausar_{time.strftime('%Y-%m-%d')}.xlsx"

# user_id -> alias (poblado de tokens_cuenta*.json)
MIS_USER_IDS: dict[int, str] = {}


def cargar_filas(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    filas = []
    for s in wb.sheetnames:
        if s == "RESUMEN":
            continue
        ws = wb[s]
        rows = ws.iter_rows(values_only=True)
        h = next(rows)
        for r in rows:
            filas.append(dict(zip(h, r)))
    return filas


def get_buy_box_winner(cuenta_query: Cuenta, cpid: str) -> dict | None:
    """Devuelve {item_id, seller_id, price, total_competitors, all_items} del ganador efectivo.

    Usa /products/{cpid}/items (el primer resultado es el ganador en el ranking).
    buy_box_winner en /products/{cpid} viene null para casi todas las fichas.
    """
    r = _request(cuenta_query, f"/products/{cpid}/items", {"limit": 50})
    if r.status_code != 200:
        return None
    j = r.json()
    results = j.get("results") or []
    if not results:
        return None
    top = results[0]
    return {
        "item_id": top.get("item_id"),
        "seller_id": top.get("seller_id"),
        "price": top.get("price"),
        "total_competitors": j.get("paging", {}).get("total", len(results)),
        "all_items": [
            {"item_id": x.get("item_id"), "seller_id": x.get("seller_id"), "price": x.get("price")}
            for x in results
        ],
    }


def fmt_money(v) -> str:
    try:
        return f"${int(v):,}".replace(",", ".")
    except Exception:
        return ""


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: no existe {SRC}")
        return 1

    # cargar mis user_ids
    for cfg in CUENTAS:
        try:
            cu = Cuenta.cargar(cfg)
            MIS_USER_IDS[cu.user_id] = cu.alias
        except FileNotFoundError:
            pass
    print(f"Mis cuentas: {MIS_USER_IDS}")

    filas = cargar_filas(SRC)
    grupos: dict[str, list[dict]] = defaultdict(list)
    for f in filas:
        cp = f.get("catalog_product_id")
        if cp:
            grupos[cp].append(f)

    # solo overlap >=2 cuentas y al menos uno con stock
    candidatos = {
        cp: its for cp, its in grupos.items()
        if len({i["cuenta"] for i in its}) >= 2
        and any((i.get("available_quantity") or 0) > 0 for i in its)
    }
    print(f"Fichas a consultar buy box: {len(candidatos)}")

    # reutilizar 1 cuenta para queries (mejor C3 que es la mas usada)
    cuenta_query = Cuenta.cargar(next(c for c in CUENTAS if c["alias"] == "C3"))

    bbox_por_cpid: dict[str, dict | None] = {}
    for i, cpid in enumerate(candidatos.keys(), 1):
        bbw = get_buy_box_winner(cuenta_query, cpid)
        bbox_por_cpid[cpid] = bbw
        if i % 10 == 0:
            print(f"  {i}/{len(candidatos)}...")
        time.sleep(0.08)
    print(f"  {len(candidatos)}/{len(candidatos)} listo")

    # clasificar
    pausar_seguro: list[dict] = []  # mio que pierde vs otro mio
    revisar_manual: list[dict] = []  # pierde vs externo, o duplicado intracuenta
    estadistica = defaultdict(int)

    for cpid, items in candidatos.items():
        bbw = bbox_por_cpid.get(cpid)
        if not bbw or not bbw.get("item_id"):
            for it in items:
                estadistica["SIN_BUY_BOX"] += 1
            continue

        ganador_iid = bbw["item_id"]
        ganador_sid = bbw.get("seller_id")
        ganador_es_mio = ganador_sid in MIS_USER_IDS
        ganador_alias = MIS_USER_IDS.get(ganador_sid, "EXTERNO")

        # detectar duplicados intracuenta (misma cuenta, mismo cpid, distintos item_id)
        from collections import Counter
        cuentas_count = Counter(it["cuenta"] for it in items)
        cuentas_duplicadas = {c for c, n in cuentas_count.items() if n > 1}

        for it in items:
            stock = it.get("available_quantity") or 0
            iid = it.get("item_id")
            if stock == 0:
                estadistica["SIN_STOCK"] += 1
                continue
            base = {
                **it,
                "ganador_cuenta": ganador_alias,
                "ganador_item_id": ganador_iid,
                "ganador_precio": bbw.get("price"),
                "diff_pct_vs_ganador": _diff_pct(it.get("price"), bbw.get("price")),
                "total_competidores": bbw.get("total_competitors"),
                "dup_intracuenta": it["cuenta"] in cuentas_duplicadas,
            }
            if iid == ganador_iid:
                estadistica["GANA_BUY_BOX"] += 1
                # si gana pero hay duplicado intracuenta, su gemelo es candidato a pausar
                continue
            # pierde
            if base["dup_intracuenta"]:
                estadistica["DUPLICADO_INTRACUENTA"] += 1
                pausar_seguro.append({**base, "ganador_seller_id": ganador_sid})
                continue
            if ganador_es_mio:
                estadistica["PIERDE_VS_NUESTRO"] += 1
                pausar_seguro.append({**base, "ganador_seller_id": ganador_sid})
            else:
                estadistica["PIERDE_VS_EXTERNO"] += 1
                revisar_manual.append({**base, "ganador_seller_id": ganador_sid})

    # ---------- consola ----------
    print("\n========== DISTRIBUCION ==========")
    for k, v in sorted(estadistica.items(), key=lambda x: -x[1]):
        print(f"  {k:>20} : {v}")
    print(f"\n  PAUSAR_SEGURO (mio pierde vs mio C3): {len(pausar_seguro)}")
    print(f"  REVISAR_MANUAL (pierde vs externo)   : {len(revisar_manual)}")

    # split de pausar_seguro por cuenta
    por_cuenta = defaultdict(int)
    for it in pausar_seguro:
        por_cuenta[it["cuenta"]] += 1
    print("\n  PAUSAR_SEGURO por cuenta:")
    for c, n in sorted(por_cuenta.items()):
        print(f"    {c}: {n}")

    # ---------- xlsx ----------
    wb = Workbook()
    wb.remove(wb.active)

    cols_pausar = [
        "cuenta", "item_id", "title", "catalog_product_id",
        "listing_type_id", "price", "ganador_cuenta", "ganador_precio",
        "diff_pct_vs_ganador", "available_quantity", "permalink",
        "ganador_item_id",
    ]

    def dump(nombre: str, data: list[dict], cols: list[str], header_fill: str = "C8E6C9") -> None:
        ws = wb.create_sheet(title=nombre)
        ws.append(cols)
        fill = PatternFill("solid", fgColor=header_fill)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True)
        for it in sorted(data, key=lambda x: -(x.get("price") or 0)):
            ws.append([it.get(c) for c in cols])
        ws.freeze_panes = "A2"
        widths = {"title": 55, "permalink": 50, "catalog_product_id": 16, "item_id": 16, "ganador_item_id": 16}
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    dump("PAUSAR_SEGURO", pausar_seguro, cols_pausar, "C8E6C9")  # verde
    dump("REVISAR_MANUAL", revisar_manual, cols_pausar + ["ganador_seller_id"], "FFE0B2")  # naranja

    # resumen
    res = wb.create_sheet(title="RESUMEN", index=0)
    res.append(["metrica", "valor"])
    for k, v in sorted(estadistica.items(), key=lambda x: -x[1]):
        res.append([k, v])
    res.append([])
    res.append(["pausar_seguro_total", len(pausar_seguro)])
    for c, n in sorted(por_cuenta.items()):
        res.append([f"pausar_seguro_{c}", n])
    res.append(["revisar_manual_total", len(revisar_manual)])
    res.column_dimensions["A"].width = 36
    res.column_dimensions["B"].width = 14

    wb.save(DST)
    print(f"\nXlsx: {DST}")
    return 0


def _diff_pct(precio_mio, precio_ganador) -> float | None:
    try:
        if not precio_ganador:
            return None
        return round((precio_mio - precio_ganador) / precio_ganador * 100, 1)
    except Exception:
        return None


if __name__ == "__main__":
    sys.exit(main())
