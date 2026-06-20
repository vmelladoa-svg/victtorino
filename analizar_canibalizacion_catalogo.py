"""
Cruza los catalog_product_id de las 3 cuentas y detecta canibalizacion.

Lee el xlsx generado por ml_catalogo_publicaciones.py y produce:
  - Resumen consola: overlaps 2-cuentas, 3-cuentas, sin canibalizacion
  - canibalizacion_<fecha>.xlsx con 3 hojas:
      OVERLAPS_3 (mismo catalog en C1+C2+C3)
      OVERLAPS_2 (en exactamente 2 cuentas)
      SOLO_1     (sin canibalizacion)
"""
from __future__ import annotations
import sys
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / f"catalogo_publicaciones_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"canibalizacion_catalogo_{time.strftime('%Y-%m-%d')}.xlsx"

# columnas en el xlsx fuente
CAMPOS = [
    "seller_id", "cuenta", "item_id", "title", "catalog_product_id",
    "listing_type_id", "price", "available_quantity", "status",
    "permalink", "thumbnail", "date_created", "last_updated",
]


def cargar_filas(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    filas = []
    for sheet in wb.sheetnames:
        if sheet == "RESUMEN":
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        for r in rows:
            filas.append(dict(zip(headers, r)))
    return filas


def agrupar_por_catalog(filas: list[dict]) -> dict[str, list[dict]]:
    g: dict[str, list[dict]] = defaultdict(list)
    for f in filas:
        cpid = f.get("catalog_product_id")
        if not cpid:
            continue
        g[cpid].append(f)
    return g


def fmt_money(v) -> str:
    try:
        return f"${int(v):,}".replace(",", ".")
    except Exception:
        return ""


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: no existe {SRC}. Corre primero ml_catalogo_publicaciones.py")
        return 1

    filas = cargar_filas(SRC)
    sin_cpid = sum(1 for f in filas if not f.get("catalog_product_id"))
    grupos = agrupar_por_catalog(filas)

    overlap3, overlap2, solo1 = [], [], []
    for cpid, items in grupos.items():
        cuentas = {it["cuenta"] for it in items}
        if len(cuentas) >= 3:
            overlap3.append((cpid, items))
        elif len(cuentas) == 2:
            overlap2.append((cpid, items))
        else:
            solo1.append((cpid, items))

    # ranking por overlap (cuanto cuesta mas la canibalizacion)
    overlap3.sort(key=lambda x: -max(it.get("price") or 0 for it in x[1]))
    overlap2.sort(key=lambda x: -max(it.get("price") or 0 for it in x[1]))

    # ---------- consola ----------
    total_listings = len(filas)
    total_cpid = len(grupos)
    print("\n========== CRUCE DE CATALOG_PRODUCT_ID ==========")
    print(f"  Total publicaciones en catalogo            : {total_listings}")
    print(f"  Total fichas (catalog_product_id) unicas   : {total_cpid}")
    print(f"  Publicaciones sin catalog_product_id       : {sin_cpid}")
    print()
    print(f"  Fichas en LAS 3 CUENTAS (C1+C2+C3)         : {len(overlap3)}")
    print(f"  Fichas en EXACTAMENTE 2 CUENTAS            : {len(overlap2)}")
    print(f"  Fichas en 1 sola cuenta                    : {len(solo1)}")

    listings_canibal = sum(len(it) for _, it in overlap3) + sum(len(it) for _, it in overlap2)
    pct = (listings_canibal / total_listings * 100) if total_listings else 0
    print(f"\n  Publicaciones involucradas en canibalizacion: {listings_canibal} ({pct:.1f}%)")

    # estimacion grosera de spend "canibalizado"
    # cada ficha duplicada: solo 1 gana buy box; las demas "sobran"
    sobrantes_3 = sum(len(it) - 1 for _, it in overlap3)
    sobrantes_2 = sum(len(it) - 1 for _, it in overlap2)
    valor_sobrantes = 0.0
    for _, items in overlap3 + overlap2:
        precios = sorted([it.get("price") or 0 for it in items], reverse=True)
        valor_sobrantes += sum(precios[1:])  # asume que la 1ra gana buy box
    print(f"\n  Listings 'sobrantes' (compiten entre si)    : {sobrantes_3 + sobrantes_2}")
    print(f"  Valor combinado de listings sobrantes       : {fmt_money(valor_sobrantes)}")
    print(f"  (si paga ~18% comision Premium y vendiera 1/mes -> {fmt_money(valor_sobrantes*0.18)} de comision evitable/mes)")

    # detalle top overlap3
    if overlap3:
        print("\n--- TOP 10 fichas en LAS 3 CUENTAS ---")
        for cpid, items in overlap3[:10]:
            print(f"  [{cpid}] {items[0]['title'][:60]}")
            for it in items:
                tipo = (it.get('listing_type_id') or '')[:11]
                print(f"     {it['cuenta']:>3} | {tipo:>11} | {fmt_money(it.get('price')):>10} | "
                      f"stock {it.get('available_quantity'):>3} | {it.get('item_id')}")

    # ---------- xlsx ----------
    wb = Workbook()
    wb.remove(wb.active)

    def dump(nombre: str, data: list[tuple[str, list[dict]]]) -> None:
        ws = wb.create_sheet(title=nombre)
        ws.append([
            "catalog_product_id", "title", "n_cuentas", "cuentas",
            "min_price", "max_price", "diff_pct",
            "items_detalle",
        ])
        for cpid, items in data:
            precios = [it.get("price") or 0 for it in items]
            mn, mx = min(precios), max(precios)
            diff = ((mx - mn) / mn * 100) if mn else 0
            detalle = " | ".join(
                f"{it['cuenta']}={it.get('item_id')}@{fmt_money(it.get('price'))}"
                f"({it.get('listing_type_id')},stk{it.get('available_quantity')})"
                for it in items
            )
            ws.append([
                cpid,
                (items[0].get("title") or "")[:120],
                len({it["cuenta"] for it in items}),
                ",".join(sorted({it["cuenta"] for it in items})),
                mn, mx, round(diff, 1),
                detalle,
            ])
        ws.freeze_panes = "A2"
        for i, w in enumerate([22, 60, 10, 12, 12, 12, 10, 90], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    dump("OVERLAPS_3", overlap3)
    dump("OVERLAPS_2", overlap2)
    dump("SOLO_1", solo1)

    # hoja resumen
    res = wb.create_sheet(title="RESUMEN", index=0)
    res.append(["metrica", "valor"])
    res.append(["publicaciones_totales_catalogo", total_listings])
    res.append(["fichas_cpid_unicas", total_cpid])
    res.append(["fichas_en_3_cuentas", len(overlap3)])
    res.append(["fichas_en_2_cuentas", len(overlap2)])
    res.append(["fichas_en_1_cuenta", len(solo1)])
    res.append(["listings_sobrantes", sobrantes_3 + sobrantes_2])
    res.append(["valor_listings_sobrantes_clp", valor_sobrantes])
    res.append(["comision_evitable_estimada_18pct", valor_sobrantes * 0.18])
    for col in (1, 2):
        res.column_dimensions[get_column_letter(col)].width = 36

    wb.save(DST)
    print(f"\nXlsx detalle: {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
