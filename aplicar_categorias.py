"""
aplicar_categorias.py
---------------------
Aplica la categoria Paris (col D) a las 38 filas del Excel
paris_productos_carga_masiva.xlsx, segun el mapeo sugerido y confirmado
por Victor el 2026-05-22.

Para evitar errores de tildes, resuelve cada substring contra la lista
maestra de categorias en la hoja 'herramientasdata' (col A) y escribe la
cadena EXACTA esperada por el dropdown.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
EXCEL = BASE / "paris_productos_carga_masiva.xlsx"
SHEET_DATA = "herramientasdata"
SHEET_MAIN = "herramientas"
COL_SKU = 2
COL_CAT = 4

MAPPING: dict[str, str] = {
    # SKU            : substring unico de la categoria Paris
    "MLC21628947":   "Monomando Lavaplatos",
    "MLC21350539":   "Grifería de Baño < Repuestos Wc",
    "MLC21394613":   "Monomando Lavaplatos",
    "MLC38612080":   "Monomando Ducha",
    "MLC25452157":   "Monomando Lavamanos",
    "MLC43500911":   "Grifería de Baño < Repuestos Wc",
    "MLC45488644":   "Monomando Lavaplatos",
    "MLC41735119":   "Monomando Ducha",
    "MLC21299614":   "Combinación Lavaplatos",
    "MLC22271656":   "Combinación Lavaplatos",
    "MLC50745282":   "Combinación Lavaplatos",
    "MLC20918885":   "Sifones Y Desagües",
    "MLC28368137":   "Sifones Y Desagües",
    "MLC24386020":   "Accesorios Y Repuestos Grifería",
    "MLC2041691925": "Grifería de Baño < Repuestos Wc",
    "MLC21394632":   "Monomando Lavamanos",
    "MLC32065874":   "Accesorios Y Repuestos Grifería",
    "MLC27509195":   "Grifería de Baño < Repuestos Wc",
    "MLC26656485":   "Sifones Y Desagües",
    "MLC67879003":   "Sifones Y Desagües",
    "MLC22271655":   "Combinación Lavaplatos",
    "MLC25264406":   "Monomando Lavaplatos",
    "MLC21394631":   "Monomando Lavamanos",
    "MLC2055571318": "Accesorios De Ducha",
    "MLC26778872":   "Grifería de Baño < Repuestos Wc",
    "MLC22736791":   "Combinación Lavaplatos",
    "MLC24044164":   "Accesorios de baño",
    "MLC22271660":   "Combinación Lavaplatos",
    "MLC32227841":   "Accesorios De Ducha",
    "MLC2046010486": "Grifería de Baño < Repuestos Wc",
    "MLC58519072":   "Monomando Lavamanos",
    "MLC22271659":   "Combinación Lavaplatos",
    "MLC2068899311": "Grifería de Baño < Repuestos Wc",
    "MLC66202132":   "Monomando Ducha",
    "MLC27107143":   "Cabinas Y Duchas < Receptáculos De Ducha",
    "MLC39244370":   "Cabinas Y Duchas < Receptáculos De Ducha",
    "MLC39080093":   "Cabinas Y Duchas < Receptáculos De Ducha",
    "MLC51764090":   "Columna de Ducha",
}


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    wb = load_workbook(EXCEL)
    ws_data = wb[SHEET_DATA]
    ws_main = wb[SHEET_MAIN]

    # Construir lista de categorias maestras
    cats = []
    for r in range(1, ws_data.max_row + 1):
        v = ws_data.cell(row=r, column=1).value
        if v:
            cats.append(v.strip())

    def resolve(substring: str) -> str:
        # Buscar match EXACTO primero, luego substring
        for c in cats:
            if c == substring:
                return c
        matches = [c for c in cats if substring.lower() in c.lower()]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            # Preferir la version sin doble nivel si hay ambiguedad
            raise RuntimeError(f"Ambiguo '{substring}': {matches}")
        raise RuntimeError(f"No encontrado '{substring}'")

    # Pre-resolver
    resolved = {sku: resolve(sub) for sku, sub in MAPPING.items()}
    print("Categorias resueltas:")
    distinct = sorted(set(resolved.values()))
    for c in distinct:
        n = sum(1 for v in resolved.values() if v == c)
        print(f"  ({n:>2}) {c}")

    # Aplicar
    applied = 0
    not_found = []
    for row in range(7, ws_main.max_row + 1):
        sku = ws_main.cell(row=row, column=COL_SKU).value
        if not sku:
            continue
        if sku not in resolved:
            not_found.append(sku)
            continue
        ws_main.cell(row=row, column=COL_CAT, value=resolved[sku])
        applied += 1

    wb.save(EXCEL)
    print()
    print(f"Aplicadas {applied} categorias. Sin mapeo: {len(not_found)} {not_found}")
    print(f"Guardado: {EXCEL}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
