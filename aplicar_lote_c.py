"""Aplica focus + meta + descripción HTML premium a los 53 productos del Lote C.
Usa la plantilla por categoría destino para generar la descripción."""
import sys, io, time, warnings, requests, re
from openpyxl import load_workbook
warnings.filterwarnings("ignore", category=UserWarning)
from seo_premium_lote import plantilla_premium
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}


def safe_put(pid, body):
    for n in range(1, 5):
        try:
            r = requests.put(f"{WC}/wp-json/wc/v3/products/{pid}", json=body, params=P, timeout=120)
            if r.status_code == 503:
                espera = 10 * n
                print(f"    503 (intento {n}), espero {espera}s")
                time.sleep(espera)
                continue
            if r.status_code >= 400:
                print(f"    HTTP {r.status_code}: {r.text[:150]}")
                return None
            try:
                return r.json()
            except Exception:
                print(f"    respuesta no-JSON; retry")
                time.sleep(5 * n)
                continue
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            espera = 5 * n
            print(f"    {type(e).__name__}; retry en {espera}s")
            time.sleep(espera)
    return None


# Leer Excel
wb = load_workbook(r"C:\Users\dell\victtorino\propuesta_lote_c.xlsx", data_only=True)
ws = wb["Lote C — 53 productos"]

items = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dec = (row[0] or "").strip().upper() if row[0] else ""
    if dec == "SALTAR":
        continue
    pid = row[1]
    name = row[3]
    focus = row[4]
    mt = row[5]
    md = row[6]
    cat_destino = row[7]
    if pid and focus and mt and md and cat_destino:
        items.append((pid, name, focus, mt, md, cat_destino))

print(f"Total a aplicar: {len(items)}\n")

ok, fallidos = 0, []
for idx, (pid, name, focus, mt, md, cat_destino) in enumerate(items, start=1):
    # Generar descripción HTML premium con plantilla
    try:
        desc_html = plantilla_premium(cat_destino, name, focus)
    except KeyError:
        print(f"({idx}/{len(items)}) {pid} cat destino '{cat_destino}' sin plantilla, salto")
        fallidos.append(pid)
        continue
    short = f"<p>{name}. <strong>{focus.capitalize()}</strong> con diseño moderno y materiales resistentes. Despacho a todo Chile. Calidad Victtorino.</p>"

    body = {
        "description": desc_html,
        "short_description": short,
        "meta_data": [
            {"key": "rank_math_title", "value": mt},
            {"key": "rank_math_description", "value": md},
            {"key": "rank_math_focus_keyword", "value": focus},
        ],
    }
    res = safe_put(pid, body)
    if res:
        plain = re.sub(r"<[^>]+>", " ", res.get("description", ""))
        pal = len(plain.split())
        print(f"({idx:2}/{len(items)}) {pid}  {focus[:30]:30}  pal={pal}")
        ok += 1
    else:
        print(f"({idx:2}/{len(items)}) {pid}  FALLO")
        fallidos.append(pid)
    time.sleep(1.5)

print(f"\n{ok}/{len(items)} aplicados, {len(fallidos)} fallidos")
if fallidos:
    print(f"Fallidos: {fallidos}")
