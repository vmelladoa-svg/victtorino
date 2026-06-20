"""Aplica focus + meta title + meta description a los 77 (Lote A) + 28 (Lote B) productos."""
import sys, io, time, warnings, requests
from openpyxl import load_workbook
warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}


def leer_propuesta(path, hoja):
    wb = load_workbook(path, data_only=True)
    ws = wb[hoja]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dec = (row[0] or "").strip().upper() if row[0] else ""
        if dec == "SALTAR":
            continue
        woo_id = row[1]
        focus = row[4]
        mt = row[5]
        md = row[6]
        if woo_id and focus and mt and md:
            items.append((woo_id, focus, mt, md))
    return items


def safe_put(pid, body):
    for n in range(1, 5):
        try:
            r = requests.put(f"{WC}/wp-json/wc/v3/products/{pid}", json=body, params=P, timeout=120)
            if r.status_code == 503:
                espera = 10 * n
                print(f"    503 (intento {n}), espero {espera}s")
                time.sleep(espera)
                continue
            if r.status_code >= 400:
                print(f"    HTTP {r.status_code}: {r.text[:150]}")
                return None
            try:
                return r.json()
            except Exception:
                print(f"    respuesta no-JSON; retry")
                time.sleep(5 * n)
                continue
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            espera = 5 * n
            print(f"    {type(e).__name__}; retry en {espera}s")
            time.sleep(espera)
    return None


todos = []
todos += leer_propuesta(r"C:\Users\dell\victtorino\propuesta_lote_a.xlsx", "Lote A — 77 productos")
todos += leer_propuesta(r"C:\Users\dell\victtorino\propuesta_lote_b.xlsx", "Lote B — 28 productos")
print(f"Total a aplicar: {len(todos)}\n")

ok, fallidos = 0, []
for idx, (pid, focus, mt, md) in enumerate(todos, start=1):
    body = {
        "meta_data": [
            {"key": "rank_math_title", "value": mt},
            {"key": "rank_math_description", "value": md},
            {"key": "rank_math_focus_keyword", "value": focus},
        ],
    }
    res = safe_put(pid, body)
    if res:
        fk = next((m["value"] for m in res.get("meta_data", [])
                   if m["key"] == "rank_math_focus_keyword"), "")
        print(f"({idx:3}/{len(todos)}) {pid}  → \"{fk[:35]}\"")
        ok += 1
    else:
        print(f"({idx:3}/{len(todos)}) {pid}  FALLO")
        fallidos.append(pid)
    time.sleep(1.2)  # pausa para no saturar Apache

print(f"\n{ok}/{len(todos)} aplicados, {len(fallidos)} fallidos")
if fallidos:
    print(f"Fallidos: {fallidos}")
