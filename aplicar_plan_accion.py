"""
Ejecuta el plan_accion_<fecha>.xlsx contra la API ML.

Filtra filas con aprobado='SI' y ejecuta segun accion:
  - PAUSAR        -> PUT /items/{id} {"status": "paused"}
  - BAJAR_PRECIO  -> PUT /items/{id} {"price": <precio_sugerido>}

Modo DRY-RUN por defecto (no toca nada). Con --apply ejecuta de verdad.
Antes de aplicar, guarda backup del estado actual en backups/.
Log de cada accion en logs/ejecucion_<timestamp>.csv

Usage:
    python aplicar_plan_accion.py                # dry-run
    python aplicar_plan_accion.py --apply        # ejecuta
    python aplicar_plan_accion.py --apply --solo PAUSAR  # solo una accion
"""
from __future__ import annotations
import argparse
import csv
import json
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

from ml_catalogo_publicaciones import CUENTAS, Cuenta, ML_BASE

ROOT = Path(__file__).parent
SRC = ROOT / f"plan_accion_{time.strftime('%Y-%m-%d')}.xlsx"
LOG_DIR = ROOT / "logs"
BACKUP_DIR = ROOT / "backups"
LOG_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)


def cargar_acciones(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb["ACCIONES"]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def get_item_current(cuenta: Cuenta, item_id: str) -> dict | None:
    """Lee estado actual del item para backup."""
    r = requests.get(
        f"{ML_BASE}/items/{item_id}",
        headers={"Authorization": f"Bearer {cuenta.access_token}"},
        params={"attributes": "id,title,price,status,available_quantity,listing_type_id"},
        timeout=20,
    )
    if r.status_code == 401:
        cuenta.refrescar()
        return get_item_current(cuenta, item_id)
    return r.json() if r.status_code == 200 else None


def put_item(cuenta: Cuenta, item_id: str, payload: dict,
             intentos: int = 3) -> tuple[int, dict | str]:
    """PUT /items/{id} con backoff. Retorna (status_code, body_or_text)."""
    url = f"{ML_BASE}/items/{item_id}"
    for n in range(1, intentos + 1):
        try:
            r = requests.put(
                url,
                headers={
                    "Authorization": f"Bearer {cuenta.access_token}",
                    "Content-Type": "application/json",
                },
                json=payload,
                timeout=30,
            )
        except requests.RequestException as e:
            time.sleep(2 ** n)
            continue
        if r.status_code == 401 and n == 1:
            cuenta.refrescar()
            continue
        if r.status_code == 429:
            wait = float(r.headers.get("Retry-After", 0)) or 2 ** n
            time.sleep(wait)
            continue
        try:
            return r.status_code, r.json()
        except Exception:
            return r.status_code, r.text
    return 0, "max_retries"


def main(argv=None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true",
                   help="Ejecuta cambios reales en ML. Sin esto, solo DRY-RUN.")
    p.add_argument("--solo", choices=["PAUSAR", "BAJAR_PRECIO"],
                   help="Solo ejecuta una accion")
    p.add_argument("--archivo", default=str(SRC),
                   help=f"Archivo plan (default: {SRC.name})")
    args = p.parse_args(argv)

    src = Path(args.archivo)
    if not src.exists():
        print(f"ERROR: no existe {src}")
        return 1

    acciones_todas = cargar_acciones(src)
    aprobadas = [a for a in acciones_todas if a.get("aprobado") == "SI"]
    if args.solo:
        aprobadas = [a for a in aprobadas if a.get("accion") == args.solo]

    modo = "APPLY (REAL)" if args.apply else "DRY-RUN (no toca nada)"
    print(f"\n{'='*60}")
    print(f"MODO: {modo}")
    print(f"Archivo: {src.name}")
    print(f"Acciones aprobadas a ejecutar: {len(aprobadas)}")
    from collections import Counter
    breakdown = Counter(a.get("accion") for a in aprobadas)
    for k, n in breakdown.items():
        print(f"  {k}: {n}")
    print(f"{'='*60}\n")

    if not aprobadas:
        print("Nada que ejecutar.")
        return 0

    # cuentas + token
    cuentas_map = {}
    for cfg in CUENTAS:
        try:
            cuentas_map[cfg["alias"]] = Cuenta.cargar(cfg)
        except FileNotFoundError:
            pass

    # backup pre-cambios
    ts = time.strftime("%Y%m%d_%H%M%S")
    backup_file = BACKUP_DIR / f"backup_{ts}.jsonl"
    log_file = LOG_DIR / f"ejecucion_{ts}.csv"

    if args.apply:
        print(f"Creando backup en {backup_file}...")

    log_rows = []
    with backup_file.open("w", encoding="utf-8") as bf:
        for i, a in enumerate(aprobadas, 1):
            iid = a.get("item_id")
            alias = a.get("cuenta")
            accion = a.get("accion")
            cu = cuentas_map.get(alias)
            if not cu:
                print(f"[{i}/{len(aprobadas)}] WARN: no hay token para {alias}, salto {iid}")
                continue

            # backup estado actual
            current = get_item_current(cu, iid) or {}
            bf.write(json.dumps({
                "ts": ts, "alias": alias, "item_id": iid,
                "accion_planeada": accion, "estado_actual": current,
            }, ensure_ascii=False) + "\n")

            payload = None
            if accion == "PAUSAR":
                payload = {"status": "paused"}
                desc = f"PAUSAR (estaba {current.get('status')})"
            elif accion == "BAJAR_PRECIO":
                nuevo = a.get("precio_sugerido")
                if not nuevo:
                    print(f"[{i}/{len(aprobadas)}] WARN: sin precio_sugerido, salto {iid}")
                    continue
                payload = {"price": int(nuevo)}
                desc = f"BAJAR ${current.get('price')} -> ${nuevo}"

            log_row = {
                "ts": ts, "n": i, "alias": alias, "item_id": iid,
                "accion": accion, "descripcion": desc,
                "precio_actual": current.get("price"),
                "status_actual": current.get("status"),
                "title": (current.get("title") or "")[:60],
            }

            if not args.apply:
                print(f"[{i}/{len(aprobadas)}] DRY-RUN {alias} {iid}: {desc}")
                log_row["resultado"] = "DRY_RUN"
                log_row["http_status"] = ""
            else:
                code, body = put_item(cu, iid, payload)
                ok = code in (200, 201)
                tag = "OK" if ok else f"FAIL_{code}"
                print(f"[{i}/{len(aprobadas)}] {tag} {alias} {iid}: {desc}")
                if not ok:
                    msg = body.get("message") if isinstance(body, dict) else str(body)[:120]
                    print(f"    -> {msg}")
                log_row["resultado"] = tag
                log_row["http_status"] = code
                log_row["respuesta"] = json.dumps(body, ensure_ascii=False)[:200] if isinstance(body, dict) else str(body)[:200]
                time.sleep(0.4)  # rate limit defensivo

            log_rows.append(log_row)

    # csv log
    if log_rows:
        all_cols = sorted({k for r in log_rows for k in r.keys()})
        with log_file.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=all_cols)
            w.writeheader()
            for r in log_rows:
                w.writerow(r)

    # resumen
    print(f"\n{'='*60}")
    print(f"Backup: {backup_file}")
    print(f"Log   : {log_file}")
    if args.apply:
        ok = sum(1 for r in log_rows if r.get("resultado") == "OK")
        fail = sum(1 for r in log_rows if str(r.get("resultado")).startswith("FAIL"))
        print(f"\nResultados:  OK={ok}  FAIL={fail}")
        if fail:
            print(f"Revisa el log para detalle de errores.")
    else:
        print("\nDRY-RUN: no se modifico nada en ML.")
        print("Para ejecutar de verdad: python aplicar_plan_accion.py --apply")
    print(f"{'='*60}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
