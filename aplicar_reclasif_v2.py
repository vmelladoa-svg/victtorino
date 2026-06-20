"""
v2: Lee la columna D del Excel (donde el usuario reescribio las categorias).
Detecta cambios comparando con la categoria actual REAL en Woo.
"""
import json
import sys
import io
import re
import time
import warnings
import requests
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

CATEGORIAS_VALIDAS = ["Accesorios", "Agarraderas y Barras", "Dispensador", "Espejos",
                     "Griferia", "Lavamanos", "Lavaplatos", "Shower/Mamparas/Receptaculos",
                     "Sifones y Desagües", "WC e Inodoros"]
CATEGORIAS_LOWER = {c.lower(): c for c in CATEGORIAS_VALIDAS}


def parsear(s):
    """Devuelve (lista_validas, lista_invalidas)."""
    if not s:
        return [], []
    s = str(s).strip()
    if s.lower() in CATEGORIAS_LOWER:
        return [CATEGORIAS_LOWER[s.lower()]], []
    encontradas, invalidas = [], []
    resto = s
    if "shower" in resto.lower() and "mamparas" in resto.lower():
        idx = resto.lower().find("shower")
        fin = idx + len("Shower/Mamparas/Receptaculos")
        encontradas.append("Shower/Mamparas/Receptaculos")
        resto = (resto[:idx] + resto[fin:]).strip(" -–—,;/")
    if resto:
        partes = re.split(r"\s*[-–—,;]\s*", resto)
        for p in partes:
            p = p.strip()
            if not p:
                continue
            if p.lower() in CATEGORIAS_LOWER:
                encontradas.append(CATEGORIAS_LOWER[p.lower()])
            else:
                invalidas.append(p)
    return encontradas, invalidas


# 1) Leer Excel
print("[1/4] Leyendo Excel...")
wb = load_workbook(r"C:\Users\dell\victtorino\reclasificar_borradores.xlsx", data_only=True)
ws = wb["Reclasificar borradores"]
filas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] is None:
        continue
    filas.append({
        "woo_id": row[1],
        "titulo": row[2] or "",
        "cat_excel_d": (row[3] or "").strip() if row[3] else "",
        "cat_excel_a": (row[0] or "").strip() if row[0] else "",
    })
print(f"      {len(filas)} filas en Excel")

# 2) Comparar con Woo en vivo
print("[2/4] Consultando categoria actual real en Woo...")
cambios = []
for f in filas:
    r = requests.get(f"{WC}/wp-json/wc/v3/products/{f['woo_id']}",
                     params=P, timeout=30)
    if r.status_code != 200:
        print(f"      Woo {f['woo_id']} no responde")
        continue
    d = r.json()
    cat_real_lista = [c["name"] for c in d.get("categories", [])]
    cat_real_str = ", ".join(cat_real_lista)

    # Tomar la mejor fuente: columna A si tiene valor, sino columna D si difiere de cat_real_str
    valor = f["cat_excel_a"] or (f["cat_excel_d"] if f["cat_excel_d"] != cat_real_str else "")
    if not valor:
        continue

    nuevas, invalidas = parsear(valor)
    if not nuevas:
        cambios.append({**f, "valor": valor, "nuevas": [], "invalidas": invalidas, "cat_real_lista": cat_real_lista, "valido": False})
        continue

    # Si las nuevas son IGUALES a la actual real, no hay nada que cambiar
    if sorted(nuevas) == sorted(cat_real_lista):
        continue

    cambios.append({**f, "valor": valor, "nuevas": nuevas, "invalidas": invalidas,
                    "cat_real_lista": cat_real_lista, "valido": True})

print(f"\n      Cambios detectados: {len(cambios)}")
validos = [c for c in cambios if c["valido"]]
invalidos = [c for c in cambios if not c["valido"]]
print(f"      Validos:   {len(validos)}")
print(f"      Invalidos: {len(invalidos)}")

if invalidos:
    print("\n      INVALIDOS (no se aplicaran):")
    for c in invalidos:
        print(f"        Woo {c['woo_id']}: \"{c['valor']}\" -> validas={c['nuevas']} invalidas={c['invalidas']}")

# 3) Cargar ids cats
print("\n[3/4] Cargando IDs categorias Woo...")
r = requests.get(f"{WC}/wp-json/wc/v3/products/categories", params={**P, "per_page": 100}, timeout=30)
cat_id_por_nombre = {c["name"]: c["id"] for c in r.json()}

# 4) Aplicar
print(f"\n[4/4] Aplicando {len(validos)} cambios validos (pausa 2s)...\n")
ok, fallidos = [], []
for idx, c in enumerate(validos, start=1):
    ids = [cat_id_por_nombre[n] for n in c["nuevas"] if n in cat_id_por_nombre]
    body = {"categories": [{"id": i} for i in ids]}
    for n in range(1, 4):
        try:
            r = requests.put(f"{WC}/wp-json/wc/v3/products/{c['woo_id']}",
                             json=body, params=P, timeout=60)
            if r.status_code >= 400:
                print(f"({idx:2}/{len(validos)}) Woo {c['woo_id']} ERR {r.status_code}: {r.text[:200]}")
                fallidos.append(c)
                break
            d = r.json()
            cats_final = [x["name"] for x in d.get("categories", [])]
            print(f"({idx:2}/{len(validos)}) Woo {c['woo_id']} {c['titulo'][:40]:40} -> {cats_final}")
            ok.append({**c, "cats_final": cats_final})
            break
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            print(f"  red intento {n}: {type(e).__name__}")
            time.sleep(2 ** n)
    else:
        fallidos.append(c)
    time.sleep(2)

print("\n" + "=" * 70)
print(f"RESUMEN: {len(ok)} aplicados, {len(fallidos)} fallidos, {len(invalidos)} con problema de formato")
print("=" * 70)

with open(r"C:\Users\dell\victtorino\reclasif_v2_resultado.json", "w", encoding="utf-8") as f:
    json.dump({"ok": ok, "fallidos": fallidos, "invalidos": invalidos}, f, ensure_ascii=False, indent=2)
print("Detalle -> reclasif_v2_resultado.json")
