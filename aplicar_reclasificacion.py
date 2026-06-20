"""
Aplica la reclasificacion del Excel `reclasificar_borradores.xlsx` a los productos Woo.

- Lee columna A "Categoria nueva" de cada fila.
- Si esta vacia: no toca el producto.
- Si tiene una categoria: PUT con esa unica.
- Si tiene varias separadas por '-' o '/' o ',': PUT con la lista.
- Valida contra las 10 categorias validas. Si hay un valor invalido, lo reporta y salta esa fila.
- Pausa de 2s entre PUTs para no saturar Apache.
"""
import json
import sys
import io
import re
import time
import warnings
import requests
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

CATEGORIAS_VALIDAS = [
    "Accesorios",
    "Agarraderas y Barras",
    "Dispensador",
    "Espejos",
    "Griferia",
    "Lavamanos",
    "Lavaplatos",
    "Shower/Mamparas/Receptaculos",
    "Sifones y Desagües",
    "WC e Inodoros",
]
CATEGORIAS_LOWER = {c.lower(): c for c in CATEGORIAS_VALIDAS}


def parsear_categorias(valor):
    """
    Recibe un string como 'Griferia - Lavaplatos' o 'Accesorios/Dispensador'.
    Devuelve lista de categorias normalizadas (nombre exacto del catalogo).
    Retorna ([validas], [invalidas]).
    """
    if not valor:
        return [], []
    # Separadores: -  –  —  ,  ;  /
    # OJO: "Shower/Mamparas/Receptaculos" tiene "/", lo trato especial
    # Estrategia: probar primero match completo, si no, separar
    s = str(valor).strip()

    # Si el valor entero matchea una categoria valida (incluyendo Shower/Mamparas/Receptaculos)
    if s.lower() in CATEGORIAS_LOWER:
        return [CATEGORIAS_LOWER[s.lower()]], []

    # Si contiene "Shower/Mamparas/Receptaculos", extraerla primero
    encontradas = []
    invalidas = []
    resto = s
    if "shower" in resto.lower() and "mamparas" in resto.lower():
        # Sacarla del string
        idx = resto.lower().find("shower")
        # buscar el fin de "Shower/Mamparas/Receptaculos"
        fin = idx + len("Shower/Mamparas/Receptaculos")
        encontradas.append("Shower/Mamparas/Receptaculos")
        resto = (resto[:idx] + resto[fin:]).strip(" -–—,;/")

    # Ahora separar por -, –, —, ',', ';', '/'
    if resto:
        partes = re.split(r"\s*[-–—,;/]\s*", resto)
        for p in partes:
            p = p.strip()
            if not p:
                continue
            if p.lower() in CATEGORIAS_LOWER:
                encontradas.append(CATEGORIAS_LOWER[p.lower()])
            else:
                invalidas.append(p)
    return encontradas, invalidas


# 1) Leer Excel
print("[1/4] Leyendo Excel...")
wb = load_workbook(r"C:\Users\dell\victtorino\reclasificar_borradores.xlsx", data_only=True)
ws = wb["Reclasificar borradores"]

cambios = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cat_nueva = (row[0] or "").strip() if row[0] else ""
    if not cat_nueva:
        continue
    cambios.append({
        "woo_id": row[1],
        "titulo": row[2],
        "cat_actual": row[3],
        "valor_raw": cat_nueva,
    })
print(f"      {len(cambios)} productos con cambios marcados")

# 2) Parsear y validar
print("[2/4] Parseando y validando...")
con_problemas = []
for c in cambios:
    validas, invalidas = parsear_categorias(c["valor_raw"])
    c["categorias_nuevas"] = validas
    c["invalidas"] = invalidas
    if invalidas or not validas:
        con_problemas.append(c)

if con_problemas:
    print(f"\n      PROBLEMAS detectados en {len(con_problemas)} filas:")
    for p in con_problemas:
        print(f"      Woo {p['woo_id']}: \"{p['valor_raw']}\" -> validas={p['categorias_nuevas']}, invalidas={p['invalidas']}")
    print("\n      Estos productos NO se van a modificar. Revisa y vuelve a guardar el Excel si hay errores.")
    cambios = [c for c in cambios if c not in con_problemas]

# 3) Cargar IDs de categorias Woo
print("[3/4] Cargando IDs de categorias Woo...")
r = requests.get(f"{WC}/wp-json/wc/v3/products/categories", params={**P, "per_page": 100}, timeout=30)
cat_id_por_nombre = {c["name"]: c["id"] for c in r.json()}

# 4) Aplicar PUTs
print(f"\n[4/4] Aplicando {len(cambios)} cambios con pausa de 2s...\n")
ok = []
fallidos = []
for idx, c in enumerate(cambios, start=1):
    ids = [cat_id_por_nombre[n] for n in c["categorias_nuevas"] if n in cat_id_por_nombre]
    body = {"categories": [{"id": i} for i in ids]}
    for n in range(1, 4):
        try:
            r = requests.put(f"{WC}/wp-json/wc/v3/products/{c['woo_id']}",
                             json=body, params=P, timeout=60)
            if r.status_code >= 400:
                print(f"({idx:2}/{len(cambios)}) Woo {c['woo_id']} ERR {r.status_code}: {r.text[:200]}")
                fallidos.append(c)
                break
            d = r.json()
            cats_final = [x["name"] for x in d.get("categories", [])]
            print(f"({idx:2}/{len(cambios)}) Woo {c['woo_id']} \"{c['titulo'][:40]}\" -> {cats_final}")
            ok.append({**c, "cats_final": cats_final})
            break
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            print(f"  red intento {n}: {type(e).__name__}; reintento")
            time.sleep(2 ** n)
    else:
        fallidos.append(c)
    time.sleep(2)

# Resumen
print("\n" + "=" * 70)
print(f"RESUMEN: {len(ok)} OK, {len(fallidos)} fallidos, {len(con_problemas)} con problemas de formato")
print("=" * 70)

if fallidos:
    print("\nFallidos:")
    for f in fallidos:
        print(f"  Woo {f['woo_id']}: {f['titulo'][:50]}")

with open(r"C:\Users\dell\victtorino\reclasificacion_resultado.json", "w", encoding="utf-8") as f:
    json.dump({"ok": ok, "fallidos": fallidos, "con_problemas": con_problemas},
              f, ensure_ascii=False, indent=2)
print("\nDetalle -> reclasificacion_resultado.json")
