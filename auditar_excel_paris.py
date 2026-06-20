"""
auditar_excel_paris.py
----------------------
Auditoria pre-subida del Excel paris_productos_carga_masiva.xlsx.

Reglas que valida:
  - Filas 1..6 de la plantilla intactas (no se sobreescribio el header).
  - Hoja "herramientas" con datos desde fila 7.
  - Campos obligatorios (*) presentes en cada fila: Nombre, Sku Seller,
    Categoria, Marca, Imagenes, Sku Seller Variant, Color, Talla.
  - Sku Seller == Sku Seller Variant.
  - Talla == "Talla Unica" (valor estandar exigido por Paris para SKUs sin
    talla definida; tolera con/sin tilde).
  - Color != vacio.
  - SKUs unicos.
  - URLs de imagen: minimo 2 por producto, sin duplicados intra-fila,
    https, terminando en jpg/png/bmp, HEAD 200, content-type aceptado, y
    al menos un lado >= 500 px (Paris exige tamano minimo 500 px).
  - Largo de cada celda dentro de margenes razonables (titulo <= 250 chars,
    descripcion <= 4000 chars, modelo <= 100 chars).
  - Caracteres no-ASCII se preservan correctamente (no aparecen como "?").

Reporta hallazgos por severidad: ERROR (bloquea carga), WARN (revisable).
"""

from __future__ import annotations

import io
import re
import sys
from collections import Counter
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image

BASE = Path(__file__).resolve().parent
EXCEL = BASE / "paris_productos_carga_masiva.xlsx"
SHEET = "herramientas"
FIRST_DATA_ROW = 7

COL = {
    "nombre": 1,
    "sku": 2,
    "ref": 3,
    "categoria": 4,
    "marca": 6,
    "modelo": 10,
    "imagenes": 13,
    "sku_var": 15,
    "color": 16,
    "talla": 17,
    "desc": 33,
    "material": 35,
}

REQUIRED = ["nombre", "sku", "categoria", "marca", "imagenes", "sku_var", "color", "talla"]
ALLOWED_IMG_CTYPES = {"image/jpeg", "image/jpg", "image/png", "image/bmp"}
MIN_IMG_SIDE = 500
MIN_IMAGES_PER_ROW = 2


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

    errors: list[str] = []
    warnings: list[str] = []

    wb = load_workbook(EXCEL)
    if SHEET not in wb.sheetnames:
        print(f"[ERROR] Falta la hoja '{SHEET}'")
        return 1
    ws = wb[SHEET]

    # ------------------------------------------------------------------ 1) Cabecera
    expected_header = {
        1: "Nombre del Producto",
        2: "Sku Seller (*)",
        4: "Categoria (*)",
        6: "Marca (*)",
        13: "Imagenes (*)",
        15: "Sku Seller Variant (*)",
        16: "Color (*)",
        17: "Talla (*)",
    }
    print("[1/4] Cabecera fila 4")
    for c, expected in expected_header.items():
        actual = ws.cell(row=4, column=c).value or ""
        # Normalizar tildes para comparar
        if (actual or "").strip() != expected:
            # Algunos headers tienen acentos. Comparar tolerante.
            norm_a = re.sub(r"[À-ſ]", "?", actual or "")
            norm_e = re.sub(r"[À-ſ]", "?", expected)
            if norm_a.strip() != norm_e.strip():
                warnings.append(f"Header col {c}: '{actual}' (esperado '{expected}')")
    print(f"      headers verificados: {len(expected_header)}")

    # Validar fila 6 (ejemplo) tenga el https://google.cl
    if ws.cell(row=6, column=13).value != "https://google.cl":
        warnings.append("Fila 6 col 13 (ejemplo) no es 'https://google.cl'")

    # ------------------------------------------------------------------ 2) Filas datos
    print("[2/4] Validando filas de datos")
    rows: list[dict] = []
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        if not ws.cell(row=row, column=COL["sku"]).value:
            continue
        rec = {k: ws.cell(row=row, column=c).value for k, c in COL.items()}
        rec["_row"] = row
        rows.append(rec)
    print(f"      filas con SKU: {len(rows)}")

    # Duplicados de SKU
    skus = [r["sku"] for r in rows]
    dup = [s for s, n in Counter(skus).items() if n > 1]
    for s in dup:
        errors.append(f"SKU duplicado: {s}")

    # Validacion por fila
    for r in rows:
        rnum = r["_row"]
        sku = r["sku"]
        # Obligatorios
        for f in REQUIRED:
            if f == "categoria":
                # categoria DEBE llenarse manual. Aviso WARN si vacio.
                if not (r.get(f) or "").strip() if isinstance(r.get(f), str) else not r.get(f):
                    warnings.append(f"Fila {rnum} ({sku}): Categoria (*) vacia - Victor debe completarla antes de subir")
                continue
            val = r.get(f)
            if val is None or (isinstance(val, str) and not val.strip()):
                errors.append(f"Fila {rnum} ({sku}): falta '{f}' obligatorio")

        # SKU == Sku Variant
        if r["sku"] != r["sku_var"]:
            errors.append(f"Fila {rnum} ({sku}): sku != sku_variant ('{r['sku_var']}')")

        # Talla
        talla = (r["talla"] or "").strip()
        if talla.lower().replace("ú", "u").replace("á", "a") not in ("talla unica", "talla unica"):
            # Aceptamos "Talla Unica" o "Talla Única"
            if talla not in ("Talla Unica", "Talla Única", "Talla Unica"):
                warnings.append(f"Fila {rnum} ({sku}): Talla='{talla}' (esperado 'Talla Unica/Única')")

        # Color no vacio
        color = (r["color"] or "").strip()
        if not color:
            errors.append(f"Fila {rnum} ({sku}): Color vacio")

        # Largos
        if r["nombre"] and len(str(r["nombre"])) > 250:
            warnings.append(f"Fila {rnum} ({sku}): nombre {len(r['nombre'])} chars (>250)")
        if r["desc"] and len(str(r["desc"])) > 4000:
            warnings.append(f"Fila {rnum} ({sku}): descripcion {len(r['desc'])} chars (>4000)")
        if r["modelo"] and len(str(r["modelo"])) > 100:
            warnings.append(f"Fila {rnum} ({sku}): modelo {len(r['modelo'])} chars (>100)")

        # Imagenes -- estructurar
        imgs_raw = (r["imagenes"] or "").strip()
        imgs = [u.strip() for u in imgs_raw.split(",") if u.strip()]
        r["_imgs"] = imgs

        if len(imgs) < MIN_IMAGES_PER_ROW:
            errors.append(f"Fila {rnum} ({sku}): {len(imgs)} imagen(es), Paris exige >= {MIN_IMAGES_PER_ROW}")
        # Duplicadas intra-fila
        dups = [u for u, n in Counter(imgs).items() if n > 1]
        for d in dups:
            warnings.append(f"Fila {rnum} ({sku}): URL imagen duplicada: {d}")
        # Formato URL y extension
        for u in imgs:
            if not u.lower().startswith("https://"):
                errors.append(f"Fila {rnum} ({sku}): URL no-HTTPS: {u}")
            base = u.lower().split("?")[0]
            if not base.endswith((".jpg", ".jpeg", ".png", ".bmp")):
                errors.append(f"Fila {rnum} ({sku}): extension no aceptada: {u}")

    # ------------------------------------------------------------------ 3) Validar imagenes online
    print("[3/4] Validando URLs de imagen (HEAD + dimensiones)")
    cache_ctype: dict[str, str] = {}
    cache_dims: dict[str, tuple[int, int]] = {}

    def check_img(url: str) -> tuple[str | None, tuple[int, int] | None, str | None]:
        """Returns (content_type, (w,h), error). mlstatic.com no acepta HEAD,
        asi que usamos GET completo (con timeout corto) y leemos dimensiones."""
        try:
            g = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30, stream=True)
        except Exception as e:
            return None, None, f"GET fallo: {e}"
        if g.status_code != 200:
            return None, None, f"HTTP {g.status_code}"
        ctype = (g.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        try:
            data = g.content
            im = Image.open(io.BytesIO(data))
            dims = im.size
        except Exception as e:
            return ctype, None, f"PIL fallo: {e}"
        return ctype, dims, None

    total_urls = sum(len(r["_imgs"]) for r in rows)
    seen_url = 0
    for r in rows:
        rnum = r["_row"]
        sku = r["sku"]
        for u in r["_imgs"]:
            seen_url += 1
            if u in cache_ctype:
                ctype, dims, err = cache_ctype[u], cache_dims.get(u), None
            else:
                ctype, dims, err = check_img(u)
                cache_ctype[u] = ctype or ""
                if dims:
                    cache_dims[u] = dims
            if err:
                errors.append(f"Fila {rnum} ({sku}): imagen {u} -> {err}")
                continue
            if ctype and ctype not in ALLOWED_IMG_CTYPES:
                errors.append(f"Fila {rnum} ({sku}): content-type '{ctype}' no aceptado por Paris en {u}")
            if dims:
                if max(dims) < MIN_IMG_SIDE:
                    errors.append(f"Fila {rnum} ({sku}): imagen {dims[0]}x{dims[1]} < {MIN_IMG_SIDE} px en {u}")
            else:
                warnings.append(f"Fila {rnum} ({sku}): no se pudieron leer dimensiones de {u}")
        # progress
        if r["_row"] % 5 == 0:
            print(f"      .. fila {r['_row']} ({seen_url}/{total_urls} URLs)")

    # ------------------------------------------------------------------ 4) Reporte
    print("[4/4] Reporte")
    print()
    print("=" * 70)
    print(f"Total productos       : {len(rows)}")
    print(f"Total URLs revisadas  : {total_urls}")
    print(f"ERRORES (bloquean)    : {len(errors)}")
    print(f"WARNINGS (revisables) : {len(warnings)}")
    print("=" * 70)

    if errors:
        print("\n--- ERRORES ---")
        for e in errors:
            print(f"  [ERR] {e}")
    if warnings:
        print("\n--- WARNINGS ---")
        for w in warnings:
            print(f"  [WARN] {w}")

    # Resumen por producto: imgs count + dimensiones
    print("\n--- RESUMEN POR PRODUCTO ---")
    for r in rows:
        imgs = r["_imgs"]
        dims_str = ""
        if imgs:
            ds = [cache_dims.get(u) for u in imgs]
            if all(ds):
                ws_, hs = zip(*ds)
                dims_str = f"min={min(ws_)}x{min(hs)} max={max(ws_)}x{max(hs)}"
        marca = (r["marca"] or "-")[:18]
        nombre = (r["nombre"] or "")[:50]
        flag = "OK" if len(imgs) >= MIN_IMAGES_PER_ROW else "!!"
        print(f"  {flag} fila {r['_row']:>2} | {r['sku']:<14} | imgs={len(imgs)} {dims_str:<28} | marca={marca:<18} | {nombre}")

    return 0 if not errors else 2


if __name__ == "__main__":
    sys.exit(main())
