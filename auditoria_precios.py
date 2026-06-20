# -*- coding: utf-8 -*-
"""Auditoria de la lista de precios Falabella vs costos Defontana.
Detecta: venta a perdida, margenes absurdos (error de precio), costo 0,
y propone un precio sugerido a margen objetivo. Solo lectura."""
import re, unicodedata, difflib, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

SNAP = r"C:\Users\dell\victtorino\fala_editor_save.md"
COSTS= r"C:\Users\dell\Downloads\costos_defontana_limpio.xlsx"
OUT  = r"C:\Users\dell\victtorino\auditoria_precios.xlsx"
TARGET_MARGIN = 0.55   # margen objetivo para precio sugerido (punto de partida)

def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().upper()
    return re.sub(r'\s+',' ',re.sub(r'[^A-Z0-9 ]',' ',s)).strip()
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else x

# catalog
prods={}
for ln in open(SNAP,encoding='utf-8').read().splitlines():
    m=re.search(r"row \"(.+?) SKU: (\d+) - EAS: - ([\d.]+) CLP\"",ln)
    if m: prods[m.group(2)]={'name':m.group(1),'sku':m.group(2),'price':int(m.group(3).replace('.',''))}
prods=list(prods.values())

# costs
wb=openpyxl.load_workbook(COSTS,read_only=True); ws=wb.active
costs=[]
for r in ws.iter_rows(values_only=True):
    if not r or r[0] in (None,'Codigo') or not isinstance(r[2],(int,float)): continue
    costs.append({'cod':r[0],'desc':r[1],'cost':int(r[2]),'n':norm(r[1])})
cmap={c['n']:c for c in costs}; cn=list(cmap.keys())
def match(name):
    n=norm(name)
    if n in cmap: return cmap[n]
    cand=difflib.get_close_matches(n,cn,n=1,cutoff=0.6)
    if cand: return cmap[cand[0]]
    nt=set(n.split()); best=None;bs=0
    for c in costs:
        ct=set(c['n'].split());
        if ct:
            j=len(nt&ct)/len(nt|ct)
            if j>bs: bs=j;best=c
    return best if bs>=0.45 else None

rows=[]
for p in prods:
    c=match(p['name'])
    cost=c['cost'] if c else None
    marg=None; flags=[]
    if cost is None: flags.append('SIN COSTO')
    elif cost==0:   flags.append('COSTO 0')
    else:
        marg=(p['price']-cost)/p['price']
        if p['price']<=cost: flags.append('VENTA A PERDIDA')
        elif marg<0.15:      flags.append('MARGEN <15%')
        if marg>=0.92:       flags.append('MARGEN ABSURDO (revisar precio)')
    if p['price']>=250000:   flags.append('PRECIO MUY ALTO')
    sug = r990(cost/(1-TARGET_MARGIN)) if (cost and cost>0) else None
    rows.append({**p,'cost':cost,'marg':marg,'flags':flags,'sug':sug})

flagged=[r for r in rows if r['flags']]
print(f"Productos: {len(rows)} | con alerta: {len(flagged)}")
print('Conteo de alertas:', Counter(f for r in rows for f in r['flags']))
print('\n--- ALERTAS (ordenadas) ---')
order={'VENTA A PERDIDA':0,'MARGEN <15%':1,'MARGEN ABSURDO (revisar precio)':2,'PRECIO MUY ALTO':2,'COSTO 0':3,'SIN COSTO':4}
for r in sorted(flagged,key=lambda x:min(order.get(f,9) for f in x['flags'])):
    mg=f"{r['marg']*100:4.0f}%" if r['marg'] is not None else "  - "
    cs=f"${r['cost']:,}" if r['cost'] is not None else "s/c"
    print(f"  [{' / '.join(r['flags'])}]  precio ${r['price']:,}  costo {cs}  marg {mg}  | {r['name'][:42]}")

# excel
wbo=openpyxl.Workbook(); w=wbo.active; w.title='Auditoria precios'
w.append(['SKU','Producto','Precio actual','Costo','Margen %','Alertas','Precio sugerido (55%)'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='C00000'); c.alignment=Alignment(horizontal='center')
for r in sorted(rows,key=lambda x:(0 if x['flags'] else 1, -(x['marg'] or 0))):
    w.append([r['sku'],r['name'],r['price'],r['cost'],
              round(r['marg'],3) if r['marg'] is not None else None,
              ' / '.join(r['flags']),r['sug']])
    if r['flags']:
        fill='FFC7CE' if any(x in r['flags'] for x in('VENTA A PERDIDA','MARGEN <15%')) else 'FFEB9C'
        for c in w[w.max_row]: c.fill=PatternFill('solid',fgColor=fill)
for col,wd in zip('ABCDEFG',[12,46,13,10,9,34,18]): w.column_dimensions[col].width=wd
for rr in range(2,w.max_row+1): w.cell(rr,5).number_format='0.0%'
wbo.save(OUT); print('\nGuardado:',OUT)
