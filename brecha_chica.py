# -*- coding: utf-8 -*-
"""Lista BRECHA CHICA: subpreciados cuyo precio real sube <=25% hasta la formula.
Propone Normal (x4.49 tachado) + Venta real (x3.14). Salida para confirmar y cargar."""
import json, re, pandas as pd, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

IVA=1.19; M_VENTA=3.14
GANCHO=0.50   # descuento exhibido objetivo (tachado): normal = venta/(1-GANCHO)
MAX_SUBIDA=0.25   # brecha chica = subida real <= 25%
REP=r"C:\Users\dell\Downloads\Informe de artículos 20250627210343860887 14062026 0314.xls"
OUT=r"C:\Users\dell\victtorino\carga_brecha_chica.xlsx"
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)
def num(x):
    s=re.sub(r'[^0-9]','',str(x)); return int(s) if s else 0

cost={}; stock={}; desc={}
t=pd.read_html(REP)[0].iloc[3:]
for _,r in t.iterrows():
    m=re.match(r'^(\d+-[A-Z0-9]+)-(.+)$',str(r.iloc[0]))
    if not m: continue
    cost[m.group(1)]=num(r.iloc[2]); stock[m.group(1)]=num(r.iloc[1]); desc[m.group(1)]=m.group(2).strip()

cur=json.load(open(r"C:\Users\dell\victtorino\precios_actuales_fala.json"))
chica=[]; grande=[]
for c in cur:
    cod=c['cod']; cv=cost.get(cod)
    if not cv or cv<=0: continue
    real=c['oferta'] if c['oferta'] else c['precio']
    base=cv*IVA; vf=r990(base*M_VENTA); nf=r990(vf/(1-GANCHO))
    if real>=vf*0.92: continue   # no es subpreciado
    subida=vf/real-1
    item={'cod':cod,'desc':desc.get(cod,''),'stock':stock.get(cod),'actual':real,
          'nueva_venta':vf,'nuevo_normal':nf,'subida':subida,
          'marg_old':1-base/real,'marg_new':1-base/vf}
    (chica if subida<=MAX_SUBIDA else grande).append(item)

chica.sort(key=lambda x:x['subida'])
print(f"BRECHA CHICA (subida<=25%): {len(chica)}  |  BRECHA GRANDE: {len(grande)}")
print(f"\n{'cod':>13} {'stock':>5} {'actual':>8} {'->venta':>8} {'subida':>6} {'normal':>8}  desc")
for x in chica:
    print(f"{x['cod']:>13} {x['stock']:>5} {x['actual']:>8,} {x['nueva_venta']:>8,} {x['subida']:>5.0%} {x['nuevo_normal']:>8,}  {x['desc'][:30]}")

wb=openpyxl.Workbook(); w=wb.active; w.title='Brecha chica - cargar'
w.append(['Codigo','Descripcion','Stock','Precio actual','NUEVO Normal (tachado)','NUEVA Venta (real)','Subida %','Margen antes','Margen despues'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
for x in chica:
    w.append([x['cod'],x['desc'],x['stock'],x['actual'],x['nuevo_normal'],x['nueva_venta'],
              round(x['subida'],3),round(x['marg_old'],3),round(x['marg_new'],3)])
for col in [3,4,5,6]:
    for rr in range(2,w.max_row+1):
        cc=w.cell(rr,col)
        if isinstance(cc.value,(int,float)): cc.number_format='#,##0'
for rr in range(2,w.max_row+1):
    for col in [7,8,9]: w.cell(rr,col).number_format='0.0%'
for i,wd in enumerate([13,40,7,13,18,18,9,12,13]): w.column_dimensions[chr(65+i)].width=wd
wb.save(OUT)
json.dump(chica, open(r"C:\Users\dell\victtorino\brecha_chica.json","w"))
print('\nGuardado:',OUT)
