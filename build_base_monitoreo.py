# -*- coding: utf-8 -*-
"""Base de monitoreo Falabella: precio, costo, margen, competencia, stock + columnas
vacias para seguir ventas/flujos. Todos los productos con stock."""
import json, glob, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def load(f):
    raw=open(f,encoding='utf-8').read().strip(); d=json.loads(raw); return json.loads(d) if isinstance(d,str) else d

# VENTA consolidado (99)
spec=open('generar_excel_master.py',encoding='utf-8').read(); ns={}
exec(spec[spec.index('VENTA = {'):spec.index('}\n# costos manuales')+1], ns)
VENTA=dict(ns['VENTA'])
l=openpyxl.load_workbook('_lote24.xlsx',data_only=True).active
for r in list(l.iter_rows(values_only=True))[1:]: VENTA[r[0]]=int(r[3])
VENTA.update({'010703004-T':54990,'010102007-T':36990,'010701012-T':33990,'040202004-T':27990,
 '040201009-T':27990,'020101004-T':37990,'040201008-T':27990,'040201007-T':23990,'040501002-T':24990,
 '010704020-T':17990,'010102003-T':17990,'010204004-T':9990,'010104001-T':16990,'020101009-T':33990})
VENTA.update({'040302001-T':23990,'010701007-T':17990,'040201006-T':19990,'040204003-T':23990,'040202001-T':23990})

admin=json.load(open('admin_full.json',encoding='utf-8'))
tpl=openpyxl.load_workbook('SellerPriceTemplate.xlsx',data_only=True).active
shop={}
for r in list(tpl.iter_rows(values_only=True))[1:]:
    shop[re.sub(r'_\d+$','',str(r[0]))]={'shop':r[1],'name':r[6]}
wv=openpyxl.load_workbook('validacion_costos.xlsx',data_only=True).active
rv=list(wv.iter_rows(values_only=True)); Hv=rv[0]
val={r[Hv.index('Codigo')]:r[Hv.index('COSTO A USAR')] for r in rv[1:] if r[Hv.index('Codigo')] and isinstance(r[Hv.index('COSTO A USAR')],(int,float))}
vig=load('costos_vigentes.json')
MANUAL={'010102012-C':15410,'010603002-T':46929,'020101014-T':40023,'020101013-T':36157,'020101009-T':12261,'010104001-T':6016,'010704020-T':4223,'040303002-T':42686}
def costo(s):
    if s in MANUAL: return MANUAL[s],'Taumm(manual)'
    if s in val: return val[s],'validado'
    if s in vig and vig[s].get('vig'): return vig[s]['vig'],'vigente'
    return None,'-'
# competidores TAUMM
taumm=[]
for f in sorted(glob.glob('taumm_[0-9].json')):
    try: taumm+=load(f)
    except: pass
comp=[c for c in taumm if 'Trade' not in c.get('seller','')]
MOD=['NOTTE','COLOMBA','DOMENICA','VANDER','MODERN','BONN','SWERTT','ANTONELLA','MILAN','PROFESIONAL','TEMPORIZADA','CISNE','VERTICAL','ALTO','BIDET','MONOBLOCK','LAVACOPA','CUELLO']
TIP=['LAVAPLATO','LAVATORIO','LAVAMANOS','DUCHA','TINA','BIDET','LAVACOPA','JARDIN','ANGULAR','DISPENSADOR','BARRA','SIFON','BANDEJA','RECEPTACULO','URINARIO','FLUXOMETRO','ESPEJO','TOALLERO','PLATO','MAMPARA','SHOWER']
MS=set(MOD)
def tk(n):
    n=n.upper(); s=set()
    for m in MOD:
        if m in n: s.add(m)
    for t in TIP:
        if t in n: s.add(t)
    return s
def bestc(desc):
    mt=tk(desc); my=mt&MS; b=None
    for c in comp:
        ct=tk(c['name'])
        if mt and len(mt&ct)>=2 and (not my or (my&ct)):
            if b is None or c['price']<b['price']: b=c
    return b

wb=openpyxl.Workbook(); w=wb.active; w.title='Base Monitoreo'
hdr=['SKU Seller','SKU Falabella','Producto','Stock','ADS','Costo neto','Costo c/IVA','Fuente costo',
     'Precio NORMAL','PRECIO OFERTA','Factor','Margen','Competidor TAUMM','Vendedor comp.','vs comp.',
     'Vigencia','Ventas (unid)','Estado special','Notas']
w.append(hdr)
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
def r990(x): x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)
rows_data=[(s,a) for s,a in admin.items() if a.get('stock') and a['stock']>0]
for s,a in sorted(rows_data, key=lambda kv:-(VENTA.get(kv[0]) or kv[1].get('special') or kv[1].get('price') or 0)):
    sm=shop.get(s,{}); cn,src=costo(s); civa=int(cn*1.19) if cn else None
    venta=VENTA.get(s)
    margen=round(1-civa/venta,3) if (civa and venta) else None
    factor=round(venta/civa,2) if (civa and venta) else None
    b=bestc(a.get('name','') or s)
    vscomp=round(venta/b['price']-1,3) if (b and venta) else None
    w.append([s, sm.get('shop'), (sm.get('name') or a.get('name'))[:48], a['stock'], 'SI' if a.get('sponsored') else '',
              cn, civa, src, r990(venta/0.5) if venta else (a.get('special') or a.get('price')), venta,
              factor, margen, b['price'] if b else None, b['seller'] if b else None, vscomp,
              '16/06-16/07/2026' if venta else '', None, 'activo' if venta else 'sin oferta', None])
# formato
for rr in range(2,w.max_row+1):
    for col in [6,7,9,10,13]:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
    c11=w.cell(rr,11)
    if isinstance(c11.value,(int,float)): c11.number_format='0.00"x"'
    for col in [12,15]:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='0%'
    w.cell(rr,10).fill=PatternFill('solid',fgColor='C6EFCE')
    for col in [17,18,19]: w.cell(rr,col).fill=PatternFill('solid',fgColor='FFF2CC')
widths=[13,13,46,6,5,11,11,13,13,13,7,8,13,16,8,16,12,13,20]
for i,wd in enumerate(widths): w.column_dimensions[chr(65+i)].width=wd
w.freeze_panes='C2'
OUT=r'C:\Users\dell\Downloads\Falabella_BASE_monitoreo.xlsx'
wb.save(OUT); print('Generado:',OUT,'| filas:',len(rows_data))
print('con oferta:',sum(1 for s,a in rows_data if VENTA.get(s)),'| con competidor:',sum(1 for s,a in rows_data if bestc(a.get('name','') or s)))
