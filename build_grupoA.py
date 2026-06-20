# -*- coding: utf-8 -*-
"""Grupo A distintivos/commodity con recomendacion aprobada por Victor.
Espejo a la pared (010703011) ajustado a 52% margen = $16.990."""
import openpyxl, re

VENTA={
 '010502005-T':219000,  # Mampara
 '010501022-T':168990,  # Shower Door
 '010201001-T':43990,   # Lavatorio Cristal
 '010102001-T':15990,   # Barra c/Jabonera
 '010703011-T':16990,   # Espejo Doble al Muro (52%)
 '010703012-T':12990,   # Espejo Redondo Aumento
 '020101002-C':13990,   # Lavaplato 37x32
 '010301004-T':14990,   # Kit Estanque WC
 '010701011-T':6990,    # Dispensador simple
 '010205004-T':4990,    # Sifon lavatorio codo
 '030101008-T':3990,    # Manguera lavadoras
 '010105001-T':2990,    # Ducha fija difusor
 '030102003-K':3990,    # Toalla interfoleada
}
def r990(x): x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)
SALE_START='2026-06-16 00:00:01'; SALE_END='2026-07-16 23:59:59'

tpl=openpyxl.load_workbook('SellerPriceTemplate.xlsx',data_only=True).active
shop={}
for r in list(tpl.iter_rows(values_only=True))[1:]:
    shop[re.sub(r'_\d+$','',str(r[0]))]={'sku':r[0],'shop':r[1],'name':r[6]}

wb=openpyxl.Workbook(); w=wb.active; w.title='Sheet'
w.append(['SellerSku','ShopSku','PriceFalabella','SalePriceFalabella','SaleStartDateFalabella','SaleEndDateFalabella','Name'])
miss=[]
for sku,venta in VENTA.items():
    sm=shop.get(sku)
    if not sm: miss.append(sku); continue
    tach=r990(venta/0.5)
    w.append([sm['sku'],sm['shop'],float(tach),float(venta),SALE_START,SALE_END,sm['name']])
wb.save('SellerPriceTemplate_UPLOAD_nuevo.xlsx')
print('Grupo A:',len(VENTA)-len(miss),'filas | faltan shopsku:',miss)
for r in list(openpyxl.load_workbook('SellerPriceTemplate_UPLOAD_nuevo.xlsx',data_only=True).active.iter_rows(values_only=True))[1:]:
    print(' ',r[0],'shop',r[1],'normal',int(r[2]),'oferta',int(r[3]))
