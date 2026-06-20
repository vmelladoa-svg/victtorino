# -*- coding: utf-8 -*-
"""Excel maestro final con TODOS los productos en oferta (~99)."""
import json, glob, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def load(f):
    raw=open(f,encoding='utf-8').read().strip(); d=json.loads(raw); return json.loads(d) if isinstance(d,str) else d

# 1) master 65
spec=open('generar_excel_master.py',encoding='utf-8').read(); ns={}
exec(spec[spec.index('VENTA = {'):spec.index('}\n# costos manuales')+1], ns)
VENTA=dict(ns['VENTA'])
# 2) lote 24 (de _lote24.xlsx, incluye espejo100x70=76990)
l=openpyxl.load_workbook('_lote24.xlsx',data_only=True).active
for r in list(l.iter_rows(values_only=True))[1:]: VENTA[r[0]]=int(r[3])
# 3) ajustes manuales posteriores
VENTA.update({'010703004-T':54990,'010102007-T':36990,'010701012-T':33990,'040202004-T':27990})
# 4) editable 8
VENTA.update({'040201009-T':27990,'020101004-T':37990,'040201008-T':27990,'040201007-T':23990,'040501002-T':24990,'010704020-T':17990,'010102003-T':17990,'010204004-T':9990})
# 5) finales 2
VENTA.update({'010104001-T':16990,'020101009-T':33990})
print('TOTAL en master:',len(VENTA))

# datos
admin=json.load(open('admin_full.json',encoding='utf-8'))
tpl=openpyxl.load_workbook('SellerPriceTemplate.xlsx',data_only=True).active
shop={}
for r in list(tpl.iter_rows(values_only=True))[1:]:
    shop[re.sub(r'_\d+$','',str(r[0]))]={'shop':r[1],'name':r[6]}
wv=openpyxl.load_workbook('validacion_costos.xlsx',data_only=True).active
rv=list(wv.iter_rows(values_only=True)); Hv=rv[0]
val={r[Hv.index('Codigo')]:r[Hv.index('COSTO A USAR')] for r in rv[1:] if r[Hv.index('Codigo')] and isinstance(r[Hv.index('COSTO A USAR')],(int,float))}
vig=load('costos_vigentes.json')
MANUAL={'010102012-C':15410,'010603002-T':46929,'020101014-T':40023,'020101013-T':36157,'020101009-T':12261,'010104001-T':6016,'010704020-T':4223}
def costo(s):
    if s in MANUAL: return MANUAL[s]
    if s in val: return val[s]
    if s in vig and vig[s].get('vig'): return vig[s]['vig']
    return None
def r990(x): x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

wb=openpyxl.Workbook(); w=wb.active; w.title='Oferta Falabella'
w.append(['SKU Seller','SKU Falabella','Producto','Stock','Precio NORMAL (tachado)','PRECIO OFERTA','Costo c/IVA','Margen','Vigencia'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
for sku,venta in sorted(VENTA.items(), key=lambda kv:-kv[1]):
    a=admin.get(sku,{}); sm=shop.get(sku,{})
    cnt=costo(sku); civa=int(cnt*1.19) if cnt else None
    margen=round(1-civa/venta,3) if civa else None
    tach= 76990*2 if sku=='010703013-T' else r990(venta/0.5)
    w.append([sku, sm.get('shop'), sm.get('name') or a.get('name'), a.get('stock'), r990(venta/0.5), venta, civa, margen, '16/06-16/07/2026'])
for rr in range(2,w.max_row+1):
    for col in [5,6,7]:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
    mc=w.cell(rr,8)
    if isinstance(mc.value,(int,float)): mc.number_format='0%'
    w.cell(rr,6).fill=PatternFill('solid',fgColor='C6EFCE')
for i,wd in enumerate([13,13,48,7,15,13,11,8,16]): w.column_dimensions[chr(65+i)].width=wd
w.freeze_panes='A2'
OUT=r'C:\Users\dell\Downloads\Oferta_Falabella_OportunidadUnica.xlsx'
wb.save(OUT); print('Generado:',OUT,'| productos:',len(VENTA))
