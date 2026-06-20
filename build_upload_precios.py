# -*- coding: utf-8 -*-
"""Construye el archivo de carga de precios Falabella sobre la plantilla oficial
SellerPriceTemplate: setea PriceFalabella=Normal, SalePriceFalabella=Venta y fechas
en los 125 productos finales (match por ShopSku). El resto queda intacto."""
import openpyxl
TPL=r'C:\Users\dell\victtorino\SellerPriceTemplate.xlsx'
FIN=r'C:\Users\dell\victtorino\PRECIOS_FINAL_falabella.xlsx'
OUT=r'C:\Users\dell\victtorino\SellerPriceTemplate_UPLOAD.xlsx'
START='2026-06-14 00:00:00'; END='2027-12-31 23:59:59'

# final: map skuF -> (normal, venta)
wf=openpyxl.load_workbook(FIN).active
m={}
for r in range(2,wf.max_row+1):
    sku=wf.cell(r,2).value      # SKU Falabella
    normal=wf.cell(r,8).value; venta=wf.cell(r,9).value
    if sku and normal and venta: m[str(sku).strip()]=(int(normal),int(venta))
print('Productos a actualizar (final):',len(m))

wb=openpyxl.load_workbook(TPL); ws=wb.active
hdr={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
cP=hdr['PriceFalabella']; cS=hdr['SalePriceFalabella']; cSt=hdr['SaleStartDateFalabella']; cEn=hdr['SaleEndDateFalabella']; cShop=hdr['ShopSku']
upd=0; notfound=set(m.keys())
for r in range(2,ws.max_row+1):
    shop=ws.cell(r,cShop).value
    if shop is None: continue
    key=str(int(shop)) if isinstance(shop,float) else str(shop).strip()
    if key in m:
        normal,venta=m[key]
        ws.cell(r,cP).value=f'{normal}.00'
        ws.cell(r,cS).value=f'{venta}.00'
        ws.cell(r,cSt).value=START
        ws.cell(r,cEn).value=END
        upd+=1; notfound.discard(key)
print('Filas actualizadas en plantilla:',upd)
print('No encontrados en plantilla (skuF):',len(notfound), list(notfound)[:10])
wb.save(OUT)
print('Guardado:',OUT)
# muestra
ws2=openpyxl.load_workbook(OUT).active
shown=0
for r in range(2,ws2.max_row+1):
    if ws2.cell(r,cS).value and str(ws2.cell(r,cSt).value)==START and shown<6:
        print('  ',ws2.cell(r,1).value, ws2.cell(r,cShop).value,'normal',ws2.cell(r,cP).value,'venta',ws2.cell(r,cS).value)
        shown+=1
