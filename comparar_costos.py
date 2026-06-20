# -*- coding: utf-8 -*-
"""Compara costos Defontana vs Taumm (cruce por descripcion) para cazar costos malos."""
import json,re,unicodedata,difflib,pandas as pd,openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
def norm(s):
    s=str(s).replace('�',' ')
    s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode().upper()
    return re.sub(r'\s+',' ',re.sub(r'[^A-Z0-9 ]',' ',s)).strip()
def n(x):
    s=re.sub(r'[^0-9]','',str(x)); return int(s) if s else 0

d=json.load(open('taumm_full.json',encoding='utf-8'))
if isinstance(d,str): d=json.loads(d)
twd={r['cod']:r for r in d['all']}
tw=[{'cod':r['cod'],'desc':r['desc'],'cost':n(r['precio']),'nd':norm(r['desc'])} for r in twd.values() if n(r['precio'])>0]
twn=[x['nd'] for x in tw]

t=pd.read_html(r'C:\Users\dell\Downloads\Informe de artículos 20250627210343860887 14062026 0314.xls')[0].iloc[3:]
dfn=[]
for _,r in t.iterrows():
    m=re.match(r'^(\d+-[A-Z0-9]+)-(.+)$',str(r.iloc[0]))
    if not m: continue
    dfn.append({'cod':m.group(1),'desc':m.group(2).strip(),'cost':n(r.iloc[2]),'nd':norm(m.group(2))})

rows=[]; match=0; big=0
for x in dfn:
    cand=difflib.get_close_matches(x['nd'],twn,n=1,cutoff=0.72)
    if not cand:
        rows.append((x['cod'],x['desc'],x['cost'],None,None,0,'')); continue
    tm=tw[twn.index(cand[0])]; sc=difflib.SequenceMatcher(None,x['nd'],cand[0]).ratio()
    diff=(x['cost']-tm['cost'])/tm['cost'] if tm['cost'] else None
    match+=1
    if diff is not None and abs(diff)>0.15: big+=1
    rows.append((x['cod'],x['desc'],x['cost'],tm['cost'],diff,round(sc,2),tm['desc']))

print('Defontana:',len(dfn),'| Taumm con precio:',len(tw))
print('Matcheados por descripcion (>=0.72):',match,'| costo discrepante >15%:',big)
print(f"\n{'desc Defontana':28} {'cDef':>7} {'cTaumm':>7} {'dif':>5} {'sc':>4}")
for r in sorted([x for x in rows if x[4] is not None and abs(x[4])>0.15],key=lambda z:-abs(z[4]))[:18]:
    print(f"{r[1][:28]:28} {r[2]:>7,} {r[3]:>7,} {r[4]*100:>4.0f}% {r[5]:>4}")

wb=openpyxl.Workbook(); w=wb.active; w.title='Defontana vs Taumm'
w.append(['Codigo Def','Desc Defontana','Costo Defontana','Costo Taumm','Diferencia %','Match score','Desc Taumm (match)'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
for r in sorted(rows,key=lambda z:(z[4] is None, -(abs(z[4]) if z[4] is not None else 0))):
    w.append([r[0],r[1],r[2],r[3],round(r[4],3) if r[4] is not None else None,r[5],r[6]])
    if r[4] is not None and abs(r[4])>0.15:
        for c in w[w.max_row]: c.fill=PatternFill('solid',fgColor='FFC7CE')
for col in [3,4]:
    for rr in range(2,w.max_row+1):
        cc=w.cell(rr,col)
        if isinstance(cc.value,(int,float)): cc.number_format='#,##0'
for rr in range(2,w.max_row+1): w.cell(rr,5).number_format='0.0%'
for i,wd in enumerate([12,34,14,12,11,10,34]): w.column_dimensions[chr(65+i)].width=wd
wb.save('comparar_costos_def_vs_taumm.xlsx')
print('\nGuardado: comparar_costos_def_vs_taumm.xlsx')
