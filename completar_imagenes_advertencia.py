"""
completar_imagenes_advertencia.py
---------------------------------
Completa las imagenes faltantes en paris_productos_carga_masiva.xlsx para los
3 productos que quedaron con menos de 2 imagenes:

  - MLC51764090 (Taumm Ducha Columna)
        Sin imagenes en su producto catalogo, pero su hijo MLC51764092 tiene 3.
        Se agregan al Excel directamente (URLs mlstatic -F.jpg).

  - MLC32065874 (Llave Lavadora 2 salidas 3/4)
        Sin imagenes adicionales en API ML. Confirmado por el usuario que es el
        mismo producto fisico que el SKU 020301002-T en victtorino.cl (Woo #799),
        que tiene 5 fotos unicas en .webp. Como Paris solo acepta .jpg/.png/.bmp
        las imagenes se descargan, se convierten a JPG con calidad 90 y se
        guardan en paris_imagenes/MLC32065874/ para subida manual.
        El Excel se deja con la imagen original de ML porque no podemos
        publicar URLs locales.

  - MLC24386020 (Flexible Lavaplatos Profesional)
        Sin match confiable en Woo. Queda con 1 imagen. Victor sube la 2a
        manualmente.
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "paris_productos_carga_masiva.xlsx"
DATA_SHEET = "herramientas"
COL_SKU = 2
COL_IMAGENES = 13
IMG_OUT_DIR = BASE_DIR / "paris_imagenes"

# Nuevas URLs mlstatic para MLC51764090 (extraidas del producto hijo MLC51764092)
MLC51764090_NEW = [
    "https://http2.mlstatic.com/D_NQ_NP_866441-MLA110294771918_052026-F.jpg",
    "https://http2.mlstatic.com/D_NQ_NP_670808-MLA99523453490_122025-F.jpg",
    "https://http2.mlstatic.com/D_NQ_NP_936923-MLA110295150786_052026-F.jpg",
]

# Imagenes webp en victtorino.cl que hay que convertir a jpg local para
# MLC32065874. La primera URL aparece duplicada en Woo (imagen principal).
MLC32065874_WEBPS = [
    "https://victtorino.cl/wp-content/uploads/2026/03/llave-lavadero-lavadora-doble-3-4-victtorino-chile.webp",
    "https://victtorino.cl/wp-content/uploads/2026/03/Llave-Lavadero-Lavadora-Doble-34-1.webp",
    "https://victtorino.cl/wp-content/uploads/2026/03/Llave-Lavadero-Lavadora-Doble-34-2.webp",
    "https://victtorino.cl/wp-content/uploads/2026/03/Llave-Lavadero-Lavadora-Doble-34-3.webp",
    "https://victtorino.cl/wp-content/uploads/2026/03/Llave-Lavadero-Lavadora-Doble-34-4.webp",
]


def find_row_by_sku(ws, sku: str) -> int | None:
    for row in range(7, ws.max_row + 2):
        if ws.cell(row=row, column=COL_SKU).value == sku:
            return row
    return None


def download_and_convert_to_jpg(url: str, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    name = url.rsplit("/", 1)[-1]
    stem = re.sub(r"\.webp$", "", name, flags=re.I)
    out = out_dir / f"{stem}.jpg"
    if out.exists():
        return out
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    r.raise_for_status()
    im = Image.open(io.BytesIO(r.content)).convert("RGB")
    im.save(out, format="JPEG", quality=90, optimize=True)
    return out


def update_excel_for_mlc51764090(ws) -> tuple[int, int]:
    row = find_row_by_sku(ws, "MLC51764090")
    if row is None:
        return (0, 0)
    current = (ws.cell(row=row, column=COL_IMAGENES).value or "").strip()
    existing = [u.strip() for u in current.split(",") if u.strip()] if current else []
    merged = existing + [u for u in MLC51764090_NEW if u not in existing]
    ws.cell(row=row, column=COL_IMAGENES, value=", ".join(merged))
    return row, len(merged)


def convert_mlc32065874_locally() -> list[Path]:
    out_dir = IMG_OUT_DIR / "MLC32065874"
    seen: set[str] = set()
    paths: list[Path] = []
    for url in MLC32065874_WEBPS:
        if url in seen:
            continue
        seen.add(url)
        try:
            p = download_and_convert_to_jpg(url, out_dir)
            paths.append(p)
        except Exception as e:
            print(f"  [WARN] No se pudo convertir {url}: {e}")
    return paths


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

    if not EXCEL_FILE.exists():
        print(f"[ERROR] No existe {EXCEL_FILE}")
        return 1

    print("[1/2] Actualizando Excel: MLC51764090 (+3 imagenes desde child MLC51764092)")
    wb = load_workbook(EXCEL_FILE)
    ws = wb[DATA_SHEET]
    row, total = update_excel_for_mlc51764090(ws)
    print(f"      fila {row} -> {total} imagenes en total")
    wb.save(EXCEL_FILE)

    print("[2/2] Convirtiendo webp -> jpg para MLC32065874 (5 imagenes)")
    paths = convert_mlc32065874_locally()
    for p in paths:
        size_kb = os.path.getsize(p) // 1024
        print(f"      {p.relative_to(BASE_DIR)}  ({size_kb} KB)")

    print()
    print("=" * 60)
    print("RESUMEN")
    print("=" * 60)
    print(f"MLC51764090  : Excel actualizado con {total} imagenes (CDN ML).")
    print(f"MLC32065874  : {len(paths)} JPG generadas en paris_imagenes/MLC32065874/.")
    print(f"               El Excel mantiene la imagen original de ML como URL.")
    print(f"               Victor debe subir manualmente las JPG locales en")
    print(f"               Seller Center al editar este SKU (Paris no acepta .webp).")
    print(f"MLC24386020  : sin cambios. 1 imagen. Sube la 2a manual.")
    print()
    print(f"Excel: {EXCEL_FILE}")
    print(f"Imagenes locales: {IMG_OUT_DIR / 'MLC32065874'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
