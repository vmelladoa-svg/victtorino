# -*- coding: utf-8 -*-
"""Cruce completo: costos Defontana vs precio web vs precio ML, + bajo costo corregido + sin publicar."""
import pandas as pd, warnings
warnings.filterwarnings('ignore')

IVA = 1.19
MARKUP_OBJETIVO = 2.2   # markup neto estandar del catalogo (mediana)
OUT = r'C:\Users\dell\victtorino\cruce_completo_costos.xlsx'

# --- costos ---
costos = pd.read_excel('costos_defontana_limpio.xlsx')
costos['Codigo'] = costos['Codigo'].astype(str).str.strip()

def num(v):
    v = str(v).strip().replace('.', '').replace(',', '')
    return int(v) if v.lstrip('-').isdigit() else None

# --- web ---
web = pd.read_csv('catalogo-trade.csv', header=None, dtype=str, keep_default_na=False)
web = web[[2, 4, 25, 26]].copy()
web.columns = ['Codigo', 'Nombre_web', 'po', 'pn']
web['Codigo'] = web['Codigo'].str.strip()
web['po'] = web['po'].map(num); web['pn'] = web['pn'].map(num)
web['Precio web (IVA)'] = web['po'].where(web['po'].fillna(0) > 0, web['pn'])
web = web.drop_duplicates('Codigo')[['Codigo', 'Nombre_web', 'Precio web (IVA)']]

# --- ML: seller_sku = codigo con letra ; varios listings por codigo -> min y nro ---
ml = pd.read_excel('matching_defontana_ml_2026-05-26.xlsx', sheet_name='MATCHED')
ml['Codigo'] = ml['seller_sku'].astype(str).str.strip()
ml = ml[ml['price_ml'].notna()]
mlg = ml.groupby('Codigo').agg(**{
    'Precio ML min (IVA)': ('price_ml', 'min'),
    'N listings ML': ('price_ml', 'size'),
}).reset_index()

# --- merge ---
m = costos.merge(web, on='Codigo', how='left').merge(mlg, on='Codigo', how='left')
costo = m['Costo compra (CLP)']

def neto(col): return (m[col] / IVA).round(0)
m['Margen web %'] = ((neto('Precio web (IVA)') - costo) / neto('Precio web (IVA)') * 100).round(1)
m['Margen ML %']  = ((neto('Precio ML min (IVA)') - costo) / neto('Precio ML min (IVA)') * 100).round(1)
m['Markup web x'] = pd.to_numeric(neto('Precio web (IVA)') / costo.replace(0, pd.NA), errors='coerce').round(2)

def alerta(r):
    msgs = []
    c = r['Costo compra (CLP)']
    mw = r['Margen web %']
    mml = r['Margen ML %']
    pw = r['Precio web (IVA)']
    # 1) costo sucio en Defontana
    if pd.isna(c) or c <= 1:
        msgs.append(f'COSTO DUDOSO en Defontana (${0 if pd.isna(c) else int(c)}) - corregir ERP')
    # 2) vende bajo costo
    if pd.notna(mw) and mw < 0:
        msgs.append(f'VENDE BAJO COSTO en web (margen {mw:.0f}%)')
    elif pd.notna(mw) and mw < 25:
        msgs.append(f'Margen web bajo ({mw:.0f}%)')
    # 3) ML bajo costo
    if pd.notna(mml) and mml < 0:
        msgs.append(f'Pierde en ML (margen {mml:.0f}%, sin comision)')
    # 4) sin publicar en web
    if pd.isna(pw):
        msgs.append('Sin precio web - no publicado')
    return ' | '.join(msgs)

cols = ['Codigo', 'Descripcion', 'Nombre_web', 'Costo compra (CLP)',
        'Precio web (IVA)', 'Margen web %', 'Markup web x',
        'Precio ML min (IVA)', 'N listings ML', 'Margen ML %']
cruce = m[cols].copy()
cruce['ALERTA'] = m.apply(alerta, axis=1)
cruce = cruce.sort_values('Margen web %', na_position='last')

# --- bajo costo (margen web negativo) con precio corregido ---
def redondear_990(x):
    if pd.isna(x): return None
    return int(round(x / 1000.0)) * 1000 - 10  # termina en .990

bajo = m[m['Margen web %'] < 0].copy()
bajo['Precio web actual (IVA)'] = bajo['Precio web (IVA)']
bajo['Precio sugerido (IVA)'] = (costo.loc[bajo.index] * MARKUP_OBJETIVO * IVA).map(redondear_990)
bajo['Margen sugerido %'] = ((bajo['Precio sugerido (IVA)'] / IVA - costo.loc[bajo.index])
                             / (bajo['Precio sugerido (IVA)'] / IVA) * 100).round(1)
bajo = bajo[['Codigo', 'Descripcion', 'Costo compra (CLP)', 'Precio web actual (IVA)',
             'Margen web %', 'Precio sugerido (IVA)', 'Margen sugerido %']]

# --- sin publicar en web ---
sin_web = m[m['Precio web (IVA)'].isna()][['Codigo', 'Descripcion', 'Costo compra (CLP)',
                                           'Precio ML min (IVA)']].copy()
sin_web['Precio sugerido web (IVA)'] = (costo.loc[sin_web.index] * MARKUP_OBJETIVO * IVA).map(redondear_990)

leyenda = pd.DataFrame({
    'Color': ['ROJO', 'NARANJO', 'AMARILLO', 'GRIS', '(sin color)'],
    'Significado': [
        'Critico: vende bajo costo en web (pierde plata en cada venta)',
        'Costo dudoso en Defontana ($0 o $1) - corregir en el ERP',
        'Margen web bajo (menor a 25%)',
        'Sin precio web - producto no publicado en el sitio',
        'Sin alertas',
    ],
})

with pd.ExcelWriter(OUT, engine='openpyxl') as xl:
    cruce.to_excel(xl, sheet_name='CRUCE', index=False)
    bajo.to_excel(xl, sheet_name='BAJO_COSTO_corregir', index=False)
    sin_web.to_excel(xl, sheet_name='SIN_PUBLICAR_WEB', index=False)
    leyenda.to_excel(xl, sheet_name='LEYENDA', index=False)

    # --- coloreo de la columna ALERTA en CRUCE ---
    from openpyxl.styles import PatternFill, Font
    ws = xl.sheets['CRUCE']
    col_alerta = list(cruce.columns).index('ALERTA') + 1  # 1-based
    ROJO = PatternFill('solid', fgColor='FFC7CE')
    NARANJO = PatternFill('solid', fgColor='FFD9A0')
    AMARILLO = PatternFill('solid', fgColor='FFEB9C')
    GRIS = PatternFill('solid', fgColor='D9D9D9')
    for i, (_, r) in enumerate(cruce.iterrows(), start=2):  # fila 1 = header
        txt = r['ALERTA']
        if not txt:
            continue
        if 'BAJO COSTO' in txt:
            fill = ROJO
        elif 'COSTO DUDOSO' in txt:
            fill = NARANJO
        elif 'Margen web bajo' in txt or 'Pierde en ML' in txt:
            fill = AMARILLO
        else:
            fill = GRIS
        ws.cell(row=i, column=col_alerta).fill = fill
    # ancho de la columna ALERTA
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_alerta)].width = 55

print('=== RESUMEN ===')
print(f'Total costos: {len(m)}')
print(f'Con precio web: {m["Precio web (IVA)"].notna().sum()}  | Con precio ML: {m["Precio ML min (IVA)"].notna().sum()}')
print(f'Sin publicar en web: {len(sin_web)}')
print(f'\nBAJO COSTO -> precio sugerido (markup {MARKUP_OBJETIVO}x):')
print(bajo.to_string(index=False))
print(f'\nSIN PUBLICAR WEB ({len(sin_web)}): primeros 12')
print(sin_web.head(12).to_string(index=False))
print('\n=== CONTEO DE ALERTAS ===')
ca = cruce['ALERTA']
print('Vende bajo costo (rojo):  ', ca.str.contains('BAJO COSTO').sum())
print('Costo dudoso (naranjo):   ', ca.str.contains('COSTO DUDOSO').sum())
print('Margen bajo / ML (amar.): ', (ca.str.contains('Margen web bajo') | ca.str.contains('Pierde en ML')).sum())
print('Sin publicar web (gris):  ', ca.str.contains('Sin precio web').sum())
print('Filas con alguna alerta:  ', (ca != '').sum(), 'de', len(ca))
print('\nGenerado:', OUT)
