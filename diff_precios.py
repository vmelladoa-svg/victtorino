# -*- coding: utf-8 -*-
"""DIFF: precio actual Falabella (scrapeado por codigo) vs precio formula (Costo neto x1.19 x mult).
Clasifica: PERDIDA / SUBIR (subpreciado) / PREMIUM-DEJAR / OK / SIN COSTO. Incluye stock."""
import json, re, pandas as pd, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

IVA=1.19; M_VENTA=3.14; M_NORMAL=4.49; CMR_OFF=0.93
OUT=r"C:\Users\dell\victtorino\diff_precios_falabella.xlsx"
REP=r"C:\Users\dell\Downloads\Informe de artículos 20250627210343860887 14062026 0314.xls"
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)
def num(x):
    s=re.sub(r'[^0-9]','',str(x)); return int(s) if s else 0

# costos + stock por codigo (del informe Defontana)
cost={}; stock={}
t=pd.read_html(REP)[0].iloc[3:]
for _,r in t.iterrows():
    m=re.match(r'^(\d+-[A-Z0-9]+)-(.+)$',str(r.iloc[0]))
    if not m: continue
    cod=m.group(1); cost[cod]=num(r.iloc[2]); stock[cod]=num(r.iloc[1])

cur=json.load(open(r"C:\Users\dell\victtorino\precios_actuales_fala.json"))

rows=[]
for c in cur:
    cod=c['cod']; precio=c['precio']; oferta=c['oferta']
    real = oferta if oferta else precio   # precio de venta real
    cv=cost.get(cod)
    st=stock.get(cod)
    if not cv or cv<=0:
        rows.append({**c,'real':real,'cost':cv,'stock':st,'base':None,'vf':None,'nf':None,'marg':None,'acc':'SIN COSTO'}); continue
    base=cv*IVA; vf=r990(base*M_VENTA); nf=r990(base*M_NORMAL)
    marg=1-base/real if real else None
    if real<=base: acc='PERDIDA'
    elif real < vf*0.92: acc='SUBIR'
    elif real > vf*1.10: acc='PREMIUM/DEJAR'
    else: acc='OK'
    rows.append({**c,'real':real,'cost':cv,'stock':st,'base':int(base),'vf':vf,'nf':nf,
                 'marg':marg,'acc':acc})

from collections import Counter
cnt=Counter(r['acc'] for r in rows)
print('Total scrapeado:',len(rows))
print('Diagnostico:',dict(cnt))
# margen perdido estimado (subir): (vf-real) por unidad, solo informativo
subir=[r for r in rows if r['acc']=='SUBIR']
perd=sum((r['vf']-r['real']) for r in subir)
print(f"\nSUBPRECIADOS: {len(subir)} | si los subes a formula, +${perd:,} por unidad vendida (suma)")
print('\nTop PERDIDA / SUBIR (mas urgentes):')
for r in sorted([x for x in rows if x['acc'] in('PERDIDA','SUBIR')],key=lambda x:(x['acc']!='PERDIDA', (x['marg'] if x['marg'] is not None else 9))):
    mg=f"{r['marg']*100:4.0f}%" if r['marg'] is not None else " - "
    print(f"  [{r['acc']:>6}] {r['cod']:>13} real ${r['real']:>7,} formula ${r['vf']:>7,} marg {mg} stock {r['stock']}")

print('\nPREMIUM (NO bajar):', len([r for r in rows if r['acc']=='PREMIUM/DEJAR']))

# excel
wb=openpyxl.Workbook(); w=wb.active; w.title='Diff precios'
w.append(['Codigo','SKU Fala','Stock','Costo neto','Precio actual','Oferta actual','Venta real',
          'Margen actual','VENTA formula','NORMAL formula','Accion'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
fill={'PERDIDA':'FF0000','SUBIR':'FFC7CE','PREMIUM/DEJAR':'C6EFCE','OK':'FFFFFF','SIN COSTO':'D9D9D9'}
order={'PERDIDA':0,'SUBIR':1,'SIN COSTO':2,'PREMIUM/DEJAR':3,'OK':4}
for r in sorted(rows,key=lambda x:(order[x['acc']], -(x['real'] or 0))):
    w.append([r['cod'],r['skuF'],r['stock'],r['cost'],r['precio'],r['oferta'],r['real'],
              round(r['marg'],3) if r['marg'] is not None else None,r['vf'],r['nf'],r['acc']])
    for c in w[w.max_row]: c.fill=PatternFill('solid',fgColor=fill[r['acc']])
for col in [3,4,5,6,7,9,10]:
    for rr in range(2,w.max_row+1):
        cc=w.cell(rr,col)
        if isinstance(cc.value,(int,float)): cc.number_format='#,##0'
for rr in range(2,w.max_row+1):
    cc=w.cell(rr,8)
    if isinstance(cc.value,(int,float)): cc.number_format='0.0%'
for i,wd in enumerate([13,11,7,10,12,12,11,12,13,13,14]): w.column_dimensions[chr(65+i)].width=wd
wb.save(OUT); print('\nGuardado:',OUT)
