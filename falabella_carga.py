# -*- coding: utf-8 -*-
"""
falabella_carga.py
==================
Genera el archivo FINAL de carga para Falabella:
  - falabella_CARGA.xlsx / .csv : planilla plana, UNA fila por SKU, con todos los
    atributos en columnas + nombres de archivo de las fotos.
  - falabella_fotos.zip : todas las fotos (carpetas por SKU) en un solo bundle.
"""
import csv, json, zipfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).parent
LISTO = ROOT / "falabella_fotos_listo"

# columna salida -> lista de atributos ML candidatos (primero que exista gana)
CAMPOS = {
    "Marca":            ["Marca"],
    "Modelo":           ["Modelo"],
    "Material":         ["Material", "Materiales"],
    "Color":            ["Color principal", "Color"],
    "Dimensiones":      ["Dimensiones"],
    "Peso":             ["Peso del paquete", "Peso del paquete del seller"],
    "Tipo_instalacion": ["Tipo de instalación", "Tipos de instalación"],
    "Garantia":         ["Garantía"],
    "Pais_origen":      ["País de origen"],
    "Acabado":          ["Acabado"],
    "Cod_universal":    ["Código universal de producto"],
    "Condicion":        ["Condición del ítem"],
}
MAXFOTOS = 8


def dims(attrs):
    if attrs.get("Dimensiones"):
        return attrs["Dimensiones"]
    # medidas de producto
    prod = [(a, attrs[a]) for a in ["Largo", "Ancho", "Altura", "Profundidad", "Diámetro"] if attrs.get(a)]
    if prod:
        return " x ".join(f"{n}: {v}" for n, v in prod)
    # respaldo: medidas del paquete (legibles)
    L = attrs.get("Largo del paquete") or attrs.get("Largo del paquete del seller")
    A = attrs.get("Ancho del paquete") or attrs.get("Ancho del paquete del seller")
    H = attrs.get("Altura del paquete") or attrs.get("Altura del paquete del seller")
    partes = [x for x in (L, A, H) if x]
    if partes:
        return "Paquete " + " x ".join(p.replace(" cm", "") for p in partes) + " cm (largo x ancho x alto)"
    return ""


def main():
    rows = list(csv.reader(open(ROOT / "falabella_atributos.csv", encoding="utf-8-sig")))[1:]
    resumen = json.loads((ROOT / "falabella_resumen.json").read_text(encoding="utf-8"))
    porsku, nombre_de, fala_de = {}, {}, {}
    for sku, fala, nombre, attr, val in rows:
        porsku.setdefault(sku, {})[attr] = val
        nombre_de[sku] = nombre; fala_de[sku] = fala

    headers = (["SKU_Seller", "SKU_Falabella", "Nombre"] + list(CAMPOS.keys())
               + ["N_fotos"] + [f"Foto_{i}" for i in range(1, MAXFOTOS + 1)] + ["Carpeta_fotos"])

    wb = Workbook(); ws = wb.active; ws.title = "Carga Falabella"
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = "D2"; ws.row_dimensions[1].height = 28

    csv_rows = [headers]
    for item in resumen:
        sku = item["sku_seller"]; attrs = porsku.get(sku, {})
        fila = [sku, fala_de.get(sku, ""), nombre_de.get(sku, "")]
        for campo, alias in CAMPOS.items():
            if campo == "Dimensiones":
                fila.append(dims(attrs))
            else:
                v = ""
                for a in alias:
                    if attrs.get(a):
                        v = attrs[a]; break
                fila.append(v)
        fotos = sorted((LISTO / sku).glob("*.jpg")) if (LISTO / sku).exists() else []
        fila.append(len(fotos))
        for i in range(MAXFOTOS):
            fila.append(fotos[i].name if i < len(fotos) else "")
        fila.append(f"falabella_fotos_listo/{sku}/")
        ws.append(fila); csv_rows.append(fila)

    # anchos
    widths = {"A": 14, "B": 13, "C": 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    from openpyxl.utils import get_column_letter
    for i, campo in enumerate(CAMPOS, start=4):
        ws.column_dimensions[get_column_letter(i)].width = 18
    fcol0 = 3 + len(CAMPOS) + 2  # primera Foto_
    for i in range(MAXFOTOS):
        ws.column_dimensions[get_column_letter(fcol0 + i)].width = 20
    ws.column_dimensions[get_column_letter(fcol0 + MAXFOTOS)].width = 28
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(ROOT / "falabella_CARGA.xlsx")
    with open(ROOT / "falabella_CARGA.csv", "w", newline="", encoding="utf-8-sig") as fh:
        csv.writer(fh).writerows(csv_rows)

    # ZIP de fotos
    nz = 0
    with zipfile.ZipFile(ROOT / "falabella_fotos.zip", "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(LISTO.rglob("*.jpg")):
            z.write(f, f"{f.parent.name}/{f.name}"); nz += 1

    print(f"OK: falabella_CARGA.xlsx / .csv ({len(resumen)} productos)  |  falabella_fotos.zip ({nz} fotos)")


if __name__ == "__main__":
    main()
