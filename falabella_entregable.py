# -*- coding: utf-8 -*-
"""
falabella_entregable.py
=======================
Construye falabella_entregable.xlsx: UN archivo con las FOTOS embebidas + los
ATRIBUTOS de los 29 productos. Una fila por SKU con hasta 5 miniaturas y los
campos minimos exigidos por Falabella. Pensado para que Victor revise todo de un
vistazo antes de cargar.

Hojas:
  - Productos : fila por SKU con fotos embebidas + atributos minimos
  - Atributos : formato largo completo (SKU_Seller|SKU_Falabella|Nombre|Atributo|Valor)
"""
import csv, json, os
from pathlib import Path
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT = Path(__file__).parent
LISTO = ROOT / "falabella_fotos_listo"
THUMBS = ROOT / "_thumbs_entregable"
THUMBS.mkdir(exist_ok=True)
TH = 88          # px miniatura
NFOTOS = 8       # columnas de foto (todas)

REQUERIDOS = {
    "Material":        ["Material", "Materiales"],
    "Dimensiones":     ["Largo", "Ancho", "Altura", "Profundidad", "Diámetro"],
    "Color":           ["Color principal", "Color"],
    "Marca":           ["Marca"],
    "Peso":            ["Peso del paquete", "Peso del paquete del seller"],
    "Tipo instalacion":["Tipo de instalación", "Tipos de instalación"],
    "Garantia":        ["Garantía"],
    "Pais origen":     ["País de origen", "Origen", "Fabricante"],
}


def thumb(src, dst):
    im = PILImage.open(src).convert("RGB")
    im.thumbnail((TH, TH))
    canvas = PILImage.new("RGB", (TH, TH), (255, 255, 255))
    canvas.paste(im, ((TH - im.width) // 2, (TH - im.height) // 2))
    canvas.save(dst, quality=85)


def main():
    rows = list(csv.reader(open(ROOT / "falabella_atributos.csv", encoding="utf-8-sig")))[1:]
    resumen = json.loads((ROOT / "falabella_resumen.json").read_text(encoding="utf-8"))

    porsku, nombre_de, fala_de = {}, {}, {}
    for sku, fala, nombre, attr, val in rows:
        porsku.setdefault(sku, {}).setdefault(attr, val)
        nombre_de[sku] = nombre
        fala_de[sku] = fala

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    gap_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Productos"
    cols = (["SKU_Seller", "SKU_Falabella", "Nombre", "Grupo", "N_fotos"]
            + [f"Foto{i}" for i in range(1, NFOTOS + 1)]
            + list(REQUERIDOS.keys()) + ["Faltantes"])
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 26

    # anchos
    widths = {"A": 14, "B": 13, "C": 34, "D": 11, "E": 7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    foto_cols = [get_column_letter(6 + i) for i in range(NFOTOS)]
    for col in foto_cols:
        ws.column_dimensions[col].width = 13
    attr_start = 6 + NFOTOS
    for i in range(len(REQUERIDOS)):
        ws.column_dimensions[get_column_letter(attr_start + i)].width = 16
    ws.column_dimensions[get_column_letter(attr_start + len(REQUERIDOS))].width = 26

    r = 2
    for item in resumen:
        sku = item["sku_seller"]
        attrs = porsku.get(sku, {})
        fotos = sorted((LISTO / sku).glob("*.jpg")) if (LISTO / sku).exists() else []
        ws.cell(r, 1, sku)
        ws.cell(r, 2, fala_de.get(sku, ""))
        ws.cell(r, 3, nombre_de.get(sku, ""))
        ws.cell(r, 4, item["grupo"])
        ws.cell(r, 5, len(fotos))
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 70

        for i in range(NFOTOS):
            if i < len(fotos):
                tp = THUMBS / f"{sku}_{i+1}.jpg"
                thumb(fotos[i], tp)
                xim = XLImage(str(tp))
                anchor = f"{foto_cols[i]}{r}"
                ws.add_image(xim, anchor)

        vacios = []
        for j, (campo, alias) in enumerate(REQUERIDOS.items()):
            val = ""
            if campo == "Dimensiones":
                if attrs.get("Dimensiones"):      # valor explicito tiene prioridad
                    val = attrs["Dimensiones"]
                else:
                    partes = [f"{a[:3]}:{attrs[a]}" for a in alias if attrs.get(a)]
                    if not partes:   # respaldo: medidas del paquete
                        for a in ["Largo del paquete", "Ancho del paquete", "Altura del paquete"]:
                            if attrs.get(a):
                                partes.append(f"{a.split()[0][:3]}pq:{attrs[a]}")
                    val = " ".join(partes)
            else:
                for a in alias:
                    if attrs.get(a):
                        val = attrs[a]; break
            cell = ws.cell(r, attr_start + j, val)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if not val:
                cell.fill = gap_fill
                vacios.append(campo)
        ws.cell(r, attr_start + len(REQUERIDOS), "; ".join(vacios))
        for c in range(1, len(cols) + 1):
            ws.cell(r, c).border = border
        r += 1

    # --- hoja Atributos (largo)
    ws2 = wb.create_sheet("Atributos")
    ws2.append(["SKU_Seller", "SKU_Falabella", "Nombre", "Atributo", "Valor"])
    for c in range(1, 6):
        ws2.cell(1, c).fill = hdr_fill; ws2.cell(1, c).font = hdr_font
    ws2.freeze_panes = "A2"
    for row in rows:
        ws2.append(row)
    for col, w in zip("ABCDE", (15, 14, 40, 28, 42)):
        ws2.column_dimensions[col].width = w

    import os as _os
    out = ROOT / "falabella_entregable.xlsx"
    if (ROOT / ".~lock.falabella_entregable.xlsx#").exists():
        out = ROOT / "falabella_entregable_ACTUALIZADO.xlsx"
    wb.save(out)
    print("OK ->", out.name, "  (", r - 2, "productos )")


if __name__ == "__main__":
    main()
