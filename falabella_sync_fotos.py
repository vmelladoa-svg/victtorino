# -*- coding: utf-8 -*-
"""
falabella_sync_fotos.py
=======================
Sincroniza las carpetas de fotos con el Excel entregable EDITADO por Victor:
si borro una miniatura de la hoja "Productos", se elimina esa foto de
falabella_fotos_listo/<SKU>/ y se renumeran las que quedan. Luego regenera los
Excel.

Uso:
  python falabella_sync_fotos.py                      # lee falabella_entregable_ACTUALIZADO.xlsx
  python falabella_sync_fotos.py mi_archivo.xlsx      # lee otro nombre

La columna Foto1..Foto8 corresponde 1:1 al orden de fotos del SKU. Una miniatura
borrada = esa foto se elimina.
"""
import sys, os, glob
from pathlib import Path
import openpyxl
ROOT = Path(__file__).parent
LISTO = ROOT / "falabella_fotos_listo"

def main():
    xlsx = ROOT / (sys.argv[1] if len(sys.argv) > 1 else "falabella_entregable_ACTUALIZADO.xlsx")
    if not xlsx.exists():
        print("No existe", xlsx); return
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Productos"]
    hdr = [c.value for c in ws[1]]
    col_foto1 = hdr.index("Foto1") + 1   # 1-indexed
    sku_de_fila = {r: ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

    # slots sobrevivientes por SKU
    vivos = {}
    for im in ws._images:
        a = im.anchor._from
        fila = a.row + 1
        slot = (a.col + 1) - col_foto1 + 1   # Foto1 -> slot 1
        sku = sku_de_fila.get(fila)
        if sku and slot >= 1:
            vivos.setdefault(sku, set()).add(slot)

    total_borradas = 0
    for sku in [sku_de_fila[r] for r in range(2, ws.max_row + 1)]:
        d = LISTO / sku
        if not d.exists():
            continue
        fotos = sorted(d.glob("*.jpg"))
        keep_slots = vivos.get(sku, set())
        keep = [f for i, f in enumerate(fotos, 1) if i in keep_slots]
        borrar = [f for i, f in enumerate(fotos, 1) if i not in keep_slots]
        if not borrar:
            continue
        for f in borrar:
            f.unlink(); total_borradas += 1
        # renumerar secuencial
        tmp = []
        for i, f in enumerate(sorted(keep), 1):
            nuevo = d / f"{sku}_{i:02d}.jpg"
            if f != nuevo:
                f.rename(d / f"_tmp_{i:02d}.jpg"); tmp.append((d / f"_tmp_{i:02d}.jpg", nuevo))
        for old, new in tmp:
            old.rename(new)
        print(f"{sku}: borradas {len(borrar)}, quedan {len(keep)}")

    print(f"\nTotal fotos borradas: {total_borradas}")
    # regenerar entregables
    import importlib, finalize_falabella, falabella_entregable
    importlib.reload(finalize_falabella).construir_excel()
    importlib.reload(falabella_entregable).main()
    print("Excel regenerados.")

if __name__ == "__main__":
    main()
