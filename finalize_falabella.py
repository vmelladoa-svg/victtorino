# -*- coding: utf-8 -*-
"""
finalize_falabella.py
=====================
1. Normaliza las fotos descargadas a cuadrado 1080x1080 con fondo blanco
   (cumple min 500x500 + fondo blanco que pide Falabella) -> falabella_fotos_listo/<SKU>/
2. Construye falabella_contenido.xlsx con 3 hojas:
   - Atributos : formato largo SKU_Seller|SKU_Falabella|Nombre|Atributo|Valor
   - Minimos   : matriz de los 8 atributos minimos exigidos por Falabella + vacios
   - Resumen   : match ML, fotos disponibles, fotos validas
"""
from __future__ import annotations
import csv, json, re
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "falabella_contenido"
DST = ROOT / "falabella_fotos_listo"
TARGET = 1080

# ---------------------------------------------------------------- 1. fotos
def normaliza_fotos():
    DST.mkdir(exist_ok=True)
    conteo = {}
    for skudir in sorted(SRC.iterdir()):
        if not skudir.is_dir():
            continue
        out = DST / skudir.name
        out.mkdir(parents=True, exist_ok=True)
        n = 0
        for img in sorted(skudir.glob("*.jpg")):
            try:
                im = Image.open(img).convert("RGB")
            except Exception as e:
                print("skip", img, e)
                continue
            w, h = im.size
            lado = max(w, h, 500)            # nunca menos de 500
            lienzo = Image.new("RGB", (lado, lado), (255, 255, 255))
            lienzo.paste(im, ((lado - w) // 2, (lado - h) // 2))
            if lado != TARGET:
                lienzo = lienzo.resize((TARGET, TARGET), Image.LANCZOS)
            n += 1
            lienzo.save(out / f"{skudir.name}_{n:02d}.jpg", quality=90)
        conteo[skudir.name] = n
    return conteo


# ---------------------------------------------------------------- 2. excel
REQUERIDOS = {
    "Material":        ["Material", "Materiales"],
    "Dimensiones":     ["Largo", "Ancho", "Altura", "Profundidad", "Diámetro",
                        "Largo del paquete", "Ancho del paquete", "Altura del paquete"],
    "Color":           ["Color principal", "Color"],
    "Marca":           ["Marca"],
    "Peso":            ["Peso del paquete", "Peso del paquete del seller"],
    "Tipo instalacion":["Tipo de instalación", "Tipos de instalación"],
    "Garantia":        ["Garantía"],
    "Pais origen":     ["País de origen", "Origen", "Fabricante"],
}


def construir_excel():
    rows = list(csv.reader(open(ROOT / "falabella_atributos.csv", encoding="utf-8-sig")))[1:]
    resumen = json.loads((ROOT / "falabella_resumen.json").read_text(encoding="utf-8"))
    listo = {d.name: len(list(d.glob("*.jpg"))) for d in DST.iterdir() if d.is_dir()}

    # agrupa atributos por sku
    porsku = {}
    nombre_de = {}
    fala_de = {}
    for sku, fala, nombre, attr, val in rows:
        porsku.setdefault(sku, {}).setdefault(attr, val)
        nombre_de[sku] = nombre
        fala_de[sku] = fala

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    gap_fill = PatternFill("solid", fgColor="FFC7CE")

    def estilo_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    # --- Hoja Atributos (largo)
    ws = wb.active
    ws.title = "Atributos"
    ws.append(["SKU_Seller", "SKU_Falabella", "Nombre", "Atributo", "Valor"])
    for sku, fala, nombre, attr, val in rows:
        ws.append([sku, fala, nombre, attr, val])
    estilo_header(ws, 5)
    for col, w in zip("ABCDE", (15, 14, 42, 30, 45)):
        ws.column_dimensions[col].width = w

    # --- Hoja Minimos (matriz)
    ws2 = wb.create_sheet("Minimos")
    cols = ["SKU_Seller", "SKU_Falabella", "Nombre"] + list(REQUERIDOS.keys()) + ["Vacios"]
    ws2.append(cols)
    for sku in [r["sku_seller"] for r in resumen]:
        attrs = porsku.get(sku, {})
        fila = [sku, fala_de.get(sku, ""), nombre_de.get(sku, "")]
        vacios = []
        valores_celda = []
        for campo, alias in REQUERIDOS.items():
            encontrado = ""
            if campo == "Dimensiones":
                if attrs.get("Dimensiones"):
                    encontrado = attrs["Dimensiones"]
                else:
                    partes = []
                    for a in ["Largo", "Ancho", "Altura", "Profundidad", "Diámetro"]:
                        if attrs.get(a):
                            partes.append(f"{a[:3]}:{attrs[a]}")
                    if not partes:
                        for a in ["Largo del paquete", "Ancho del paquete", "Altura del paquete"]:
                            if attrs.get(a):
                                partes.append(f"{a.split()[0][:3]}pq:{attrs[a]}")
                    encontrado = " ".join(partes)
            else:
                for a in alias:
                    if attrs.get(a):
                        encontrado = attrs[a]
                        break
            valores_celda.append(encontrado)
            if not encontrado:
                vacios.append(campo)
        fila += valores_celda + ["; ".join(vacios)]
        ws2.append(fila)
        # pintar celdas vacias
        r = ws2.max_row
        for idx, campo in enumerate(REQUERIDOS.keys()):
            if not valores_celda[idx]:
                ws2.cell(r, 4 + idx).fill = gap_fill
    estilo_header(ws2, len(cols))
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 42
    for i in range(len(REQUERIDOS)):
        ws2.column_dimensions[get_column_letter(4 + i)].width = 18
    ws2.column_dimensions[get_column_letter(4 + len(REQUERIDOS))].width = 30

    # --- Hoja Resumen
    ws3 = wb.create_sheet("Resumen")
    ws3.append(["SKU_Seller", "SKU_Falabella", "Nombre", "Grupo", "ML_item",
                "Cuenta", "Fotos_ML", "Fotos_listas", "Atributos_ML"])
    for r in resumen:
        sku = r["sku_seller"]
        ws3.append([sku, r["sku_fala"], r["nombre"], r["grupo"], r["ml_item"],
                    r["cuenta"], r["fotos"], listo.get(sku, 0), r["attrs"]])
    estilo_header(ws3, 9)
    for col, w in zip("ABCDEFGHI", (15, 14, 42, 12, 14, 8, 9, 11, 12)):
        ws3.column_dimensions[col].width = w

    wb.save(ROOT / "falabella_contenido.xlsx")
    print("Excel: falabella_contenido.xlsx")


if __name__ == "__main__":
    c = normaliza_fotos()
    print("Fotos normalizadas por SKU:", sum(c.values()), "imgs en", len(c), "carpetas")
    construir_excel()
