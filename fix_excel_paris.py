"""
fix_excel_paris.py
------------------
Aplica los fixes detectados por el auditor sobre paris_productos_carga_masiva.xlsx:

  * MLC2055571318 (fila 30): remueve la imagen 481x499 (lado max < 500 px).

  * MLC2046010486 (fila 36): remueve la imagen 370x448, y agrega 2 imagenes
    nuevas extraidas del item MLC3724904324, ambas validadas JPG >= 500 px.

  * MLC2068899311 (fila 39): el catalogo ML solo expone GIF placeholders.
    Se reemplazan las 3 URLs (que servian 'image/gif' pese a la extension
    .jpg) por las 5 imagenes reales del item MLC3724867792, todas validadas
    JPG con max-side >= 500 px. Ademas se completan los campos Marca/Modelo
    /Color/Material con los valores que el catalogo ML expone hoy
    (Marca: 'Taumm', Modelo: 'Valvula Lateral', Color: 'Blanco', Material:
    'Plastico'); el campo Marca queda asi cumpliendo el obligatorio (*).
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
EXCEL = BASE / "paris_productos_carga_masiva.xlsx"
SHEET = "herramientas"

COL_SKU = 2
COL_MARCA = 6
COL_MODELO = 10
COL_IMAGENES = 13
COL_COLOR = 16
COL_DESC = 33
COL_MATERIAL = 35

REMOVE_FROM = {
    "MLC2055571318": [
        "https://http2.mlstatic.com/D_922761-MLC53437349501_012023-F.jpg",
    ],
    "MLC2046010486": [
        "https://http2.mlstatic.com/D_956819-MLC51110704354_082022-F.jpg",
    ],
    "MLC2068899311": [
        "https://http2.mlstatic.com/D_943530-MLC108380842847_032026-F.jpg",
        "https://http2.mlstatic.com/D_942334-MLC108382214205_032026-F.jpg",
        "https://http2.mlstatic.com/D_978373-MLC108381439917_032026-F.jpg",
    ],
}

ADD_TO = {
    "MLC2046010486": [
        "https://http2.mlstatic.com/D_704045-MLC107481345592_032026-F.jpg",
        "https://http2.mlstatic.com/D_607135-MLC107481375470_032026-F.jpg",
    ],
    "MLC2068899311": [
        "https://http2.mlstatic.com/D_866374-MLC108817551572_032026-F.jpg",
        "https://http2.mlstatic.com/D_712060-MLC109638492239_032026-F.jpg",
        "https://http2.mlstatic.com/D_815444-MLC108817195608_032026-F.jpg",
        "https://http2.mlstatic.com/D_849397-MLC109638372983_032026-F.jpg",
        "https://http2.mlstatic.com/D_834861-MLC108817225204_032026-F.jpg",
    ],
}

# Campos no-imagen a setear (origen: catalogo ML actual)
SET_FIELDS = {
    "MLC2068899311": {
        COL_MARCA: "Täumm",
        COL_MODELO: "Valvula Lateral",
        COL_COLOR: "Blanco",
        COL_MATERIAL: "Plástico",
    },
}


def find_row(ws, sku: str) -> int | None:
    for row in range(7, ws.max_row + 1):
        if ws.cell(row=row, column=COL_SKU).value == sku:
            return row
    return None


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]

    skus = sorted(set(REMOVE_FROM) | set(ADD_TO) | set(SET_FIELDS))
    for sku in skus:
        row = find_row(ws, sku)
        if row is None:
            print(f"[WARN] {sku}: no encontrado en Excel")
            continue
        before = ws.cell(row=row, column=COL_IMAGENES).value or ""
        urls = [u.strip() for u in before.split(",") if u.strip()]
        n_before = len(urls)

        remove = set(REMOVE_FROM.get(sku, []))
        urls = [u for u in urls if u not in remove]
        for u in ADD_TO.get(sku, []):
            if u not in urls:
                urls.append(u)

        ws.cell(row=row, column=COL_IMAGENES, value=", ".join(urls))
        print(f"  {sku} (fila {row}): imagenes {n_before} -> {len(urls)} "
              f"(removidas={len(remove)}, agregadas={len(ADD_TO.get(sku, []))})")

        for col, val in SET_FIELDS.get(sku, {}).items():
            prev = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col, value=val)
            print(f"    col {col}: {prev!r} -> {val!r}")

    wb.save(EXCEL)
    print(f"\nGuardado: {EXCEL}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
