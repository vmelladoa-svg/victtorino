# -*- coding: utf-8 -*-
"""Genera la planilla de carga masiva de precios Falabella (tachado + oferta)
para TODOS los productos que calzan con la lista maestra de precios.
  PriceFalabella       = tachado del nivel elegido (normal/referencia)
  SalePriceFalabella   = VENTA real (lo que paga el cliente)
  Sale dates           = vigencia (hoy -> +1 mes)
Solo incluye SKUs que calzan por codigo completo con la lista maestra; los demas
NO se tocan. Lee precios desde precios_descuentos_falabella.xlsx.
"""
import re
import openpyxl

MAESTRA  = r"C:\Users\dell\victtorino\precios_descuentos_falabella.xlsx"
TEMPLATE = r"C:\Users\dell\victtorino\SellerPriceTemplate.xlsx"
OUT      = r"C:\Users\dell\victtorino\SellerPriceTemplate_UPLOAD_nuevo.xlsx"

NIVEL_COL = "Tachado 50%"          # columna de la maestra a usar como precio normal
SALE_START = "2026-06-16 00:00:01"
SALE_END   = "2026-07-16 23:59:59"

# Trabajamos POR PARTES y SOLO CON STOCK.
# BATCH: lista de SellerSku a incluir en esta parte (vacio = todos los que calzan).
BATCH = {"040203017-C", "040201006-T", "040202005-T"}   # Parte 3: griferias margen sano
# SKUs sin stock a excluir siempre (agotados conocidos / verificados):
SIN_STOCK = set()

# Cargar maestra
mw = openpyxl.load_workbook(MAESTRA, data_only=True).active
mrows = list(mw.iter_rows(values_only=True))
H = list(mrows[0])
i_venta = H.index("VENTA real (oferta)")
i_tach  = H.index(NIVEL_COL)
master = {}
for r in mrows[1:]:
    if isinstance(r[i_venta], int) and isinstance(r[i_tach], int):
        master[r[0]] = (r[i_tach], r[i_venta], r[1])

def norm(sku):
    return re.sub(r"_\d+$", "", str(sku))

# Recorrer template y construir upload solo con los que calzan
tw = openpyxl.load_workbook(TEMPLATE, data_only=True).active
trows = list(tw.iter_rows(values_only=True))
hdr = list(trows[0])

wb = openpyxl.Workbook(); w = wb.active; w.title = "Sheet"
w.append(hdr)

incl = 0; skip = 0; anomalias = []
for r in trows[1:]:
    sku_orig = r[0]
    sku = norm(sku_orig)
    if sku not in master:
        skip += 1
        continue
    if BATCH and sku not in BATCH:
        skip += 1
        continue
    if sku in SIN_STOCK:
        skip += 1
        continue
    tach, venta, desc = master[sku]
    cur = float(r[2]) if r[2] else 0
    # anomalia: la nueva venta sube respecto al precio normal actual del template
    if cur and venta > cur:
        anomalias.append((sku_orig, int(cur), venta, desc[:30]))
    w.append([sku_orig, r[1], float(tach), float(venta), SALE_START, SALE_END, r[6]])
    incl += 1

wb.save(OUT)
print(f"Generado: {OUT}")
print(f"Filas incluidas (calzan): {incl} | omitidas (no tocar): {skip}")
print(f"Nivel: {NIVEL_COL} | vigencia {SALE_START} -> {SALE_END}")
print(f"\nAnomalias (venta nueva > precio normal actual del template): {len(anomalias)}")
for a in anomalias[:20]:
    print(f"  {a[0]:16} actual {a[1]:>8,} -> venta {a[2]:>8,}  {a[3]}")
