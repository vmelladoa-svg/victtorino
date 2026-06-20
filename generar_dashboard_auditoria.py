"""
Dashboard Excel multi-hoja a partir de analisis.pkl.
Hojas:
  1. Resumen ejecutivo (KPIs por cuenta y comparativa)
  2. Publicaciones (todas) — con filtros, ranking, score
  3. Top sellers 180d (por cuenta)
  4. Críticas — alto/medio severidad
  5. Quick wins — items con potencial alto y esfuerzo bajo
  6. Cross-Defontana — items con SKU cruzado
  7. Plan de acción priorizado
"""
import pickle
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

ROOT = Path(__file__).parent
ANALISIS = ROOT / "data" / "auditoria" / "analisis.pkl"
TODAY = date.today().isoformat()
OUT = ROOT / f"auditoria_ml_{TODAY}.xlsx"

# Estilos
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True, color="1F4E78")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEV_COLORS = {
    "CRITICO": "FFC7CE",
    "ALTO": "FFD68A",
    "MEDIO": "FFEB9C",
    "BAJO": "C6EFCE",
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border = BORDER


def autosize(ws, max_w=50):
    # Recolectar anchos por índice de columna evitando MergedCell
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if hasattr(c, "column_letter") and c.value is not None:
                widths[c.column_letter] = max(widths.get(c.column_letter, 0), len(str(c.value)))
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = min(w + 2, max_w)


def write_df(ws, df, start_row=1, freeze=True):
    ws.append(list(df.columns))
    style_header(ws, start_row)
    for r in df.itertuples(index=False):
        row = [v if not isinstance(v, (list, tuple, dict)) else str(v) for v in r]
        ws.append(row)
    if freeze:
        ws.freeze_panes = "A2"
    autosize(ws)


def hoja_resumen(wb, df_res):
    ws = wb.create_sheet("1. Resumen Ejecutivo")
    ws["A1"] = "AUDITORÍA MERCADOLIBRE — 3 CUENTAS"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws["A2"] = f"Fecha: {TODAY}    |    Activas totales: {int(df_res['Activas'].sum())}    |    GMV 180d: ${int(df_res['GMV180d'].sum()):,} CLP"
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    # Pivot: KPI vs cuenta
    pivot_cols = [
        ("Activas","Activas"),
        ("Visitas 30d","Visitas30d"),
        ("Items con tráfico","ItemsConTrafico"),
        ("Órdenes 180d","Ordenes180d"),
        ("GMV 180d CLP","GMV180d"),
        ("Ticket promedio","TicketProm"),
        ("Conv global %","ConvGlobal%"),
        ("Health avg /100","HealthAvg"),
        ("Listings Gold Pro","ListingsGoldPro"),
        ("Listings Gold Special","ListingsGoldSpecial"),
        ("Listings Free","ListingsFree"),
        ("Fotos <4","FotosMenos4"),
        ("Títulos <40 chars","TitulosCortos<40"),
        ("Sin stock","SinStock"),
        ("Stock crítico (1-2)","StockCritico<3"),
        ("Reputación","RepNivel"),
        ("Power Seller","PowerSeller"),
        ("Tx completadas","TxCompletadas"),
        ("Tx canceladas","TxCanceladas"),
        ("Preguntas resp 180d","PreguntasResp180d"),
        ("Preguntas pendientes","PreguntasPendientes"),
        ("Claims rate %","MetricaClaimsRate%"),
        ("Cancel rate %","MetricaCancRate%"),
        ("Delayed handling %","MetricaDelayedRate%"),
    ]
    header = ["KPI"] + df_res["Cuenta"].tolist() + ["Total / Comparativa"]
    ws.append([])
    ws.append(header)
    style_header(ws, ws.max_row)
    for label, col in pivot_cols:
        vals = df_res[col].tolist()
        total = ""
        try:
            nums = [v for v in vals if isinstance(v,(int,float))]
            if len(nums)==len(vals):
                if "%" in label or "Health" in label:
                    total = round(sum(nums)/len(nums),2)
                else:
                    total = sum(nums)
        except Exception:
            total = ""
        ws.append([label] + vals + [total])
        last = ws.max_row
        ws.cell(last,1).font = SUB_FONT
        ws.cell(last,1).fill = SUB_FILL
        # Resaltar liderato
        if all(isinstance(v,(int,float)) for v in vals):
            mx = max(vals); mn = min(vals)
            for i, v in enumerate(vals):
                cell = ws.cell(last, 2+i)
                # invertir lógica para métricas donde menos es mejor
                bad_when_high = label in ("Claims rate %","Cancel rate %","Delayed handling %",
                                          "Tx canceladas","Sin stock","Stock crítico (1-2)",
                                          "Preguntas pendientes","Fotos <4","Títulos <40 chars",
                                          "Listings Free")
                if mx != mn:
                    if (v == mx and not bad_when_high) or (v == mn and bad_when_high):
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    if (v == mn and not bad_when_high) or (v == mx and bad_when_high):
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
    autosize(ws, 30)
    ws.freeze_panes = "B5"


def hoja_publicaciones(wb, df):
    ws = wb.create_sheet("2. Publicaciones (todas)")
    cols = ["Cuenta","ItemID","SKU","Título","Categoría","Precio","Stock","ListingType",
            "Fotos","Visitas30d","Vendidos180d","Revenue180d","ConvRate%","HealthCalc",
            "Severidad","NumIssues","Issues","Acciones","Permalink"]
    write_df(ws, df[cols].sort_values(["Cuenta","Severidad","Visitas30d"], ascending=[True,False,False]))
    # Color scales: HealthCalc, Visitas, Severidad
    n = ws.max_row
    rule_health = ColorScaleRule(start_type="num", start_value=20, start_color="F8696B",
                                 mid_type="num", mid_value=60, mid_color="FFEB84",
                                 end_type="num", end_value=100, end_color="63BE7B")
    ws.conditional_formatting.add(f"N2:N{n}", rule_health)
    rule_sev = ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                              mid_type="num", mid_value=5, mid_color="FFEB84",
                              end_type="num", end_value=12, end_color="F8696B")
    ws.conditional_formatting.add(f"O2:O{n}", rule_sev)


def hoja_top_sellers(wb, df):
    ws = wb.create_sheet("3. Top sellers 180d")
    top = (df[df["Vendidos180d"]>0]
           .sort_values(["Cuenta","Vendidos180d"], ascending=[True,False])
           .groupby("Cuenta", group_keys=False)
           .head(30))
    write_df(ws, top[["Cuenta","ItemID","Título","Precio","Stock","Visitas30d","Vendidos180d",
                      "Revenue180d","ConvRate%","HealthCalc","Permalink"]])


def hoja_criticas(wb, df):
    ws = wb.create_sheet("4. Críticas (acción urgente)")
    crit = df[df["Issues"].str.contains("CRITICO", na=False)].copy()
    crit = crit.sort_values(["Cuenta","Severidad","Visitas30d"], ascending=[True,False,False])
    write_df(ws, crit[["Cuenta","ItemID","Título","Stock","Visitas30d","Vendidos180d",
                       "ConvRate%","HealthCalc","Severidad","Issues","Acciones","Permalink"]])


def hoja_quick_wins(wb, df):
    ws = wb.create_sheet("5. Quick wins")
    """
    Quick win = item con tráfico (>=10 visitas), score <70, vendido 180d >0, accion barata
    (fotos, título, descripción).
    """
    qw = df[(df["Visitas30d"]>=10) & (df["HealthCalc"]<70) & (df["Vendidos180d"]>0)].copy()
    qw["Potencial+50%Conv"] = (qw["Vendidos180d"] * 0.5 * (qw["Revenue180d"]/qw["Vendidos180d"].replace(0,1))).round(0)
    write_df(ws, qw[["Cuenta","ItemID","Título","Visitas30d","Vendidos180d","ConvRate%",
                     "HealthCalc","Revenue180d","Potencial+50%Conv","Issues","Acciones","Permalink"]]
             .sort_values("Potencial+50%Conv", ascending=False))


def hoja_traffic_no_sales(wb, df):
    ws = wb.create_sheet("6. Tráfico sin venta")
    """Items con >=20 visitas y 0 ventas: arreglar precio/oferta/copy."""
    tns = df[(df["Visitas30d"]>=20) & (df["Vendidos180d"]==0)].copy()
    write_df(ws, tns[["Cuenta","ItemID","Título","Precio","Stock","Visitas30d","Fotos",
                      "HealthCalc","Issues","Permalink"]]
             .sort_values("Visitas30d", ascending=False))


def hoja_categorias(wb, df):
    ws = wb.create_sheet("7. Categorías top")
    cat = (df.groupby(["Cuenta","Categoría"])
              .agg(Activas=("ItemID","count"),
                   Visitas=("Visitas30d","sum"),
                   Vendidos=("Vendidos180d","sum"),
                   Revenue=("Revenue180d","sum"),
                   HealthMean=("HealthCalc","mean"))
              .reset_index()
              .sort_values(["Cuenta","Revenue"], ascending=[True,False]))
    cat["HealthMean"] = cat["HealthMean"].round(1)
    write_df(ws, cat.groupby("Cuenta", group_keys=False).head(15))


def hoja_defontana(wb, df):
    ws = wb.create_sheet("8. Cross Defontana")
    cross = df[df["Margen_%"].notna()].copy()
    if cross.empty:
        ws["A1"] = "No hay coincidencias SKU ↔ ML (los SKU en publicaciones ML no coinciden con muestra Defontana)"
        return
    write_df(ws, cross[["Cuenta","ItemID","SKU","Título","Precio","Precio_Venta","Costo",
                        "Margen_%","Comisión_ML_%","Margen_Neto_ML","Vendidos180d","MargenTotal180d",
                        "Stock","Stock_Actual","Permalink"]])


def hoja_plan(wb, df_res, df_pub):
    ws = wb.create_sheet("9. Plan de Acción Priorizado")
    activas = int(df_res["Activas"].sum())
    gmv = int(df_res["GMV180d"].sum())
    sin_stock = int(df_res["SinStock"].sum())
    pocas_fotos = int(df_res["FotosMenos4"].sum())
    sin_visitas = int((df_pub["Visitas30d"]==0).sum())
    alto_traf_sin_venta = int(((df_pub["Visitas30d"]>=50) & (df_pub["Vendidos180d"]==0)).sum())
    preg_pend = int(df_res["PreguntasPendientes"].sum())
    items_free = int(df_res["ListingsFree"].sum())

    rows = [
        ("Bucket","Prioridad","Problema","Acción concreta","Items afectados","Plazo","Impacto estimado"),
        ("QUICK WIN","CRÍTICO","Sin stock activas",
         "Pausar publicaciones con stock=0 o reabastecer ya",
         sin_stock,"1-3 días","Evita ventas frustradas, mejora reputación"),
        ("QUICK WIN","ALTO","Pocas fotos (<4)",
         "Subir hasta 6+ fotos en items top de tráfico — empezar por C1/C2 que están más bajos",
         pocas_fotos,"5-7 días","+10-20% conversión esperada en mejorados"),
        ("QUICK WIN","ALTO","Preguntas pendientes",
         "Responder pendientes hoy + activar bot para preguntas frecuentes (FAQ)",
         preg_pend,"1 día","Evita pérdida de reputación + ventas inmediatas"),
        ("QUICK WIN","ALTO","Listings Free",
         "Migrar a Gold Pro/Special — al menos los que tienen tráfico",
         items_free,"3-7 días","+50-100% visibilidad para items migrados"),
        ("QUICK WIN","CRÍTICO","Tráfico sin venta (>=50 visitas, 0 ventas 180d)",
         "Revisar precio vs competencia, calidad de fotos, descripción, oferta",
         alto_traf_sin_venta,"1-2 semanas","Recupera 10-30% como ventas reales"),
        ("TÁCTICO","ALTO","Sin visitas 30d",
         "Optimizar SEO (título palabra clave), categorías, atributos, fotos primarias",
         sin_visitas,"2-4 semanas","Activa tráfico orgánico ML"),
        ("TÁCTICO","ALTO","Replicar C3 → C1 y C2",
         "C3 tiene 5.4% conv global vs ~2% en C1/C2. Replicar sus fichas y atributos en C1/C2",
         "varias","3-4 semanas","Igualar conversión ~5% en otras cuentas"),
        ("TÁCTICO","ALTO","Activar Mercado Ads",
         "C3 tiene $10k presupuesto pero 0 campañas. Activar campaña inicial con top 10 vendedores",
         "10 SKUs","1 semana","ROAS típico 5-8x en categorías baño/cocina"),
        ("TÁCTICO","MEDIO","Health ML <60 (~80% del catálogo)",
         "Completar atributos pendientes y descripciones (script actualizar_atributos.py)",
         f"{(df_pub['HealthCalc']<60).sum()} items","2-4 semanas","Sube health, mejora ranking interno ML"),
        ("ESTRATÉGICO","ALTO","Catálogo equivalente entre cuentas",
         "Asegurar que C1, C2 y C3 publiquen el mismo top de bestsellers (lavaplatos, llaves, accesorios)",
         "varias","1-2 meses","Maximiza GMV total"),
        ("ESTRATÉGICO","CRÍTICO","Reputación procesing_time C1",
         "Cambiar tiempo de procesamiento en Seller Center (3 métricas sobre límite)",
         "C1","2 semanas","Sale de bandera amarilla / sube nivel"),
        ("ESTRATÉGICO","MEDIO","Pipeline KPIs semanal",
         "Cron del audit + dashboard live (mismo proceso este snapshot ejecutado weekly)",
         "Infra","1 mes","Toma de decisiones data-driven"),
        ("ESTRATÉGICO","MEDIO","Video por publicación",
         "Grabar 1 video por SKU bestseller — +10-15% conversión típica en ML",
         "Top 30 SKUs","2 meses","Diferenciación vs competencia"),
    ]
    for r in rows:
        ws.append(list(r))
    style_header(ws, 1)
    # color severidad
    for row in ws.iter_rows(min_row=2):
        prio = row[1].value
        if prio in SEV_COLORS:
            for c in row:
                c.fill = PatternFill("solid", fgColor=SEV_COLORS[prio])
    autosize(ws, 60)
    ws.column_dimensions["D"].width = 70
    ws.row_dimensions[1].height = 35
    for r in range(2, ws.max_row+1):
        ws.row_dimensions[r].height = 30


def hoja_kpis_objetivo(wb, df_res):
    ws = wb.create_sheet("10. KPIs Objetivo 90 días")
    base = {r["Cuenta"]: r for _, r in df_res.iterrows()}
    rows = [["Cuenta","KPI","Actual","Meta 30d","Meta 90d","Lift estimado"]]
    for c in ("C1","C2","C3"):
        r = base[c]
        gmv = r["GMV180d"]
        gmv_90d = gmv/2
        # Metas
        rows += [
            [c, "Health avg",           r["HealthAvg"],         60,                 75,                 f"+{(75-r['HealthAvg'])/r['HealthAvg']*100:.0f}%"],
            [c, "Conv global %",        r["ConvGlobal%"],       max(r["ConvGlobal%"]*1.3, 3.0), 5.0,        "+30 a 50%"],
            [c, "GMV 90d CLP",          int(gmv_90d),           int(gmv_90d*1.15),  int(gmv_90d*1.4),   "+15 a 40%"],
            [c, "Items con tráfico",    r["ItemsConTrafico"],   int(r["Activas"]*0.85), int(r["Activas"]*0.95), "+15 a 25 p.p."],
            [c, "Preguntas pendientes", r["PreguntasPendientes"], 0,                0,                  "0"],
            [c, "Sin stock",            r["SinStock"],          0,                  0,                  "0"],
            [c, "Listings Free",        r["ListingsFree"],      0,                  0,                  "Migrar todos"],
        ]
    for r in rows:
        ws.append(r)
    style_header(ws, 1)
    # alterna color por cuenta
    cuenta_colors = {"C1":"FFF2CC","C2":"DDEBF7","C3":"E2EFDA"}
    for row in ws.iter_rows(min_row=2):
        c = row[0].value
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=cuenta_colors.get(c,"FFFFFF"))
    autosize(ws, 30)


def main():
    data = pickle.loads(ANALISIS.read_bytes())
    df_res = data["df_resumen"]
    df_pub = data["df_publicaciones"]
    wb = Workbook()
    wb.remove(wb.active)
    hoja_resumen(wb, df_res)
    hoja_publicaciones(wb, df_pub)
    hoja_top_sellers(wb, df_pub)
    hoja_criticas(wb, df_pub)
    hoja_quick_wins(wb, df_pub)
    hoja_traffic_no_sales(wb, df_pub)
    hoja_categorias(wb, df_pub)
    hoja_defontana(wb, df_pub)
    hoja_plan(wb, df_res, df_pub)
    hoja_kpis_objetivo(wb, df_res)
    wb.save(OUT)
    print(f"OK Dashboard: {OUT}  ({OUT.stat().st_size//1024} KB)")
    print(f"Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
