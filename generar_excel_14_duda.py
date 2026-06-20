"""
Genera Excel para revisar los 14 productos marcados DUDA en dudosos_revision.xlsx.
Muestra mas info lado a lado para facilitar la decision: fotos, precios, stock,
descripcion corta de ML y de Woo.
"""
import json
import sys
import io
import re
import warnings
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

with open(r"C:\Users\dell\victtorino\tokens_cuenta3.json", encoding="utf-8") as f:
    tk = json.load(f)
ML_HEAD = {"Authorization": f"Bearer {tk['access_token']}"}


def strip_html(s):
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# 1) Leer DUDA del Excel revisado
print("[1/3] Cargando DUDA...")
wb_in = load_workbook(r"C:\Users\dell\victtorino\dudosos_revision.xlsx", data_only=True)
ws_in = wb_in["Dudosos ML vs Web"]
duda = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    dec = (row[0] or "").strip().upper() if row[0] else ""
    if dec != "DUDA":
        continue
    duda.append({
        "ml_id": row[2],
        "ml_titulo": row[3],
        "ml_cat": row[6],
        "woo_id_match": row[9],
        "woo_titulo_match": row[10],
        "score": row[1],
    })
print(f"      {len(duda)} DUDA")

# 2) Enriquecer con ML y Woo
print("[2/3] Trayendo info de ML y Woo...")
for d in duda:
    # ML
    item = requests.get(f"https://api.mercadolibre.com/items/{d['ml_id']}",
                        headers=ML_HEAD, timeout=30).json()
    d["ml_precio"] = item.get("price")
    d["ml_stock"] = item.get("available_quantity")
    d["ml_permalink"] = item.get("permalink", "")
    d["ml_thumbnail"] = (item.get("thumbnail") or "").replace("http://", "https://")
    # Atributos clave (medidas, color, material)
    attrs = item.get("attributes", [])
    attrs_dict = {a.get("id"): a.get("value_name") for a in attrs if a.get("value_name")}
    d["ml_attr"] = "; ".join(f"{k}={v}" for k, v in attrs_dict.items() if k in
                              ("WIDTH", "HEIGHT", "LENGTH", "DEPTH", "COLOR", "MATERIAL", "MODEL", "BRAND"))
    try:
        desc = requests.get(f"https://api.mercadolibre.com/items/{d['ml_id']}/description",
                            headers=ML_HEAD, timeout=20).json()
        d["ml_desc_corta"] = (desc.get("plain_text", "") or "")[:300]
    except Exception:
        d["ml_desc_corta"] = ""

    # Woo
    if d["woo_id_match"]:
        try:
            r = requests.get(f"{WC}/wp-json/wc/v3/products/{d['woo_id_match']}",
                             params=P, timeout=30)
            wp = r.json()
            d["woo_precio"] = wp.get("regular_price") or wp.get("price")
            d["woo_stock"] = wp.get("stock_quantity")
            d["woo_permalink"] = wp.get("permalink")
            d["woo_sku"] = wp.get("sku")
            d["woo_thumbnail"] = wp["images"][0]["src"] if wp.get("images") else ""
            d["woo_desc_corta"] = strip_html(wp.get("description", ""))[:300]
            d["woo_status"] = wp.get("status", "")
        except Exception:
            pass

# 3) Generar Excel
print("[3/3] Generando Excel...")
wb = Workbook()
ws = wb.active
ws.title = "DUDA (14)"

HEADERS = [
    ("DECISION", 14),
    ("Score", 7),
    ("ML Foto", 8),
    ("ML Título", 35),
    ("ML Precio", 10),
    ("ML Stock", 7),
    ("ML Atributos", 28),
    ("ML Desc (300 ch)", 45),
    ("ML Link", 7),
    ("|", 3),
    ("Woo Foto", 8),
    ("Woo Título", 35),
    ("Woo Precio", 10),
    ("Woo Stock", 7),
    ("Woo SKU", 14),
    ("Woo Desc (300 ch)", 45),
    ("Woo Link", 7),
    ("Notas", 25),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 25
ws.freeze_panes = "A2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

for r_idx, d in enumerate(duda, start=2):
    score = d.get("score", 0)
    fill_score = "C6EFCE" if score >= 0.80 else ("FFEB9C" if score >= 0.70 else "FFC7CE")
    # Comparacion visual de precios y stock
    misma_precio = ""
    if d.get("ml_precio") and d.get("woo_precio"):
        try:
            diff = abs(float(d["ml_precio"]) - float(d["woo_precio"]))
            misma_precio = "=" if diff < 100 else f"Δ${int(diff)}"
        except Exception:
            pass

    fila = [
        "",  # DECISION
        score,
        "Foto",
        d["ml_titulo"],
        d.get("ml_precio"),
        d.get("ml_stock"),
        d.get("ml_attr", ""),
        d.get("ml_desc_corta", ""),
        "Ver",
        "",
        "Foto",
        d.get("woo_titulo_match", ""),
        d.get("woo_precio"),
        d.get("woo_stock"),
        d.get("woo_sku", ""),
        d.get("woo_desc_corta", ""),
        "Ver",
        misma_precio,
    ]
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 2:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=fill_score)
            c.number_format = "0.00"
        if c_idx in (5, 13):
            c.number_format = "#,##0"
    # Hyperlinks
    if d.get("ml_thumbnail"):
        ws.cell(row=r_idx, column=3).hyperlink = d["ml_thumbnail"]
        ws.cell(row=r_idx, column=3).font = Font(color="0563C1", underline="single")
    if d.get("ml_permalink"):
        ws.cell(row=r_idx, column=9).hyperlink = d["ml_permalink"]
        ws.cell(row=r_idx, column=9).font = Font(color="0563C1", underline="single")
    if d.get("woo_thumbnail"):
        ws.cell(row=r_idx, column=11).hyperlink = d["woo_thumbnail"]
        ws.cell(row=r_idx, column=11).font = Font(color="0563C1", underline="single")
    if d.get("woo_permalink"):
        ws.cell(row=r_idx, column=17).hyperlink = d["woo_permalink"]
        ws.cell(row=r_idx, column=17).font = Font(color="0563C1", underline="single")
    ws.row_dimensions[r_idx].height = 110

# Dropdown
dv = DataValidation(type="list", formula1='"DUP,NO"', allow_blank=True)
dv.add(f"A2:A{len(duda) + 1}")
ws.add_data_validation(dv)

# Instrucciones
ws2 = wb.create_sheet("Instrucciones", 0)
inst = [
    ("Cómo revisar los 14 DUDA", True),
    ("", False),
    ("Cada fila muestra el producto de ML (izquierda) vs el match en la web (derecha).", False),
    ("Compara medidas, precios, fotos y descripción para decidir si son el MISMO producto.", False),
    ("", False),
    ("Decisiones disponibles (columna A):", True),
    ("  DUP = es el mismo producto -> NO importar", False),
    ("  NO  = son productos distintos -> SÍ importar el de ML como nuevo", False),
    ("", False),
    ("Tips:", True),
    ("  - Las medidas (cm, mm) son el dato más decisivo: 80x44 ≠ 100x44.", False),
    ("  - Si los precios son MUY distintos, casi siempre son productos diferentes.", False),
    ("  - Click en 'Foto' abre la imagen para comparación visual directa.", False),
    ("  - Click en 'Ver' (ML / Web) abre la ficha completa.", False),
    ("", False),
    ("Cuando termines:", True),
    ("  Guarda el archivo (Ctrl+S) y dile a Claude: \"ya revisé las 14\".", False),
    ("  Claude importará solo los marcados NO.", False),
]
for r, (t, b) in enumerate(inst, start=1):
    c = ws2.cell(row=r, column=1, value=t)
    if b:
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 110

out = r"C:\Users\dell\victtorino\duda_revision.xlsx"
wb.save(out)
print(f"\nGuardado: {out}")
print(f"Filas: {len(duda)}")
