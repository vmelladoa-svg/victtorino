"""
Genera Excel para revision manual de los 110 dudosos.
Una fila por pareja ML <-> Woo. Victor marca DECISION = DUP (es el mismo) / NO (son distintos) / DUDA.
"""
import json
import sys
import io
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

with open(r"C:\Users\dell\victtorino\tokens_cuenta3.json", encoding="utf-8") as f:
    tk = json.load(f)
ML_HEAD = {"Authorization": f"Bearer {tk['access_token']}"}

with open(r"C:\Users\dell\victtorino\cruce_ml_woo_resultado.json", encoding="utf-8") as f:
    cruce = json.load(f)
DUDOSOS = cruce["dudosos"]

print(f"Procesando {len(DUDOSOS)} dudosos...")

# 1) Traer permalink y thumbnail de cada item ML (multiget batch 20)
print("[1/2] Trayendo permalinks ML (multiget)...")
ml_extra = {}
ids = [d["ml_id"] for d in DUDOSOS]
for i in range(0, len(ids), 20):
    batch = ",".join(ids[i:i + 20])
    r = requests.get(
        "https://api.mercadolibre.com/items",
        headers=ML_HEAD,
        params={"ids": batch, "attributes": "id,permalink,price,thumbnail,available_quantity"},
        timeout=30,
    )
    if r.status_code != 200:
        continue
    for entry in r.json():
        if entry.get("code") == 200:
            b = entry["body"]
            ml_extra[b["id"]] = {
                "permalink": b.get("permalink", ""),
                "price": b.get("price"),
                "thumbnail": b.get("thumbnail", ""),
                "stock": b.get("available_quantity"),
            }

# 2) Traer permalink y SKU de cada producto Woo dudoso
print("[2/2] Trayendo permalinks Woo de los matches...")
woo_ids_unicos = list({d["woo_match"][0] for d in DUDOSOS if d.get("woo_match")})
woo_extra = {}
for i in range(0, len(woo_ids_unicos), 10):
    batch_ids = woo_ids_unicos[i:i + 10]
    r = requests.get(
        f"{WC}/wp-json/wc/v3/products",
        params={**P, "include": ",".join(map(str, batch_ids)), "per_page": 100},
        timeout=30,
    )
    if r.status_code != 200:
        continue
    for prod in r.json():
        woo_extra[prod["id"]] = {
            "permalink": prod.get("permalink", ""),
            "sku": prod.get("sku", ""),
            "status": prod.get("status", ""),
            "price": prod.get("regular_price", ""),
        }

# 3) Construir Excel
wb = Workbook()
ws = wb.active
ws.title = "Dudosos ML vs Web"

HEADERS = [
    ("DECISION", 14),
    ("Score", 8),
    ("ML ID", 18),
    ("ML Título", 50),
    ("ML Precio", 12),
    ("ML Stock", 9),
    ("ML Categoría", 25),
    ("ML Link", 14),
    ("Foto ML", 14),
    ("Woo ID", 9),
    ("Woo Título", 50),
    ("Woo Precio", 12),
    ("Woo SKU", 16),
    ("Woo Status", 11),
    ("Woo Link", 14),
    ("Notas", 30),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

for r_idx, d in enumerate(DUDOSOS, start=2):
    woo = d.get("woo_match")
    ml_e = ml_extra.get(d["ml_id"], {})
    woo_e = woo_extra.get(woo[0], {}) if woo else {}

    score = d.get("score", 0)
    fill_color = "C6EFCE" if score >= 0.80 else ("FFEB9C" if score >= 0.70 else "FFC7CE")

    fila = [
        "",  # DECISION (vacio, dropdown)
        score,
        d["ml_id"],
        d["ml_titulo"],
        ml_e.get("price"),
        ml_e.get("stock", d.get("ml_stock")),
        d.get("ml_cat", ""),
        "ML",
        "Ver",
        woo[0] if woo else "",
        woo[1] if woo else "",
        woo_e.get("price", ""),
        woo_e.get("sku", ""),
        woo_e.get("status", ""),
        "Web",
        "",
    ]
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if c_idx == 2:  # score
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.number_format = "0.00"
        if c_idx in (5, 12):  # precios
            c.number_format = "#,##0"
    # Hyperlinks
    if ml_e.get("permalink"):
        ws.cell(row=r_idx, column=8).hyperlink = ml_e["permalink"]
        ws.cell(row=r_idx, column=8).font = Font(color="0563C1", underline="single")
    if ml_e.get("thumbnail"):
        # convertir http a https
        thumb = ml_e["thumbnail"].replace("http://", "https://")
        ws.cell(row=r_idx, column=9).hyperlink = thumb
        ws.cell(row=r_idx, column=9).font = Font(color="0563C1", underline="single")
    if woo_e.get("permalink"):
        ws.cell(row=r_idx, column=15).hyperlink = woo_e["permalink"]
        ws.cell(row=r_idx, column=15).font = Font(color="0563C1", underline="single")
    ws.row_dimensions[r_idx].height = 50

# Dropdown en columna DECISION
dv = DataValidation(type="list", formula1='"DUP,NO,DUDA"', allow_blank=True,
                    showErrorMessage=True, errorTitle="Valor inválido",
                    error="Solo se permite DUP / NO / DUDA")
dv.add(f"A2:A{len(DUDOSOS) + 1}")
ws.add_data_validation(dv)

# Hoja de instrucciones
ws2 = wb.create_sheet("Instrucciones", 0)
instrucciones = [
    ("Cómo revisar este Excel", True),
    ("", False),
    ("Para cada fila, decide si el producto de Mercado Libre (columna ML) y el de la web (columna Woo)", False),
    ("son el MISMO producto o son productos DISTINTOS:", False),
    ("", False),
    ("  DUP   = Es el mismo producto. Ya está en la web. NO importar.", False),
    ("  NO    = Son productos distintos (modelos, medidas o variantes diferentes). SÍ importar a la web.", False),
    ("  DUDA  = No estás seguro. Lo dejamos para una segunda revisión.", False),
    ("", False),
    ("Tips de revisión:", True),
    ("", False),
    ("  - El color del Score te ayuda: verde (>=0.80) = muy probable duplicado.", False),
    ("                                amarillo (0.70-0.80) = revisar bien.", False),
    ("                                rosa (<0.70) = más probable que sean distintos.", False),
    ("", False),
    ("  - Click en 'ML' abre el anuncio en Mercado Libre para ver fotos y detalle.", False),
    ("  - Click en 'Foto ML' abre solo la imagen.", False),
    ("  - Click en 'Web' abre el producto en victtorino.cl.", False),
    ("", False),
    ("  - Si dudas, mira las medidas y el modelo, no el título genérico.", False),
    ("    Ej: 'Lavaplatos 80x44' y 'Lavaplatos 100x44' suenan parecidos pero son DISTINTOS.", False),
    ("", False),
    ("Cuando termines:", True),
    ("", False),
    ("  Guarda el archivo y dile a Claude: \"ya revisé el Excel\".", False),
    ("  Claude importará a la web solo las filas marcadas NO (es decir, las que NO están duplicadas).", False),
    ("", False),
    (f"Total filas a revisar: {len(DUDOSOS)}", True),
]
for r, (txt, bold) in enumerate(instrucciones, start=1):
    c = ws2.cell(row=r, column=1, value=txt)
    if bold:
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 120

# Guardar
out_path = r"C:\Users\dell\victtorino\dudosos_revision.xlsx"
wb.save(out_path)
print(f"\nGuardado: {out_path}")
print(f"Filas a revisar: {len(DUDOSOS)}")
print(f"  Score >=0.80 (verde, casi-seguro duplicado): {sum(1 for d in DUDOSOS if d['score']>=0.80)}")
print(f"  Score 0.70-0.80 (amarillo, revisar)        : {sum(1 for d in DUDOSOS if 0.70<=d['score']<0.80)}")
print(f"  Score <0.70 (rosa, prob. distintos)        : {sum(1 for d in DUDOSOS if d['score']<0.70)}")
