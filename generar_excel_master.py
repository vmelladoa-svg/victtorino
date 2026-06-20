# -*- coding: utf-8 -*-
"""Excel maestro de todo lo puesto en oferta en Falabella (Oportunidad unica)."""
import json, glob, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

VENTA = {
 # OU / griferia
 '010701007-T':13990,'020101010-T':57990,'020101011-T':57990,'020101001-T':38990,'020101002-T':38990,
 '020101003-T':28990,'040303003-T':81990,'040203017-C':21990,'040201006-T':10990,'040202005-T':10990,
 # commodity
 '040303004-T':91990,'010701001-T':46990,'010205002-T':13990,'010104003-T':15990,'010301011-T':11990,
 '020103006-T':11990,'010301003-T':10990,'030102002-T':9990,'030102003-T':7990,'030101002-T':6990,
 '010104002-T':7990,'030102004-T':6990,'010301010-T':6990,'010301009-T':5990,'020301001-T':13990,
 '210000990-T':5990,'020202008-T':4990,'010704002-T':4990,'030101009-T':4990,'010602002-T':8990,
 '020203001-T':3990,'010103002-T':4990,'010301001-T':1990,
 # lote 2
 '040302001-T':12990,'010403001-T':18990,'010603003-T':36990,'010103003-T':5990,'040301001-T':18990,
 '040202001-T':17990,'020102002-T':29990,'010701008-T':18990,'020102001-T':29990,'040201001-T':21990,
 '020103001-T':29990,'020201003-T':30990,'010102002-T':50990,
 # grupo A
 '010502005-T':219000,'010501022-T':168990,'010201001-T':43990,'010102001-T':15990,'010703011-T':16990,
 '010703012-T':12990,'020101002-C':13990,'010301004-T':14990,'010701011-T':6990,'010205004-T':4990,
 '030101008-T':3990,'010105001-T':2990,'030102003-K':3990,
 # singles
 '010102012-C':37990,'010603002-T':139990,'020101014-T':136990,'020101013-T':116990,
 '010102007-C':37990,'010604003-T':9990,
}
# costos manuales (neto) que dio Victor para los sin-costo
COST_MANUAL={'010102012-C':15410,'010603002-T':46929,'020101014-T':40023,'020101013-T':36157}

def load(f):
    raw=open(f,encoding='utf-8').read().strip(); d=json.loads(raw); return json.loads(d) if isinstance(d,str) else d

# admin (nombre, stock, precio viejo)
admin={}
for f in sorted(glob.glob('admin_p*.json')):
    d=load(f); rows=d['rows'] if isinstance(d,dict) and 'rows' in d else d
    for r in rows: admin[r['seller']]=r
# template (shopsku, nombre oficial)
tpl=openpyxl.load_workbook('SellerPriceTemplate.xlsx',data_only=True).active
import re
shop={}
for r in list(tpl.iter_rows(values_only=True))[1:]:
    shop[re.sub(r'_\d+$','',str(r[0]))]={'shop':r[1],'name':r[6]}
# costos
wv=openpyxl.load_workbook('validacion_costos.xlsx',data_only=True).active
rv=list(wv.iter_rows(values_only=True)); Hv=rv[0]
val={r[Hv.index('Codigo')]:r[Hv.index('COSTO A USAR')] for r in rv[1:] if r[Hv.index('Codigo')] and isinstance(r[Hv.index('COSTO A USAR')],(int,float))}
vig=load('costos_vigentes.json')
def costo_neto(s):
    if s in COST_MANUAL: return COST_MANUAL[s]
    if s in val: return val[s]
    if s in vig and vig[s].get('vig'): return vig[s]['vig']
    return None
def r990(x): x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

wb=openpyxl.Workbook(); w=wb.active; w.title='Oferta Falabella'
hdr=['SKU Seller','SKU Falabella','Producto','Stock','Precio ANTES','Precio NORMAL (tachado)','PRECIO OFERTA','Costo c/IVA','Margen','Vigencia']
w.append(hdr)
for c in w[1]:
    c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
for sku,venta in VENTA.items():
    a=admin.get(sku,{}); sm=shop.get(sku,{})
    cn=costo_neto(sku); civa=int(cn*1.19) if cn else None
    margen=round(1-civa/venta,3) if civa else None
    w.append([sku, sm.get('shop'), sm.get('name') or a.get('name'), a.get('stock'),
              a.get('special') or a.get('price'), r990(venta/0.5), venta, civa, margen, '16/06-16/07/2026'])
# formato
for rr in range(2,w.max_row+1):
    for col in [5,6,7,8]:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
    mc=w.cell(rr,9)
    if isinstance(mc.value,(int,float)): mc.number_format='0%'
    w.cell(rr,7).fill=PatternFill('solid',fgColor='C6EFCE')
for i,wd in enumerate([13,13,46,7,13,15,13,11,8,16]):
    w.column_dimensions[chr(65+i)].width=wd
w.freeze_panes='A2'
OUT=r'C:\Users\dell\Downloads\Oferta_Falabella_OportunidadUnica.xlsx'
wb.save(OUT)
print('Generado:',OUT)
print('Productos:',len(VENTA))
