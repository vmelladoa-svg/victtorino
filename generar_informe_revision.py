"""
Genera el informe ejecutivo final para revisar manualmente:
  - Resumen ejecutivo de la estrategia
  - Proyeccion de ventas esperadas (con supuestos editables)
  - Las 32 acciones detalladas
  - Lo que NO se toca (scope)
  - Proximos pasos

Output: informe_revision_<fecha>.xlsx
"""
from __future__ import annotations
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
PLAN = ROOT / f"plan_accion_{time.strftime('%Y-%m-%d')}.xlsx"
DIAG = ROOT / f"margen_diagnostico_{time.strftime('%Y-%m-%d')}.xlsx"
MATCH = ROOT / f"matching_defontana_ml_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"informe_revision_{time.strftime('%Y-%m-%d')}.xlsx"

# ============================================================
# SUPUESTOS para proyeccion de venta (editables en hoja SUPUESTOS)
# ============================================================
SUPUESTOS = {
    "ventas_mes_si_gana_bb": 3.0,        # unidades/mes promedio si listing gana buy box
    "ventas_mes_si_pierde_bb": 0.3,      # unidades/mes si pierde (clicks/conversiones residuales)
    "ventas_mes_si_sin_competencia": 2.0, # unidades/mes si no hay overlap
    "iva": 0.19,
    "ml_channel_fee": 0.32,
}

NETO_FACTOR = (1 / (1 + SUPUESTOS["iva"])) * (1 - SUPUESTOS["ml_channel_fee"])  # 0.5714


def cargar_hoja(path: Path, hoja: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def calcular_margen_clp(price: float, costo: float) -> float:
    if not price or not costo or price <= 0 or costo <= 0:
        return 0
    venta_neta = price / (1 + SUPUESTOS["iva"])
    return venta_neta * (1 - SUPUESTOS["ml_channel_fee"]) - costo


def proyectar(accion: dict) -> dict:
    """Calcula proyeccion mensual y anual para una accion."""
    precio_act = accion.get("precio_actual") or 0
    precio_nuevo = accion.get("precio_sugerido") or precio_act
    costo = accion.get("costo") or 0
    tipo = accion.get("accion")

    margen_unidad_actual = calcular_margen_clp(precio_act, costo)
    margen_unidad_nuevo = calcular_margen_clp(precio_nuevo, costo)

    if tipo == "PAUSAR":
        # estado actual: pierde BB (la otra cuenta gana). Ventas residuales bajas.
        u_mes_actual = SUPUESTOS["ventas_mes_si_pierde_bb"]
        u_mes_nuevo = 0  # pausado
        # IMPORTANTE: las ventas que pierde este listing se TRANSFIEREN a la cuenta ganadora interna
        # asumimos que la cuenta ganadora absorbe el 80% de las ventas pausadas
        venta_transferida_mes = u_mes_actual * 0.8
        venta_transferida_clp_mes = venta_transferida_mes * precio_act
        margen_transferido_mes = venta_transferida_mes * margen_unidad_actual
        # tambien hay ahorro de "ruido" — listings pausados no consumen atencion del algoritmo
        # pero no hay impacto monetario directo
        return {
            "ventas_actuales_mes": round(u_mes_actual, 2),
            "ventas_proyectadas_mes": 0,
            "delta_ventas_mes": round(-u_mes_actual, 2),
            "ingresos_actuales_mes": round(u_mes_actual * precio_act),
            "ingresos_proyectados_mes": 0,
            "margen_actual_mes": round(u_mes_actual * margen_unidad_actual),
            "margen_proyectado_mes": 0,
            "venta_transferida_mes": round(venta_transferida_clp_mes),
            "margen_transferido_mes": round(margen_transferido_mes),
            "impacto_neto_margen_mes": round(margen_transferido_mes - (u_mes_actual * margen_unidad_actual)),
        }
    elif tipo == "BAJAR_PRECIO":
        # gana buy box despues del cambio
        u_mes_actual = SUPUESTOS["ventas_mes_si_pierde_bb"]
        u_mes_nuevo = SUPUESTOS["ventas_mes_si_gana_bb"]
        return {
            "ventas_actuales_mes": round(u_mes_actual, 2),
            "ventas_proyectadas_mes": round(u_mes_nuevo, 2),
            "delta_ventas_mes": round(u_mes_nuevo - u_mes_actual, 2),
            "ingresos_actuales_mes": round(u_mes_actual * precio_act),
            "ingresos_proyectados_mes": round(u_mes_nuevo * precio_nuevo),
            "margen_actual_mes": round(u_mes_actual * margen_unidad_actual),
            "margen_proyectado_mes": round(u_mes_nuevo * margen_unidad_nuevo),
            "venta_transferida_mes": 0,
            "margen_transferido_mes": 0,
            "impacto_neto_margen_mes": round(u_mes_nuevo * margen_unidad_nuevo - u_mes_actual * margen_unidad_actual),
        }
    return {}


def fmt_clp(v) -> str:
    try:
        return f"${int(v):,}".replace(",", ".")
    except Exception:
        return "-"


# ============================================================
# Estilos
# ============================================================
H1 = Font(bold=True, size=14, color="FFFFFF")
H2 = Font(bold=True, size=12, color="FFFFFF")
H3 = Font(bold=True, size=11)
TXT = Font(size=10)
TXT_BOLD = Font(size=10, bold=True)
FILL_H1 = PatternFill("solid", fgColor="1A237E")
FILL_H2 = PatternFill("solid", fgColor="455A64")
FILL_PAUSAR = PatternFill("solid", fgColor="C5E1A5")
FILL_BAJAR = PatternFill("solid", fgColor="FFE082")
FILL_INFO = PatternFill("solid", fgColor="E1F5FE")
FILL_WARN = PatternFill("solid", fgColor="FFF9C4")
FILL_DANGER = PatternFill("solid", fgColor="FFCDD2")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def main() -> int:
    if not PLAN.exists():
        print(f"ERROR: no existe {PLAN.name}")
        return 1

    acciones = [a for a in cargar_hoja(PLAN, "ACCIONES") if a.get("aprobado") == "SI"]
    todos_listings = cargar_hoja(DIAG, "TODOS_LOS_LISTINGS") if DIAG.exists() else []
    matched = cargar_hoja(MATCH, "MATCHED") if MATCH.exists() else []
    sin_match_def = cargar_hoja(MATCH, "UNMATCHED_DEF") if MATCH.exists() else []

    # proyectar cada accion
    for a in acciones:
        a.update(proyectar(a))

    # totales agregados
    total_margen_actual_mes = sum(a.get("margen_actual_mes", 0) for a in acciones)
    total_margen_proyectado_mes = sum(a.get("margen_proyectado_mes", 0) for a in acciones)
    total_margen_transferido_mes = sum(a.get("margen_transferido_mes", 0) for a in acciones)
    total_impacto_mes = sum(a.get("impacto_neto_margen_mes", 0) for a in acciones)
    total_impacto_anual = total_impacto_mes * 12

    # diagnostico global
    n_a_perdida = sum(1 for r in todos_listings if r.get("estado_margen") == "a_perdida")
    n_sin_chance = sum(1 for r in todos_listings
                       if r.get("estado_bb") == "PIERDE_VS_EXTERNO"
                       and not r.get("puede_competir_vs_ext"))
    n_pendientes = sum(1 for r in cargar_hoja(PLAN, "ACCIONES") if r.get("aprobado") == "PENDIENTE")

    # ==========================================================
    # construir xlsx
    # ==========================================================
    wb = Workbook()
    wb.remove(wb.active)

    # ===== HOJA 0: GLOSARIO Y DEFINICIONES =====
    wsg = wb.create_sheet("0_GLOSARIO")
    rowg = 1

    def cellg(r, c, val, font=TXT, fill=None, align=None):
        cl = wsg.cell(row=r, column=c, value=val)
        cl.font = font
        if fill: cl.fill = fill
        if align: cl.alignment = align
        return cl

    cellg(rowg, 1, "GLOSARIO Y DEFINICIONES — Lee esto primero", H1, FILL_H1)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    wsg.row_dimensions[rowg].height = 28
    rowg += 2

    # ----------- ACCION PAUSAR -----------
    cellg(rowg, 1, "QUE SIGNIFICA 'PAUSAR' EN ESTE PLAN", H2, FILL_H2)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    rowg += 1
    cellg(rowg, 1, "En MercadoLibre hay 4 formas distintas de 'dejar de promover un item'. Cada una tiene efecto distinto:", TXT_BOLD)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    rowg += 2

    cellg(rowg, 1, "Forma", TXT_BOLD, FILL_INFO)
    cellg(rowg, 2, "API call", TXT_BOLD, FILL_INFO)
    cellg(rowg, 3, "Efecto visible", TXT_BOLD, FILL_INFO)
    cellg(rowg, 4, "Reversibilidad / Notas", TXT_BOLD, FILL_INFO)
    rowg += 1

    formas = [
        ("1. PAUSAR PUBLICACION  ⬅ LO QUE HAREMOS",
         'PUT /items/{id} {"status":"paused"}',
         "Item DESAPARECE del buscador ML. Nadie lo ve. No genera ventas. No genera visitas. No paga comision (no hay venta).",
         "100% reversible: PUT status='active' lo vuelve a publicar tal cual. Mantiene historial, ventas pasadas, reputacion del item, ranking SEO interno. Es lo MAS conservador."),
        ("2. SALIR DEL CATALOGO (ALTERNATIVA)",
         'PUT /items/{id} {"catalog_listing":false}',
         "Item SIGUE VISIBLE en busqueda como producto propio. NO compite por buy box en la ficha de catalogo compartida. Vende como tu publicacion individual.",
         "Reversible. Util cuando el item es identico al catalogo pero quieres venderlo bajo tu propia ficha (sin pelear con otros). NO aplica al patron de canibalizacion interna (estamos peleando con cuentas nuestras, no externos)."),
        ("3. CERRAR PUBLICACION",
         'PUT /items/{id} {"status":"closed"}',
         "Item se cierra DEFINITIVAMENTE en ese item_id. No se puede reactivar sin crear publicacion nueva (con item_id distinto).",
         "IRREVERSIBLE practico — pierdes historial, reputacion del item, ranking. NO LO USAREMOS."),
        ("4. PAUSAR PUBLICIDAD (Product Ads)",
         "PUT /advertising/.../campaigns (otro endpoint)",
         "Solo aplica si el item tiene anuncios pagados en ML Ads. Pausa solo la publicidad, el item sigue activo y visible en busqueda organica.",
         "Reversible. NO APLICA: tus 3 cuentas tienen ML Ads en VIEWER (no podemos modificar via API segun tu memoria). Habria que hacerlo manual en ads.mercadolibre.cl/productAds."),
    ]
    for forma in formas:
        fill = FILL_PAUSAR if forma[0].startswith("1.") else None
        for i, val in enumerate(forma, start=1):
            c = cellg(rowg, i, val, TXT, fill)
            c.alignment = WRAP
        wsg.row_dimensions[rowg].height = 75
        rowg += 1
    rowg += 1

    cellg(rowg, 1, "¿POR QUE elegimos 'pausar publicacion' (#1) y no las otras?", H3, FILL_INFO)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    rowg += 1
    razones = [
        "• Los 28 casos son items donde OTRA cuenta tuya ya gana el buy box. Pausar el perdedor no quita ventas (ya estaban yendo al ganador).",
        "• 'Pausar publicacion' es 100% reversible — si maniana cambian condiciones, reactivar es 1 API call.",
        "• 'Salir del catalogo' (#2) no resuelve nuestra canibalizacion interna — el item perdedor seguiria visible y compitiendo en busqueda, solo dejaria de pelear en la ficha. Sigue consumiendo atencion del algoritmo.",
        "• 'Cerrar' (#3) es irreversible y pierdes el historial del item. Mucho riesgo para una decision exploratoria.",
        "• 'Pausar publicidad' (#4) solo aplica si tuvieras ML Ads activos en esos items (no es el caso, ML Ads esta en VIEWER y bloqueado por API en tu setup).",
    ]
    for r in razones:
        cellg(rowg, 1, r, TXT)
        wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
        wsg.row_dimensions[rowg].height = 25
        rowg += 1
    rowg += 2

    # ----------- ACCION BAJAR_PRECIO -----------
    cellg(rowg, 1, "QUE SIGNIFICA 'BAJAR_PRECIO'", H2, FILL_H2)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    rowg += 1
    bajar_lines = [
        "• API call: PUT /items/{id} {\"price\": <nuevo_valor>}",
        "• El cambio aplica INMEDIATAMENTE al item ML (no afecta otros canales como Web, Tienda, Falabella, Paris, Walmart).",
        "• Defontana sigue con su 'precio sugerido' que no se modifica.",
        "• Si tu Defontana sincroniza precios hacia ML, ese cambio podria revertirse en la siguiente sincronizacion. Verifica si Defontana hace push automatico de precios a ML.",
        "• 100% reversible: PUT con el precio anterior lo restaura.",
        "• ML respeta el cambio en buy box dentro de 1-6 horas tipicamente.",
    ]
    for line in bajar_lines:
        cellg(rowg, 1, line, TXT)
        wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
        wsg.row_dimensions[rowg].height = 22
        rowg += 1
    rowg += 2

    # ----------- BUY BOX -----------
    cellg(rowg, 1, "GLOSARIO BASICO", H2, FILL_H2)
    wsg.merge_cells(start_row=rowg, start_column=1, end_row=rowg, end_column=4)
    rowg += 1
    glos = [
        ("Buy Box", "El boton 'Comprar' destacado en una ficha de catalogo compartida. Quien gana el buy box recibe ~80-90% de las ventas de esa ficha. Lo gana el seller con mejor combinacion de precio, reputacion, FULL, ranking."),
        ("Catalog listing (catalog_listing=true)", "Modo donde tu publicacion compite junto a otros sellers por el mismo producto bajo una unica ficha. ML decide quien gana el buy box."),
        ("Listing type (gold_pro)", "Tipo Premium. Comision ~17.5%. Mayor exposicion, 12 cuotas sin interes, postularse a buy box con prioridad."),
        ("Listing type (gold_special)", "Tipo Clasica. Comision ~13-15%. Menor exposicion. Buy box menos competitivo."),
        ("Listing type (free)", "Gratuita. 0% comision pero limite de items, sin envio Full, no compite por buy box Premium."),
        ("Status='active'", "Item publicado y visible."),
        ("Status='paused'", "Item invisible para clientes. No genera ventas. Mantiene su item_id, historial, reputacion. Es REVERSIBLE."),
        ("Status='closed'", "Item cerrado definitivamente. No puede reactivarse — habria que crear nueva publicacion con nuevo item_id."),
        ("Status='under_review'", "ML lo tiene en revision (politicas, fotos, descripcion). Sigue visible. Algunas acciones pueden quedar pendientes hasta resolucion."),
        ("Canibalizacion interna", "Cuando 2 o 3 de tus propias cuentas (C1/C2/C3) compiten en la MISMA ficha de catalogo. Una gana buy box; las demas pagan exposicion para perder."),
        ("Lookalike encubierto", "Externo vende producto IDENTICO al tuyo a precio mucho menor. Detectado en investigacion previa: 12 de 13 casos eran tu mismo producto vendido por tus propios distribuidores con menor margen."),
        ("PIERDE_VS_NUESTRO", "El buy box lo gana otra cuenta tuya (C1, C2 o C3). Pausar evita autoperdida."),
        ("PIERDE_VS_EXTERNO", "El buy box lo gana un seller externo. Bajar precio puede recuperarlo SI el margen lo permite."),
        ("GANA_BUY_BOX", "Tu listing tiene el buy box. No tocar."),
        ("PUEDE_COMPETIR", "Calculamos tu precio break-even (costo / 0.5714). Si es menor al precio del externo, puedes bajar precio y ganar manteniendo margen positivo."),
        ("SIN_CHANCE", "Tu precio break-even es MAYOR al precio externo. Bajar te lleva a perdida. No tiene sentido competir."),
    ]
    cellg(rowg, 1, "Termino", TXT_BOLD, FILL_INFO)
    cellg(rowg, 2, "Definicion", TXT_BOLD, FILL_INFO)
    wsg.merge_cells(start_row=rowg, start_column=2, end_row=rowg, end_column=4)
    rowg += 1
    for termino, defin in glos:
        cellg(rowg, 1, termino, TXT_BOLD)
        cellg(rowg, 2, defin, TXT)
        wsg.merge_cells(start_row=rowg, start_column=2, end_row=rowg, end_column=4)
        wsg.cell(row=rowg, column=2).alignment = WRAP
        wsg.row_dimensions[rowg].height = max(25, len(defin) // 4)
        rowg += 1

    wsg.column_dimensions["A"].width = 32
    wsg.column_dimensions["B"].width = 38
    wsg.column_dimensions["C"].width = 45
    wsg.column_dimensions["D"].width = 50

    # ===== HOJA 1: RESUMEN EJECUTIVO =====
    ws = wb.create_sheet("1_RESUMEN_EJECUTIVO")
    row = 1

    def cell(r, c, val, font=TXT, fill=None, align=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = font
        if fill: cl.fill = fill
        if align: cl.alignment = align
        return cl

    cell(row, 1, "INFORME EJECUTIVO — OPTIMIZACION ML CATALOGO", H1, FILL_H1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 1
    cell(row, 1, f"Generado: {time.strftime('%Y-%m-%d %H:%M')}  |  Cuentas: C1, C2, C3  |  533 publicaciones analizadas", TXT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # diagnostico
    cell(row, 1, "DIAGNOSTICO ACTUAL", H2, FILL_H2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    diag_lines = [
        "• 533 publicaciones en catalog listing repartidas entre las 3 cuentas (C1: 170, C2: 176, C3: 187).",
        "• 152 fichas (catalog_product_id) tienen overlap entre cuentas tuyas → canibalizacion interna.",
        "• De esas, 40 tienen al menos una cuenta con stock real → la canibalizacion 'duele' en 40 fichas.",
        "• 51 publicaciones pierden buy box vs sellers externos (FACTORYNETCL, VIBRA TOOLS, OUTLETSEVENCL concentran 41%).",
        "• 12 de 13 'lookalikes sospechosos' resultaron ser EL MISMO producto vendido por distribuidores tuyos (control de canal, no marketing).",
        f"• {n_a_perdida} items vendiendose a perdida (revision urgente — probable error de costos o precios).",
        "• Los costos de Defontana (netos) cruzaron al 95.3% con los items ML via SELLER_SKU.",
    ]
    for line in diag_lines:
        cell(row, 1, line, TXT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    # estrategia
    cell(row, 1, "ESTRATEGIA QUE IMPLEMENTAMOS", H2, FILL_H2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    estrat = [
        ("Principio 1", "Una cuenta nuestra por ficha de catalogo. La que YA gana buy box se queda; las demas pausan."),
        ("Principio 2", "Bajamos precio SOLO si margen final >=15% y caida <=10 pp. Nunca a costa de margen miserable."),
        ("Principio 3", "No tocamos lo que funciona (429 fichas = 84% del catalogo donde ganamos o no hay competencia)."),
        ("Principio 4", "Decisiones comerciales (canal, distribuidores, salir del catalogo) quedan separadas — necesitan tu input."),
    ]
    for nombre, txt in estrat:
        cell(row, 1, nombre, TXT_BOLD, FILL_INFO)
        cell(row, 2, txt, TXT, FILL_INFO)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 30
        ws.cell(row=row, column=2).alignment = WRAP
        row += 1
    row += 1

    # plan
    cell(row, 1, "PLAN DE ACCION PRE-APROBADO (32 acciones)", H2, FILL_H2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    cell(row, 1, "Accion", TXT_BOLD); cell(row, 2, "Cuenta", TXT_BOLD)
    cell(row, 3, "Items", TXT_BOLD); cell(row, 4, "Reversibilidad", TXT_BOLD)
    cell(row, 5, "Riesgo", TXT_BOLD); cell(row, 6, "Detalle", TXT_BOLD)
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = FILL_INFO
    row += 1
    from collections import Counter
    accion_cuenta = Counter((a["accion"], a["cuenta"]) for a in acciones)
    plan_rows = [
        ("PAUSAR", "C1", accion_cuenta.get(("PAUSAR", "C1"), 0), "100% (PUT status=active)", "Nulo", "Pierde vs cuenta tuya"),
        ("PAUSAR", "C2", accion_cuenta.get(("PAUSAR", "C2"), 0), "100% (PUT status=active)", "Nulo", "Pierde vs cuenta tuya"),
        ("PAUSAR", "C3", accion_cuenta.get(("PAUSAR", "C3"), 0), "100% (PUT status=active)", "Nulo", "Pierde vs cuenta tuya"),
        ("BAJAR_PRECIO", "C1", accion_cuenta.get(("BAJAR_PRECIO", "C1"), 0), "100% (PUT price=anterior)", "Bajo", "Recuperar buy box vs externo"),
        ("BAJAR_PRECIO", "C2", accion_cuenta.get(("BAJAR_PRECIO", "C2"), 0), "100% (PUT price=anterior)", "Bajo", "Recuperar buy box vs externo"),
        ("BAJAR_PRECIO", "C3", accion_cuenta.get(("BAJAR_PRECIO", "C3"), 0), "100% (PUT price=anterior)", "Bajo", "Recuperar buy box vs externo"),
    ]
    for r_ in plan_rows:
        fill = FILL_PAUSAR if r_[0] == "PAUSAR" else FILL_BAJAR
        for i, v in enumerate(r_, start=1):
            c = cell(row, i, v, TXT, fill)
            if i == 3: c.alignment = CENTER
        row += 1
    row += 1

    # proyeccion
    cell(row, 1, "PROYECCION DE IMPACTO (con supuestos en hoja 4)", H2, FILL_H2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    cell(row, 1, "Concepto", TXT_BOLD, FILL_INFO)
    cell(row, 2, "Mensual", TXT_BOLD, FILL_INFO)
    cell(row, 3, "Anual", TXT_BOLD, FILL_INFO)
    cell(row, 4, "Nota", TXT_BOLD, FILL_INFO)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    row += 1

    metricas = [
        ("Margen actual en items afectados", total_margen_actual_mes,
         "Estado actual de los 32 items (mayoria PIERDE buy box, ventas bajas)"),
        ("Margen post-cambios en mismos items", total_margen_proyectado_mes,
         "BAJAR_PRECIO gana BB (mas ventas), PAUSAR deja de generar"),
        ("Margen transferido a cuenta ganadora", total_margen_transferido_mes,
         "Ventas que dejan de hacer los pausados las absorbe la cuenta interna ganadora (80%)"),
        ("IMPACTO NETO ESPERADO", total_impacto_mes,
         "Diferencia neta: gana ventas el ganador + reduccion ruido competitivo interno"),
    ]
    for nombre, val_mes, nota in metricas:
        is_total = "IMPACTO NETO" in nombre
        fill = FILL_WARN if is_total else None
        font = TXT_BOLD if is_total else TXT
        cell(row, 1, nombre, font, fill)
        cell(row, 2, fmt_clp(val_mes), font, fill)
        cell(row, 3, fmt_clp(val_mes * 12), font, fill)
        cell(row, 4, nota, TXT, fill)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4).alignment = WRAP
        ws.row_dimensions[row].height = 30
        row += 1
    row += 2

    cell(row, 1, "ATENCION: los supuestos de ventas son ESTIMACIONES (3 u/mes si gana BB, 0.3 u/mes si pierde).", TXT_BOLD, FILL_DANGER)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    cell(row, 1, "Edita la hoja '4_SUPUESTOS' con tus tasas reales y recalculamos. Los numeros actuales son orden de magnitud.", TXT, FILL_DANGER)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # lo que NO se toca
    cell(row, 1, "LO QUE NO SE TOCA (decisiones pendientes - ver hoja 5)", H2, FILL_H2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    no_se_toca = [
        f"{n_a_perdida} items vendiendose a perdida → revision urgente (¿error precio o costo desactualizado?)",
        f"{n_sin_chance} items donde externo te gana SIN CHANCE → decision: salir del catalogo, aceptar perdida, o migrar a otro canal",
        f"{n_pendientes} BAJAR_PRECIO PENDIENTES (margen final <15%) → revision 1 a 1 si quieres bajarlos igual",
        "5 sellers externos canibalizandote (FACTORYNETCL, VIBRA TOOLS, OUTLETSEVENCL...) → decision de canal/distribucion",
        "112 fichas con stock 0 (estrategia 'precio bloqueo $400.000') → operacional",
        "74 codigos Defontana sin publicacion ML → oportunidad de listado nuevo",
        "1 lookalike real (Kit Estanque WC) → reportar a soporte ML para que separe la ficha",
    ]
    for line in no_se_toca:
        cell(row, 1, "•", TXT_BOLD)
        cell(row, 2, line, TXT)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    for c in ["C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 18

    # ===== HOJA 2: PROYECCION DE VENTAS DETALLADA =====
    ws2 = wb.create_sheet("2_PROYECCION_VENTAS")
    cols = [
        "accion", "cuenta", "item_id", "title",
        "precio_actual", "precio_sugerido", "costo",
        "margen_actual_pct", "margen_proyectado_pct",
        "ventas_actuales_mes", "ventas_proyectadas_mes",
        "ingresos_actuales_mes", "ingresos_proyectados_mes",
        "margen_actual_mes", "margen_proyectado_mes",
        "venta_transferida_mes", "margen_transferido_mes",
        "impacto_neto_margen_mes", "impacto_neto_margen_anual",
        "ganador_actual",
    ]
    ws2.append(cols)
    for c in ws2[1]:
        c.font = H3; c.fill = FILL_H2; c.font = Font(bold=True, color="FFFFFF")
    for a in sorted(acciones, key=lambda x: (-x.get("impacto_neto_margen_mes", 0))):
        a["impacto_neto_margen_anual"] = a.get("impacto_neto_margen_mes", 0) * 12
        ws2.append([a.get(c) for c in cols])
        last = ws2.max_row
        fill = FILL_PAUSAR if a["accion"] == "PAUSAR" else FILL_BAJAR
        ws2.cell(row=last, column=1).fill = fill
    # total
    last = ws2.max_row + 1
    ws2.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=last, column=1).fill = FILL_WARN
    for c in ["ventas_actuales_mes", "ventas_proyectadas_mes",
              "ingresos_actuales_mes", "ingresos_proyectados_mes",
              "margen_actual_mes", "margen_proyectado_mes",
              "venta_transferida_mes", "margen_transferido_mes",
              "impacto_neto_margen_mes", "impacto_neto_margen_anual"]:
        idx = cols.index(c) + 1
        total = sum(a.get(c) or 0 for a in acciones)
        ws2.cell(row=last, column=idx, value=round(total))
        ws2.cell(row=last, column=idx).font = Font(bold=True)
        ws2.cell(row=last, column=idx).fill = FILL_WARN

    ws2.freeze_panes = "E2"
    widths = {"title": 50, "ganador_actual": 25}
    for i, c in enumerate(cols, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    # ===== HOJA 3: ACCIONES (DETALLE COMPLETO) =====
    ws3 = wb.create_sheet("3_ACCIONES_DETALLE")
    detalle_cols = [
        "aprobado", "accion", "cuenta", "item_id", "title", "listing_type_id",
        "stock_actual", "precio_actual", "costo", "margen_actual_pct",
        "precio_sugerido", "margen_proyectado_pct", "delta_margen_pct",
        "ganador_actual", "ganador_precio", "razon",
        "catalog_product_id", "permalink",
    ]
    ws3.append(detalle_cols)
    for c in ws3[1]:
        c.fill = FILL_H2; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = CENTER
    ws3.row_dimensions[1].height = 25

    todas_acciones = cargar_hoja(PLAN, "ACCIONES")
    for a in sorted(todas_acciones, key=lambda x: (x.get("accion") or "", -(x.get("precio_actual") or 0))):
        ws3.append([a.get(c) for c in detalle_cols])
        last = ws3.max_row
        accion = a.get("accion")
        aprobado = a.get("aprobado")
        if aprobado == "SI" and accion == "PAUSAR":
            ws3.cell(row=last, column=1).fill = FILL_PAUSAR
        elif aprobado == "SI" and accion == "BAJAR_PRECIO":
            ws3.cell(row=last, column=1).fill = FILL_BAJAR
        elif aprobado == "PENDIENTE":
            ws3.cell(row=last, column=1).fill = FILL_WARN
    ws3.freeze_panes = "E2"
    widths3 = {"title": 50, "razon": 50, "permalink": 40, "ganador_actual": 22}
    for i, c in enumerate(detalle_cols, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = widths3.get(c, 13)

    # ===== HOJA 4: SUPUESTOS (editables) =====
    ws4 = wb.create_sheet("4_SUPUESTOS")
    ws4.append(["parametro", "valor", "explicacion", "fuente"])
    for c in ws4[1]:
        c.fill = FILL_H2; c.font = Font(bold=True, color="FFFFFF")
    sup_rows = [
        ("ventas_mes_si_gana_bb", SUPUESTOS["ventas_mes_si_gana_bb"],
         "Unidades por mes promedio cuando un listing gana buy box",
         "Estimacion. Ajustar con datos reales por categoria."),
        ("ventas_mes_si_pierde_bb", SUPUESTOS["ventas_mes_si_pierde_bb"],
         "Unidades por mes cuando listing pierde (clicks residuales, busquedas directas)",
         "Estimacion. ML reporta visitas en /items/{id}/visits"),
        ("ventas_mes_si_sin_competencia", SUPUESTOS["ventas_mes_si_sin_competencia"],
         "Unidades por mes cuando no hay overlap (no aplica a este analisis)",
         "Referencia"),
        ("iva", SUPUESTOS["iva"], "IVA Chile", "SII"),
        ("ml_channel_fee", SUPUESTOS["ml_channel_fee"],
         "Cargo canal ML sobre venta neta",
         "Reporte 'Margen por canal' Victor 2026-05-26 (ajustado de 42% a 32%)"),
        ("factor_neto_efectivo", round(NETO_FACTOR, 4),
         "(1/1.19) * (1 - 0.32) = lo que queda del precio publicado tras IVA + comision ML",
         "Derivado"),
    ]
    for r_ in sup_rows:
        ws4.append(r_)
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 55
    ws4.column_dimensions["D"].width = 50
    ws4.append([])
    ws4.append(["NOTA: Si editas los supuestos, debes correr nuevamente generar_informe_revision.py"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, italic=True)
    ws4.cell(row=ws4.max_row, column=1).fill = FILL_WARN

    # ===== HOJA 5: PENDIENTES Y FUERA DE SCOPE =====
    ws5 = wb.create_sheet("5_PENDIENTES_SCOPE")
    ws5.append(["categoria", "cantidad", "descripcion", "recomendacion"])
    for c in ws5[1]:
        c.fill = FILL_H2; c.font = Font(bold=True, color="FFFFFF")
    pend = [
        ("Items a perdida", n_a_perdida,
         "Items vendiendose por debajo del costo (margen negativo)",
         "URGENTE: revisar caso por caso. Posibles causas: error de carga, costo desactualizado, liquidacion deliberada."),
        ("PIERDE_VS_EXT SIN CHANCE", n_sin_chance,
         "El externo te vence incluso si bajas a tu precio break-even",
         "Decision: salir del catalogo (publicar como producto propio), aceptar perdida, o migrar trafico a Web (72% margen vs 30% ML)."),
        ("BAJAR_PRECIO PENDIENTE", n_pendientes,
         "Casos donde bajar precio para ganar BB deja margen <15% o caida >10pp",
         "Revisar 1 a 1 en hoja '3_ACCIONES_DETALLE'. Marca SI o NO en columna 'aprobado'."),
        ("Sellers externos a investigar", 5,
         "FACTORYNETCL, VIBRA TOOLS, OUTLETSEVENCL, MDWIS, GRIFERIALDAYCL",
         "Verificar si son clientes/distribuidores tuyos. Si SI: politica de RRP. Si NO: brand protection ML."),
        ("Stock 0 con precio bloqueo", 112,
         "Fichas mantenidas activas con precio $400.000 (no venden, mantienen posicion)",
         "Operacional: ¿reponen pronto? Mantener. ¿Discontinuados? Cerrar."),
        ("Defontana sin publicacion ML", len(sin_match_def),
         "SKUs en tu maestro de costos que no tienen publicacion en ML",
         "Oportunidad: ¿son productos para listar? ¿O son B2B/internos?"),
        ("Lookalike real (Kit Estanque WC)", 1,
         "Externo vende producto DISTINTO bajo el mismo catalog_product_id",
         "Reportar a soporte ML (programa brand protection) para que separe la ficha."),
    ]
    for r_ in pend:
        ws5.append(r_)
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 10
    ws5.column_dimensions["C"].width = 50
    ws5.column_dimensions["D"].width = 60
    for r in ws5.iter_rows(min_row=2):
        for c in r:
            c.alignment = WRAP

    # ===== HOJA 6: COMO EJECUTAR =====
    ws6 = wb.create_sheet("6_COMO_EJECUTAR")
    instrucciones = [
        ("PASO", "ACCION"),
        ("1", "Revisar hojas 1 (resumen), 2 (proyeccion), 3 (acciones), 5 (pendientes)."),
        ("2", "Si quieres ajustar alguna fila pre-aprobada: editar columna 'aprobado' en hoja 3 (SI/NO/PENDIENTE)."),
        ("3", "Si quieres modificar supuestos de proyeccion: editar hoja 4 y recorrer 'generar_informe_revision.py'."),
        ("4", "Validar el plan: las 32 acciones aprobadas son: 28 PAUSAR + 4 BAJAR_PRECIO."),
        ("5", "Cuando estes listo, ejecutar DRY-RUN: python aplicar_plan_accion.py"),
        ("6", "Verifica que el dry-run muestra lo correcto."),
        ("7", "Ejecutar de verdad: python aplicar_plan_accion.py --apply"),
        ("8", "El script crea backup en backups/ y log CSV en logs/."),
        ("9", "Para revertir: cargar backup_<timestamp>.jsonl y volver al precio/status anterior."),
        ("", ""),
        ("ROLLBACK", "Cada accion es 100% reversible. PAUSAR <-> ACTIVAR via PUT status. BAJAR_PRECIO <-> precio anterior via PUT price."),
    ]
    for r_ in instrucciones:
        ws6.append(r_)
    ws6.column_dimensions["A"].width = 12
    ws6.column_dimensions["B"].width = 90
    for r in ws6.iter_rows(min_row=2, max_col=2):
        for c in r:
            c.alignment = WRAP
    ws6.cell(row=1, column=1).fill = FILL_H2
    ws6.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    ws6.cell(row=1, column=2).fill = FILL_H2
    ws6.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")

    wb.save(DST)
    print(f"\nInforme generado: {DST}")
    print(f"\nResumen:")
    print(f"  Acciones aprobadas         : {len(acciones)}")
    print(f"  Margen actual mes (estim)  : {fmt_clp(total_margen_actual_mes)}")
    print(f"  Margen proyectado mes      : {fmt_clp(total_margen_proyectado_mes)}")
    print(f"  Margen transferido a ganador: {fmt_clp(total_margen_transferido_mes)}")
    print(f"  IMPACTO NETO MES           : {fmt_clp(total_impacto_mes)}")
    print(f"  IMPACTO NETO ANUAL         : {fmt_clp(total_impacto_anual)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
