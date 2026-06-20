# -*- coding: utf-8 -*-
"""Lista de precios FINAL Falabella desde costos validados por Victor.
3 niveles: Normal (tachado -50%) / Venta (x3.14, 68%) / CMR (-7%). Join con precio actual."""
import json,openpyxl,re
from openpyxl.styles import Font,PatternFill,Alignment

VAL=r'C:\Users\dell\OneDrive\Desktop\validacion_costos2.xlsx'
ORIG=r'C:\Users\dell\victtorino\validacion_costos.xlsx'
OUT=r'C:\Users\dell\victtorino\PRECIOS_FINAL_falabella.xlsx'
IVA=1.19; GANCHO=0.50; CMR_OFF=0.93
OVERRIDE={'010205004-T':1077,'010701011-T':2222,   # sifon, dispensador (Victor)
          '010204004-T':1000,'010701010-T':2800}   # desague, disp.simple (costo estimado tras ver foto)
# topes por competencia (venta fijada manual tras chequeo de mercado)
VENTA_OVERRIDE={'010501022-T':199990,'040303004-T':104990,'040303003-T':99990,
                '010703004-T':54990,'010703013-T':54990,'010703007-T':54990,'010701001-T':33990,
                '010204004-T':2990,'010701010-T':8990}
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)
def factor_por_ticket(base):
    prov=base*3.14   # venta provisional para clasificar tramo
    if prov<15000: return 3.14
    if prov<50000: return 2.90
    if prov<120000: return 2.60
    return 2.30

origU={}
wo=openpyxl.load_workbook(ORIG).active
for r in range(2,wo.max_row+1): origU[wo.cell(r,1).value]=wo.cell(r,9).value

cur={c['cod']:c for c in json.load(open(r'C:\Users\dell\victtorino\precios_actuales_fala.json'))}

wb=openpyxl.load_workbook(VAL).active
final=[]; excl=[]
for r in range(2,wb.max_row+1):
    cod=wb.cell(r,1).value; desc=str(wb.cell(r,2).value)
    cdef=wb.cell(r,3).value; ctw=wb.cell(r,4).value; usar=wb.cell(r,9).value
    ok=str(wb.cell(r,11).value or '').strip().lower()
    if cod in OVERRIDE:
        costo=OVERRIDE[cod]; src='Victor'
    elif ok=='agotado':
        excl.append((cod,desc,'descontinuado')); continue
    elif ok in ('no','no lo tengo'):
        excl.append((cod,desc,ok)); continue
    else:
        changed = origU.get(cod) is not None and usar is not None and str(origU[cod])!=str(usar)
        if changed: costo=usar; src='manual'
        elif 'taumm' in ok: costo=ctw if ctw else usar; src='taumm'
        elif 'defontana' in ok: costo=cdef if cdef is not None else usar; src='defontana'
        else: costo=usar; src='sugerido'
    try: costo=int(costo)
    except: excl.append((cod,desc,'costo invalido')); continue
    base=costo*IVA; fac=factor_por_ticket(base); venta=r990(base*fac)
    if cod in VENTA_OVERRIDE: venta=VENTA_OVERRIDE[cod]; fac=round(venta/base,2)   # tope competencia
    normal=r990(venta/(1-GANCHO)); cmr=r990(venta*CMR_OFF)
    c=cur.get(cod,{})
    actual=c.get('oferta') or c.get('precio')
    final.append({'cod':cod,'sku':c.get('skuF',''),'desc':desc,'costo':costo,'src':src,'factor':fac,
                  'actual':actual,'venta':venta,'normal':normal,'cmr':cmr,
                  'marg':1-base/venta,'subida':(venta/actual-1) if actual else None,
                  'ou':'SI' if normal>=19990 else 'no'})

print('FINAL:',len(final),'| excluidos:',len(excl))
sub=[f for f in final if f['subida'] is not None and f['subida']>0.01]
baj=[f for f in final if f['subida'] is not None and f['subida']<-0.01]
nou=sum(1 for f in final if f['ou']=='SI')
print('Suben:',len(sub),'| Bajan:',len(baj),'| Elegibles Oport.Unica:',nou)
print('Muestra:')
for f in final[:8]:
    a=f['actual'] or 0
    print('  ',f['cod'],'costo',f['costo'],'| actual',a,'-> venta',f['venta'],'| normal',f['normal'],'| margen',round(f['marg']*100),'%')

# lista de ticket alto para chequeo de competencia
alto=[f for f in final if f['venta']>=50000]
alto.sort(key=lambda x:-x['venta'])
print('\nTICKET ALTO (venta>=50.000) para chequear competencia:',len(alto))
for f in alto:
    print('  ',f['sku'],'|',f['desc'][:36],'| factor',f['factor'],'| venta',f['venta'])
json.dump([{'cod':f['cod'],'sku':f['sku'],'desc':f['desc'],'venta':f['venta'],'costo':f['costo']} for f in alto],
          open(r'C:\Users\dell\victtorino\ticket_alto.json','w'))

wb2=openpyxl.Workbook(); w=wb2.active; w.title='Precios FINAL'
w.append(['Codigo','SKU Falabella','Descripcion','Costo real','Fuente','Factor','Precio actual','NORMAL (tachado)','VENTA (real)','CMR','Margen venta','Subida %','Elegible OU'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
for f in sorted(final,key=lambda x:-(x['subida'] or -9)):
    w.append([f['cod'],f['sku'],f['desc'],f['costo'],f['src'],f['factor'],f['actual'],f['normal'],f['venta'],f['cmr'],
              round(f['marg'],3),round(f['subida'],3) if f['subida'] is not None else None,f['ou']])
for col in [4,7,8,9,10]:
    for rr in range(2,w.max_row+1):
        cc=w.cell(rr,col)
        if isinstance(cc.value,(int,float)): cc.number_format='#,##0'
for rr in range(2,w.max_row+1):
    w.cell(rr,11).number_format='0.0%'
    if isinstance(w.cell(rr,12).value,(int,float)): w.cell(rr,12).number_format='0.0%'
for i,wd in enumerate([13,13,36,10,10,7,12,15,13,11,11,9,10]): w.column_dimensions[chr(65+i)].width=wd
wb2.save(OUT)
print('\nGuardado:',OUT)
print('\nExcluidos:')
for e in excl: print('   ',e[0],str(e[1])[:34],'->',e[2])
