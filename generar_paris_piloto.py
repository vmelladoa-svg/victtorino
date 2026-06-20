"""
Genera un Excel piloto con 10 productos del catalogo Woo mapeados a la hoja
gasfiteria de la plantilla de Paris (Cencommerce).
"""
import sys
import io
import json
import re
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent

# Cargar productos
prods = json.loads((ROOT / "woo_products_all.json").read_text(encoding="utf-8"))

# Filtrar viables: SKU + fotos + precio + stock, en categorias relevantes
CATS_GASFITERIA = {"Griferia", "Lavaplatos", "Accesorios", "Lavamanos",
                    "Shower/Mamparas/Receptaculos", "Dispensador"}

def es_viable(p):
    if not p.get("sku"): return False
    if not p.get("images"): return False
    if not p.get("price"): return False
    if p.get("stock_status") != "instock": return False
    cats = {c["name"] for c in (p.get("categories") or [])}
    return bool(cats & CATS_GASFITERIA)


viables = [p for p in prods if es_viable(p)]
print(f"Total viables para gasfiteria: {len(viables)}")

# Tomar 10: 4 grifería, 3 lavaplatos, 3 accesorios
def cat_principal(p):
    cats = {c["name"] for c in (p.get("categories") or [])}
    for k in ["Griferia", "Lavaplatos", "Lavamanos", "Shower/Mamparas/Receptaculos",
              "Dispensador", "Accesorios"]:
        if k in cats: return k
    return "Otros"

# Agrupar
by_cat = {}
for p in viables:
    by_cat.setdefault(cat_principal(p), []).append(p)

piloto = []
piloto += (by_cat.get("Griferia") or [])[:4]
piloto += (by_cat.get("Lavaplatos") or [])[:3]
piloto += (by_cat.get("Accesorios") or [])[:3]
print(f"\nPiloto seleccionado: {len(piloto)} productos")
for p in piloto:
    print(f"  · [{cat_principal(p)}] {p['sku']} — {p['name'][:60]}")


# === Helpers de mapeo ===

def detectar_color(name: str) -> str:
    """Extrae color comercial del nombre."""
    name_low = name.lower()
    mapa = {
        "negro mate": "Negro Mate", "negro": "Negro",
        "cromad": "Cromado", "cromo": "Cromado",
        "dorad": "Dorado", "oro": "Dorado",
        "bronce": "Bronce", "cobre": "Cobre",
        "platead": "Plateado", "blanco": "Blanco",
        "gris": "Gris", "chocolate": "Chocolate",
        "inoxidable": "Acero Inoxidable", "acero": "Acero Inoxidable",
    }
    for k, v in mapa.items():
        if k in name_low: return v
    return "Plateado"  # default


def detectar_material(name: str) -> str:
    nl = name.lower()
    if "acero inoxidable" in nl or "inoxidable" in nl: return "Acero Inoxidable"
    if "bronce" in nl: return "Bronce"
    if "cobre" in nl: return "Cobre"
    if "plástico" in nl or "plastico" in nl: return "Plástico"
    if "cerámic" in nl or "ceramic" in nl: return "Cerámica"
    return "Metal"


def strip_html(t: str) -> str:
    if not t: return ""
    t = re.sub(r"<[^>]+>", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def fotos_csv(p, max_fotos=5) -> str:
    urls = [img["src"] for img in (p.get("images") or [])][:max_fotos]
    return ",".join(urls)


def mapear_producto(p):
    """Devuelve dict con todos los campos de la hoja gasfiteria."""
    name = p.get("name", "")
    sku = p.get("sku", "")
    desc_corta = strip_html(p.get("short_description") or "")[:300]
    desc_larga = strip_html(p.get("description") or "")[:2000]
    cats = [c["name"] for c in (p.get("categories") or [])]

    # Categoria Cencommerce — aproximacion
    if "Griferia" in cats or "Lavamanos" in cats:
        cat_paris = "Repuestos Y Accesorios De Gasfitería < Llaves y Grifería"
    elif "Lavaplatos" in cats:
        cat_paris = "Repuestos Y Accesorios De Gasfitería < Lavaplatos"
    elif "Shower/Mamparas/Receptaculos" in cats:
        cat_paris = "Repuestos Y Accesorios De Gasfitería < Duchas y Receptáculos"
    elif "Dispensador" in cats:
        cat_paris = "Repuestos Y Accesorios De Gasfitería < Dispensadores"
    else:
        cat_paris = "Repuestos Y Accesorios De Gasfitería < Accesorios"

    weight = p.get("weight") or ""
    dim = p.get("dimensions") or {}
    alto = dim.get("height") or ""
    ancho = dim.get("width") or ""
    largo = dim.get("length") or ""

    color = detectar_color(name)
    material = detectar_material(name)

    return {
        "Nombre del Producto": name,
        "Sku Seller (*)": sku,
        "Ref Product": sku,
        "Categoria (*)": cat_paris,
        "Marca (*)": "Victtorino",
        "MEASUREMENT_UNIT": "Unidad",
        "UNIT_MULTIPLIER": 1,
        "Descripción corta": desc_corta or name[:200],
        "Garantía Mínima Legal": "6 meses",
        "Garantía Proveedor": "6 meses",
        "Embalaje": "Caja individual",
        "Imagenes (*)": fotos_csv(p),
        "Sku Seller Variant (*)": sku,
        "Color (*)": color,
        "Talla (*)": "Única",
        "Descripción Larga/Emocional": desc_larga or name,
        "Material": material,
        "Tono": color,
        "Origen": "Importado",
        "Peso": weight,
        "Alto": alto,
        "Ancho": ancho,
        "Profundidad": largo,
        "Modelo": sku,
        "Notas": "",
        "Tiempo de entrega al Courier (en días hábiles)": "2",
    }


# === Generar Excel ===
filas = [mapear_producto(p) for p in piloto]
df = pd.DataFrame(filas)

out = ROOT / "paris_piloto_10.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="gasfiteria_piloto", index=False)

    # Hoja con instrucciones
    inst_df = pd.DataFrame({
        "Paso": [
            "1. Abrir paris_plantilla.xlsx original (Cencommerce)",
            "2. Ir a la hoja 'gasfiteria'",
            "3. Copiar las 10 filas de datos de paris_piloto_10.xlsx (hoja 'gasfiteria_piloto', desde la fila 2)",
            "4. Pegar en la hoja 'gasfiteria' de la plantilla, debajo de las cabeceras (fila 6 en adelante)",
            "5. Validar:",
            "   - Imagenes (*): URLs separadas por coma, minimo 2",
            "   - Color, Talla: confirmar que el valor existe en el dropdown (revisar hoja 'gasfiteriadata')",
            "   - Categoria (*): puede que el dropdown tenga otro nombre — ajustar segun lo que Paris acepte",
            "6. Subir el archivo a Cencommerce y validar errores",
        ]
    })
    inst_df.to_excel(writer, sheet_name="INSTRUCCIONES", index=False)

# Formato bonito
from openpyxl import load_workbook
wb = load_workbook(out)
ws = wb["gasfiteria_piloto"]
header_fill = PatternFill("solid", fgColor="0F172A")
for j, _ in enumerate(df.columns, 1):
    c = ws.cell(row=1, column=j)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Ancho columnas
for j, col in enumerate(df.columns, 1):
    ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = min(len(col)+3, 35)
ws.freeze_panes = "A2"
wb.save(out)

print(f"\nOK  {out}")
print(f"   {len(filas)} productos mapeados en la hoja 'gasfiteria_piloto'")
print(f"   Hoja 'INSTRUCCIONES' incluida con los pasos.")
