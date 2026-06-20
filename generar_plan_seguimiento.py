"""
Genera plan_seguimiento_2026-05-23.xlsx + .md con:
  - Cronograma con fechas concretas (hitos Day +7 / +14 / +28)
  - KPIs a medir por cohorte
  - Criterios de éxito / decisión
  - Acciones de respaldo por escenario (revivió / parcial / no revivió)
  - Checklist de tareas semanales
"""
import json
from datetime import date, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).parent
OUT_XLSX = ROOT / "plan_seguimiento_2026-05-23.xlsx"
OUT_MD = ROOT / "plan_seguimiento_2026-05-23.md"

BASELINE_DATE = date(2026, 5, 23)
CHECK_DATES = {
    "D+7":  BASELINE_DATE + timedelta(days=7),
    "D+14": BASELINE_DATE + timedelta(days=14),
    "D+28": BASELINE_DATE + timedelta(days=28),
}


def main():
    # ===== Hoja 1: Cronograma =====
    cronograma = pd.DataFrame([
        {"Fecha": str(BASELINE_DATE), "Hito": "BASELINE (hoy)",
         "Acción": "Cambios aplicados: 24 upgrades gold_pro + 2 sync attrs + 14 fotos + 1 título + 7 bajadas precio. Baseline JSON capturado.",
         "Responsable": "Claude (ya hecho)", "Estado": "✓ Hecho"},
        {"Fecha": str(BASELINE_DATE), "Hito": "Activar Mercado Ads C3",
         "Acción": "Crear campaña 10 SKUs en Seller Center con bids individuales del instructivo",
         "Responsable": "Victor (manual)", "Estado": "⏳ Pendiente"},
        {"Fecha": str(BASELINE_DATE + timedelta(days=3)), "Hito": "D+3 — Check rápido Ads",
         "Acción": "Sólo Mercado Ads: confirmar que la campaña recibe impresiones (>500/día) y clicks (>10/día)",
         "Responsable": "Victor en Seller Center", "Estado": "⏳"},
        {"Fecha": str(CHECK_DATES["D+7"]), "Hito": "D+7 — Primer check",
         "Acción": "Correr: auditoria_ml.py + fix_visitas_snapshots.py + analisis_auditoria_ml.py + seguimiento_comparar.py",
         "Responsable": "Claude (ejecutar)", "Estado": "⏳"},
        {"Fecha": str(CHECK_DATES["D+14"]), "Hito": "D+14 — Check principal",
         "Acción": "Re-correr scripts. Decisión: escalar / mantener / revertir. Reporte ejecutivo.",
         "Responsable": "Claude + Victor", "Estado": "⏳"},
        {"Fecha": str(CHECK_DATES["D+28"]), "Hito": "D+28 — Cierre y plan v2",
         "Acción": "Análisis final 30d post. Decisión sobre los 'muertos' (republicar / pausar). Plan próximas acciones.",
         "Responsable": "Claude + Victor", "Estado": "⏳"},
    ])

    # ===== Hoja 2: KPIs por cohorte =====
    kpis = pd.DataFrame([
        {"Cohorte": "upgrade_listing (24)", "KPI principal": "Visitas 30d", "Baseline": "157 total",
         "Meta D+14": "+50% (~235 visitas)", "Meta D+28": "+100% (~315 visitas)",
         "Si falla": "Re-evaluar listing_type vs categoría / atributos"},
        {"Cohorte": "upgrade_listing (24)", "KPI principal": "Ventas 30d", "Baseline": "estimar 4-5",
         "Meta D+14": "8-10", "Meta D+28": "12-15",
         "Si falla": "Revisar precio competencia + categoría"},
        {"Cohorte": "bajada_precio (7)", "KPI principal": "Conversión (Vendidos/Visitas)", "Baseline": "0%",
         "Meta D+14": ">1% (al menos 5-10 ventas)", "Meta D+28": ">3%",
         "Si falla": "Investigar competencia, calidad ficha, fotos"},
        {"Cohorte": "bajada_precio (7)", "KPI principal": "Ventas absolutas", "Baseline": "0",
         "Meta D+14": "3-5 ventas", "Meta D+28": "10+ ventas",
         "Si falla": "Probar otra bajada de 5% más o pausar"},
        {"Cohorte": "fotos_cross (14)", "KPI principal": "Health avg", "Baseline": "37.6",
         "Meta D+14": ">50", "Meta D+28": ">60",
         "Si falla": "Completar atributos faltantes en estos items"},
        {"Cohorte": "sync_atributos (2)", "KPI principal": "Visitas 30d", "Baseline": "6",
         "Meta D+14": "20+", "Meta D+28": "50+",
         "Si falla": "Sync atributos RISKY (FINISH, dimensiones) si Victor valida físicamente"},
        {"Cohorte": "titulo_playwright (1)", "KPI principal": "Visitas 30d", "Baseline": "0",
         "Meta D+14": ">10 (señal de revivir)", "Meta D+28": ">25",
         "Si falla": "No escalar a más cambios de título. Considerar republicar."},
        {"Cohorte": "ads_top10 (10)", "KPI principal": "ROAS (Revenue/Gasto Ads)", "Baseline": "N/A (sin Ads)",
         "Meta D+14": ">3x", "Meta D+28": "5-8x",
         "Si falla": "Bajar bids o eliminar SKUs no rentables. Pausar campaña si ROAS < 2x"},
        {"Cohorte": "ads_top10 (10)", "KPI principal": "Ventas atribuidas semanales", "Baseline": "N/A",
         "Meta D+14": "10-25/sem total", "Meta D+28": "20-50/sem",
         "Si falla": "Revisar segmentación y bid"},
        {"Cohorte": "control (30)", "KPI principal": "Visitas 30d", "Baseline": "393",
         "Meta D+14": "~393 (sin cambio esperado)", "Meta D+28": "~393",
         "Si falla": "Grupo control no debería moverse mucho — si sube similar a intervenidos, atribución es ruido (no nuestros cambios)"},
    ])

    # ===== Hoja 3: Criterios de decisión =====
    decisiones = pd.DataFrame([
        {"Escenario en D+14": "ÉXITO TOTAL — todas las cohortes intervenidas suben +30%+ vs control",
         "Lectura": "La estrategia funciona. Escalar.",
         "Acción": "1. Aplicar upgrade gold_pro a items ROI+ pendientes (los que excluimos por margen). 2. Bajar precio a otros 5-10 items con tráfico bajo. 3. Mantener Ads."},
        {"Escenario en D+14": "ÉXITO PARCIAL — Ads + bajada precio suben pero upgrade no se mueve",
         "Lectura": "Upgrade solo no alcanza. El precio + visibilidad pagada son los drivers.",
         "Acción": "1. Mantener Ads. 2. Escalar bajada precio a más items. 3. Para items sin movimiento, considerar republicar o pausar."},
        {"Escenario en D+14": "Ads funciona, resto no",
         "Lectura": "El algoritmo orgánico es resistente. Solo presencia pagada mueve la aguja.",
         "Acción": "1. Re-balancear inversión a Ads. 2. Pausar items 'muertos' orgánicamente. 3. Plan estratégico de catálogo (menos items pero mejor curados)."},
        {"Escenario en D+14": "Nada se mueve significativamente",
         "Lectura": "Cambios fueron muy chicos para mover el algoritmo, O existe bloqueo estructural (catálogo, categoría, reputación).",
         "Acción": "1. Diagnóstico profundo de 5 items mejorados: por qué siguen invisibles. 2. Considerar republicar como reset. 3. Revisar reputación general de cuenta."},
        {"Escenario en D+14": "Algunos items se MUEVEN HACIA ABAJO (regresión)",
         "Lectura": "Cambios fueron contraproducentes (raro pero posible).",
         "Acción": "1. Identificar cuáles. 2. Revertir el cambio específico via API. 3. Análisis de causa raíz."},
    ])

    # ===== Hoja 4: Checklist semanal =====
    checklist = pd.DataFrame([
        {"Semana": "1 (D+0 a D+7)", "Tarea": "Activar Mercado Ads C3 (10 SKUs)", "Tiempo": "15 min", "Responsable": "Victor", "Hecho?": ""},
        {"Semana": "1", "Tarea": "Verificar piloto de título quedó aplicado (MLC1858919499)", "Tiempo": "1 min", "Responsable": "Victor (visual)", "Hecho?": ""},
        {"Semana": "1", "Tarea": "D+3: chequear primeras métricas de campaña Ads", "Tiempo": "5 min", "Responsable": "Victor", "Hecho?": ""},
        {"Semana": "1", "Tarea": "D+7: correr seguimiento (re-snapshot + comparar)", "Tiempo": "5 min", "Responsable": "Claude (ejec)", "Hecho?": ""},
        {"Semana": "2 (D+7 a D+14)", "Tarea": "Ajustar bids Ads según ROAS preliminar", "Tiempo": "10 min", "Responsable": "Victor", "Hecho?": ""},
        {"Semana": "2", "Tarea": "D+14: re-snapshot completo + análisis ejecutivo", "Tiempo": "15 min", "Responsable": "Claude", "Hecho?": ""},
        {"Semana": "2", "Tarea": "Tomar decisión: escalar / mantener / pausar según escenario", "Tiempo": "30 min", "Responsable": "Victor + Claude", "Hecho?": ""},
        {"Semana": "3 (D+14 a D+21)", "Tarea": "Ejecutar acciones decididas en D+14", "Tiempo": "1-2h", "Responsable": "Claude", "Hecho?": ""},
        {"Semana": "4 (D+21 a D+28)", "Tarea": "Re-snapshot final 28d", "Tiempo": "10 min", "Responsable": "Claude", "Hecho?": ""},
        {"Semana": "4", "Tarea": "Reporte ejecutivo final + plan próxima ronda", "Tiempo": "1h", "Responsable": "Claude + Victor", "Hecho?": ""},
    ])

    # ===== Hoja 5: Comandos exactos para ejecutar =====
    comandos = pd.DataFrame([
        {"Fecha": str(CHECK_DATES["D+7"]),
         "Comando": "python auditoria_ml.py",
         "Descripción": "Re-extrae snapshots ML de C1/C2/C3 con publicaciones, órdenes, reputación, preguntas",
         "Duración": "~3 min"},
        {"Fecha": str(CHECK_DATES["D+7"]),
         "Comando": "python fix_visitas_snapshots.py",
         "Descripción": "Re-extrae visitas 30d por item (concurrente, 8 threads)",
         "Duración": "~1 min"},
        {"Fecha": str(CHECK_DATES["D+7"]),
         "Comando": "python analisis_auditoria_ml.py",
         "Descripción": "Re-construye dataframe consolidado con health scores e issues",
         "Duración": "~10 seg"},
        {"Fecha": str(CHECK_DATES["D+7"]),
         "Comando": "python seguimiento_comparar.py",
         "Descripción": "Compara baseline vs snapshot actual. Genera seguimiento_diff_<fecha>.xlsx",
         "Duración": "~5 seg"},
        {"Fecha": "Recurrente",
         "Comando": "python generar_dashboard_auditoria.py + generar_informe_ejecutivo.py",
         "Descripción": "Re-genera dashboard Excel + DOCX ejecutivo con datos frescos",
         "Duración": "~30 seg"},
    ])

    # Escribir Excel
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        cronograma.to_excel(writer, sheet_name="1. Cronograma", index=False)
        kpis.to_excel(writer, sheet_name="2. KPIs y metas", index=False)
        decisiones.to_excel(writer, sheet_name="3. Criterios decisión", index=False)
        checklist.to_excel(writer, sheet_name="4. Checklist semanal", index=False)
        comandos.to_excel(writer, sheet_name="5. Comandos a ejecutar", index=False)

    # Formato
    wb = load_workbook(OUT_XLSX)
    HDR = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = HDR; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        widths = {}
        for row in ws.iter_rows():
            for c in row:
                if hasattr(c, "column_letter") and c.value is not None:
                    widths[c.column_letter] = max(widths.get(c.column_letter, 0), min(len(str(c.value)), 90))
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = min(w + 2, 60)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)

    # Markdown
    md = ["# Plan de Seguimiento — Optimización ML 2026-05-23", ""]
    md.append(f"Baseline: **{BASELINE_DATE}** · Próximos hitos: **D+7 ({CHECK_DATES['D+7']})**, **D+14 ({CHECK_DATES['D+14']})**, **D+28 ({CHECK_DATES['D+28']})**")
    md.append("")
    md.append("## 1. Cronograma")
    md.append("")
    for _, r in cronograma.iterrows():
        md.append(f"### {r['Fecha']} — {r['Hito']}")
        md.append(f"- **Acción**: {r['Acción']}")
        md.append(f"- **Responsable**: {r['Responsable']}")
        md.append(f"- **Estado**: {r['Estado']}")
        md.append("")

    md.append("## 2. KPIs y metas por cohorte")
    md.append("")
    md.append("| Cohorte | KPI | Baseline | Meta D+14 | Meta D+28 | Si falla |")
    md.append("|---|---|---|---|---|---|")
    for _, r in kpis.iterrows():
        md.append(f"| {r['Cohorte']} | {r['KPI principal']} | {r['Baseline']} | {r['Meta D+14']} | {r['Meta D+28']} | {r['Si falla']} |")

    md.append("")
    md.append("## 3. Criterios de decisión D+14")
    md.append("")
    for _, r in decisiones.iterrows():
        md.append(f"### {r['Escenario en D+14']}")
        md.append(f"- **Lectura**: {r['Lectura']}")
        md.append(f"- **Acción**: {r['Acción']}")
        md.append("")

    md.append("## 4. Comandos a ejecutar en D+7 / D+14 / D+28")
    md.append("")
    md.append("```bash")
    md.append("# Re-snapshot completo (3 cuentas)")
    md.append("python auditoria_ml.py")
    md.append("")
    md.append("# Fix visitas (endpoint multi-item de ML acepta solo 1 ID por call)")
    md.append("python fix_visitas_snapshots.py")
    md.append("")
    md.append("# Re-construir dataframe")
    md.append("python analisis_auditoria_ml.py")
    md.append("")
    md.append("# Comparar baseline vs hoy → genera seguimiento_diff_<fecha>.xlsx")
    md.append("python seguimiento_comparar.py")
    md.append("")
    md.append("# Opcional: regenerar dashboard ejecutivo con datos frescos")
    md.append("python generar_dashboard_auditoria.py")
    md.append("python generar_informe_ejecutivo.py")
    md.append("```")

    md.append("")
    md.append("## 5. Archivos clave de la sesión")
    md.append("")
    md.append("- `data/auditoria/seguimiento_baseline_2026-05-23.json` — baseline inmutable (NO modificar)")
    md.append("- `data/auditoria/analisis.pkl` — dataframe consolidado (se regenera con cada análisis)")
    md.append("- `data/auditoria/snapshot_c1/c2/c3.json` — snapshots crudos por cuenta")
    md.append("- `data/auditoria/upgrade_listing_*.json` — logs de upgrades aplicados")
    md.append("- `data/auditoria/bajar_precio_*.json` — log bajada precios")
    md.append("- `data/auditoria/sync_atributos_*.json` — log sync atributos")
    md.append("- `data/auditoria/fotos_cross_*.json` — log cambios de fotos")
    md.append("- `data/auditoria/titulos_playwright_*.json` — log cambios de título")
    md.append("- `instructivo_mercado_ads_c3.xlsx` + `.md` — guía para activar Ads")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")
    print(f"OK Excel: {OUT_XLSX}")
    print(f"OK MD:    {OUT_MD}")


if __name__ == "__main__":
    main()
