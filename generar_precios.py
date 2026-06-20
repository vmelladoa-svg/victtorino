# -*- coding: utf-8 -*-
"""Genera la lista de precios canonica por CODIGO INTERNO desde los costos
Defontana aplicando la formula por canal (mayo 2026). Redondeo a 990."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COSTS=r"C:\Users\dell\Downloads\costos_defontana_limpio.xlsx"
OUT=r"C:\Users\dell\victtorino\precios_por_codigo.xlsx"

# canal: (etiqueta, multiplicador)
CANALES=[
    ('ML Ticket Alto',1.57),
    ('ML Normal',2.85),
    ('Falabella',3.14),
    ('Paris/Walmart',2.99),
    ('Web',2.26),
    ('Tienda Fisica',2.06),
    ('Mayorista',1.85),
]
IVA=1.19   # costos del archivo son NETOS -> Base c/IVA = neto x 1.19
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

wb=openpyxl.load_workbook(COSTS,read_only=True); ws=wb.active
data=[]
for r in ws.iter_rows(values_only=True):
    if not r or r[0] in (None,'Codigo') or not isinstance(r[2],(int,float)): continue
    data.append((r[0],r[1],int(r[2])))

wbo=openpyxl.Workbook(); w=wbo.active; w.title='Precios por canal'
hdr=['Codigo','Descripcion','Costo (c/IVA)']+[c[0] for c in CANALES]
w.append(hdr)
for i,c in enumerate(w[1]):
    c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(horizontal='center')
    c.fill=PatternFill('solid',fgColor='1F4E79' if hdr[i]!='Falabella' else 'C00000')
sincosto=0
for cod,desc,cost in data:
    if cost<=0:
        w.append([cod,desc,cost]+['s/c']*len(CANALES)); sincosto+=1
    else:
        base=cost*IVA
        w.append([cod,desc,cost]+[r990(base*m) for _,m in CANALES])
# formato moneda
for col in range(3,3+1+len(CANALES)):
    for rr in range(2,w.max_row+1):
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
widths=[12,46,13]+[14]*len(CANALES)
for i,wd in enumerate(widths): w.column_dimensions[chr(65+i)].width=wd
# resaltar columna Falabella
fcol=4+[c[0] for c in CANALES].index('Falabella')
for rr in range(2,w.max_row+1):
    w.cell(rr,fcol).fill=PatternFill('solid',fgColor='FCE4D6')

wbo.save(OUT)
print(f"Generado: {OUT}")
print(f"Productos: {len(data)} | sin costo (s/c): {sincosto}")
print("\nMuestra (Codigo | Desc | Neto | Falabella x3.14 sobre c/IVA):")
for cod,desc,cost in data[:8]:
    fp=r990(cost*IVA*3.14) if cost>0 else 's/c'
    print(f"  {cod:>13} | {desc[:34]:34} | ${cost:>6,} | ${fp if isinstance(fp,str) else format(fp,',')}")
