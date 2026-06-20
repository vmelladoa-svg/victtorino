# -*- coding: utf-8 -*-
"""Lista de precios Falabella en 3 niveles por CODIGO INTERNO:
  - Precio NORMAL (tachado/lista) = Costo x 4.49  -> alza moderada (~30% dcto exhibido)
  - Precio VENTA (oferta real)    = Costo x 3.14  -> formula, margen 68% (lo que de verdad cobras)
  - Precio CMR (Oportunidad unica)= oferta x 0.93 -> -7% extra solo CMR
Redondeo a 990. Marca elegibilidad Oportunidad unica (normal >= $19.990)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COSTS=r"C:\Users\dell\Downloads\costos_defontana_limpio.xlsx"
OUT=r"C:\Users\dell\victtorino\precios_3niveles_falabella.xlsx"
M_NORMAL=4.49; M_VENTA=3.14; CMR_OFF=0.93   # CMR = -7% sobre la venta
IVA=1.19   # costos del archivo son NETOS -> Base (c/IVA) = costo_neto x 1.19

def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

wb=openpyxl.load_workbook(COSTS,read_only=True); ws=wb.active
data=[(r[0],r[1],int(r[2])) for r in ws.iter_rows(values_only=True)
      if r and r[0] not in (None,'Codigo') and isinstance(r[2],(int,float))]

wbo=openpyxl.Workbook(); w=wbo.active; w.title='Precios 3 niveles'
hdr=['Codigo','Descripcion','Costo','Precio NORMAL','Precio VENTA (real)','% dcto exhibido',
     'Precio CMR','% dcto CMR','Margen venta','Margen CMR','Elegible Oport.Unica']
w.append(hdr)
for c in w[1]:
    c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(horizontal='center',wrap_text=True)
    c.fill=PatternFill('solid',fgColor='1F4E79')
nelig=0
for cod,desc,cost in data:
    if cost<=0:
        w.append([cod,desc,cost]+['s/c']*8); continue
    base=cost*IVA   # costo neto -> con IVA
    normal=r990(base*M_NORMAL); venta=r990(base*M_VENTA); cmr=r990(venta*CMR_OFF)
    dexh=1-venta/normal; dcmr=1-cmr/normal
    mv=1-base/venta; mc=1-base/cmr
    elig = 'SI' if normal>=19990 else 'no (<19.990)'
    if normal>=19990: nelig+=1
    w.append([cod,desc,cost,normal,venta,round(dexh,3),cmr,round(dcmr,3),
              round(mv,3),round(mc,3),elig])

# formato
money_cols=[3,4,5,7]; pct_cols=[6,8,9,10]
for rr in range(2,w.max_row+1):
    for col in money_cols:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
    for col in pct_cols:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
    # resaltar venta (real) y CMR
    w.cell(rr,5).fill=PatternFill('solid',fgColor='C6EFCE')
    w.cell(rr,7).fill=PatternFill('solid',fgColor='FCE4D6')
for i,wd in enumerate([12,42,9,13,15,12,12,11,11,11,16]):
    w.column_dimensions[chr(65+i)].width=wd
wbo.save(OUT)
print(f"Generado: {OUT}")
print(f"Productos: {len(data)} | elegibles Oportunidad unica (normal>=19.990): {nelig}")
print("(costos NETOS x1.19 = Base c/IVA antes de la formula)")
print(f"\n{'Desc':30} {'Neto':>6} {'c/IVA':>6} {'NORMAL':>8} {'VENTA':>8} {'dcto':>5} {'CMR':>8}")
for cod,desc,cost in data[:7]:
    if cost<=0: continue
    base=cost*IVA
    normal=r990(base*M_NORMAL); venta=r990(base*M_VENTA); cmr=r990(venta*CMR_OFF)
    print(f"{desc[:30]:30} {cost:>6,} {int(base):>6,} {normal:>8,} {venta:>8,} {1-venta/normal:>4.0%} {cmr:>8,}")
