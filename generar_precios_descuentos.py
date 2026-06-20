# -*- coding: utf-8 -*-
"""Excel maestro de precios Falabella para estrategia tachado+oferta.
Fija el PRECIO DE VENTA REAL (Costo c/IVA x 3.14, margen 68%) y entrega, por
nivel de descuento exhibido, el PRECIO BASE/TACHADO necesario para mostrar ese %
manteniendo constante lo que paga el cliente.
  base_tachado(d) = venta_real / (1 - d)   (redondeo 990)
Se usa para monitoreo diario: el 'precio normal' en Falabella debe ser el base
del nivel elegido y el 'precio oferta/special' debe ser la VENTA real.
"""
import csv, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COSTS = r"C:\Users\dell\victtorino\costos_finales.json"
OUT   = r"C:\Users\dell\victtorino\precios_descuentos_falabella.xlsx"
IVA   = 1.19
M_VENTA = 1.8           # factor competitivo Falabella (margen ~44%, a mercado vs TAUMM)
CMR_OFF = 0.93          # -7% extra solo CMR
NIVELES = [0.20, 0.30, 0.40, 0.50, 0.60]   # descuentos exhibidos

# Factor 1.8 base. OVERRIDE = precio de venta a mercado (estandar: 1 escalon bajo
# el competidor TAUMM mas barato). El tachado se calcula = venta/(1-dcto).
OVERRIDE = {
    "010701007-T": 13990,   # Dispensador (P1) - BAUMART 13.490
    "020101010-T": 57990,   # Lavaplatos doble der (P1) - GYS 59.990
    "020101011-T": 57990,   # Lavaplatos doble izq (P1) - GYS 59.990
    "020101001-T": 38990,   # Lavaplatos simple 100x44 der - GYS 39.990
    "020101002-T": 38990,   # Lavaplatos simple 100x44 izq - GYS 39.990
    "020101003-T": 28990,   # Lavaplatos simple 80x44 - GYS 29.990
    "040303003-T": 81990,   # Fluxometro palanca - Inversiones MP 82.990
    "040203017-C": 21990,   # Monomando Lavatorio Alto Moderna - comp 22.380 (m21%)
    "040201006-T": 10990,   # Monomando Lavaplatos Vertical - comp 11.990 (m33%)
    "040202005-T": 10990,   # Monomando Lavatorio Colomba - comp 11.990 (m42%)
}

def r990(x):
    x = int(round(x))
    return (x // 1000) * 1000 + 990 if x >= 1000 else max(0, x)

# Leer costos curados (neto) desde costos_finales.json
data = []
with open(COSTS, encoding="utf-8") as f:
    for x in json.load(f):
        data.append((x["cod"], x["desc"], int(x["costo"]), x.get("src", ""), bool(x.get("agot"))))

wb = openpyxl.Workbook()
w = wb.active
w.title = "Precios y descuentos"

hdr = ["Codigo", "Descripcion", "Fuente", "Agotado", "Costo neto", "Costo c/IVA",
       "VENTA real (oferta)", "Margen", "CMR (-7%)"]
hdr += [f"Tachado {int(d*100)}%" for d in NIVELES]
w.append(hdr)
NC = len(hdr)   # total columnas

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in w[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.border = border

for cod, desc, costo, src, agot in data:
    agot_txt = "AGOTADO" if agot else ""
    if costo <= 0:
        w.append([cod, desc, src, agot_txt, costo, "s/c"] + ["s/c"] * (3 + len(NIVELES)))
        continue
    base_iva = costo * IVA
    venta = OVERRIDE.get(cod) or r990(base_iva * M_VENTA)
    margen = 1 - base_iva / venta
    cmr = r990(venta * CMR_OFF)
    tachados = [r990(venta / (1 - d)) for d in NIVELES]
    w.append([cod, desc, src, agot_txt, costo, round(base_iva), venta, round(margen, 3), cmr] + tachados)

# Formato (col: 1 Cod 2 Desc 3 Fuente 4 Agot 5 Neto 6 c/IVA 7 VENTA 8 Margen 9 CMR 10.. tachados)
money_cols = [5, 6, 7, 9] + list(range(10, 10 + len(NIVELES)))
for rr in range(2, w.max_row + 1):
    for col in money_cols:
        cell = w.cell(rr, col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"
    mc = w.cell(rr, 8)
    if isinstance(mc.value, (int, float)):
        mc.number_format = "0.0%"
    w.cell(rr, 7).fill = PatternFill("solid", fgColor="C6EFCE")   # venta real
    w.cell(rr, 9).fill = PatternFill("solid", fgColor="FCE4D6")   # CMR
    if w.cell(rr, 1).value in OVERRIDE:
        w.cell(rr, 1).fill = PatternFill("solid", fgColor="FFF2CC")  # override marcado
    if w.cell(rr, 4).value == "AGOTADO":
        w.cell(rr, 4).font = Font(color="C00000", bold=True)
    for col in range(1, NC + 1):
        w.cell(rr, col).border = border

widths = [13, 42, 11, 9, 10, 11, 16, 9, 12] + [12] * len(NIVELES)
for i, wd in enumerate(widths):
    w.column_dimensions[chr(65 + i)].width = wd
w.freeze_panes = "A2"

wb.save(OUT)
print(f"Generado: {OUT}")
print(f"Productos con costo valido: {sum(1 for r in data if r[2]>0)} / {len(data)}")
print(f"Niveles de descuento: {[int(d*100) for d in NIVELES]}%")
print(f"\nEjemplo (Dispensador 010701007-T):")
for cod, desc, costo, *_ in data:
    if cod == "010701007-T":
        base_iva = costo * IVA
        venta = OVERRIDE.get(cod) or r990(base_iva * M_VENTA)
        print(f"  costo {costo:,} -> c/IVA {int(base_iva):,} -> VENTA {venta:,}")
        for d in NIVELES:
            print(f"    tachado {int(d*100)}% = {r990(venta/(1-d)):,}  (cliente paga {venta:,})")
