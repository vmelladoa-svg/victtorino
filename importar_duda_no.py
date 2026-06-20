"""
Importa los productos marcados NO en duda_revision.xlsx.
Cruza con dudosos_revision.xlsx por titulo para recuperar ml_id y ml_cat.
Mismo flujo SEO premium + pausa entre POSTs.
"""
import json
import sys
import io
import time
import re
import unicodedata
import warnings
import requests
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)
from seo_premium_lote import plantilla_premium, meta_desc_premium, meta_title_premium
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

with open(r"C:\Users\dell\victtorino\tokens_cuenta3.json", encoding="utf-8") as f:
    tk = json.load(f)
ML_HEAD = {"Authorization": f"Bearer {tk['access_token']}"}

MAPEO = {
    "Lavaplatos de Cocina":          ("Lavaplatos",                    "lavaplatos cocina"),
    "Griferías Convencionales":      ("Griferia",                      "llave grifería"),
    "Sifones de Desague":            ("Sifones y Desagües",            "sifón desagüe"),
    "Agarraderas":                   ("Agarraderas y Barras",          "agarradera baño"),
    "De Papel":                      ("Accesorios",                    "dispensador papel higiénico"),
    "Otros":                         ("Accesorios",                    "accesorios baño"),
    "De Jabón y Alcohol en Gel":     ("Dispensador",                   "dispensador de jabón"),
    "Sets de Baño":                  ("Accesorios",                    "set accesorios baño"),
    "Mamparas y Cabinas":            ("Shower/Mamparas/Receptaculos",  "mampara ducha"),
    "Duchas Higiénicas":             ("Griferia",                      "ducha higiénica"),
    "Mangos para Ducha":             ("Griferia",                      "ducha teléfono"),
    "Válvulas de Descarga para WC":  ("WC e Inodoros",                 "válvula descarga WC"),
    "Mangueras":                     ("Accesorios",                    "manguera baño"),
    "Papeles Higiénicos":            ("Accesorios",                    "dispensador papel higiénico"),
    "Tapas de WC":                   ("WC e Inodoros",                 "tapa WC"),
    "Válvulas de Descarga":          ("WC e Inodoros",                 "válvula descarga"),
    "Mezcladores":                   ("Griferia",                      "mezcladora baño"),
    "Espejos":                       ("Espejos",                       "espejo baño"),
}


def slugify(s):
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:80]


def woo_request(method, path, **kwargs):
    url = f"{WC}/wp-json/wc/v3{path}"
    kwargs.setdefault("timeout", 120)
    kwargs.setdefault("params", {}).update(P)
    for n in range(1, 5):
        try:
            r = requests.request(method, url, **kwargs)
            if r.status_code == 503:
                espera = 8 * n
                print(f"    503 -> espera {espera}s")
                time.sleep(espera)
                continue
            if r.status_code >= 400:
                print(f"    HTTP {r.status_code}: {r.text[:200]}")
                return None
            return r.json()
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            espera = 4 * n
            print(f"    RED {type(e).__name__}; espera {espera}s")
            time.sleep(espera)
    return None


# 1) Leer duda_revision.xlsx -> titulos NO
print("[1/4] Leyendo duda_revision.xlsx...")
wb1 = load_workbook(r"C:\Users\dell\victtorino\duda_revision.xlsx", data_only=True)
ws1 = wb1["DUDA (14)"]
titulos_no = set()
for row in ws1.iter_rows(min_row=2, values_only=True):
    dec = (row[0] or "").strip().upper() if row[0] else ""
    if dec == "NO" and row[3]:
        titulos_no.add(row[3].strip())
print(f"      {len(titulos_no)} titulos NO")

# 2) Cruzar con dudosos_revision.xlsx para obtener ml_id + ml_cat
print("[2/4] Cruzando con dudosos_revision.xlsx...")
wb2 = load_workbook(r"C:\Users\dell\victtorino\dudosos_revision.xlsx", data_only=True)
ws2 = wb2["Dudosos ML vs Web"]
a_importar = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    titulo = (row[3] or "").strip()
    if titulo in titulos_no:
        a_importar.append({"ml_id": row[2], "ml_titulo": titulo, "ml_cat": row[6]})
print(f"      Cruzados: {len(a_importar)} de {len(titulos_no)}")

# 3) Cargar IDs cats Woo
print("[3/4] Cargando categorias Woo...")
cats = woo_request("GET", "/products/categories", params={"per_page": 100})
cat_id_por_nombre = {c["name"]: c["id"] for c in cats}

# 4) Importar
print(f"\n[4/4] Importando {len(a_importar)} productos con pausa de 5s...\n")
resultados, omitidos = [], []
for idx, f in enumerate(a_importar, start=1):
    mid = f["ml_id"]
    ml_cat = f["ml_cat"]
    sku = f"ML-{mid}"

    ya = woo_request("GET", "/products", params={"sku": sku})
    if ya:
        print(f"({idx:2}/{len(a_importar)}) {mid} -> YA EXISTE Woo {ya[0]['id']}. Skip.")
        resultados.append({"ml_id": mid, "woo_id": ya[0]["id"], "reused": True})
        continue

    mapeo = MAPEO.get(ml_cat)
    if not mapeo:
        print(f"({idx:2}/{len(a_importar)}) {mid} -> SIN MAPEO para '{ml_cat}', uso Accesorios")
        mapeo = ("Accesorios", "accesorios baño")
    cat_destino, focus = mapeo

    item = requests.get(f"https://api.mercadolibre.com/items/{mid}", headers=ML_HEAD, timeout=30).json()
    titulo = item.get("title", f["ml_titulo"])
    precio = item.get("price")
    stock = int(item.get("available_quantity") or 0)
    pictures = [p["url"] for p in item.get("pictures", []) if p.get("url")]

    desc_html = plantilla_premium(cat_destino, titulo, focus)
    meta_t = meta_title_premium(titulo, focus)
    meta_d = meta_desc_premium(titulo, cat_destino, focus)
    short_d = (f"<p>{titulo}. <strong>{focus.capitalize()}</strong> con diseño "
               "contemporáneo, materiales resistentes y despacho a todo Chile. "
               "Renueva tu baño con piezas que duran y se ven bien.</p>")
    images_payload = [{"src": u, "alt": (titulo if i == 0 else f"{titulo} - foto {i+1}")[:120]}
                      for i, u in enumerate(pictures)]
    body = {
        "name": titulo, "slug": slugify(titulo), "type": "simple", "status": "draft",
        "regular_price": str(int(precio)) if precio is not None else "",
        "manage_stock": True, "stock_quantity": stock,
        "description": desc_html, "short_description": short_d, "sku": sku,
        "categories": [{"id": cat_id_por_nombre[cat_destino]}],
        "images": images_payload,
        "meta_data": [
            {"key": "rank_math_title", "value": meta_t},
            {"key": "rank_math_description", "value": meta_d},
            {"key": "rank_math_focus_keyword", "value": focus},
        ],
    }
    nuevo = woo_request("POST", "/products", json=body)
    if not nuevo:
        print(f"({idx:2}/{len(a_importar)}) {mid} FALLO")
        omitidos.append({"ml_id": mid, "razon": "post fallo"})
    else:
        palabras = len(re.sub(r"<[^>]+>", " ", nuevo["description"]).split())
        print(f"({idx:2}/{len(a_importar)}) {mid} -> Woo {nuevo['id']} [{cat_destino[:20]:20}] palabras={palabras:4} {titulo[:45]}")
        resultados.append({
            "ml_id": mid, "ml_titulo": titulo, "ml_cat": ml_cat,
            "woo_id": nuevo["id"], "woo_cat": cat_destino,
            "permalink": nuevo.get("permalink"),
            "edit_url": f"{WC}/wp-admin/post.php?post={nuevo['id']}&action=edit",
        })

    if idx < len(a_importar):
        time.sleep(5)

print("\n" + "=" * 60)
print(f"RESUMEN: {len(resultados)} importados, {len(omitidos)} omitidos")
print("=" * 60)

with open(r"C:\Users\dell\victtorino\importar_duda_resultado.json", "w", encoding="utf-8") as f:
    json.dump({"resultados": resultados, "omitidos": omitidos}, f, ensure_ascii=False, indent=2)
print("Detalle -> importar_duda_resultado.json")
