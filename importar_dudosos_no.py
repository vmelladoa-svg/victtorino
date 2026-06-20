"""
Importa a Woo (draft) los productos del Excel `dudosos_revision.xlsx` marcados como NO.

- SEO premium completo desde el primer POST (plantillas de seo_premium_lote.py)
- Idempotencia por SKU `ML-{id}`
- Retries con backoff en errores de red
- Categoria asignada segun mapeo cat ML hoja -> cat Woo
"""
import json
import sys
import io
import time
import re
import unicodedata
import requests
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# Reusar plantillas (import primero, antes de reasignar stdout, porque ese modulo
# tambien reasigna stdout y cerraria el nuestro)
from seo_premium_lote import (
    plantilla_premium, meta_desc_premium, meta_title_premium, FOCUS_POR_CAT,
)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

with open(r"C:\Users\dell\victtorino\tokens_cuenta3.json", encoding="utf-8") as f:
    tk = json.load(f)
ML_HEAD = {"Authorization": f"Bearer {tk['access_token']}"}

# Cat ML hoja -> (cat Woo destino, focus_keyword sugerido)
MAPEO = {
    "Lavaplatos de Cocina":          ("Lavaplatos",                    "lavaplatos cocina"),
    "Griferías Convencionales":      ("Griferia",                      "llave grifería"),
    "Sifones de Desague":            ("Sifones y Desagües",            "sifón desagüe"),
    "Agarraderas":                   ("Agarraderas y Barras",          "agarradera baño"),
    "De Papel":                      ("Accesorios",                    "dispensador papel higiénico"),
    "Otros":                         ("Accesorios",                    "accesorios baño"),
    "De Jabón y Alcohol en Gel":     ("Dispensador",                   "dispensador de jabón"),
    "Sets de Baño":                  ("Accesorios",                    "set accesorios baño"),
    "Mamparas y Cabinas":            ("Shower/Mamparas/Receptaculos",  "mampara ducha"),
    "Duchas Higiénicas":             ("Griferia",                      "ducha higiénica"),
    "Mangos para Ducha":             ("Griferia",                      "ducha teléfono"),
    "Válvulas de Descarga para WC":  ("WC e Inodoros",                 "válvula descarga WC"),
    "Mangueras":                     ("Accesorios",                    "manguera baño"),
    "Papeles Higiénicos":            ("Accesorios",                    "dispensador papel higiénico"),
}


def slugify(s):
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:80]


def woo_request(method, path, **kwargs):
    url = f"{WC}/wp-json/wc/v3{path}"
    kwargs.setdefault("timeout", 120)
    kwargs.setdefault("params", {}).update(P)
    for n in range(1, 4):
        try:
            r = requests.request(method, url, **kwargs)
            if r.status_code == 401:
                print(f"  ERR 401 en {method} {path}")
                return None
            if r.status_code >= 400:
                print(f"  ERR {r.status_code}: {r.text[:300]}")
                return None
            return r.json()
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            espera = 2 ** n
            print(f"  RED intento {n}/3 {type(e).__name__}; reintento en {espera}s")
            time.sleep(espera)
    return None


def existe_sku(sku):
    r = woo_request("GET", "/products", params={"sku": sku})
    if r:
        return r[0] if r else None
    return None


def detalle_ml(mid):
    item = requests.get(f"https://api.mercadolibre.com/items/{mid}",
                        headers=ML_HEAD, timeout=30).json()
    return item


# ============================================================
# 1) Leer Excel y filtrar NO
# ============================================================
print("[1/4] Leyendo Excel...")
wb = load_workbook(r"C:\Users\dell\victtorino\dudosos_revision.xlsx", data_only=True)
ws = wb["Dudosos ML vs Web"]

a_importar = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dec = (row[0] or "").strip().upper() if row[0] else ""
    if dec != "NO":
        continue
    a_importar.append({
        "ml_id": row[2],
        "ml_titulo": row[3],
        "ml_cat": row[6],
    })
print(f"      {len(a_importar)} productos marcados NO")

# ============================================================
# 2) Cargar IDs de categorias Woo
# ============================================================
print("[2/4] Cargando IDs de categorias Woo...")
cats_woo = woo_request("GET", "/products/categories", params={"per_page": 100})
cat_id_por_nombre = {c["name"]: c["id"] for c in cats_woo}

# Verificar que todas las categorias destino existen
faltan_cats = set()
for f in a_importar:
    m = MAPEO.get(f["ml_cat"])
    if not m:
        print(f"  WARN: ML cat '{f['ml_cat']}' sin mapeo")
        continue
    if m[0] not in cat_id_por_nombre:
        faltan_cats.add(m[0])
if faltan_cats:
    print(f"  ABORTAR: faltan categorias Woo: {faltan_cats}")
    sys.exit(1)
print(f"      OK, {len(cat_id_por_nombre)} categorias Woo cargadas")

# ============================================================
# 3) Procesar cada NO
# ============================================================
print(f"\n[3/4] Importando {len(a_importar)} productos como DRAFT con SEO premium...\n")
resultados = []
omitidos = []

for idx, f in enumerate(a_importar, start=1):
    mid = f["ml_id"]
    ml_cat = f["ml_cat"]
    sku = f"ML-{mid}"

    # Idempotencia
    ya = existe_sku(sku)
    if ya:
        print(f"({idx:2}/{len(a_importar)}) {mid} -> YA EXISTE Woo {ya['id']} status={ya['status']}. Skip.")
        resultados.append({"ml_id": mid, "woo_id": ya["id"], "reused": True,
                           "permalink": ya.get("permalink")})
        continue

    mapeo = MAPEO.get(ml_cat)
    if not mapeo:
        print(f"({idx:2}/{len(a_importar)}) {mid} -> SIN MAPEO para '{ml_cat}'")
        omitidos.append({"ml_id": mid, "razon": f"sin mapeo {ml_cat}"})
        continue
    cat_destino, focus = mapeo

    # Detalle ML
    item = detalle_ml(mid)
    titulo = item.get("title", f["ml_titulo"] or "")
    precio = item.get("price")
    stock = int(item.get("available_quantity") or 0)
    pictures = [p["url"] for p in item.get("pictures", []) if p.get("url")]

    # Construir body con SEO premium
    desc_html = plantilla_premium(cat_destino, titulo, focus)
    meta_t = meta_title_premium(titulo, focus)
    meta_d = meta_desc_premium(titulo, cat_destino, focus)
    short_d = (f"<p>{titulo}. <strong>{focus.capitalize()}</strong> con diseño "
               "contemporáneo, materiales resistentes y despacho a todo Chile. "
               "Renueva tu baño con piezas que duran y se ven bien.</p>")

    images_payload = []
    for i, url in enumerate(pictures, start=1):
        alt = titulo if i == 1 else f"{titulo} - foto {i}"
        images_payload.append({"src": url, "alt": alt[:120]})

    body = {
        "name": titulo,
        "slug": slugify(titulo),
        "type": "simple",
        "status": "draft",
        "regular_price": str(int(precio)) if precio is not None else "",
        "manage_stock": True,
        "stock_quantity": stock,
        "description": desc_html,
        "short_description": short_d,
        "sku": sku,
        "categories": [{"id": cat_id_por_nombre[cat_destino]}],
        "images": images_payload,
        "meta_data": [
            {"key": "rank_math_title", "value": meta_t},
            {"key": "rank_math_description", "value": meta_d},
            {"key": "rank_math_focus_keyword", "value": focus},
        ],
    }

    nuevo = woo_request("POST", "/products", json=body)
    if not nuevo:
        print(f"({idx:2}/{len(a_importar)}) {mid} FALLO post")
        omitidos.append({"ml_id": mid, "razon": "post fallo"})
        continue

    palabras = len(re.sub(r"<[^>]+>", " ", nuevo["description"]).split())
    print(f"({idx:2}/{len(a_importar)}) {mid} -> Woo {nuevo['id']} [{cat_destino[:22]:22}] palabras={palabras:4} {titulo[:50]}")
    resultados.append({
        "ml_id": mid, "ml_titulo": titulo, "ml_cat": ml_cat,
        "woo_id": nuevo["id"], "woo_cat": cat_destino,
        "permalink": nuevo.get("permalink"),
        "edit_url": f"{WC}/wp-admin/post.php?post={nuevo['id']}&action=edit",
        "palabras": palabras, "focus": focus, "stock": stock, "precio": precio,
    })

# ============================================================
# 4) Reporte
# ============================================================
print("\n" + "=" * 70)
print(f"[4/4] RESUMEN: {len(resultados)} importados, {len(omitidos)} omitidos")
print("=" * 70)

print(f"\nDistribucion por categoria Woo destino:")
from collections import Counter
for cat, n in Counter(r.get("woo_cat") for r in resultados if r.get("woo_cat")).most_common():
    print(f"  {n:3}  {cat}")

if omitidos:
    print(f"\nOmitidos:")
    for o in omitidos:
        print(f"  {o}")

with open(r"C:\Users\dell\victtorino\importar_no_resultado.json", "w", encoding="utf-8") as f:
    json.dump({"importados": resultados, "omitidos": omitidos}, f, ensure_ascii=False, indent=2)
print(f"\nDetalle -> importar_no_resultado.json")
print("\nTodos los productos quedaron en DRAFT.")
