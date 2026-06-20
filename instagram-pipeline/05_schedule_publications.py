"""
Bloque 5 — Planificacion de publicaciones en Instagram.

Genera un calendario con 1 post/dia durante ~4 meses (117 posts).
Rota los tipos de caption (tecnico → emocional → oferta) para variar el feed.

Output:
  data/calendario_instagram.csv  — plan completo de posts
  data/calendario_instagram.xlsx — version visual con formato

Despues, 06_publish_instagram.py se encarga de la publicacion real via Meta API
(necesita Access Token + Instagram Business Account ID).
"""
import sys
import io
import csv
import os
from pathlib import Path
from datetime import datetime, timedelta

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd

ROOT = Path(__file__).parent
DATA = ROOT / "data"
CAPTIONS = ROOT / "captions"

# Configuracion del calendario
START_DATE = datetime(2026, 5, 20)   # arranca manana
POST_HOUR = 19                        # 19:00 horario Chile (peak engagement)
POSTS_PER_DAY = 1
SKIP_SUNDAYS = False                  # Instagram funciona los domingos tambien

# Orden de rotacion para variar el feed (no 3 oferta seguidos)
ROTACION = ["emocional", "tecnico", "oferta"]


def cargar_shortlist():
    """Lee CSV de catalogo y filtra los aprobados con foto."""
    items = []
    with open(DATA / "catalogo_instagram.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["pass_filter"].lower() == "true" \
               and int(r["filter_score"]) >= 4 \
               and r["image_url"]:
                items.append(r)
    return items


def hashtags_de(caption_text):
    """Extrae la linea de hashtags del final del caption."""
    lines = [l.strip() for l in caption_text.splitlines() if l.strip()]
    for line in reversed(lines):
        if line.startswith("#"):
            return line
    return ""


def caption_sin_hashtags(caption_text):
    """Cuerpo del caption sin la linea final de hashtags."""
    lines = caption_text.splitlines()
    while lines and (not lines[-1].strip() or lines[-1].strip().startswith("#")):
        lines.pop()
    return "\n".join(lines).strip()


def main():
    shortlist = cargar_shortlist()
    print(f"Shortlist: {len(shortlist)} SKUs disponibles\n")

    # Generar 3 entradas por SKU = 117 posts
    # Mezclar para variar feed: ciclamos rotacion y SKU
    plan = []
    for i, tipo in enumerate(ROTACION):
        for prod in shortlist:
            caption_file = CAPTIONS / f"{prod['sku']}_{tipo}.txt"
            if not caption_file.exists():
                continue
            txt = caption_file.read_text(encoding="utf-8")
            plan.append({
                "sku": prod["sku"],
                "nombre": prod["name"],
                "tipo_caption": tipo,
                "image_url": prod["image_url"],
                "product_url": prod["product_url"],
                "category": prod["category"],
                "precio": prod["price"],
                "caption_body": caption_sin_hashtags(txt),
                "hashtags": hashtags_de(txt),
                "caption_full": txt,
                "caption_chars": len(txt),
            })

    # Asignar fechas
    fecha = START_DATE
    for entry in plan:
        if SKIP_SUNDAYS and fecha.weekday() == 6:
            fecha += timedelta(days=1)
        entry["fecha"] = fecha.strftime("%Y-%m-%d")
        entry["hora"] = f"{POST_HOUR:02d}:00"
        entry["dia_semana"] = ["lunes","martes","miercoles","jueves",
                               "viernes","sabado","domingo"][fecha.weekday()]
        fecha += timedelta(days=1)

    # CSV
    csv_path = DATA / "calendario_instagram.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        cols = ["fecha","hora","dia_semana","sku","nombre","categoria",
                "precio","tipo_caption","image_url","product_url",
                "caption_chars","hashtags"]
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for e in plan:
            w.writerow({
                "fecha": e["fecha"], "hora": e["hora"],
                "dia_semana": e["dia_semana"],
                "sku": e["sku"], "nombre": e["nombre"],
                "categoria": e["category"], "precio": e["precio"],
                "tipo_caption": e["tipo_caption"],
                "image_url": e["image_url"],
                "product_url": e["product_url"],
                "caption_chars": e["caption_chars"],
                "hashtags": e["hashtags"],
            })

    # Excel con formato bonito
    df = pd.DataFrame([{
        "Fecha": e["fecha"],
        "Día": e["dia_semana"],
        "Hora": e["hora"],
        "SKU": e["sku"],
        "Categoría": e["category"],
        "Producto": e["nombre"][:60],
        "Tipo": e["tipo_caption"],
        "Precio": int(e["precio"]),
        "Caption (preview)": e["caption_body"][:200] + "...",
        "Imagen URL": e["image_url"],
        "Producto URL": e["product_url"],
    } for e in plan])
    xlsx_path = DATA / "calendario_instagram.xlsx"
    df.to_excel(xlsx_path, index=False, sheet_name="Calendario")

    # Format basico
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = load_workbook(xlsx_path)
    ws = wb["Calendario"]
    for j, _ in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F172A")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, w_ in zip(ws.columns, [12, 12, 8, 14, 14, 35, 12, 10, 50, 35, 35]):
        ws.column_dimensions[col[0].column_letter].width = w_
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)

    # Resumen
    print(f"=== Calendario generado ===")
    print(f"  Posts totales: {len(plan)}")
    print(f"  Periodo:       {plan[0]['fecha']} → {plan[-1]['fecha']}")
    print(f"  Hora diaria:   {POST_HOUR:02d}:00 (Chile)")
    print(f"  Rotacion:      {' → '.join(ROTACION)}")
    print()
    print(f"  CSV:   {csv_path}")
    print(f"  Excel: {xlsx_path}")
    print()
    print("Primeras 5 fechas:")
    for e in plan[:5]:
        print(f"  {e['fecha']} {e['dia_semana']:<10} {e['hora']}  "
              f"[{e['tipo_caption']:<9}] {e['sku']:<14} {e['nombre'][:55]}")


if __name__ == "__main__":
    main()
