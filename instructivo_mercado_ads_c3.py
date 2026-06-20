"""
Genera instructivo Excel + Markdown para activar Mercado Ads en C3.
Top 10 SKUs con bid sugerido individual + configuración recomendada de campaña.
"""
import json
import pickle
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).parent
ANALISIS = ROOT / "data" / "auditoria" / "analisis.pkl"
OUT_XLSX = ROOT / "instructivo_mercado_ads_c3.xlsx"
OUT_MD = ROOT / "instructivo_mercado_ads_c3.md"


def main():
    data = pickle.loads(ANALISIS.read_bytes())
    df = data["df_publicaciones"]
    c3 = df[df["Cuenta"] == "C3"].copy()

    # Top 10: alto health, stock OK, vendidos>0
    candidatos = c3[(c3["Stock"] >= 3) & (c3["HealthCalc"] >= 70) & (c3["Vendidos180d"] > 0)].copy()
    candidatos["Rev/Visita"] = candidatos["Revenue180d"] / candidatos["Visitas30d"].replace(0, 1)
    candidatos["Bid sugerido CLP/click"] = (candidatos["Rev/Visita"] * 0.20).round(0).astype(int)
    top = candidatos.sort_values("Vendidos180d", ascending=False).head(10).reset_index(drop=True)

    # DataFrame final
    rows = []
    for i, r in top.iterrows():
        rows.append({
            "#": i+1,
            "Item ID": r["ItemID"],
            "Producto": (r["Título"] or "")[:60],
            "Precio CLP": int(r["Precio"]),
            "Stock": int(r["Stock"]),
            "Visitas 30d": int(r["Visitas30d"]),
            "Vendidos 180d": int(r["Vendidos180d"]),
            "Revenue 180d CLP": int(r["Revenue180d"]),
            "Conv estimada %": round(r["ConvRate%"], 2),
            "Bid CLP/click": int(r["Bid sugerido CLP/click"]),
            "Permalink": r["Permalink"],
            "Estado actual ads": "Sin campaña",
        })
    df_out = pd.DataFrame(rows)

    # Resumen / config recomendada
    total_revenue = sum(r["Revenue 180d CLP"] for r in rows)
    config_recomendada = pd.DataFrame([
        {"Parámetro": "Cuenta", "Valor": "C3 NOVAGRIFERIAS3 (advertiser_id 79197)"},
        {"Parámetro": "Tipo de campaña", "Valor": "Product Ads (productos patrocinados)"},
        {"Parámetro": "Estrategia de puja", "Valor": "Manual por SKU (recomendado para empezar) — o ACOS objetivo 20%"},
        {"Parámetro": "Presupuesto diario total", "Valor": "$10.000 CLP (ya configurado en cuenta)"},
        {"Parámetro": "Distribución sugerida", "Valor": "Compartido entre los 10 SKUs (no individual)"},
        {"Parámetro": "Período inicial", "Valor": "Sin fecha de fin — pausar a los 14 días para evaluar"},
        {"Parámetro": "Items a incluir", "Valor": f"Los 10 SKUs de la primera hoja (revenue 180d ${total_revenue:,})"},
        {"Parámetro": "Bid por item", "Valor": "Usar la columna 'Bid CLP/click' como bid manual"},
        {"Parámetro": "ACOS objetivo (si elegís automático)", "Valor": "20% — equilibrio entre volumen y rentabilidad"},
        {"Parámetro": "ROAS objetivo equivalente", "Valor": "5x (1 / 0.20)"},
    ])

    # Instructivo paso a paso
    pasos = pd.DataFrame([
        {"Paso": 1, "Acción": "Entrar a Seller Center C3",
         "Detalle": "https://www.mercadolibre.cl/ con cuenta NOVAGRIFERIAS3 (C3) logueado"},
        {"Paso": 2, "Acción": "Ir a Mercado Ads",
         "Detalle": "Menú superior → Publicidad → Mercado Ads → Product Ads. O directamente: https://www.mercadolibre.cl/ads/"},
        {"Paso": 3, "Acción": "Verificar presupuesto",
         "Detalle": "Sección 'Mi presupuesto' debe mostrar $10.000 CLP/día (ya configurado). Si no, configurarlo ahora."},
        {"Paso": 4, "Acción": "Crear nueva campaña",
         "Detalle": "Click 'Crear campaña'. Nombre sugerido: 'Top 10 Bestsellers C3 — 2026-05'"},
        {"Paso": 5, "Acción": "Agregar los 10 productos",
         "Detalle": "Buscar por Item ID o título. Agregar los 10 de la hoja 'Top 10 SKUs'. Verificar que todos quedan en la campaña."},
        {"Paso": 6, "Acción": "Configurar bid por item",
         "Detalle": "Para cada producto, ingresar manualmente el valor de columna 'Bid CLP/click'. (Si la UI no permite individual, usar bid uniforme = promedio ~$1000/click)"},
        {"Paso": 7, "Acción": "Activar campaña",
         "Detalle": "Confirmar y activar. Empieza a generar tráfico pagado en pocas horas."},
        {"Paso": 8, "Acción": "Monitorear",
         "Detalle": "A los 3 días: revisar primeras métricas (impresiones, clicks, CTR, conversión). A los 7 días: ROAS preliminar. A los 14 días: decisión de escalar/pausar/ajustar bids."},
        {"Paso": 9, "Acción": "Reportar a Claude",
         "Detalle": "Una vez activada, pasame el ID de campaña + métricas semanales para análisis."},
    ])

    # Métricas objetivo
    metas = pd.DataFrame([
        {"KPI": "Impresiones / día (suma 10 SKUs)", "Meta semana 1": "1.000+", "Meta semana 4": "3.000+"},
        {"KPI": "Clicks / día (suma 10 SKUs)", "Meta semana 1": "20-50", "Meta semana 4": "50-150"},
        {"KPI": "CTR promedio", "Meta semana 1": "1.5-3%", "Meta semana 4": "2-4%"},
        {"KPI": "Ventas atribuidas / semana", "Meta semana 1": "3-7", "Meta semana 4": "10-25"},
        {"KPI": "ROAS (revenue / gasto)", "Meta semana 1": "3x+", "Meta semana 4": "5-8x"},
        {"KPI": "ACOS (gasto / revenue)", "Meta semana 1": "<30%", "Meta semana 4": "12-20%"},
        {"KPI": "Gasto diario", "Meta semana 1": "$3-5k", "Meta semana 4": "$8-10k (usar todo el budget)"},
    ])

    # Escribir Excel
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Top 10 SKUs", index=False)
        config_recomendada.to_excel(writer, sheet_name="Configuración", index=False)
        pasos.to_excel(writer, sheet_name="Paso a paso", index=False)
        metas.to_excel(writer, sheet_name="KPIs objetivo", index=False)

    # Formato
    wb = load_workbook(OUT_XLSX)
    HDR = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(*[Side(border_style="thin", color="AAAAAA")] * 4)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = HDR
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # autosize
        widths = {}
        for row in ws.iter_rows():
            for c in row:
                if hasattr(c, "column_letter") and c.value is not None:
                    widths[c.column_letter] = max(widths.get(c.column_letter, 0), min(len(str(c.value)), 80))
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = min(w + 2, 70)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = BORDER
        ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)

    # Markdown alternativo (más fácil de leer en mensaje)
    md = []
    md.append("# Instructivo Mercado Ads C3 — Campaña Top 10 SKUs")
    md.append("")
    md.append(f"Cuenta: **NOVAGRIFERIAS3 (C3)** · Advertiser ID: `79197` · Presupuesto: `$10.000/día`")
    md.append(f"Total revenue 180d de los 10 SKUs seleccionados: **${total_revenue:,} CLP**")
    md.append(f"Si Ads agrega +30% conversión en estos items: **+${int(total_revenue*0.3/2):,} CLP/90d adicional**")
    md.append("")
    md.append("## Top 10 SKUs a publicitar")
    md.append("")
    md.append("| # | Item | Producto | Precio | Stock | Visitas 30d | Vendidos 180d | Bid sugerido |")
    md.append("|---|---|---|---|---|---|---|---|")
    for r in rows:
        md.append(f"| {r['#']} | `{r['Item ID']}` | {r['Producto']} | ${r['Precio CLP']:,} | {r['Stock']} | {r['Visitas 30d']} | {r['Vendidos 180d']} | **${r['Bid CLP/click']:,} CLP/click** |")
    md.append("")
    md.append("## Pasos para activar (en Seller Center)")
    md.append("")
    for _, p in pasos.iterrows():
        md.append(f"**{p['Paso']}. {p['Acción']}**")
        md.append(f"   - {p['Detalle']}")
        md.append("")
    md.append("## KPIs objetivo para medir")
    md.append("")
    md.append("| KPI | Semana 1 | Semana 4 |")
    md.append("|---|---|---|")
    for _, k in metas.iterrows():
        md.append(f"| {k['KPI']} | {k['Meta semana 1']} | {k['Meta semana 4']} |")
    OUT_MD.write_text("\n".join(md), encoding="utf-8")

    print(f"OK Excel: {OUT_XLSX}")
    print(f"OK Markdown: {OUT_MD}")
    print(f"\nTotal revenue 180d top 10: ${total_revenue:,}")
    print(f"Upside esperado 90d (+30% conv): ${int(total_revenue*0.3/2):,}")


if __name__ == "__main__":
    main()
