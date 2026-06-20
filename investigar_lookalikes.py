"""
Para los casos donde un externo nos gana por >100% en una ficha de catalogo,
chequea si el externo realmente vende EL MISMO producto o un "lookalike"
(mismo cpid pero distinta marca/modelo — clasificacion incorrecta en ML).

Como /items/{id} esta bloqueado para items ajenos (y para mios con otro token),
para externos scrapeamos el permalink publico y extraemos title del HTML.
Para mis items usamos el token de la cuenta dueña.

Output:
  - Consola: por cada caso, side-by-side mio vs externo
  - lookalikes_investigacion_<fecha>.xlsx con veredicto coloreado
"""
from __future__ import annotations
import re
import sys
import time
from html import unescape
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from ml_catalogo_publicaciones import CUENTAS, Cuenta, _request

ROOT = Path(__file__).parent
SRC = ROOT / f"candidatos_pausar_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"lookalikes_investigacion_{time.strftime('%Y-%m-%d')}.xlsx"

ATTRS_CLAVE = ["BRAND", "MODEL", "MODEL_NUMBER", "SELLER_SKU",
               "UNITS_PER_PACKAGE", "MAIN_COLOR", "ITEM_CONDITION"]

UMBRAL_DIFF_PCT = 100.0
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


def cargar_revisar(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    if "REVISAR_MANUAL" not in wb.sheetnames:
        return []
    ws = wb["REVISAR_MANUAL"]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def get_item_propio(cuentas_map: dict[str, Cuenta], alias: str, item_id: str) -> dict | None:
    cu = cuentas_map.get(alias)
    if not cu:
        return None
    r = _request(cu, f"/items/{item_id}")
    if r.status_code != 200:
        print(f"  WARN /items/{item_id} con token {alias}: {r.status_code}")
        return None
    return r.json()


def scrape_externo(item_id: str) -> dict:
    """Devuelve {title, price_str, marca_breadcrumb, raw_url, status}."""
    out = {"title": None, "marca_breadcrumb": None, "raw_url": None, "status": None}
    url = f"https://articulo.mercadolibre.cl/MLC-{item_id[3:]}-_JM"
    try:
        r = requests.get(url, headers=UA, allow_redirects=True, timeout=20)
        out["status"] = r.status_code
        out["raw_url"] = r.url
        if r.status_code != 200:
            return out
        html = r.text
        m_t = re.search(r"<title>(.+?)</title>", html, re.IGNORECASE | re.DOTALL)
        if m_t:
            out["title"] = unescape(m_t.group(1).strip())
            # ML pone " | MercadoLibre" o " - Mercado Libre" al final
            out["title"] = re.split(r"\s+\|\s+Mercado", out["title"])[0]
        # marca en breadcrumb o en atributo Marca
        m_brand = re.search(r">\s*Marca\s*</[^>]+>\s*<[^>]+>\s*([^<\n]{2,60})\s*<", html, re.IGNORECASE)
        if m_brand:
            out["marca_breadcrumb"] = m_brand.group(1).strip()
        # alt: extraer del JSON-LD si esta
        m_json_brand = re.search(r'"brand"\s*:\s*\{?[^}]*?"name"\s*:\s*"([^"]+)"', html)
        if m_json_brand:
            out["marca_breadcrumb"] = out["marca_breadcrumb"] or m_json_brand.group(1)
    except Exception as e:
        out["status"] = f"ERR:{e}"
    return out


def extract_attrs(item: dict) -> dict:
    out: dict = {}
    for a in (item.get("attributes") or []):
        aid = a.get("id")
        if aid in ATTRS_CLAVE:
            out[aid] = a.get("value_name")
    return out


def diagnostico(mi_title: str | None, mi_brand: str | None,
                ext_title: str | None, ext_brand: str | None) -> str:
    """Heuristica para decidir veredicto."""
    def norm(s):
        return re.sub(r"\s+", " ", (s or "").lower().strip())
    mt, et = norm(mi_title), norm(ext_title)
    mb, eb = norm(mi_brand), norm(ext_brand)

    if not et:
        return "DATOS_INSUFICIENTES"

    # marca dura: si ambas tienen marca y son distintas -> lookalike
    if mb and eb and mb != eb:
        return "LOOKALIKE_MARCA_DISTINTA"

    # marca menciondada en titulo del competidor distinta a la mia
    marcas_mi = {"tüumm", "tuumm", "victtorino", "notte", "stella"}
    marca_en_mi = next((m for m in marcas_mi if m in mt or (mb and m in mb)), None)
    if marca_en_mi and marca_en_mi not in et and not (eb and marca_en_mi in eb):
        return "LOOKALIKE_MARCA_AUSENTE_EN_EXT"

    # titulos identicos
    if mt and mt == et:
        return "MISMO_PRODUCTO"

    # match parcial: tokens en comun
    if mt and et:
        tm = set(re.findall(r"\w{4,}", mt))
        te = set(re.findall(r"\w{4,}", et))
        if tm and te:
            overlap = len(tm & te) / max(len(tm | te), 1)
            if overlap >= 0.7:
                return "MISMO_PRODUCTO"
            if overlap < 0.3:
                return "LOOKALIKE_TITULO_DISTINTO"
    return "INCIERTO"


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: no existe {SRC}.")
        return 1

    cuentas_map: dict[str, Cuenta] = {}
    for cfg in CUENTAS:
        try:
            cuentas_map[cfg["alias"]] = Cuenta.cargar(cfg)
        except FileNotFoundError:
            pass

    cuenta_q = cuentas_map["C3"]
    revisar = cargar_revisar(SRC)
    sospechosos = [r for r in revisar if (r.get("diff_pct_vs_ganador") or 0) > UMBRAL_DIFF_PCT]
    print(f"Casos con diff >{UMBRAL_DIFF_PCT}% vs externo: {len(sospechosos)}")
    if not sospechosos:
        return 0

    seller_cache: dict[int, dict] = {}
    resultados: list[dict] = []

    for i, caso in enumerate(sospechosos, 1):
        mi_alias = caso.get("cuenta")
        mi_iid = caso.get("item_id")
        ext_iid = caso.get("ganador_item_id")
        ext_sid = caso.get("ganador_seller_id")
        cpid = caso.get("catalog_product_id")
        print(f"\n[{i}/{len(sospechosos)}] {cpid} | {(caso.get('title') or '')[:65]}")

        mi_full = get_item_propio(cuentas_map, mi_alias, mi_iid) or {}
        mi_attrs = extract_attrs(mi_full)
        mi_title = mi_full.get("title")
        mi_brand = mi_attrs.get("BRAND")

        ext = scrape_externo(ext_iid)
        ext_title = ext.get("title")
        ext_brand = ext.get("marca_breadcrumb")

        # seller externo
        if ext_sid and ext_sid not in seller_cache:
            rs = _request(cuenta_q, f"/users/{ext_sid}")
            seller_cache[ext_sid] = rs.json() if rs.status_code == 200 else {}
            time.sleep(0.05)
        ext_seller = seller_cache.get(ext_sid, {})

        veredicto = diagnostico(mi_title, mi_brand, ext_title, ext_brand)
        print(f"  veredicto: {veredicto}")
        print(f"  MIO  ({mi_alias}): ${int(caso.get('price') or 0):,}".replace(",", ".") +
              f" | brand={mi_brand!r}")
        print(f"        {mi_title[:90] if mi_title else '???'}")
        print(f"  EXT       : ${int(caso.get('ganador_precio') or 0):,}".replace(",", ".") +
              f" | brand={ext_brand!r} | seller={ext_seller.get('nickname','?')} "
              f"(reput {(ext_seller.get('seller_reputation') or {}).get('power_seller_status') or 'sin'})")
        print(f"        {ext_title[:90] if ext_title else '???'}")

        resultados.append({
            "veredicto": veredicto,
            "diff_pct": caso.get("diff_pct_vs_ganador"),
            "catalog_product_id": cpid,
            "cuenta": mi_alias,
            "mi_precio": caso.get("price"),
            "ext_precio": caso.get("ganador_precio"),
            "mi_titulo": mi_title,
            "ext_titulo": ext_title,
            "mi_brand": mi_brand,
            "ext_brand": ext_brand,
            "mi_model": mi_attrs.get("MODEL"),
            "mi_sku": mi_attrs.get("SELLER_SKU"),
            "ext_seller_nick": ext_seller.get("nickname"),
            "ext_seller_reputation": (ext_seller.get("seller_reputation") or {}).get("power_seller_status"),
            "ext_seller_ventas": (ext_seller.get("seller_reputation") or {}).get("transactions", {}).get("total"),
            "mi_item_id": mi_iid,
            "ext_item_id": ext_iid,
            "mi_permalink": mi_full.get("permalink"),
            "ext_permalink": ext.get("raw_url"),
        })

    # ---------- xlsx ----------
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("RESULTADOS")
    cols = [
        "veredicto", "diff_pct", "catalog_product_id", "cuenta",
        "mi_precio", "ext_precio",
        "mi_brand", "ext_brand",
        "mi_model", "mi_sku",
        "ext_seller_nick", "ext_seller_reputation", "ext_seller_ventas",
        "mi_titulo", "ext_titulo",
        "mi_item_id", "ext_item_id",
        "mi_permalink", "ext_permalink",
    ]
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="455A64")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    fills = {
        "LOOKALIKE_MARCA_DISTINTA":    PatternFill("solid", fgColor="EF9A9A"),
        "LOOKALIKE_MARCA_AUSENTE_EN_EXT": PatternFill("solid", fgColor="FFAB91"),
        "LOOKALIKE_TITULO_DISTINTO":   PatternFill("solid", fgColor="FFCC80"),
        "MISMO_PRODUCTO":              PatternFill("solid", fgColor="A5D6A7"),
        "DATOS_INSUFICIENTES":         PatternFill("solid", fgColor="FFF59D"),
        "INCIERTO":                    PatternFill("solid", fgColor="CFD8DC"),
    }
    for r in sorted(resultados, key=lambda x: (x["veredicto"], -(x.get("diff_pct") or 0))):
        ws.append([r.get(c) for c in cols])
        fill = fills.get(r["veredicto"])
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill
    ws.freeze_panes = "C2"
    widths = {
        "mi_titulo": 60, "ext_titulo": 60,
        "mi_permalink": 50, "ext_permalink": 50,
        "catalog_product_id": 16, "mi_item_id": 16, "ext_item_id": 16,
        "ext_seller_nick": 18, "ext_seller_reputation": 16,
    }
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    res = wb.create_sheet("RESUMEN", index=0)
    res.append(["veredicto", "n_casos", "interpretacion"])
    from collections import Counter
    cnt = Counter(r["veredicto"] for r in resultados)
    interp = {
        "LOOKALIKE_MARCA_DISTINTA":     "Distinta marca declarada -> reportar ficha a ML",
        "LOOKALIKE_MARCA_AUSENTE_EN_EXT": "Externo NO menciona tu marca -> probable lookalike, reportar",
        "LOOKALIKE_TITULO_DISTINTO":    "Titulo muy distinto -> revisar manualmente, posible reportar",
        "MISMO_PRODUCTO":               "Mismo producto, decision comercial pura (margen vs salir catalogo)",
        "DATOS_INSUFICIENTES":          "No pude leer datos del externo -> revisar permalink manual",
        "INCIERTO":                     "Match parcial -> revisar 1 a 1",
    }
    for v in ["LOOKALIKE_MARCA_DISTINTA", "LOOKALIKE_MARCA_AUSENTE_EN_EXT",
              "LOOKALIKE_TITULO_DISTINTO", "MISMO_PRODUCTO",
              "DATOS_INSUFICIENTES", "INCIERTO"]:
        res.append([v, cnt.get(v, 0), interp[v]])
    res.column_dimensions["A"].width = 32
    res.column_dimensions["B"].width = 8
    res.column_dimensions["C"].width = 70

    wb.save(DST)
    print(f"\n========== RESUMEN ==========")
    for v, n in sorted(cnt.items(), key=lambda x: -x[1]):
        print(f"  {v}: {n}")
    print(f"\nXlsx: {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
