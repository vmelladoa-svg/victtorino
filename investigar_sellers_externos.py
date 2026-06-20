"""
Cruza A + C de la conversacion:
  A) Para cada seller externo que nos gana buy box, traer info enriquecida
     (/users/{id}: nick, reputacion, ventas, ubicacion, fecha registro, brand)
  C) Para los 51 REVISAR_MANUAL completos (no solo los >100%), agruparlos
     por seller externo y ver patrones sistematicos (¿quien aparece varias
     veces? ¿quien es el "competidor estructural"?)

Output: sellers_externos_<fecha>.xlsx con 3 hojas:
  - SELLERS_RANKING: 1 fila por seller, ordenado por # apariciones
  - DETALLE_POR_LISTING: las 51 lineas con seller info embebida
  - RESUMEN: estadistica global
"""
from __future__ import annotations
import sys
import time
from collections import defaultdict, Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from ml_catalogo_publicaciones import CUENTAS, Cuenta, _request

ROOT = Path(__file__).parent
SRC = ROOT / f"candidatos_pausar_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"sellers_externos_{time.strftime('%Y-%m-%d')}.xlsx"


def cargar_revisar(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb["REVISAR_MANUAL"]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def get_user(cuenta: Cuenta, sid: int) -> dict | None:
    r = _request(cuenta, f"/users/{sid}")
    return r.json() if r.status_code == 200 else None


def fmt_money(v) -> str:
    try:
        return f"${int(v):,}".replace(",", ".")
    except Exception:
        return ""


def bucket_diff(d: float | None) -> str:
    if d is None:
        return "?"
    if d <= 0:
        return "ganador?"
    if d <= 3:
        return "<=3%"
    if d <= 7:
        return "3-7%"
    if d <= 15:
        return "7-15%"
    if d <= 50:
        return "15-50%"
    if d <= 100:
        return "50-100%"
    return ">100%"


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: no existe {SRC}")
        return 1

    cuenta_q = Cuenta.cargar(next(c for c in CUENTAS if c["alias"] == "C3"))
    revisar = cargar_revisar(SRC)
    print(f"Listings REVISAR_MANUAL: {len(revisar)}")

    # agrupar por seller externo
    por_seller: dict[int, list[dict]] = defaultdict(list)
    for it in revisar:
        sid = it.get("ganador_seller_id")
        if sid:
            por_seller[sid].append(it)

    print(f"Sellers externos unicos: {len(por_seller)}")

    # enriquecer cada seller con /users/{id}
    seller_info: dict[int, dict] = {}
    for i, sid in enumerate(por_seller.keys(), 1):
        u = get_user(cuenta_q, sid) or {}
        seller_info[sid] = u
        if i % 5 == 0:
            print(f"  {i}/{len(por_seller)}...")
        time.sleep(0.06)
    print(f"  {len(por_seller)}/{len(por_seller)} listo")

    # ---------- SELLERS RANKING ----------
    rows_sellers = []
    for sid, items in por_seller.items():
        u = seller_info.get(sid) or {}
        rep = u.get("seller_reputation") or {}
        trx = rep.get("transactions") or {}
        addr = u.get("address") or {}

        # precio promedio descuento que aplica vs nosotros
        diffs = [it.get("diff_pct_vs_ganador") for it in items if it.get("diff_pct_vs_ganador") is not None]
        diff_avg = round(sum(diffs) / len(diffs), 1) if diffs else None
        diff_max = round(max(diffs), 1) if diffs else None

        # cuentas afectadas
        cuentas_aff = sorted({it["cuenta"] for it in items})
        # listing types donde aparecen
        tipos = Counter(it.get("listing_type_id") for it in items)

        rows_sellers.append({
            "seller_id": sid,
            "nickname": u.get("nickname"),
            "n_apariciones": len(items),
            "ratings_positive_pct": rep.get("transactions", {}).get("ratings", {}).get("positive") if isinstance(rep.get("transactions", {}).get("ratings"), dict) else None,
            "reputation_status": rep.get("power_seller_status") or "sin",
            "level_id": rep.get("level_id"),
            "total_transactions": trx.get("total"),
            "completed_transactions": trx.get("completed"),
            "canceled_transactions": trx.get("canceled"),
            "country_id": u.get("country_id"),
            "city": addr.get("city"),
            "state": addr.get("state"),
            "registration_date": u.get("registration_date"),
            "user_type": u.get("user_type"),
            "tags": ",".join(u.get("tags") or []),
            "cuentas_nuestras_afectadas": ",".join(cuentas_aff),
            "diff_pct_promedio": diff_avg,
            "diff_pct_maximo": diff_max,
            "listing_types_compitiendo": ",".join(f"{k}:{v}" for k, v in tipos.items()),
            "items_que_gana": ",".join(it.get("ganador_item_id") for it in items if it.get("ganador_item_id")),
        })

    rows_sellers.sort(key=lambda x: -x["n_apariciones"])

    # ---------- DETALLE POR LISTING ----------
    rows_detalle = []
    for it in revisar:
        sid = it.get("ganador_seller_id")
        u = seller_info.get(sid) or {}
        rep = u.get("seller_reputation") or {}
        rows_detalle.append({
            **it,
            "ext_seller_nickname": u.get("nickname"),
            "ext_seller_reputation": rep.get("power_seller_status") or "sin",
            "ext_seller_ventas_total": (rep.get("transactions") or {}).get("total"),
            "ext_seller_pais": u.get("country_id"),
            "ext_seller_registro": u.get("registration_date"),
            "bucket_diff": bucket_diff(it.get("diff_pct_vs_ganador")),
        })

    # ---------- XLSX ----------
    wb = Workbook()
    wb.remove(wb.active)

    # SELLERS_RANKING
    ws1 = wb.create_sheet("SELLERS_RANKING")
    cols1 = [
        "seller_id", "nickname", "n_apariciones",
        "reputation_status", "level_id", "total_transactions",
        "completed_transactions", "canceled_transactions",
        "diff_pct_promedio", "diff_pct_maximo",
        "cuentas_nuestras_afectadas", "listing_types_compitiendo",
        "country_id", "city", "state",
        "registration_date", "user_type", "tags",
        "items_que_gana",
    ]
    ws1.append(cols1)
    for cell in ws1[1]:
        cell.fill = PatternFill("solid", fgColor="455A64")
        cell.font = Font(bold=True, color="FFFFFF")

    fill_critical = PatternFill("solid", fgColor="FFCDD2")
    fill_warn = PatternFill("solid", fgColor="FFE0B2")
    for r in rows_sellers:
        ws1.append([r.get(c) for c in cols1])
        last = ws1.max_row
        # colorear si aparece mas de 3 veces
        if r["n_apariciones"] >= 5:
            ws1.cell(row=last, column=3).fill = fill_critical
        elif r["n_apariciones"] >= 2:
            ws1.cell(row=last, column=3).fill = fill_warn

    ws1.freeze_panes = "C2"
    widths1 = {
        "nickname": 22, "n_apariciones": 14, "reputation_status": 16,
        "cuentas_nuestras_afectadas": 22, "listing_types_compitiendo": 24,
        "items_que_gana": 50, "registration_date": 22, "city": 12, "state": 16,
    }
    for i, c in enumerate(cols1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = widths1.get(c, 14)

    # DETALLE_POR_LISTING
    ws2 = wb.create_sheet("DETALLE_POR_LISTING")
    cols2 = [
        "cuenta", "item_id", "title", "catalog_product_id",
        "listing_type_id", "price", "ganador_precio", "diff_pct_vs_ganador", "bucket_diff",
        "ext_seller_nickname", "ext_seller_reputation", "ext_seller_ventas_total",
        "ganador_seller_id", "ganador_item_id",
        "available_quantity", "ext_seller_registro", "permalink",
    ]
    ws2.append(cols2)
    for cell in ws2[1]:
        cell.fill = PatternFill("solid", fgColor="455A64")
        cell.font = Font(bold=True, color="FFFFFF")
    for r in sorted(rows_detalle, key=lambda x: (-(x.get("diff_pct_vs_ganador") or 0))):
        ws2.append([r.get(c) for c in cols2])
    ws2.freeze_panes = "C2"
    widths2 = {
        "title": 50, "ext_seller_nickname": 22, "permalink": 50,
        "catalog_product_id": 16, "item_id": 16, "ganador_item_id": 16,
    }
    for i, c in enumerate(cols2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = widths2.get(c, 14)

    # RESUMEN
    ws3 = wb.create_sheet("RESUMEN", index=0)
    ws3.append(["metrica", "valor"])
    ws3.append(["listings_perdiendo_vs_externo", len(revisar)])
    ws3.append(["sellers_externos_unicos", len(por_seller)])

    # top 5 sellers
    ws3.append([])
    ws3.append(["TOP 5 SELLERS por # apariciones", ""])
    for r in rows_sellers[:5]:
        ws3.append([f"  {r['nickname']} ({r['seller_id']})", r["n_apariciones"]])

    # distribucion de buckets
    ws3.append([])
    ws3.append(["DISTRIBUCION drop% necesario para ganar", ""])
    buckets = Counter(bucket_diff(it.get("diff_pct_vs_ganador")) for it in revisar)
    for k in ["<=3%", "3-7%", "7-15%", "15-50%", "50-100%", ">100%", "?"]:
        if buckets.get(k):
            ws3.append([f"  {k}", buckets[k]])

    # reputacion ext
    ws3.append([])
    ws3.append(["DISTRIBUCION reputacion externos", ""])
    reps = Counter(seller_info.get(sid, {}).get("seller_reputation", {}).get("power_seller_status") or "sin"
                   for sid in por_seller)
    for k, n in reps.most_common():
        ws3.append([f"  {k}", n])

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 14

    wb.save(DST)

    # ---------- consola ----------
    print(f"\n========== TOP 10 SELLERS EXTERNOS ==========")
    print(f"{'nickname':<22} {'apariciones':>11} {'reput':<10} {'ventas':>10} {'diff prom':>10}  cuentas")
    print("-" * 90)
    for r in rows_sellers[:10]:
        print(f"{(r['nickname'] or '?'):<22} {r['n_apariciones']:>11} "
              f"{(r['reputation_status'] or 'sin')[:10]:<10} "
              f"{(r['total_transactions'] or 0):>10} "
              f"{(str(r['diff_pct_promedio'] or '?')+'%'):>10}  "
              f"{r['cuentas_nuestras_afectadas']}")

    print(f"\nXlsx: {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
