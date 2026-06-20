"""Lee las decisiones de Lote A y B desde los Excels."""
import sys, io, warnings
from openpyxl import load_workbook
warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

def leer(path, hoja):
    wb = load_workbook(path, data_only=True)
    ws = wb[hoja]
    decisiones = {"OK": [], "Editar": [], "Saltar": [], "(vacio)": []}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dec = (row[0] or "").strip() if row[0] else ""
        dec_upper = dec.upper() if dec else ""
        woo_id = row[1]
        producto = row[3]
        nota = (row[7] or "").strip() if row[7] else ""
        if dec_upper == "OK":
            decisiones["OK"].append((woo_id, producto))
        elif dec_upper == "EDITAR":
            decisiones["Editar"].append((woo_id, producto, nota))
        elif dec_upper == "SALTAR":
            decisiones["Saltar"].append((woo_id, producto))
        elif not dec:
            decisiones["(vacio)"].append((woo_id, producto))
        else:
            decisiones.setdefault(f"otro:{dec}", []).append((woo_id, producto))
    return decisiones

for label, path, hoja in [("Lote A", r"C:\Users\dell\victtorino\propuesta_lote_a.xlsx", "Lote A — 77 productos"),
                          ("Lote B", r"C:\Users\dell\victtorino\propuesta_lote_b.xlsx", "Lote B — 28 productos")]:
    print(f"=== {label} ===")
    d = leer(path, hoja)
    for k, v in d.items():
        print(f"  {k}: {len(v)}")
    if d.get("Editar"):
        print("  EDITAR detalles:")
        for woo_id, prod, nota in d["Editar"]:
            print(f"    {woo_id} {prod[:50]} | nota: {nota[:80]}")
    if d.get("Saltar"):
        print("  SALTAR:")
        for woo_id, prod in d["Saltar"]:
            print(f"    {woo_id} {prod[:50]}")
    print()
