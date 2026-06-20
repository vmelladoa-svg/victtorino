"""
Cruza la lista de costos Defontana (HTML-xls) con los items ML de las 3 cuentas
usando SELLER_SKU.

Pipeline:
1. Parsea el informe Defontana -> diccionario {codigo, descripcion, costo, stock}
2. Para cada cuenta C1/C2/C3, multi-get /items con attributes=id,title,attributes
   y extrae SELLER_SKU de cada item
3. Intenta matchear codigo Defontana <-> SELLER_SKU ML (3 variantes)
4. Reporta cobertura de match y escribe matching_defontana_ml_<fecha>.xlsx
   con 3 hojas:
     COBERTURA      : resumen
     MATCHED        : items ML con costo asociado
     UNMATCHED_ML   : items ML sin match
     UNMATCHED_DEF  : codigos Defontana sin match en ML
"""
from __future__ import annotations
import re
import sys
import time
from io import StringIO
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from ml_catalogo_publicaciones import CUENTAS, Cuenta, _request

ROOT = Path(__file__).parent
DEFONTANA_XLS = Path(r"C:/Users/dell/Downloads/Informe de artículos 20250627210343860887 26052026 2033.xls")
CATALOGO_XLSX = ROOT / f"catalogo_publicaciones_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"matching_defontana_ml_{time.strftime('%Y-%m-%d')}.xlsx"


def parse_defontana(path: Path) -> pd.DataFrame:
    raw = path.read_bytes().decode("latin-1")
    tablas = pd.read_html(StringIO(raw))
    df = max(tablas, key=lambda t: t.shape[0] * t.shape[1])
    df.columns = ["codigo_desc", "stock", "costo_vigente", "costo_reposicion",
                  "pend_recep_u", "pend_recep_fecha", "pend_entrega", "stock_futuro"]
    # 3 filas de header
    df = df.iloc[3:].reset_index(drop=True)
    # separar codigo - descripcion
    parts = df["codigo_desc"].str.split("-", n=1, expand=True)
    df["codigo"] = parts[0].str.strip()
    df["descripcion"] = parts[1].str.strip() if 1 in parts.columns else None
    # tipos numericos
    for c in ["stock", "costo_vigente", "costo_reposicion", "stock_futuro"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df[["codigo", "descripcion", "stock", "costo_vigente",
             "costo_reposicion", "stock_futuro"]]
    df = df[df["codigo"].notna() & (df["codigo"] != "")]
    return df


def cargar_items_xlsx(path: Path) -> dict[str, list[dict]]:
    """Devuelve {cuenta: [{item_id, title, ...}, ...]}"""
    wb = load_workbook(path, read_only=True)
    out: dict[str, list[dict]] = {}
    for s in wb.sheetnames:
        if s == "RESUMEN":
            continue
        ws = wb[s]
        rows = ws.iter_rows(values_only=True)
        h = next(rows)
        out[s] = [dict(zip(h, r)) for r in rows]
    return out


def pull_seller_skus(cuenta: Cuenta, item_ids: list[str]) -> dict[str, dict]:
    """Multi-get /items con attributes=id,title,attributes para extraer SELLER_SKU.

    Devuelve {item_id: {title, seller_sku, brand, model}}.
    """
    out: dict[str, dict] = {}
    fails = 0
    for i in range(0, len(item_ids), 20):
        chunk = item_ids[i:i+20]
        r = _request(cuenta, "/items", {
            "ids": ",".join(chunk),
            "attributes": "id,title,attributes",
        })
        if r.status_code != 200:
            fails += len(chunk)
            continue
        for entry in r.json():
            if entry.get("code") != 200 or not entry.get("body"):
                fails += 1
                continue
            body = entry["body"]
            attrs = {a.get("id"): (a.get("value_name") or "") for a in (body.get("attributes") or [])}
            out[body.get("id")] = {
                "title": body.get("title"),
                "seller_sku": attrs.get("SELLER_SKU"),
                "brand": attrs.get("BRAND"),
                "model": attrs.get("MODEL"),
            }
        time.sleep(0.05)
    if fails:
        print(f"  [{cuenta.alias}] {fails} items sin data")
    return out


def normalizar(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def match_strategy(codigo_def: str, descripcion_def: str | None,
                   sku_ml: str | None, title_ml: str | None) -> str | None:
    """Devuelve la estrategia que funciono, o None."""
    if not sku_ml:
        return None
    sku_n = normalizar(sku_ml)
    cod_n = normalizar(codigo_def)
    # 1) exact normalizado
    if cod_n and cod_n == sku_n:
        return "exact"
    # 2) codigo contenido en sku
    if cod_n and cod_n in sku_n:
        return "contained_in_sku"
    # 3) sku contenido en codigo (caso raro)
    if sku_n and sku_n in cod_n and len(sku_n) >= 6:
        return "sku_in_codigo"
    return None


def main() -> int:
    if not DEFONTANA_XLS.exists():
        print(f"ERROR: no existe {DEFONTANA_XLS}")
        return 1
    if not CATALOGO_XLSX.exists():
        print(f"ERROR: no existe {CATALOGO_XLSX}")
        return 1

    df_def = parse_defontana(DEFONTANA_XLS)
    print(f"Defontana: {len(df_def)} productos cargados")
    print(f"  con costo_vigente > 0: {(df_def['costo_vigente'] > 0).sum()}")
    print(f"  con stock > 0       : {(df_def['stock'] > 0).sum()}")

    items_por_cuenta = cargar_items_xlsx(CATALOGO_XLSX)
    print(f"\nItems ML por cuenta:")
    for c, items in items_por_cuenta.items():
        print(f"  {c}: {len(items)}")

    # pull SKUs por cuenta (usando token correcto)
    sku_map: dict[str, dict] = {}  # item_id -> {title, seller_sku, ...}
    for cfg in CUENTAS:
        alias = cfg["alias"]
        if alias not in items_por_cuenta:
            continue
        cu = Cuenta.cargar(cfg)
        ids = [it["item_id"] for it in items_por_cuenta[alias] if it.get("item_id")]
        print(f"\n[{alias}] pull SELLER_SKU de {len(ids)} items...")
        data = pull_seller_skus(cu, ids)
        print(f"  recibidos: {len(data)}")
        n_con_sku = sum(1 for v in data.values() if v.get("seller_sku"))
        print(f"  con SELLER_SKU: {n_con_sku} ({n_con_sku/max(len(data),1)*100:.0f}%)")
        sku_map.update(data)

    # construir lookup defontana por codigo normalizado
    def_por_cod = {normalizar(r["codigo"]): r.to_dict() for _, r in df_def.iterrows()}

    # matchear cada item ML contra defontana
    matched: list[dict] = []
    unmatched_ml: list[dict] = []
    matched_codigos: set[str] = set()

    for cuenta_alias, items in items_por_cuenta.items():
        for it in items:
            iid = it.get("item_id")
            ml_data = sku_map.get(iid, {})
            sku = ml_data.get("seller_sku")
            # probar match por SKU contra cada codigo
            best = None
            for cod_norm, def_row in def_por_cod.items():
                strat = match_strategy(def_row["codigo"], def_row["descripcion"], sku, ml_data.get("title"))
                if strat:
                    best = (strat, def_row)
                    break
            if best:
                strat, def_row = best
                matched_codigos.add(normalizar(def_row["codigo"]))
                matched.append({
                    "cuenta": cuenta_alias,
                    "item_id": iid,
                    "title_ml": ml_data.get("title") or it.get("title"),
                    "seller_sku": sku,
                    "match_strategy": strat,
                    "codigo_defontana": def_row["codigo"],
                    "descripcion_defontana": def_row["descripcion"],
                    "costo_vigente": def_row["costo_vigente"],
                    "stock_defontana": def_row["stock"],
                    "price_ml": it.get("price"),
                    "available_quantity_ml": it.get("available_quantity"),
                    "listing_type_id": it.get("listing_type_id"),
                    "catalog_product_id": it.get("catalog_product_id"),
                    "permalink": it.get("permalink"),
                })
            else:
                unmatched_ml.append({
                    "cuenta": cuenta_alias,
                    "item_id": iid,
                    "title_ml": ml_data.get("title") or it.get("title"),
                    "seller_sku": sku,
                    "brand": ml_data.get("brand"),
                    "model": ml_data.get("model"),
                    "price_ml": it.get("price"),
                    "catalog_product_id": it.get("catalog_product_id"),
                    "permalink": it.get("permalink"),
                })

    # codigos defontana sin match en ML
    unmatched_def = []
    for cod_norm, def_row in def_por_cod.items():
        if cod_norm not in matched_codigos:
            unmatched_def.append(def_row)

    total_ml = sum(len(v) for v in items_por_cuenta.values())
    cobertura_ml = len(matched) / total_ml * 100 if total_ml else 0
    cobertura_def = (len(df_def) - len(unmatched_def)) / len(df_def) * 100 if len(df_def) else 0

    print(f"\n========== COBERTURA ==========")
    print(f"  Items ML totales       : {total_ml}")
    print(f"  Items ML con match     : {len(matched)} ({cobertura_ml:.1f}%)")
    print(f"  Items ML sin SELLER_SKU: {sum(1 for v in sku_map.values() if not v.get('seller_sku'))}")
    print()
    print(f"  Codigos Defontana totales : {len(df_def)}")
    print(f"  Defontana con match en ML : {len(df_def) - len(unmatched_def)} ({cobertura_def:.1f}%)")
    print(f"  Defontana sin match en ML : {len(unmatched_def)}")

    # ---------- xlsx ----------
    wb = Workbook()
    wb.remove(wb.active)

    cob = wb.create_sheet("COBERTURA")
    cob.append(["metrica", "valor"])
    for k, v in [
        ("items_ml_totales", total_ml),
        ("items_ml_con_match", len(matched)),
        ("items_ml_sin_match", len(unmatched_ml)),
        ("cobertura_ml_pct", round(cobertura_ml, 1)),
        ("items_ml_sin_seller_sku", sum(1 for v in sku_map.values() if not v.get("seller_sku"))),
        ("defontana_total", len(df_def)),
        ("defontana_con_match", len(df_def) - len(unmatched_def)),
        ("defontana_sin_match", len(unmatched_def)),
        ("cobertura_defontana_pct", round(cobertura_def, 1)),
    ]:
        cob.append([k, v])
    cob.column_dimensions["A"].width = 30
    cob.column_dimensions["B"].width = 12

    def dump(nombre: str, data: list[dict], cols: list[str]) -> None:
        ws = wb.create_sheet(nombre)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="455A64")
            cell.font = Font(bold=True, color="FFFFFF")
        for r in data:
            ws.append([r.get(c) for c in cols])
        ws.freeze_panes = "A2"
        widths = {"title_ml": 55, "descripcion_defontana": 40, "permalink": 45,
                  "catalog_product_id": 16, "item_id": 16,
                  "codigo_defontana": 14, "seller_sku": 18}
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    dump("MATCHED", matched, [
        "cuenta", "item_id", "title_ml", "seller_sku", "match_strategy",
        "codigo_defontana", "descripcion_defontana",
        "costo_vigente", "stock_defontana", "price_ml", "available_quantity_ml",
        "listing_type_id", "catalog_product_id", "permalink",
    ])
    dump("UNMATCHED_ML", unmatched_ml, [
        "cuenta", "item_id", "title_ml", "seller_sku", "brand", "model",
        "price_ml", "catalog_product_id", "permalink",
    ])
    dump("UNMATCHED_DEF", unmatched_def, [
        "codigo", "descripcion", "stock", "costo_vigente",
        "costo_reposicion", "stock_futuro",
    ])

    wb.save(DST)
    print(f"\nXlsx: {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
