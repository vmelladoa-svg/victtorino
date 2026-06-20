"""
ml_a_paris_excel.py
-------------------
Extrae 38 productos de catalogo de MercadoLibre Chile (URLs /p/MLC...) y
escribe un Excel con el formato exacto requerido por Paris Marketplace
(Seller Center), usando como base la plantilla descargada.

Fuente de datos:
  - API publica de MercadoLibre /products/{MLC_ID}  (requiere token).
  - Se reutiliza tokens_cuenta3.json (la cuenta C3 ya autorizada).

Salida:
  - paris_productos_carga_masiva.xlsx en el directorio actual.

La plantilla original contiene un dataValidation invalido (errorStyle="error")
en la hoja "herramientasdata" que openpyxl rechaza; se parcha en memoria al
descomprimir el .xlsx.
"""

from __future__ import annotations

import io
import json
import re
import shutil
import sys
import time
import zipfile
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
TOKEN_FILE = BASE_DIR / "tokens_cuenta3.json"
TEMPLATE_FILE = BASE_DIR / "create_products_template_2026-05-22T14_32_35.210Z.xlsx"
OUTPUT_FILE = BASE_DIR / "paris_productos_carga_masiva.xlsx"
DATA_SHEET = "herramientas"
FIRST_DATA_ROW = 7
SLEEP_BETWEEN_REQUESTS = 1.0
MAX_IMAGES = 6
DESC_MAX_CHARS = 400

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Columnas a completar (1-indexed) -- mapeo verificado contra fila 4 de la plantilla.
COL_NOMBRE = 1
COL_SKU_SELLER = 2
COL_REF_PRODUCT = 3
COL_CATEGORIA = 4
COL_MARCA = 6
COL_MODELO = 10
COL_IMAGENES = 13
COL_SKU_VARIANT = 15
COL_COLOR = 16
COL_TALLA = 17
COL_DESC_LARGA = 33
COL_MATERIAL = 35

URLS = [
    "https://www.mercadolibre.cl/llave-monomando-lavaplato-inox-acabado-cepillado-color-plateado/p/MLC21628947",
    "https://www.mercadolibre.cl/asiento-y-tapa-para-wc-cierre-lento-modern-tumm/p/MLC21350539",
    "https://www.mercadolibre.cl/llave-lavacopas-acero-inoxidable-acabado-cepillado-color-plateado/p/MLC21394613",
    "https://www.mercadolibre.cl/llave-monomando-tina-ducha-colomba-acabado-cromado-color-plateado/p/MLC38612080",
    "https://www.mercadolibre.cl/llave-monomando-lavatorio-notte-acabado-mate-color-negro/p/MLC25452157",
    "https://www.mercadolibre.cl/tumm-asiento-con-tapa-wc-eco-ovalada-blancopara-wc-eco-color-blanco/p/MLC43500911",
    "https://www.mercadolibre.cl/llave-tumm-monomando-lavaplato-negra-notte/p/MLC45488644",
    "https://www.mercadolibre.cl/monomando-ducha-empotrado-notte-acabado-mate-color-negro/p/MLC41735119",
    "https://www.mercadolibre.cl/lavaplatos-sobreponer-simple-80x50-inoxidable-derecho-color-plateado/p/MLC21299614",
    "https://www.mercadolibre.cl/lavaplatos-empotrado-80x44-inox-secderecho-color-plateado/p/MLC22271656",
    "https://www.mercadolibre.cl/lavaplatos-empotrados-simple-80x44-inox-derecho-198432513-color-plateado/p/MLC50745282",
    "https://www.mercadolibre.cl/desague-para-tina-con-tubo-y-campana-de-rebalse/p/MLC20918885",
    "https://www.mercadolibre.cl/desague-lavaplatos-3-12-o-114-cm-o-114-mm-con-rebalse-color-plateado/p/MLC28368137",
    "https://www.mercadolibre.cl/flexible-para-llave-monomando-lavaplatos-profesional-color-plateado/p/MLC24386020",
    "https://www.mercadolibre.cl/kit-para-estanque-wc-bano-universal-valvula-admision/p/MLC2041691925",
    "https://www.mercadolibre.cl/llave-monomando-lavatorio-lavamanos-modern-plateado-cromado/p/MLC21394632",
    "https://www.mercadolibre.cl/llave-lavadora-lavadero-o-jardin-simultanea-2-salidas-34-color-plateado/p/MLC32065874",
    "https://www.mercadolibre.cl/valvula-de-descarga-dual-flush-para-estanque-tradicional/p/MLC27509195",
    "https://www.mercadolibre.cl/sifon-plastico-con-tubo-para-lavaplatos-1-12-color-blanco/p/MLC26656485",
    "https://www.mercadolibre.cl/desague-sifon-de-90-mm-para-receptaculo-plateado/p/MLC67879003",
    "https://www.mercadolibre.cl/lavaplatos-empotrados-simple-100x44-inox-izquierdo-191432514/p/MLC22271655",
    "https://www.mercadolibre.cl/grifo-monomando-profesional-tumm-con-cuello-extensible-para-lavaplatos/p/MLC25264406",
    "https://www.mercadolibre.cl/llave-monomando-lavatorio-lavamanos-modern-plateado-cepillado/p/MLC21394631",
    "https://www.mercadolibre.cl/barra-de-soporte-ducha-deslizable-65-cm-acero-inoxidable/p/MLC2055571318",
    "https://www.mercadolibre.cl/valvula-de-descarga-para-wc-con-flapper-y-cadena/p/MLC26778872",
    "https://www.mercadolibre.cl/pack-lavaplato-80x44-izquierdo-empotradollavesifondesague-color-plateado/p/MLC22736791",
    "https://www.mercadolibre.cl/lavamanos-sobrepuesto-de-vidrio-transparente-redondo-tumm-de-415/p/MLC24044164",
    "https://www.mercadolibre.cl/lavaplatos-simple-empotrable-ancho-80x50x22-cm-hannover-color-plateado/p/MLC22271660",
    "https://www.mercadolibre.cl/ducha-fija-al-muro-con-difusor-color-plateado/p/MLC32227841",
    "https://www.mercadolibre.cl/valvula-de-carga-con-flotador-para-estanque-wc-taumm/p/MLC2046010486",
    "https://www.mercadolibre.cl/griferia-bano-monocomando-llave-lavatorio-victtorino-cromado-acero-inoxidable/p/MLC58519072",
    "https://www.mercadolibre.cl/lavaplatos-simple-empotrable-ancho-686x456x21-cm-hamburg-color-plateado/p/MLC22271659",
    "https://www.mercadolibre.cl/valvula-lateral-para-wc-eficiente-y-durable/p/MLC2068899311",
    "https://www.mercadolibre.cl/mezcladora-de-bano-victtorino-monocomando-de-acero-inoxidable-cromado-brillante-para-ducha-empotrada/p/MLC66202132",
    "https://www.mercadolibre.cl/shower-door-y-receptaculo-80x80x195-vidrio-templado/p/MLC27107143",
    "https://www.mercadolibre.cl/mampara-ducha-de-120-a-140-cms-ancho-x-190-cms-vidrio-templ/p/MLC39244370",
    "https://www.mercadolibre.cl/shower-door-receptaculo-90x90-curvo-transparente-negro/p/MLC39080093",
    "https://www.mercadolibre.cl/p/MLC51764090",
]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def extract_ml_id(url: str) -> Optional[str]:
    m = re.search(r"/p/(MLC\d+)", url)
    return m.group(1) if m else None


def get_attr(attrs: list, attr_id: str) -> Optional[str]:
    for a in attrs or []:
        if a.get("id") == attr_id:
            value = a.get("value_name")
            if value:
                return value.strip()
    return None


def clean_image_url(url: str) -> str:
    """Garantiza el sufijo -F.jpg y elimina query strings."""
    if not url:
        return ""
    url = url.split("?", 1)[0]
    # Reemplaza cualquier -[letra].(jpg|png|webp) por -F.jpg
    url = re.sub(r"-[A-Z]\.(jpg|jpeg|png|webp)$", "-F.jpg", url, flags=re.I)
    return url


def patch_template(template_path: Path) -> Path:
    """
    Devuelve la ruta a una copia de la plantilla con el dataValidation
    invalido corregido. Si openpyxl ya puede abrirla, devuelve la original.
    """
    try:
        load_workbook(template_path)
        return template_path
    except Exception:
        pass

    patched = template_path.with_suffix(".fixed.xlsx")
    with zipfile.ZipFile(template_path, "r") as zin, \
         zipfile.ZipFile(patched, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="replace")
                # ML/Paris exporta errorStyle="error" que no es valido.
                text = text.replace('errorStyle="error"', 'errorStyle="stop"')
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return patched


def load_token() -> str:
    with open(TOKEN_FILE, "r", encoding="utf-8") as f:
        tk = json.load(f)
    return tk["access_token"]


# ---------------------------------------------------------------------------
# Extraccion ML
# ---------------------------------------------------------------------------


def fetch_product(ml_id: str, token: str) -> dict:
    headers = {
        "Authorization": f"Bearer {token}",
        "User-Agent": USER_AGENT,
    }
    r = requests.get(
        f"https://api.mercadolibre.com/products/{ml_id}",
        headers=headers,
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def parse_product(ml_id: str, data: dict) -> dict:
    attrs = data.get("attributes") or []

    titulo = (data.get("name") or "").strip()

    marca = get_attr(attrs, "BRAND")
    modelo = get_attr(attrs, "MODEL")
    color = get_attr(attrs, "COLOR") or get_attr(attrs, "MAIN_COLOR") or "Unico color"
    material = (
        get_attr(attrs, "MATERIALS")
        or get_attr(attrs, "MATERIAL")
        or get_attr(attrs, "PRIMARY_MATERIAL")
        or ""
    )

    pics = data.get("pictures") or []
    images: list[str] = []
    for p in pics:
        u = clean_image_url(p.get("url") or "")
        if u and "mlstatic" in u and u not in images:
            images.append(u)
        if len(images) >= MAX_IMAGES:
            break

    # Descripcion
    short = data.get("short_description") or {}
    if isinstance(short, dict):
        desc = (short.get("content") or "").strip()
    else:
        desc = str(short).strip()
    if not desc:
        feats = data.get("main_features") or []
        desc = " ".join(
            (f.get("text") or "").strip()
            for f in feats
            if isinstance(f, dict) and f.get("text")
        ).strip()
    if len(desc) > DESC_MAX_CHARS:
        desc = desc[: DESC_MAX_CHARS - 1].rstrip() + "..."

    # Precio (no se escribe en Excel, solo informativo)
    bbw = data.get("buy_box_winner") or {}
    precio = bbw.get("price")

    return {
        "ml_id": ml_id,
        "titulo": titulo,
        "precio": precio,
        "marca": (marca or "").strip(),
        "modelo": (modelo or "").strip(),
        "color": color.strip() if color else "Unico color",
        "material": material.strip(),
        "imagenes": images,
        "descripcion": desc,
    }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def write_excel(productos: list[dict]) -> None:
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"No se encuentra la plantilla: {TEMPLATE_FILE}")

    src = patch_template(TEMPLATE_FILE)
    if src != OUTPUT_FILE:
        shutil.copy(src, OUTPUT_FILE)

    wb = load_workbook(OUTPUT_FILE)
    if DATA_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"La hoja '{DATA_SHEET}' no existe. Hojas disponibles: {wb.sheetnames}"
        )
    ws = wb[DATA_SHEET]

    for i, p in enumerate(productos):
        row = FIRST_DATA_ROW + i
        ws.cell(row=row, column=COL_NOMBRE, value=p["titulo"])
        ws.cell(row=row, column=COL_SKU_SELLER, value=p["ml_id"])
        ws.cell(row=row, column=COL_REF_PRODUCT, value=None)
        ws.cell(row=row, column=COL_CATEGORIA, value=None)
        ws.cell(row=row, column=COL_MARCA, value=p["marca"] or None)
        ws.cell(row=row, column=COL_MODELO, value=p["modelo"] or None)
        ws.cell(row=row, column=COL_IMAGENES, value=", ".join(p["imagenes"]))
        ws.cell(row=row, column=COL_SKU_VARIANT, value=p["ml_id"])
        ws.cell(row=row, column=COL_COLOR, value=p["color"] or "Unico color")
        ws.cell(row=row, column=COL_TALLA, value="Talla Unica")
        ws.cell(row=row, column=COL_DESC_LARGA, value=p["descripcion"] or None)
        ws.cell(row=row, column=COL_MATERIAL, value=p["material"] or None)

    wb.save(OUTPUT_FILE)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    # Forzar UTF-8 en stdout (Windows + tildes/diéresis)
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

    print(f"[INFO] Cargando token desde {TOKEN_FILE.name}")
    try:
        token = load_token()
    except Exception as e:
        print(f"[ERROR] No se pudo leer el token: {e}")
        return 1

    productos: list[dict] = []
    fallidos: list[tuple[str, str]] = []

    print(f"[INFO] Extrayendo {len(URLS)} productos de MercadoLibre...")
    for idx, url in enumerate(URLS, 1):
        ml_id = extract_ml_id(url)
        if not ml_id:
            print(f"  [{idx}/{len(URLS)}] URL sin MLC ID: {url}")
            fallidos.append((url, "sin ML ID"))
            continue
        try:
            data = fetch_product(ml_id, token)
            p = parse_product(ml_id, data)
            productos.append(p)
            print(
                f"  [{idx}/{len(URLS)}] {ml_id} OK "
                f"({len(p['imagenes'])} img, marca='{p['marca'] or '-'}', "
                f"modelo='{p['modelo'] or '-'}')"
            )
        except requests.HTTPError as e:
            sc = e.response.status_code if e.response is not None else "?"
            print(f"  [{idx}/{len(URLS)}] {ml_id} FAIL HTTP {sc}")
            fallidos.append((ml_id, f"HTTP {sc}"))
        except Exception as e:
            print(f"  [{idx}/{len(URLS)}] {ml_id} FAIL {type(e).__name__}: {e}")
            fallidos.append((ml_id, f"{type(e).__name__}: {e}"))
        time.sleep(SLEEP_BETWEEN_REQUESTS)

    if not productos:
        print("[ERROR] No se extrajo ningun producto. Aborto sin generar Excel.")
        return 2

    print(f"\n[INFO] Generando Excel: {OUTPUT_FILE.name}")
    write_excel(productos)

    # ---------------------------------------------------------------- Reporte
    print("\n" + "=" * 60)
    print("REPORTE FINAL")
    print("=" * 60)
    print(f"Total URLs procesadas    : {len(URLS)}")
    print(f"Productos extraidos OK   : {len(productos)}")
    print(f"Productos con fallo      : {len(fallidos)}")

    pocos_imgs = [p for p in productos if len(p["imagenes"]) < 2]
    sin_marca = [p for p in productos if not p["marca"]]

    print(f"Productos con <2 imagenes: {len(pocos_imgs)}")
    for p in pocos_imgs:
        print(f"  - {p['ml_id']} ({len(p['imagenes'])} img) :: {p['titulo'][:70]}")

    print(f"Productos sin marca      : {len(sin_marca)}")
    for p in sin_marca:
        print(f"  - {p['ml_id']} :: {p['titulo'][:70]}")

    if fallidos:
        print("Detalle de fallos:")
        for ml_id, reason in fallidos:
            print(f"  - {ml_id}: {reason}")

    print(f"\nArchivo generado: {OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
