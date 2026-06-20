"""
Rescata todas las publicaciones en estado catalogo (catalog_listing=true)
de las cuentas ML conectadas (C1, C2, C3) y exporta a Excel + CSV.

Uso:
    python ml_catalogo_publicaciones.py
    python ml_catalogo_publicaciones.py --cuenta C1
    python ml_catalogo_publicaciones.py --salida catalogo_2026-05-26.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import random
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

ML_BASE = "https://api.mercadolibre.com"
OAUTH_URL = f"{ML_BASE}/oauth/token"

CLIENT_ID = os.getenv("ML_CLIENT_ID") or "3959231945649654"
CLIENT_SECRET = os.getenv("ML_CLIENT_SECRET") or "PVvpcsugyGfL7DPeQTVi71iCbRkqdNtG"

CUENTAS = [
    {"alias": "C1", "nick": "PREMIUMGRIFERIAS1",  "tokens_file": ROOT / "tokens_cuenta1.json"},
    {"alias": "C2", "nick": "VICTTORINOOFICIAL2", "tokens_file": ROOT / "tokens_cuenta2.json"},
    {"alias": "C3", "nick": "NOVAGRIFERIAS3",     "tokens_file": ROOT / "tokens_cuenta3.json"},
]

CAMPOS_SALIDA = [
    "seller_id", "cuenta", "item_id", "title", "catalog_product_id",
    "listing_type_id", "price", "available_quantity", "status",
    "permalink", "thumbnail", "date_created", "last_updated",
]

# Atributos pedidos al endpoint /items multi-get (limita payload)
ITEM_ATTRS = ",".join([
    "id", "title", "catalog_product_id", "catalog_listing",
    "listing_type_id", "price", "available_quantity", "status",
    "permalink", "thumbnail", "date_created", "last_updated", "seller_id",
])

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ml-catalogo")


# ----------------------------------------------------------------------------
# Tokens
# ----------------------------------------------------------------------------
@dataclass
class Cuenta:
    alias: str
    nick: str
    tokens_file: Path
    access_token: str
    refresh_token: str
    user_id: int

    @classmethod
    def cargar(cls, cfg: dict) -> "Cuenta":
        f = cfg["tokens_file"]
        if not f.exists():
            raise FileNotFoundError(
                f"No existe {f}. Ejecuta primero: python ml_auth.py {cfg['alias'][-1]}"
            )
        data = json.loads(f.read_text(encoding="utf-8"))
        return cls(
            alias=cfg["alias"],
            nick=cfg["nick"],
            tokens_file=f,
            access_token=data["access_token"],
            refresh_token=data["refresh_token"],
            user_id=int(data["user_id"]),
        )

    def refrescar(self) -> None:
        log.info("[%s] refresh token...", self.alias)
        r = requests.post(
            OAUTH_URL,
            data={
                "grant_type": "refresh_token",
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "refresh_token": self.refresh_token,
            },
            timeout=30,
        )
        r.raise_for_status()
        d = r.json()
        self.access_token = d["access_token"]
        self.refresh_token = d.get("refresh_token", self.refresh_token)
        # persistir
        merged = json.loads(self.tokens_file.read_text(encoding="utf-8"))
        merged.update({
            "access_token": self.access_token,
            "refresh_token": self.refresh_token,
            "expires_in": d.get("expires_in"),
            "user_id": d.get("user_id", self.user_id),
        })
        self.tokens_file.write_text(json.dumps(merged, indent=2), encoding="utf-8")
        # sync .env si existe la clave
        _sync_env_token(self.alias, self.access_token, self.refresh_token)
        log.info("[%s] token refrescado", self.alias)


def _sync_env_token(alias: str, access: str, refresh: str) -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    txt = env_path.read_text(encoding="utf-8")
    for key, val in (
        (f"ML_ACCESS_TOKEN_{alias}", access),
        (f"ML_REFRESH_TOKEN_{alias}", refresh),
    ):
        pat = rf"^{re.escape(key)}=.*$"
        if re.search(pat, txt, re.MULTILINE):
            txt = re.sub(pat, f"{key}={val}", txt, flags=re.MULTILINE)
    env_path.write_text(txt, encoding="utf-8")


# ----------------------------------------------------------------------------
# HTTP con retry/backoff y refresh transparente
# ----------------------------------------------------------------------------
def _request(cuenta: Cuenta, path: str, params: dict | None = None,
             max_retries: int = 5) -> requests.Response:
    url = f"{ML_BASE}{path}"
    for intento in range(1, max_retries + 1):
        headers = {"Authorization": f"Bearer {cuenta.access_token}"}
        try:
            r = requests.get(url, headers=headers, params=params or {}, timeout=30)
        except requests.RequestException as e:
            wait = min(60, 2 ** intento) + random.random()
            log.warning("net error %s -> retry %.1fs", e, wait)
            time.sleep(wait)
            continue

        if r.status_code == 401 and intento == 1:
            cuenta.refrescar()
            continue

        if r.status_code == 429:
            wait = float(r.headers.get("Retry-After", 0)) or min(60, 2 ** intento)
            wait += random.random()
            log.warning("[%s] 429 rate-limit -> espera %.1fs", cuenta.alias, wait)
            time.sleep(wait)
            continue

        if r.status_code >= 500:
            wait = min(30, 2 ** intento) + random.random()
            log.warning("[%s] %s -> retry %.1fs", cuenta.alias, r.status_code, wait)
            time.sleep(wait)
            continue

        return r

    r.raise_for_status()
    return r


# ----------------------------------------------------------------------------
# Logica catalogo
# ----------------------------------------------------------------------------
def ids_via_scroll(cuenta: Cuenta, extra_params: dict | None = None) -> list[str]:
    """Scroll search completo para una cuenta."""
    ids: list[str] = []
    scroll = None
    while True:
        params = {"search_type": "scan", "limit": 100}
        if extra_params:
            params.update(extra_params)
        if scroll:
            params["scroll_id"] = scroll
        r = _request(cuenta, f"/users/{cuenta.user_id}/items/search", params)
        if r.status_code != 200:
            log.error("[%s] search fallo: %s %s", cuenta.alias, r.status_code, r.text[:200])
            break
        j = r.json()
        results = j.get("results") or []
        ids.extend(results)
        scroll = j.get("scroll_id")
        if not scroll or not results:
            break
        time.sleep(0.05)
    return ids


def ids_via_offset(cuenta: Cuenta, extra_params: dict | None = None,
                   limit: int = 50) -> list[str]:
    """Paginacion clasica offset/limit (la API capa ~1000)."""
    ids: list[str] = []
    offset = 0
    while True:
        params = {"limit": limit, "offset": offset}
        if extra_params:
            params.update(extra_params)
        r = _request(cuenta, f"/users/{cuenta.user_id}/items/search", params)
        if r.status_code != 200:
            log.error("[%s] search offset fallo: %s %s", cuenta.alias, r.status_code, r.text[:200])
            break
        j = r.json()
        results = j.get("results") or []
        ids.extend(results)
        total = j.get("paging", {}).get("total", 0)
        offset += limit
        if not results or offset >= min(total, 1000):
            break
        time.sleep(0.05)
    return ids


def multiget_items(cuenta: Cuenta, ids: list[str]) -> tuple[list[dict], list[str]]:
    """Detalle de items con multi-get. Devuelve (items, ids_fallidos)."""
    items: list[dict] = []
    fallos: list[str] = []
    for i in range(0, len(ids), 20):
        chunk = ids[i:i + 20]
        r = _request(cuenta, "/items", {"ids": ",".join(chunk), "attributes": ITEM_ATTRS})
        if r.status_code != 200:
            fallos.extend(chunk)
            log.error("[%s] multiget fallo: %s", cuenta.alias, r.status_code)
            continue
        for entry in r.json():
            iid = entry.get("body", {}).get("id") if entry.get("body") else None
            if entry.get("code") == 200 and entry.get("body"):
                items.append(entry["body"])
            else:
                fallos.append(iid or "?")
                log.warning("[%s] item %s code=%s", cuenta.alias, iid, entry.get("code"))
        time.sleep(0.05)
    return items, fallos


def fila_publicacion(it: dict, cuenta: Cuenta) -> dict:
    return {
        "seller_id": it.get("seller_id", cuenta.user_id),
        "cuenta": cuenta.alias,
        "item_id": it.get("id"),
        "title": it.get("title"),
        "catalog_product_id": it.get("catalog_product_id"),
        "listing_type_id": it.get("listing_type_id"),
        "price": it.get("price"),
        "available_quantity": it.get("available_quantity"),
        "status": it.get("status"),
        "permalink": it.get("permalink"),
        "thumbnail": it.get("thumbnail"),
        "date_created": it.get("date_created"),
        "last_updated": it.get("last_updated"),
    }


def procesar_cuenta(cuenta: Cuenta) -> list[dict]:
    log.info("[%s] inicio (user_id=%s, %s)", cuenta.alias, cuenta.user_id, cuenta.nick)

    # 1) ids con scroll (todos los items activos)
    ids_scroll = set(ids_via_scroll(cuenta, {"status": "active"}))
    log.info("[%s] scroll status=active -> %d ids", cuenta.alias, len(ids_scroll))

    # 2) ids con filtro listing_type_id=gold_pro (clasificados premium, top tier)
    ids_gold = set(ids_via_offset(cuenta, {
        "listing_type_id": "gold_pro", "status": "active",
    }))
    log.info("[%s] listing_type_id=gold_pro -> %d ids", cuenta.alias, len(ids_gold))

    # 3) ids con filtro catalog_listing=true (filtro nativo)
    ids_catflag = set(ids_via_offset(cuenta, {"catalog_listing": "true"}))
    log.info("[%s] catalog_listing=true -> %d ids", cuenta.alias, len(ids_catflag))

    todos = list(ids_scroll | ids_gold | ids_catflag)
    log.info("[%s] union -> %d ids unicos a inspeccionar", cuenta.alias, len(todos))

    if not todos:
        return []

    items, fallos = multiget_items(cuenta, todos)
    if fallos:
        log.warning("[%s] %d items con error en multi-get", cuenta.alias, len(fallos))
        _dump_log_errores(cuenta, fallos)

    catalogo = [it for it in items if it.get("catalog_listing") is True]
    log.info("[%s] %d items con catalog_listing=true", cuenta.alias, len(catalogo))

    return [fila_publicacion(it, cuenta) for it in catalogo]


def _dump_log_errores(cuenta: Cuenta, ids: list[str]) -> None:
    log_dir = ROOT / "data" / "catalogo_logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    f = log_dir / f"errores_{cuenta.alias}_{int(time.time())}.txt"
    f.write_text("\n".join(ids), encoding="utf-8")
    log.info("[%s] log de fallos -> %s", cuenta.alias, f)


# ----------------------------------------------------------------------------
# Export
# ----------------------------------------------------------------------------
def exportar_xlsx(por_cuenta: dict[str, list[dict]], destino: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for alias, filas in por_cuenta.items():
        ws = wb.create_sheet(title=alias)
        ws.append(CAMPOS_SALIDA)
        for fila in filas:
            ws.append([fila.get(c) for c in CAMPOS_SALIDA])
        for i, _ in enumerate(CAMPOS_SALIDA, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        ws.freeze_panes = "A2"

    # hoja resumen
    res = wb.create_sheet(title="RESUMEN", index=0)
    res.append(["cuenta", "publicaciones_en_catalogo"])
    for alias, filas in por_cuenta.items():
        res.append([alias, len(filas)])
    res.append(["TOTAL", sum(len(v) for v in por_cuenta.values())])

    wb.save(destino)


def exportar_csv(por_cuenta: dict[str, list[dict]], destino: Path) -> None:
    with destino.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CAMPOS_SALIDA)
        w.writeheader()
        for filas in por_cuenta.values():
            for fila in filas:
                w.writerow(fila)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main(argv: Iterable[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Rescata catalog listings ML")
    p.add_argument("--cuenta", choices=["C1", "C2", "C3"],
                   help="Procesa una sola cuenta (default: todas)")
    fecha = time.strftime("%Y-%m-%d")
    p.add_argument("--salida", default=str(ROOT / f"catalogo_publicaciones_{fecha}.xlsx"))
    args = p.parse_args(argv)

    cuentas_cfg = [c for c in CUENTAS if not args.cuenta or c["alias"] == args.cuenta]
    por_cuenta: dict[str, list[dict]] = {}
    for cfg in cuentas_cfg:
        try:
            cu = Cuenta.cargar(cfg)
        except FileNotFoundError as e:
            log.error(str(e))
            por_cuenta[cfg["alias"]] = []
            continue
        try:
            por_cuenta[cu.alias] = procesar_cuenta(cu)
        except Exception as e:
            log.exception("[%s] error fatal: %s", cu.alias, e)
            por_cuenta[cu.alias] = []

    xlsx = Path(args.salida)
    csv_path = xlsx.with_suffix(".csv")
    exportar_xlsx(por_cuenta, xlsx)
    exportar_csv(por_cuenta, csv_path)

    print("\n========== RESUMEN ==========")
    for alias, filas in por_cuenta.items():
        print(f"  {alias}: {len(filas)} publicaciones en catalogo")
    print(f"  TOTAL: {sum(len(v) for v in por_cuenta.values())}")
    print(f"\nExcel: {xlsx}")
    print(f"CSV  : {csv_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
