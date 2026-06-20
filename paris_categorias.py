"""
paris_categorias.py
-------------------
Genera dos archivos:

  - paris_categorias_relevantes.md  : las ~21 categorias del dropdown Paris
    que aplican a productos de griferia/sanitarios.

  - paris_categorias_sugeridas.md   : tabla SKU -> producto -> categoria
    sugerida (mi mejor inferencia), basada en el nombre del producto y los
    atributos extraidos de ML.

Las categorias se leen literalmente de la hoja oculta 'herramientasdata'
(col A) del Excel, manteniendo los caracteres acentuados originales para
copiar/pegar tal cual al dropdown.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
EXCEL = BASE / "paris_productos_carga_masiva.xlsx"
OUT_RELEV = BASE / "paris_categorias_relevantes.md"
OUT_SUGER = BASE / "paris_categorias_sugeridas.md"


# Mapeo manual SKU -> categoria sugerida (clave entera = numero en la lista
# relevante). Las marcadas con TODO requieren decision de Victor.
SUGGESTIONS: dict[str, tuple[str, str]] = {
    # SKU              : (categoria_sugerida_substring, comentario)
    "MLC21628947": ("Monomando Lavaplatos",         "llave monomando lavaplato inox"),
    "MLC21350539": ("Repuestos Wc",                 "asiento y tapa WC cierre lento"),
    "MLC21394613": ("Monomando Lavaplatos",         "llave lavacopas (asimilado a lavaplatos)"),
    "MLC38612080": ("Monomando Ducha",              "llave monomando tina/ducha Colomba"),
    "MLC25452157": ("Monomando Lavamanos",          "llave monomando lavatorio Notte"),
    "MLC43500911": ("Repuestos Wc",                 "asiento con tapa WC Eco"),
    "MLC45488644": ("Monomando Lavaplatos",         "llave monomando lavaplato negra Notte"),
    "MLC41735119": ("Monomando Ducha",              "monomando ducha empotrado Notte"),
    "MLC21299614": ("Combinación Lavaplatos",       "lavaplatos sobreponer 80x50"),
    "MLC22271656": ("Combinación Lavaplatos",       "lavaplatos empotrado 80x44"),
    "MLC50745282": ("Combinación Lavaplatos",       "lavaplatos empotrado 80x44"),
    "MLC20918885": ("Sifones Y Desagües",           "desagüe tina con tubo y campana"),
    "MLC28368137": ("Sifones Y Desagües",           "desagüe lavaplatos 3 1/2"),
    "MLC24386020": ("Accesorios Y Repuestos Grifería", "flexible para llave monomando"),
    "MLC2041691925": ("Repuestos Wc",               "kit estanque WC válvula admisión - TODO confirmar (repuestos WC vs accesorios estanques)"),
    "MLC21394632": ("Monomando Lavamanos",          "llave monomando lavatorio Modern"),
    "MLC32065874": ("Accesorios Y Repuestos Grifería", "llave lavadora/lavadero/jardín 3/4 simultanea - TODO confirmar"),
    "MLC27509195": ("Repuestos Wc",                 "válvula descarga dual flush"),
    "MLC26656485": ("Sifones Y Desagües",           "sifón plástico lavaplatos"),
    "MLC67879003": ("Sifones Y Desagües",           "desagüe sifón 90mm para receptáculo"),
    "MLC22271655": ("Combinación Lavaplatos",       "lavaplatos empotrado 100x44"),
    "MLC25264406": ("Monomando Lavaplatos",         "grifo monomando profesional cuello extensible"),
    "MLC21394631": ("Monomando Lavamanos",          "llave monomando lavatorio Modern (cepillado)"),
    "MLC2055571318": ("Accesorios De Ducha",        "barra soporte ducha deslizable 65cm"),
    "MLC26778872": ("Repuestos Wc",                 "válvula descarga WC flapper y cadena"),
    "MLC22736791": ("Combinación Lavaplatos",       "pack lavaplato + llave + sifón + desagüe"),
    "MLC24044164": ("Accesorios de baño",           "lavamanos sobrepuesto de vidrio - TODO no hay categoria 'lavamanos receptaculo' explicita"),
    "MLC22271660": ("Combinación Lavaplatos",       "lavaplatos empotrable 80x50 Hannover"),
    "MLC32227841": ("Accesorios De Ducha",          "ducha fija al muro con difusor"),
    "MLC2046010486": ("Repuestos Wc",               "válvula carga con flotador estanque WC"),
    "MLC58519072": ("Monomando Lavamanos",          "grifería monocomando lavatorio Victtorino"),
    "MLC22271659": ("Combinación Lavaplatos",       "lavaplatos empotrable 686x456 Hamburg"),
    "MLC2068899311": ("Repuestos Wc",               "válvula lateral WC"),
    "MLC66202132": ("Monomando Ducha",              "mezcladora ducha empotrada acero inox"),
    "MLC27107143": ("Receptáculos",                 "shower door + receptáculo 80x80 (TODO: 'Receptaculos de Ducha' vs 'Cabinas y Duchas < Receptaculos')"),
    "MLC39244370": ("Receptáculos",                 "mampara ducha 120-140cm vidrio templado (TODO)"),
    "MLC39080093": ("Receptáculos",                 "shower door 90x90 curvo (TODO)"),
    "MLC51764090": ("Columna de Ducha",             "Täumm ducha columna"),
}


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    wb = load_workbook(EXCEL)
    ws_data = wb["herramientasdata"]
    ws_main = wb["herramientas"]

    # Lista completa de categorias desde herramientasdata col A
    cats = []
    for r in range(1, ws_data.max_row + 1):
        v = ws_data.cell(row=r, column=1).value
        if v:
            cats.append(v.strip())

    # Filtrar relevantes
    patterns = [
        "grifer", "wc y san", "ducha", "lavaplato", "lavamano", "lavator",
        "mampara", "tina", "asiento", "estanque", "valvula", "flexible",
        "sifon", "desague", "fregad", "combinaci", "columna", "accesorios de ba",
    ]
    relevant = sorted({c for c in cats if any(p in c.lower() for p in patterns)})

    # Archivo 1: lista relevante
    lines = ["# Categorias Paris relevantes para grifería / sanitarios",
             "",
             f"Total disponible en la plantilla: {len(cats)} categorias",
             f"Filtradas como relevantes a estos 38 productos: {len(relevant)}",
             "",
             "Copia la cadena EXACTA al dropdown de la columna D (Categoria) del Excel.",
             "",
             "| # | Categoria Paris |",
             "|---|---|"]
    for i, c in enumerate(relevant, 1):
        lines.append(f"| {i} | `{c}` |")
    OUT_RELEV.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[OK] {OUT_RELEV.name} ({len(relevant)} categorias)")

    # Archivo 2: sugerencias por SKU
    # Resolver substring -> categoria completa
    def find_cat(substring: str) -> str:
        for c in cats:
            if substring.lower() in c.lower():
                return c
        return f"<NO ENCONTRADA: {substring}>"

    rows_data = []
    for row in range(7, ws_main.max_row + 1):
        sku = ws_main.cell(row=row, column=2).value
        if not sku:
            continue
        nombre = ws_main.cell(row=row, column=1).value or ""
        marca = ws_main.cell(row=row, column=6).value or ""
        sug = SUGGESTIONS.get(sku)
        if sug:
            cat = find_cat(sug[0])
            comentario = sug[1]
        else:
            cat = "<sin sugerencia>"
            comentario = ""
        flag = "TODO" if "TODO" in comentario else ""
        rows_data.append((row, sku, marca, nombre, cat, comentario, flag))

    lines2 = ["# Sugerencia de categoria por producto",
              "",
              "Tabla de mapeo SKU -> categoria Paris. Las marcadas TODO son dudosas y requieren decision.",
              "",
              "| Fila | SKU | Marca | Producto | Categoria sugerida | Nota |",
              "|---|---|---|---|---|---|"]
    for r, sku, marca, nombre, cat, com, flag in rows_data:
        flag_mark = " **TODO**" if flag else ""
        lines2.append(f"| {r} | `{sku}` | {marca} | {nombre[:55]} | `{cat}` |{flag_mark} {com} |")
    OUT_SUGER.write_text("\n".join(lines2) + "\n", encoding="utf-8")
    print(f"[OK] {OUT_SUGER.name} ({len(rows_data)} filas)")

    # Resumen en consola por categoria
    print("\nDistribucion por categoria sugerida:")
    from collections import Counter
    counter = Counter()
    for r, sku, marca, nombre, cat, com, flag in rows_data:
        short = cat.rsplit("<", 1)[-1].strip()
        counter[short] += 1
    for cat, n in counter.most_common():
        print(f"  {n:>2}  {cat}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
