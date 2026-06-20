"""
Genera el plan de accion accionable para los 38 listings:
  - 10 PUEDE_COMPETIR -> BAJAR_PRECIO a ganador_precio - 1%
  - 28 PAUSAR_SEGURO  -> PAUSAR (pierden vs otra cuenta nuestra)

Lee margen_diagnostico_<fecha>.xlsx y produce plan_accion_<fecha>.xlsx
listo para que Victor revise antes de autorizar la ejecucion via API.
"""
from __future__ import annotations
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / f"margen_diagnostico_{time.strftime('%Y-%m-%d')}.xlsx"
DST = ROOT / f"plan_accion_{time.strftime('%Y-%m-%d')}.xlsx"

IVA = 0.19
ML_CHANNEL_FEE = 0.32
NETO_FACTOR = (1 / (1 + IVA)) * (1 - ML_CHANNEL_FEE)  # 0.4874


def cargar(nombre: str) -> list[dict]:
    wb = load_workbook(SRC, read_only=True)
    if nombre not in wb.sheetnames:
        return []
    ws = wb[nombre]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    return [dict(zip(h, r)) for r in rows]


def cargar_seller_nicks() -> dict[str, str]:
    """Lee sellers_externos_<fecha>.xlsx para mapear ganador_item_id -> nickname."""
    out: dict[str, str] = {}
    fecha = time.strftime("%Y-%m-%d")
    f = ROOT / f"sellers_externos_{fecha}.xlsx"
    if not f.exists():
        return out
    wb = load_workbook(f, read_only=True)
    if "DETALLE_POR_LISTING" not in wb.sheetnames:
        return out
    ws = wb["DETALLE_POR_LISTING"]
    rows = ws.iter_rows(values_only=True)
    h = next(rows)
    for r in rows:
        d = dict(zip(h, r))
        iid = d.get("item_id")
        nick = d.get("ext_seller_nickname")
        if iid and nick:
            out[iid] = nick
    return out


def calcular_margen(price: float, costo: float) -> dict:
    if not price or not costo or price <= 0 or costo <= 0:
        return {"margen_clp": None, "margen_pct": None}
    venta_neta = price / (1 + IVA)
    ingreso = venta_neta * (1 - ML_CHANNEL_FEE)
    margen = ingreso - costo
    return {
        "margen_clp": round(margen),
        "margen_pct": round(margen / venta_neta * 100, 1),
    }


def precio_sugerido(ganador_precio: float | None) -> int | None:
    """Recomienda precio para ganar buy box: 1% bajo el externo, redondeado a $10."""
    if not ganador_precio or ganador_precio <= 0:
        return None
    candidato = ganador_precio * 0.99
    return int(candidato // 10 * 10)


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: no existe {SRC}")
        return 1

    rows = cargar("TODOS_LOS_LISTINGS")
    puede_competir = [r for r in rows
                      if r.get("estado_bb") == "PIERDE_VS_EXTERNO"
                      and r.get("puede_competir_vs_ext") in (True, "True", "VERDADERO", 1)]
    pausar = [r for r in rows if r.get("estado_bb") == "PAUSAR_SEGURO"]

    print(f"PUEDE_COMPETIR (bajar precio): {len(puede_competir)}")
    print(f"PAUSAR_SEGURO (pausar)       : {len(pausar)}")

    # -------- generar acciones --------
    acciones: list[dict] = []

    # PAUSAR: prioridad 1 (riesgo nulo) - pre-aprobado
    for it in pausar:
        m_actual = calcular_margen(it.get("price_ml"), it.get("costo_vigente"))
        acciones.append({
            "aprobado": "SI",
            "accion": "PAUSAR",
            "prioridad": 1,
            "cuenta": it.get("cuenta"),
            "item_id": it.get("item_id"),
            "title": it.get("title_ml"),
            "listing_type_id": it.get("listing_type_id"),
            "stock_actual": it.get("available_quantity_ml"),
            "precio_actual": it.get("price_ml"),
            "costo": it.get("costo_vigente"),
            "margen_actual_clp": m_actual["margen_clp"],
            "margen_actual_pct": m_actual["margen_pct"],
            "precio_sugerido": None,
            "margen_proyectado_clp": None,
            "margen_proyectado_pct": None,
            "delta_margen_pct": None,
            "ganador_actual": f"NUESTRO: {it.get('ganador_cuenta')}",
            "ganador_precio": it.get("ganador_precio"),
            "razon": f"Pierde buy box vs cuenta tuya {it.get('ganador_cuenta')} "
                     f"({((it.get('diff_pct_vs_ganador') or 0))}% mas caro). Pausar evita pagar exposicion para perder contra ti mismo.",
            "catalog_product_id": it.get("catalog_product_id"),
            "permalink": it.get("permalink"),
            "diff_pct_vs_ganador": it.get("diff_pct_vs_ganador"),
        })

    # BAJAR_PRECIO: prioridad 2 (pre-aprobado solo si margen final >=15% y delta >=-10pp)
    for it in puede_competir:
        m_actual = calcular_margen(it.get("price_ml"), it.get("costo_vigente"))
        sugerido = precio_sugerido(it.get("ganador_precio"))
        m_proy = calcular_margen(sugerido, it.get("costo_vigente"))
        delta = (
            (m_proy["margen_pct"] - m_actual["margen_pct"])
            if (m_proy["margen_pct"] is not None and m_actual["margen_pct"] is not None) else None
        )
        # criterio pre-aprobacion: margen final >=15% Y delta >=-10pp
        if (m_proy["margen_pct"] is not None and m_proy["margen_pct"] >= 15
                and delta is not None and delta >= -10):
            aprobado = "SI"
        else:
            aprobado = "PENDIENTE"
        acciones.append({
            "aprobado": aprobado,
            "accion": "BAJAR_PRECIO",
            "prioridad": 2 if aprobado == "SI" else 3,
            "cuenta": it.get("cuenta"),
            "item_id": it.get("item_id"),
            "title": it.get("title_ml"),
            "listing_type_id": it.get("listing_type_id"),
            "stock_actual": it.get("available_quantity_ml"),
            "precio_actual": it.get("price_ml"),
            "costo": it.get("costo_vigente"),
            "margen_actual_clp": m_actual["margen_clp"],
            "margen_actual_pct": m_actual["margen_pct"],
            "precio_sugerido": sugerido,
            "margen_proyectado_clp": m_proy["margen_clp"],
            "margen_proyectado_pct": m_proy["margen_pct"],
            "delta_margen_pct": round(delta, 1) if delta is not None else None,
            "ganador_actual": f"EXT: {it.get('ext_seller_nick')}",
            "ganador_precio": it.get("ganador_precio"),
            "razon": f"Externo {it.get('ext_seller_nick')} gana a ${int(it.get('ganador_precio') or 0):,}".replace(",", ".") +
                     f". Bajando a ${sugerido:,}".replace(",", ".") +
                     f" recuperas buy box manteniendo margen {m_proy['margen_pct']}%.",
            "catalog_product_id": it.get("catalog_product_id"),
            "permalink": it.get("permalink"),
            "diff_pct_vs_ganador": it.get("diff_pct_vs_ganador"),
        })

    # ordenar por prioridad luego por impacto
    acciones.sort(key=lambda x: (x["prioridad"], -(x.get("precio_actual") or 0)))

    # -------- xlsx --------
    wb = Workbook()
    wb.remove(wb.active)

    # RESUMEN
    res = wb.create_sheet("RESUMEN")
    res.append(["==== PLAN DE ACCION ===="])
    res.append([f"Generado: {time.strftime('%Y-%m-%d %H:%M')}"])
    res.append([])
    res.append(["accion", "n_items", "impacto"])
    n_pausar = sum(1 for a in acciones if a["accion"] == "PAUSAR")
    n_bajar = sum(1 for a in acciones if a["accion"] == "BAJAR_PRECIO")
    res.append(["PAUSAR (riesgo nulo)", n_pausar, "Detiene exposicion contra ti mismo"])
    res.append(["BAJAR_PRECIO", n_bajar, "Recupera buy box manteniendo margen"])
    res.append([])
    # impacto detallado bajar_precio
    total_costo_invol = sum(a.get("costo") or 0 for a in acciones if a["accion"] == "BAJAR_PRECIO")
    total_margen_actual = sum(a.get("margen_actual_clp") or 0 for a in acciones if a["accion"] == "BAJAR_PRECIO")
    total_margen_nuevo = sum(a.get("margen_proyectado_clp") or 0 for a in acciones if a["accion"] == "BAJAR_PRECIO")
    res.append(["==== Si vendes 1 unidad de cada BAJAR_PRECIO ===="])
    res.append(["Margen total ACTUAL (no ganas buy box)", f"${total_margen_actual:,}".replace(",", ".")])
    res.append(["Margen total con precio nuevo (asumes buy box)", f"${total_margen_nuevo:,}".replace(",", ".")])
    res.append(["Delta (ganas margen aunque sea menor) si recuperas ventas", f"${total_margen_nuevo - total_margen_actual:,}".replace(",", ".")])
    res.append([])
    res.append(["==== INSTRUCCIONES ===="])
    res.append(["1. Revisa la hoja ACCIONES y marca 'aprobado' (SI/NO) en cada fila"])
    res.append(["2. Para BAJAR_PRECIO: si el margen proyectado no te convence, marca NO o ajusta el precio_sugerido"])
    res.append(["3. Avisa cuando este lista y corro el script de ejecucion via API"])
    res.append(["4. El script ejecutara solo las filas con aprobado=SI"])
    res.column_dimensions["A"].width = 50
    res.column_dimensions["B"].width = 20
    res.column_dimensions["C"].width = 50

    # ACCIONES
    ws = wb.create_sheet("ACCIONES")
    cols = [
        "aprobado",  # columna que Victor llena con SI/NO
        "accion", "prioridad", "cuenta", "item_id", "title",
        "listing_type_id", "stock_actual",
        "precio_actual", "costo",
        "margen_actual_clp", "margen_actual_pct",
        "precio_sugerido", "margen_proyectado_clp", "margen_proyectado_pct",
        "delta_margen_pct",
        "ganador_actual", "ganador_precio", "diff_pct_vs_ganador",
        "razon", "catalog_product_id", "permalink",
    ]
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="455A64")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    fill_pausar = PatternFill("solid", fgColor="C5E1A5")  # verde claro
    fill_bajar = PatternFill("solid", fgColor="FFE082")   # ambar claro
    fill_aprobado_col = PatternFill("solid", fgColor="FFF59D")  # amarillo input

    for a in acciones:
        ws.append([a.get(c) for c in cols])
        last = ws.max_row
        accion_fill = fill_pausar if a["accion"] == "PAUSAR" else fill_bajar
        ws.cell(row=last, column=2).fill = accion_fill
        ws.cell(row=last, column=1).fill = fill_aprobado_col
        ws.cell(row=last, column=1).alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C2"
    widths = {
        "aprobado": 9, "accion": 14, "prioridad": 8, "cuenta": 7,
        "item_id": 15, "title": 50, "listing_type_id": 13, "stock_actual": 7,
        "precio_actual": 12, "costo": 10, "margen_actual_clp": 13,
        "margen_actual_pct": 13, "precio_sugerido": 13,
        "margen_proyectado_clp": 15, "margen_proyectado_pct": 15,
        "delta_margen_pct": 12, "ganador_actual": 20, "ganador_precio": 12,
        "diff_pct_vs_ganador": 12, "razon": 65,
        "catalog_product_id": 15, "permalink": 45,
    }
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)

    wb.save(DST)
    print(f"\nXlsx: {DST}")

    # consola: muestra preview
    print("\n=== PREVIEW PAUSAR_SEGURO (top 5 por valor) ===")
    for a in sorted([x for x in acciones if x["accion"] == "PAUSAR"],
                    key=lambda x: -(x.get("precio_actual") or 0))[:5]:
        print(f"  {a['cuenta']} {a['item_id']} | ${int(a['precio_actual'] or 0):,}".replace(",", ".") +
              f" | pierde {a['diff_pct_vs_ganador']}% vs {a['ganador_actual']}")
        print(f"     {a['title'][:80]}")

    print("\n=== PREVIEW BAJAR_PRECIO (los 10) ===")
    for a in [x for x in acciones if x["accion"] == "BAJAR_PRECIO"]:
        print(f"  {a['cuenta']} {a['item_id']} | "
              f"${int(a['precio_actual'] or 0):,}".replace(",", ".") +
              f" -> ${a['precio_sugerido']:,}".replace(",", ".") +
              f" | margen {a['margen_actual_pct']}% -> {a['margen_proyectado_pct']}% "
              f"(delta {a['delta_margen_pct']}%)")
        print(f"     {a['title'][:80]}")
        print(f"     externo: {a['ganador_actual']} @ ${int(a['ganador_precio'] or 0):,}".replace(",", "."))

    return 0


if __name__ == "__main__":
    sys.exit(main())
