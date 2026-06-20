"""
Pobla el template Paris (create_products_template_*.xlsx) con los 181 productos
del export ML enriquecidos vía API. No modifica la estructura, solo agrega filas
de datos a partir de la fila 7 de cada hoja categoría.

Inputs:
  - paris_template_fixed.xlsx (copia con errorStyle parcheado)
  - paris_cache_ml_items.json (enriquecimiento API ML)
  - Publicaciones-2026_05_24-23_02.xlsx (export ML base)

Output:
  - paris_carga_181.xlsx
  - paris_carga_181_report.md (qué quedó completo / qué falta)
"""
import json, re, sys, io, shutil
from pathlib import Path
import pandas as pd
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r'C:\Users\dell\victtorino')
SRC_TEMPLATE = ROOT / 'paris_template_fixed.xlsx'
ML_EXPORT    = Path(r'C:\Users\dell\Downloads\Publicaciones-2026_05_24-23_02.xlsx')
ML_CACHE     = ROOT / 'paris_cache_ml_items.json'
OUT_XLSX     = ROOT / 'paris_carga_v2.xlsx'
REPORT_MD    = ROOT / 'paris_carga_v2_report.md'
SKUS_EXCLUIR = ROOT / 'paris_skus_duplicados.json'  # SKUs ya en Paris

# --- Mapeo ML category -> (Paris sheet, PIM path) -------------------------
MAP = {
    'Grifería y mezcladores de baño':               ('herramientas',       'Ferretería < Grifería < Grifería de Baño < Lavamanos < Monomando Lavamanos MKP'),
    'Grifería de cocina':                            ('herramientas',       'Ferretería < Grifería < Grifería de Cocina < Combinación Lavaplatos MKP'),
    'Lavaplatos de cocina':                          ('herramientas',       'Ferretería < Grifería < Grifería de Cocina < Lavaplatos < Monomando Lavaplatos MKP'),
    'Sifones de desagüe':                            ('herramientas',       'Ferretería < Gasfiteria < Complementos de Gasfitería < Sifones Y Desagües MKP'),
    'Mamparas y cabinas de ducha':                   ('herramientas',       'Construcción < Baño < Cabinas y Duchas < Receptáculos MKP'),
    'Asientos WC':                                   ('herramientas',       'Sanitarios < Tapas Y Asientos De Inodoro'),
    'Válvulas de descarga para WC':                  ('herramientas',       'Ferretería < Gasfiteria < Complementos de Gasfitería < Repuestos Wc MKP'),
    'Platos de ducha':                               ('herramientas',       'Cabinas Y Duchas < Receptáculos De Ducha'),
    'Duchas higiénicas':                             ('herramientas',       'Ferretería < Grifería < Grifería de Baño < Ducha < Accesorios De Ducha MKP'),
    'Lavamanos para baño':                           ('herramientas',       'Ferretería < Grifería < Grifería de Baño < Lavamanos < Monomando Lavamanos MKP'),
    'Mangos para ducha':                             ('herramientas',       'Ferretería < Grifería < Grifería de Baño < Ducha < Accesorios De Ducha MKP'),
    'Mangueras para lavadoras':                      ('herramientas',       'Complementos Hidrolavadoras < Mangueras, Lanzadores Y Pistolas'),
    'Válvulas de cierre':                            ('herramientas',       'Ferretería < Gasfiteria < Complementos de Gasfitería < Accesorios De Reparación MKP'),
    'WC':                                            ('herramientas',       'Construcción < WC y Sanitarios < WC y Estanques MKP'),
    'Columnas de duchas':                            ('herramientas',       'Ferretería < Grifería < Grifería de Baño < Ducha < Columna de Ducha MKP'),
    'Espejos':                                       ('decoracion',         'Hogar < Decoración < Espejos MKP'),
    'Toalleros':                                     ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Toalleros MKP'),
    'Jaboneras':                                     ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Dispensadores MKP'),
    'Dispensadores manuales de jabón y detergente':  ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Dispensadores MKP'),
    'Dispensadores de papel':                        ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Dispensadores MKP'),
    'Papeles higiénicos':                            ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Papeleros MKP'),
    'Escobillas para WC':                            ('accDeBanoYToallas',  'Hogar < Baño < Accesorios de baño MKP'),
    'Barras de seguridad y agarraderas para baño':   ('accDeBanoYToallas',  'Hogar < Baño < Accesorios de baño MKP'),
    'Kits de accesorios para baño':                  ('accDeBanoYToallas',  'Hogar < Baño < Accesorios de baño MKP'),
    'Artículos para baño':                           ('accDeBanoYToallas',  'Hogar < Baño < Accesorios de baño MKP'),
    'Organizadores de ducha para baño':              ('accDeBanoYToallas',  'Hogar < Baño < Organizadores de baño < Organizadores de Ducha MKP'),
    'Basureros':                                     ('hogarOrganizacion',  'Organización Cocina < Contenedores'),
    'Herramientas y artículos de construcción':      ('herramientas',       'Herramientas y maquinarias < Herramientas manuales < Accesorios herramientas manuales MKP'),
    'Selladores y siliconas':                        ('herramientas',       'Ferretería < Adhesivos y Selladores < Silicona MKP'),
    'Artículos de cocina':                           ('herramientas',       'Ferretería < Gasfiteria < Complementos de Gasfitería < Accesorios De Reparación MKP'),
    'Toallas de cocina':                             ('accDeBanoYToallas',  'Hogar < Baño < Accesorios de baño MKP'),
}

# --- Helpers ---------------------------------------------------------------
def get_attr(item, *ids, default=None):
    """Devuelve el value_name del primer attribute id encontrado."""
    for a in item.get('attributes') or []:
        if a.get('id') in ids:
            v = a.get('value_name')
            if v not in (None, ''): return v
    return default

def get_attr_num(item, *ids):
    """Devuelve solo la parte numérica del attribute (ej '30 cm' -> 30)."""
    v = get_attr(item, *ids)
    if v is None: return None
    m = re.search(r'-?\d+(?:[\.,]\d+)?', str(v))
    if not m: return None
    return float(m.group().replace(',', '.'))

def pics(item, limit=10):
    urls = [p.get('secure_url') or p.get('url') for p in (item.get('pictures') or [])]
    urls = [u for u in urls if u]
    return ','.join(urls[:limit])

def short_desc(item, max_len=160):
    """Descripción corta: primeros caracteres del title o un slice del long description."""
    return (item.get('title') or '')[:max_len]

# Normalización de marcas para que matcheen el dropdown de Paris
BRAND_NORMALIZE = {
    'Täumm': 'Taumm',
    'TÄUMM': 'Taumm',
    'TAUMM': 'Taumm',
    'V': 'Victtorino',
    'VICTTORINO': 'Victtorino',
    'Ducha Extensible HI-HI': 'Victtorino',
    'Colomba': 'Victtorino',
}
def normalize_brand(brand):
    if not brand: return 'Victtorino'
    return BRAND_NORMALIZE.get(brand, brand)

# Inferencia de color desde TITLE (orden importa: chequea más específico primero)
COLOR_PATTERNS = [
    ('acero inoxidable', 'Plata'),
    ('inoxidable',       'Plata'),
    ('cromado',          'Plata'),
    ('plateado',         'Plata'),
    ('plata',            'Plata'),
    ('dorado',           'Dorado'),
    ('vidrio templado',  'Transparente'),
    ('transparente',     'Transparente'),
    ('blanco',           'Blanco'),
    ('negro',            'Negro'),
    ('gris',             'Gris'),
    ('azul',             'Azul'),
    ('verde',            'Verde'),
    ('rojo',             'Rojo'),
    ('amarillo',         'Amarillo'),
    ('beige',            'Beige'),
    ('café',             'Café'),
    ('cafe',             'Café'),
    ('marrón',           'Café'),
    ('marron',           'Café'),
]

def infer_color_from_title(title):
    if not title: return None
    t = title.lower()
    for pat, color in COLOR_PATTERNS:
        if pat in t:
            return color
    # Defaults por tipo de producto cuando el título no menciona color
    if 'papel' in t or 'toalla interfoliada' in t or 'toalla de papel' in t:
        return 'Blanco'
    if 'flexible' in t:
        return 'Plata'
    if 'goma' in t:
        return 'Negro'
    if any(k in t for k in ['llave', 'grifería', 'griferia', 'válvula', 'valvula',
                             'kit anclaje', 'set de ducha', 'monomando',
                             'ducha bidet', 'kit completo ducha']):
        return 'Plata'
    return None

def value_for_label(label, item):
    """Resuelve el valor a poner en una celda dado el label de la columna."""
    if not label: return None
    L = label.lower().replace('(*)', '').strip()

    # Marca (normalizada al dropdown de Paris)
    if 'marca' in L: return normalize_brand(get_attr(item, 'BRAND'))
    # Modelo (Automotriz): es un dropdown de modelos de auto — no aplica a plomería
    if 'automotriz' in L: return None
    # Modelo regular
    if 'modelo' in L and 'fábrica' not in L: return get_attr(item, 'MODEL') or (item.get('title') or '')[:60]
    if L == 'ean': return get_attr(item, 'GTIN')

    # Descripción corta
    if 'descripción corta' in L: return short_desc(item)

    # Color (con fallback de inferencia desde el TITLE)
    if 'color' in L:
        v = get_attr(item, 'COLOR', 'MAIN_COLOR')
        if v: return v
        return infer_color_from_title(item.get('title'))
    if 'tono' in L: return get_attr(item, 'FINISH', 'COLOR')

    # Material
    if 'material' in L: return get_attr(item, 'MATERIAL', 'MATERIALS')

    # Dimensiones — peso
    if any(k in L for k in ['peso bruto']): return get_attr_num(item, 'SELLER_PACKAGE_WEIGHT', 'PACKAGE_WEIGHT', 'WEIGHT')
    if L in ('peso', 'weight', 'realweight'): return get_attr_num(item, 'WEIGHT', 'SELLER_PACKAGE_WEIGHT', 'PACKAGE_WEIGHT')

    # Dimensiones — alto
    if L in ('alto', 'height', 'realheight'): return get_attr_num(item, 'SELLER_PACKAGE_HEIGHT', 'PACKAGE_HEIGHT', 'HEIGHT')
    # ancho
    if L in ('ancho', 'width', 'realwidth'): return get_attr_num(item, 'SELLER_PACKAGE_WIDTH', 'PACKAGE_WIDTH', 'WIDTH')
    # largo / longitud
    if L in ('largo', 'length', 'reallength', 'profundidad', 'depth'): return get_attr_num(item, 'SELLER_PACKAGE_LENGTH', 'PACKAGE_LENGTH', 'LENGTH', 'DEPTH')

    # Garantía
    if 'garantía' in L and 'legal' in L: return '6 meses'

    # KeyWords (palabras claves)
    if 'keyword' in L:
        b = get_attr(item, 'BRAND') or ''
        m = get_attr(item, 'MODEL') or ''
        return ', '.join(x for x in [b, m, item.get('title','').split()[0]] if x)

    # Tipo de producto baño y plomería (etc)
    if 'tipo de producto' in L: return get_attr(item, 'PRODUCT_TYPE', 'FAUCET_CONTROL_TYPE')

    # Diámetro
    if 'diámetro' in L: return get_attr_num(item, 'DIAMETER', 'INNER_DIAMETER', 'OUTER_DIAMETER')

    # Talla (no aplica para baño)
    if L == 'talla': return 'Único'

    return None


# --- Main ------------------------------------------------------------------
def main():
    # 1. Copy template
    shutil.copy(SRC_TEMPLATE, OUT_XLSX)

    # 2. Cargar enriquecimiento
    enriched = json.loads(ML_CACHE.read_text(encoding='utf-8'))

    # 3. Cargar ML export base (para CATEGORY y SKU oficiales)
    base = pd.read_excel(ML_EXPORT, sheet_name='Publicaciones')
    base = base[base['ITEM_ID'].astype(str).str.startswith('MLC')].copy()
    base = base.set_index('ITEM_ID')

    # 3b. SKUs a excluir (ya existen en Paris)
    excluir = set()
    if SKUS_EXCLUIR.exists():
        excluir = set(json.loads(SKUS_EXCLUIR.read_text(encoding='utf-8')))
    print(f'Excluyendo {len(excluir)} SKUs ya registrados en Paris')

    # 4. Abrir workbook destino
    wb = openpyxl.load_workbook(OUT_XLSX)

    # 5. Construir índice de columnas por hoja
    SHEETS = list({mp[0] for mp in MAP.values()})
    sheet_cols = {}
    for sh in SHEETS:
        ws = wb[sh]
        cols = []
        for c in range(1, ws.max_column + 1):
            cols.append({
                'col':   c,
                'code':  ws.cell(3, c).value,
                'label': ws.cell(4, c).value,
            })
        sheet_cols[sh] = cols

    # 6. Trackeo
    report = {
        'total': 0, 'written': 0, 'skipped': [],
        'by_sheet': {}, 'missing_categories': set(),
        'fields_filled': {}, 'fields_missing_required': {}
    }

    # 7. Poblar
    sheet_next_row = {sh: 7 for sh in SHEETS}

    skus_escritos = set()
    for item_id, ml_base_row in base.iterrows():
        report['total'] += 1
        ml_cat = ml_base_row.get('CATEGORY')
        sku    = ml_base_row.get('SKU')
        if sku in excluir:
            report['skipped'].append((item_id, f'ya_registrado_en_paris:{sku}'))
            continue
        if sku in skus_escritos:
            report['skipped'].append((item_id, f'sku_duplicado_en_archivo:{sku}'))
            continue
        skus_escritos.add(sku)
        item   = enriched.get(item_id)
        if not item:
            report['skipped'].append((item_id, 'no_enriched'))
            continue
        if ml_cat not in MAP:
            report['skipped'].append((item_id, f'unmapped_category:{ml_cat}'))
            report['missing_categories'].add(ml_cat)
            continue

        sheet, pim = MAP[ml_cat]
        ws = wb[sheet]
        row = sheet_next_row[sheet]
        sheet_next_row[sheet] += 1
        report['by_sheet'].setdefault(sheet, 0)
        report['by_sheet'][sheet] += 1

        for col in sheet_cols[sheet]:
            c = col['col']; code = col['code']; label = col['label']
            v = None
            if code == 'title':              v = item.get('title')
            elif code == 'skuSeller':        v = sku
            elif code == 'skuSellerVariant': v = sku
            elif code == 'refProduct':       v = get_attr(item, 'GTIN') or sku
            elif code == 'category':         v = pim
            elif code == 'image':            v = pics(item)
            elif code == 'ean':              v = get_attr(item, 'GTIN')
            else:
                v = value_for_label(label, item)
            if v is not None and v != '':
                ws.cell(row, c, v)
                report['fields_filled'].setdefault(sheet, {}).setdefault(label or code, 0)
                report['fields_filled'][sheet][label or code] += 1
        report['written'] += 1

    # 8. Detectar obligatorios sin llenar
    for sh in sheet_cols:
        if sh not in report['by_sheet']: continue
        for col in sheet_cols[sh]:
            label = col['label'] or ''
            if '(*)' not in label: continue
            filled = report['fields_filled'].get(sh, {}).get(label, 0)
            total  = report['by_sheet'][sh]
            if filled < total:
                report['fields_missing_required'].setdefault(sh, []).append(
                    (label, f'{filled}/{total}')
                )

    # 9. Guardar
    wb.save(OUT_XLSX)
    print(f'\n✓ Guardado: {OUT_XLSX}')

    # 10. Reporte
    lines = ['# Reporte carga Paris (181 productos ML → template Paris)\n']
    lines.append(f'**Procesados:** {report["written"]}/{report["total"]}\n')
    lines.append('## Distribución por hoja Paris')
    lines.append('| Hoja | Cantidad |')
    lines.append('|---|---|')
    for sh, n in sorted(report['by_sheet'].items(), key=lambda x: -x[1]):
        lines.append(f'| {sh} | {n} |')

    lines.append('\n## Cobertura de campos obligatorios')
    for sh in report['by_sheet']:
        gaps = report['fields_missing_required'].get(sh, [])
        total = report['by_sheet'][sh]
        if gaps:
            lines.append(f'\n### {sh} ({total} productos)')
            lines.append('| Campo obligatorio | Cobertura |')
            lines.append('|---|---|')
            for label, cov in gaps:
                lines.append(f'| {label} | {cov} |')
        else:
            lines.append(f'\n### {sh} — ✅ todos los obligatorios cubiertos')

    if report['skipped']:
        lines.append('\n## Productos no procesados')
        for item_id, reason in report['skipped']:
            lines.append(f'- {item_id}: {reason}')

    REPORT_MD.write_text('\n'.join(lines), encoding='utf-8')
    print(f'✓ Reporte: {REPORT_MD}')

    # Summary stdout
    print(f'\nResumen:')
    print(f'  Procesados: {report["written"]}/{report["total"]}')
    for sh, n in report['by_sheet'].items():
        print(f'  {sh}: {n}')
    if report['skipped']:
        print(f'  Saltados: {len(report["skipped"])}')

if __name__ == '__main__':
    main()
