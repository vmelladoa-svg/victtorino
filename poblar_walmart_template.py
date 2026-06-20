"""
Pobla la plantilla Walmart Marketplace con los 181 productos ML enriquecidos.
Respeta estructura del template (no modifica filas 1-9 ni hidden sheets).

Inputs:
  - omniintl-marketplace-es_cl-external (1).xlsx  (template Walmart)
  - paris_cache_ml_items.json                       (enriquecimiento API ML)
  - Publicaciones-2026_05_24-23_02.xlsx             (export ML base)

Output:
  - walmart_carga.xlsx
  - walmart_carga_report.md
"""
import json, re, sys, io, shutil
from pathlib import Path
import pandas as pd
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r'C:\Users\dell\victtorino')
SRC_TEMPLATE = Path(r'C:\Users\dell\Downloads\omniintl-marketplace-es_cl-external (1).xlsx')
ML_EXPORT    = Path(r'C:\Users\dell\Downloads\Publicaciones-2026_05_24-23_02.xlsx')
ML_CACHE     = ROOT / 'paris_cache_ml_items.json'
OUT_XLSX     = ROOT / 'walmart_carga.xlsx'
REPORT_MD    = ROOT / 'walmart_carga_report.md'
SKUS_EXISTENTES = ROOT / 'walmart_skus_existentes.json'
EANS_GENERADOS = ROOT / 'walmart_eans_generados.json'

# ML category -> Walmart sheet
# Mapeo alineado con la categorización real de los 37 productos ya publicados en Walmart.
# Cuando hay data publicada, se usa la categoría mayoritaria. Cuando no hay productos
# publicados de esa categoría ML, se extrapola por analogía con productos similares.
MAP = {
    # Grifería / sanitarios / llaves → Ferretería (mayoría publicada va ahí)
    'Grifería y mezcladores de baño':               'Ferretería',
    'Grifería de cocina':                            'Ferretería',
    'Lavaplatos de cocina':                          'Ferretería',
    'Sifones de desagüe':                            'Ferretería',
    'Mamparas y cabinas de ducha':                   'Ferretería',
    'Asientos WC':                                   'Ferretería',
    'Válvulas de descarga para WC':                  'Ferretería',
    'Platos de ducha':                               'Ferretería',
    'Duchas higiénicas':                             'Ferretería',
    'Lavamanos para baño':                           'Ferretería',
    'Mangos para ducha':                             'Ferretería',
    'Mangueras para lavadoras':                      'Ferretería',
    'Válvulas de cierre':                            'Ferretería',
    'WC':                                            'Ferretería',
    'Columnas de duchas':                            'Ferretería',
    'Artículos para baño':                           'Ferretería',
    'Barras de seguridad y agarraderas para baño':   'Ferretería',

    # Escobillas WC: las 2 publicadas están en Plomería
    'Escobillas para WC':                            'Plomería y HVAC',

    # Decoración (espejos publicados van ahí)
    'Espejos':                                       'Decoración de Hogar, Cocina y ',
    'Dispensadores de papel':                        'Decoración de Hogar, Cocina y ',

    # Otros ferretería (accesorios sin destino claro en Ferretería)
    'Toalleros':                                     'Otros ferretería',
    'Jaboneras':                                     'Otros ferretería',
    'Dispensadores manuales de jabón y detergente':  'Otros ferretería',
    'Papeles higiénicos':                            'Otros ferretería',
    'Kits de accesorios para baño':                  'Otros ferretería',
    'Organizadores de ducha para baño':              'Otros ferretería',
    'Artículos de cocina':                           'Otros ferretería',
    'Toallas de cocina':                             'Otros ferretería',

    # Selladores y herramientas → Ferretería
    'Herramientas y artículos de construcción':      'Ferretería',
    'Selladores y siliconas':                        'Ferretería',

    # Basureros: los publicados están en "Limpieza..." (categoría que NO viene en
    # esta plantilla). Los enviamos a Otros ferretería como mejor opción disponible.
    'Basureros':                                     'Otros ferretería',
}

# ---- helpers de extracción de atributos ML --------------------------------
def get_attr(item, *ids, default=None):
    for a in item.get('attributes') or []:
        if a.get('id') in ids:
            v = a.get('value_name')
            if v not in (None, ''): return v
    return default

def get_attr_num(item, *ids):
    v = get_attr(item, *ids)
    if v is None: return None
    m = re.search(r'-?\d+(?:[\.,]\d+)?', str(v))
    if not m: return None
    return float(m.group().replace(',', '.'))

# Normalización de marcas (más permisivo en Walmart que en Paris)
BRAND_NORMALIZE = {
    'Täumm': 'Taumm',
    'TÄUMM': 'Taumm',
    'TAUMM': 'Taumm',
    'V': 'Victtorino',
    'VICTTORINO': 'Victtorino',
    'Ducha Extensible HI-HI': 'Victtorino',
    'Colomba': 'Victtorino',
}
def normalize_brand(brand):
    return BRAND_NORMALIZE.get(brand, brand) if brand else 'Victtorino'

# Inferencia de color desde TITLE
COLOR_PATTERNS = [
    ('acero inoxidable', 'Plateado'), ('inoxidable', 'Plateado'),
    ('cromado', 'Cromado'), ('plateado', 'Plateado'), ('plata', 'Plateado'),
    ('dorado', 'Dorado'), ('vidrio templado', 'Transparente'),
    ('transparente', 'Transparente'),
    ('blanco', 'Blanco'), ('negro', 'Negro'), ('gris', 'Gris'),
    ('azul', 'Azul'), ('verde', 'Verde'), ('rojo', 'Rojo'),
    ('amarillo', 'Amarillo'), ('beige', 'Beige'),
    ('café', 'Café'), ('cafe', 'Café'),
]
def infer_color_from_title(title):
    if not title: return None
    t = title.lower()
    for pat, color in COLOR_PATTERNS:
        if pat in t: return color
    if 'papel' in t or 'toalla interfoliada' in t or 'toalla de papel' in t: return 'Blanco'
    if 'flexible' in t: return 'Plateado'
    if 'goma' in t: return 'Negro'
    if any(k in t for k in ['llave','grifería','griferia','válvula','valvula',
                             'kit anclaje','set de ducha','monomando',
                             'ducha bidet','kit completo ducha']):
        return 'Plateado'
    return None

def force_jpg(url):
    """ML CDN sirve JPEG válido cuando cambiamos .webp por .jpg en la URL."""
    if not url: return url
    return url.replace('.webp', '.jpg')

def pic_main(item):
    pics = item.get('pictures') or []
    if pics: return force_jpg(pics[0].get('secure_url') or pics[0].get('url'))
    return None

def pic_secondaries(item):
    pics = item.get('pictures') or []
    return [force_jpg(p.get('secure_url') or p.get('url')) for p in pics[1:] if (p.get('secure_url') or p.get('url'))]

# ---- Schema parsing -------------------------------------------------------
def parse_sheet_schema(ws):
    """Devuelve lista de cols con {col, code, label, group, required}.
    En Walmart: row 1=version, row 2=Required marker, row 3=group(EN), row 4=label(EN),
    row 5=code (camelCase), row 6=help(EN), row 7=group(ES), row 8=label(ES), row 9=help(ES).
    Data desde row 10.
    El group label solo aparece en la PRIMERA col del grupo (ej: 'measure'),
    así que lo propagamos a las siguientes cols del mismo grupo (ej: 'unit').
    """
    cols = []
    last_group = None
    for c in range(1, ws.max_column + 1):
        required_marker = ws.cell(2, c).value
        group_en = ws.cell(3, c).value
        code = ws.cell(5, c).value
        label_es = ws.cell(8, c).value
        if group_en:
            last_group = group_en
        # Para code=unit, propagar el group del measure anterior si esta col no tiene
        effective_group = group_en if group_en else (last_group if code == 'unit' else None)
        cols.append({
            'col': c,
            'code': code,
            'label': label_es,
            'group': effective_group,
            'required': bool(required_marker and 'Required' in str(required_marker)),
        })
    return cols

def group_to_attr_code(group_str):
    """Extrae el código en paréntesis de un group como 'Shipping Dimensions Height (cm) (shippingDimensionsHeight)'"""
    if not group_str: return None
    m = re.findall(r'\(([^)]+)\)', group_str)
    if m: return m[-1]  # último paréntesis = el código
    return None

# ---- Defaults / value derivation -----------------------------------------
COUNTRY_DEFAULT = 'Chile'
CONDITION_DEFAULT = 'Nuevo'
WARRANTY_TEXT_DEFAULT = '6 meses de garantía'
SELLER_WARRANTY_DEFAULT = 'Garantía del vendedor por defectos de fabricación'
SELLER_WARRANTY_CONDITION_DEFAULT = 'Defecto de fabricación'
SELLER_WARRANTY_PERIOD_DEFAULT = 6  # meses

def derive_value(col_meta, item, sku, ml_price, eans_generados=None):
    """Devuelve el valor para una celda. None = dejar vacío."""
    code = col_meta['code']
    label = (col_meta['label'] or '').lower()
    group = group_to_attr_code(col_meta['group'])

    # Campos directos por código
    if code == 'sku': return sku
    if code == 'productIdType':
        gtin = get_attr(item, 'GTIN')
        if not gtin and eans_generados:
            gtin = eans_generados.get(sku)
        if not gtin: return None
        # UPC-A: 12 dígitos, EAN-13: 13 dígitos
        if len(str(gtin).strip()) == 12: return 'UPC'
        return 'EAN'
    if code == 'productId':
        gtin = get_attr(item, 'GTIN')
        if not gtin and eans_generados:
            gtin = eans_generados.get(sku)
        return gtin
    if code == 'productName':
        return (item.get('title') or '')[:200]
    if code == 'brand':
        return normalize_brand(get_attr(item, 'BRAND'))
    if code == 'price':
        return float(ml_price) if ml_price else None
    if code == 'ShippingWeight':
        # En kg. SELLER_PACKAGE_WEIGHT viene en g (mostly)
        w = get_attr_num(item, 'SELLER_PACKAGE_WEIGHT', 'PACKAGE_WEIGHT', 'WEIGHT')
        if w is None: return 0.5  # default 500g
        # Si el valor es > 100 asumimos gramos → convertir a kg
        return round(w/1000, 3) if w > 100 else round(w, 3)
    if code == 'pricePerUnitQuantity': return 1
    if code == 'pricePerUnitUom': return 'Unit'
    if code == 'countryOfOriginAssembly': return COUNTRY_DEFAULT
    if code == 'manufacturer': return normalize_brand(get_attr(item, 'BRAND'))
    if code == 'warrantyText': return WARRANTY_TEXT_DEFAULT
    if code == 'mainImageUrl': return pic_main(item)
    if code == 'productSecondaryImageURL':
        # Concatenar todas las secundarias separadas por coma
        sec = pic_secondaries(item)
        return ','.join(sec) if sec else None
    if code == 'keyFeatures':
        # Descripción larga: bullets de atributos clave
        bullets = []
        for attr_id, prefix in [('BRAND','Marca'), ('MODEL','Modelo'),
                                 ('COLOR','Color'), ('MATERIAL','Material'),
                                 ('FINISH','Acabado')]:
            v = get_attr(item, attr_id)
            if v: bullets.append(f'{prefix}: {v}')
        return ' | '.join(bullets) if bullets else (item.get('title') or '')
    if code == 'shortDescription':
        return (item.get('title') or '')[:200]
    if code == 'condition': return CONDITION_DEFAULT
    if code == 'sellerWarranty': return SELLER_WARRANTY_DEFAULT
    if code == 'sellerWarrantyCondition': return SELLER_WARRANTY_CONDITION_DEFAULT
    if code == 'sellerWarrantyPeriod': return SELLER_WARRANTY_PERIOD_DEFAULT
    if code == 'color':
        return get_attr(item, 'COLOR', 'MAIN_COLOR') or infer_color_from_title(item.get('title'))
    if code == 'colorCategory':
        return get_attr(item, 'COLOR', 'MAIN_COLOR') or infer_color_from_title(item.get('title'))
    if code == 'material':
        return get_attr(item, 'MATERIAL', 'MATERIALS')
    if code == 'modelNumber':
        return get_attr(item, 'MODEL') or (item.get('title') or '')[:60]
    if code == 'finish':
        return get_attr(item, 'FINISH')
    if code == 'keywords':
        b = get_attr(item, 'BRAND') or ''
        m = get_attr(item, 'MODEL') or ''
        kw = [x for x in [b, m, (item.get('title') or '').split()[0]] if x]
        return ', '.join(kw)

    # Pares measure / unit por grupo
    if code in ('measure', 'unit') and group:
        if group == 'shippingDimensionsHeight':
            h = get_attr_num(item, 'SELLER_PACKAGE_HEIGHT', 'PACKAGE_HEIGHT', 'HEIGHT') or 10
            return h if code=='measure' else 'cm'
        if group == 'ShippingDimensionsWidth':
            w = get_attr_num(item, 'SELLER_PACKAGE_WIDTH', 'PACKAGE_WIDTH', 'WIDTH') or 10
            return w if code=='measure' else 'cm'
        if group == 'ShippingDimensionsDepth':
            d = get_attr_num(item, 'SELLER_PACKAGE_LENGTH', 'PACKAGE_LENGTH', 'LENGTH', 'DEPTH') or 10
            return d if code=='measure' else 'cm'
        if group == 'assembledProductHeight':
            h = get_attr_num(item, 'HEIGHT', 'SELLER_PACKAGE_HEIGHT', 'PACKAGE_HEIGHT') or 10
            return h if code=='measure' else 'cm'
        if group == 'assembledProductWidth':
            w = get_attr_num(item, 'WIDTH', 'SELLER_PACKAGE_WIDTH', 'PACKAGE_WIDTH') or 10
            return w if code=='measure' else 'cm'
        if group == 'assembledProductLength':
            d = get_attr_num(item, 'LENGTH', 'DEPTH', 'SELLER_PACKAGE_LENGTH', 'PACKAGE_LENGTH') or 10
            return d if code=='measure' else 'cm'
        if group == 'assembledProductWeight':
            wt = get_attr_num(item, 'WEIGHT', 'SELLER_PACKAGE_WEIGHT', 'PACKAGE_WEIGHT') or 500
            wt = wt/1000 if wt > 100 else wt
            return round(wt, 3) if code=='measure' else 'kg'
    return None

# ---- Main -----------------------------------------------------------------
def main():
    shutil.copy(SRC_TEMPLATE, OUT_XLSX)
    enriched = json.loads(ML_CACHE.read_text(encoding='utf-8'))

    # SKUs ya publicados en Walmart - excluir
    excluir = set()
    if SKUS_EXISTENTES.exists():
        excluir = set(json.loads(SKUS_EXISTENTES.read_text(encoding='utf-8')))
    print(f'Excluyendo {len(excluir)} SKUs ya publicados en Walmart')

    # UPC-A generados para SKUs que ML no tiene GTIN
    eans_generados = {}
    if EANS_GENERADOS.exists():
        eans_generados = json.loads(EANS_GENERADOS.read_text(encoding='utf-8'))
    print(f'EANs generados disponibles: {len(eans_generados)}')

    base = pd.read_excel(ML_EXPORT, sheet_name='Publicaciones')
    base = base[base['ITEM_ID'].astype(str).str.startswith('MLC')].copy()
    base['PRICE'] = pd.to_numeric(base['PRICE'], errors='coerce')
    base = base.set_index('ITEM_ID')

    wb = openpyxl.load_workbook(OUT_XLSX)

    # Schema por hoja Walmart
    SHEETS = list({mp for mp in MAP.values()})
    sheet_schema = {sh: parse_sheet_schema(wb[sh]) for sh in SHEETS}
    sheet_next_row = {sh: 10 for sh in SHEETS}  # data starts at row 10

    report = {'total': 0, 'written': 0, 'skipped': [], 'by_sheet': {}, 'required_gaps': {}}

    skus_escritos = set()
    for item_id, ml_row in base.iterrows():
        report['total'] += 1
        ml_cat = ml_row.get('CATEGORY')
        sku = ml_row.get('SKU')
        ml_price = ml_row.get('PRICE')
        item = enriched.get(item_id)
        if not item:
            report['skipped'].append((item_id, 'no_enriched')); continue
        if ml_cat not in MAP:
            report['skipped'].append((item_id, f'unmapped:{ml_cat}')); continue
        if sku in excluir:
            report['skipped'].append((item_id, f'ya_publicado:{sku}')); continue
        if sku in skus_escritos:
            report['skipped'].append((item_id, f'sku_duplicado:{sku}')); continue
        skus_escritos.add(sku)

        sheet = MAP[ml_cat]
        ws = wb[sheet]
        row = sheet_next_row[sheet]
        sheet_next_row[sheet] += 1
        report['by_sheet'].setdefault(sheet, 0); report['by_sheet'][sheet] += 1

        # Para productSecondaryImageURL hay 2 columnas — distribuir URLs sin repetir
        secondary_idx = 0
        secondaries = pic_secondaries(item)
        for col_meta in sheet_schema[sheet]:
            code = col_meta['code']
            if code == 'productSecondaryImageURL':
                # Primer col: junta TODAS las secundarias con coma
                # Segundo col: vacío (evitar duplicar)
                if secondary_idx == 0 and secondaries:
                    v = ','.join(secondaries)
                else:
                    v = None
                secondary_idx += 1
            else:
                v = derive_value(col_meta, item, sku, ml_price, eans_generados)
            if v is not None and v != '':
                ws.cell(row, col_meta['col'], v)

        # Trackear gaps
        for col_meta in sheet_schema[sheet]:
            if col_meta['required']:
                v = ws.cell(row, col_meta['col']).value
                if v in (None, ''):
                    label = col_meta['label'] or col_meta['code']
                    report['required_gaps'].setdefault(sheet, {}).setdefault(label, 0)
                    report['required_gaps'][sheet][label] += 1
        report['written'] += 1

    # Eliminar pares visible+hidden de categorías sin productos.
    # Walmart valida: si Hidden_X existe, el visible X debe existir y tener data.
    # Solución: borrar AMBAS (visible y hidden) para las categorías sin productos.
    par_hojas_walmart = [
        ('Artículos para la construcción',      'Hidden_building_supply'),
        ('Decoración de Hogar, Cocina y ',      'Hidden_home_other'),
        ('Ferretería',                            'Hidden_hardware'),
        ('Muebles',                               'Hidden_furniture_other'),
        ('Otros ferretería',                      'Hidden_tools_and_hardware_othe'),
        ('Parrillas y cocina al aire lib',       'Hidden_grills_and_outdoor_cook'),
        ('Plomería y HVAC',                       'Hidden_plumbing_and_hvac'),
    ]
    eliminadas = []
    for visible, hidden in par_hojas_walmart:
        if visible in report['by_sheet']:
            continue  # tiene productos, conservar par
        if visible in wb.sheetnames: del wb[visible]; eliminadas.append(visible)
        if hidden in wb.sheetnames: del wb[hidden]; eliminadas.append(hidden)
    print(f'Hojas eliminadas (pares visible+hidden sin data): {eliminadas}')

    wb.save(OUT_XLSX)
    print(f'\n✓ Guardado: {OUT_XLSX}')

    # Reporte
    lines = ['# Reporte carga Walmart\n', f'**Procesados:** {report["written"]}/{report["total"]}\n']
    lines.append('## Distribución por hoja')
    lines.append('| Hoja | Cantidad |')
    lines.append('|---|---|')
    for sh, n in sorted(report['by_sheet'].items(), key=lambda x: -x[1]):
        lines.append(f'| {sh} | {n} |')

    lines.append('\n## Campos obligatorios sin cubrir')
    if not report['required_gaps']:
        lines.append('✅ Todos los obligatorios cubiertos en todas las hojas.')
    for sh, gaps in report['required_gaps'].items():
        total = report['by_sheet'][sh]
        lines.append(f'\n### {sh} ({total} productos)')
        lines.append('| Campo | Faltan |')
        lines.append('|---|---|')
        for label, n in sorted(gaps.items(), key=lambda x: -x[1]):
            lines.append(f'| {label} | {n}/{total} |')

    if report['skipped']:
        lines.append(f'\n## Productos saltados: {len(report["skipped"])}')
        from collections import Counter
        reasons = Counter(r.split(':',1)[0] for _, r in report['skipped'])
        for r, n in reasons.most_common():
            lines.append(f'- {r}: {n}')

    REPORT_MD.write_text('\n'.join(lines), encoding='utf-8')
    print(f'✓ Reporte: {REPORT_MD}')
    print(f'\nResumen:')
    print(f'  Procesados: {report["written"]}/{report["total"]}')
    for sh, n in report['by_sheet'].items():
        print(f'  {sh}: {n}')
    if report['required_gaps']:
        print('  ⚠️ Hay obligatorios sin cubrir (ver reporte)')

if __name__ == '__main__':
    main()
