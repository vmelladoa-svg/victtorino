# -*- coding: utf-8 -*-
"""Master de precios Falabella desde el Informe de Articulos Defontana (HTML .xls).
Trae Codigo, Descripcion, Stock, Costo Vigente (NETO). Aplica IVA x1.19 + formula
3 niveles. Marca duplicados, costo 0, alertas de stock y elegibilidad Oport.Unica."""
import pandas as pd, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC=r"C:\Users\dell\Downloads\Informe de artículos 20250627210343860887 14062026 0314.xls"
OUT=r"C:\Users\dell\victtorino\precios_master_falabella.xlsx"
IVA=1.19; M_NORMAL=4.49; M_VENTA=3.14; CMR_OFF=0.93
def num(x):
    s=re.sub(r'[^0-9]','',str(x)); return int(s) if s else 0
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

t=pd.read_html(SRC)[0].iloc[3:].reset_index(drop=True)
t.columns=['art','stock','cv','cr','pu','pf','pe','sf'][:t.shape[1]]
rows=[]
for _,r in t.iterrows():
    art=str(r['art'])
    m=re.match(r'^(\d+-[A-Z0-9]+)-(.+)$',art)
    cod,desc=(m.group(1),m.group(2)) if m else ('',art)
    rows.append({'cod':cod,'desc':desc.strip(),'stock':num(r['stock']),'cv':num(r['cv'])})

# detectar duplicados por descripcion
from collections import Counter
dcount=Counter(x['desc'].upper() for x in rows)

wb=openpyxl.Workbook(); w=wb.active; w.title='Master precios'
hdr=['Codigo','Descripcion','Stock','Costo neto','Base c/IVA','Precio NORMAL','Precio VENTA',
     '% dcto','Precio CMR','Margen venta','Elegible OU','Alertas']
w.append(hdr)
for c in w[1]:
    c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(horizontal='center',wrap_text=True)
    c.fill=PatternFill('solid',fgColor='1F4E79')

n_ok=n_c0=n_dup=n_oos=n_low=n_ou=0
for x in rows:
    al=[]
    if x['cv']<=0: al.append('SIN COSTO'); n_c0+=1
    if dcount[x['desc'].upper()]>1: al.append('DUPLICADO'); n_dup+=1
    if x['stock']<=0: al.append('SIN STOCK'); n_oos+=1
    elif x['stock']<=3: al.append('STOCK BAJO'); n_low+=1
    if x['cv']>0:
        base=x['cv']*IVA
        normal=r990(base*M_NORMAL); venta=r990(base*M_VENTA); cmr=r990(venta*CMR_OFF)
        dexh=1-venta/normal; mv=1-base/venta
        elig='SI' if normal>=19990 else 'no'
        if normal>=19990: n_ou+=1
        if not al: n_ok+=1
        w.append([x['cod'],x['desc'],x['stock'],x['cv'],int(base),normal,venta,
                  round(dexh,3),cmr,round(mv,3),elig,' / '.join(al)])
    else:
        w.append([x['cod'],x['desc'],x['stock'],x['cv'],None,None,None,None,None,None,'no',' / '.join(al)])
    rr=w.max_row
    if 'SIN COSTO' in al or 'SIN STOCK' in al:
        for c in w[rr]: c.fill=PatternFill('solid',fgColor='FFC7CE')
    elif al:
        for c in w[rr]: c.fill=PatternFill('solid',fgColor='FFEB9C')

for col in [4,5,6,7,9]:
    for rr in range(2,w.max_row+1):
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for rr in range(2,w.max_row+1):
    for col in [8,10]:
        cell=w.cell(rr,col)
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
for i,wd in enumerate([13,40,7,10,10,13,12,8,11,11,10,22]): w.column_dimensions[chr(65+i)].width=wd
wb.save(OUT)
print(f"Generado: {OUT}")
print(f"Articulos: {len(rows)}")
print(f"  OK (con costo, sin alerta): {n_ok}")
print(f"  SIN COSTO (vigente 0): {n_c0}")
print(f"  DUPLICADOS (misma desc): {n_dup}")
print(f"  SIN STOCK: {n_oos} | STOCK BAJO (<=3): {n_low}")
print(f"  Elegibles Oportunidad unica (normal>=19.990): {n_ou}")
