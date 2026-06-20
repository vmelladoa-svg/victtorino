"""Genera Excel con la propuesta SEO Lote A (77 productos) para revisión visual."""
import sys, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# Lista completa: (id, producto, categoria, focus, meta_title, meta_description)
DATOS = [
    # Receptáculos y Mamparas (9)
    (2051, "Receptor Ducha Esquinero Curvo 80x120 Izq", "Shower/Mamparas/Receptaculos", "receptor ducha esquinero 80x120 izquierdo", "Receptor Ducha Esquinero 80x120 Izquierdo | Victtorino", "Receptor de ducha esquinero curvo 80x120 izquierdo, acrílico antideslizante. Despacho a todo Chile. Calidad Victtorino."),
    (2050, "Receptor Ducha Esquinero Curvo 80x120 Der", "Shower/Mamparas/Receptaculos", "receptor ducha esquinero 80x120 derecho", "Receptor Ducha Esquinero 80x120 Derecho | Victtorino", "Receptor de ducha esquinero curvo 80x120 derecho, acrílico antideslizante. Despacho a todo Chile. Calidad Victtorino."),
    (2049, "Receptor Ducha Esquinero Curvo 90x90", "Shower/Mamparas/Receptaculos", "receptor ducha esquinero 90x90", "Receptor Ducha Esquinero 90x90 Curvo | Victtorino", "Receptor de ducha esquinero curvo 90x90 acrílico antideslizante blanco. Despacho a todo Chile. Calidad Victtorino."),
    (2043, "Receptor Ducha Esquinero Curvo 80x80", "Shower/Mamparas/Receptaculos", "receptor ducha esquinero 80x80", "Receptor Ducha Esquinero 80x80 Curvo | Victtorino", "Receptor de ducha esquinero curvo 80x80 acrílico antideslizante blanco. Despacho a todo Chile. Calidad Victtorino."),
    (2037, "Receptor Ducha Cuadrado 90x90", "Shower/Mamparas/Receptaculos", "receptor ducha cuadrado 90x90", "Receptor Ducha Cuadrado 90x90 | Victtorino", "Receptor de ducha cuadrado 90x90 acrílico antideslizante blanco. Despacho a todo Chile. Calidad Victtorino."),
    (1401, "Receptáculo Ducha 70x70", "Shower/Mamparas/Receptaculos", "receptáculo ducha 70x70", "Receptáculo Ducha 70x70 cm | Victtorino", "Receptáculo de ducha 70x70 cm, base sólida para shower compacto. Despacho a todo Chile. Calidad Victtorino."),
    (1375, "Receptáculo Ducha con Patas 140x70", "Shower/Mamparas/Receptaculos", "receptáculo ducha 140x70", "Receptáculo Ducha con Patas 140x70 cm | Victtorino", "Receptáculo de ducha con patas 140x70 cm, ideal para tinas y duchas amplias. Despacho a todo Chile. Victtorino."),
    (1374, "Mampara Ducha 120-140 x 190", "Shower/Mamparas/Receptaculos", "mampara ducha 120 140", "Mampara Ducha 120-140x190 cm | Victtorino", "Mampara de ducha 120-140x190 cm vidrio templado. Renueva tu shower con seguridad. Despacho a todo Chile. Victtorino."),
    (1286, "Receptáculo Ducha con Patas 120x70", "Shower/Mamparas/Receptaculos", "receptáculo ducha 120x70", "Receptáculo Ducha con Patas 120x70 cm | Victtorino", "Receptáculo de ducha con patas 120x70 cm, robusto y duradero. Despacho a todo Chile. Calidad Victtorino."),

    # Grifería (17)
    (1559, "Cuello Cisne Lavaplatos", "Griferia", "cuello cisne lavaplatos", "Cuello Cisne para Lavaplatos | Victtorino", "Cuello cisne para lavaplatos con tuerca hilo, repuesto universal. Despacho a todo Chile. Calidad Victtorino."),
    (1499, "Monomando Tina Ducha Colomba", "Griferia", "monomando tina ducha colomba", "Monomando Tina Ducha Colomba | Victtorino", "Llave monomando tina ducha modelo Colomba, diseño elegante y duradero. Despacho a todo Chile. Calidad Victtorino."),
    (1498, "Monomando Tina Ducha Doménica", "Griferia", "monomando tina ducha doménica", "Monomando Tina Ducha Doménica | Victtorino", "Llave monomando tina ducha modelo Doménica, acabado clásico. Despacho a todo Chile. Calidad Victtorino."),
    (1497, "Monomando Ducha Colomba", "Griferia", "monomando ducha colomba", "Monomando Ducha Colomba | Victtorino", "Llave monomando ducha modelo Colomba, mecanismo cerámico de larga duración. Despacho a todo Chile. Victtorino."),
    (1496, "Monomando Ducha Domenica", "Griferia", "monomando ducha domenica", "Monomando Ducha Doménica | Victtorino", "Llave monomando ducha modelo Doménica, instalación accesible. Despacho a todo Chile. Calidad Victtorino."),
    (1495, "Monomando Ducha Notte Negro", "Griferia", "monomando ducha notte negro", "Monomando Ducha Notte Negro al Muro | Victtorino", "Llave monomando ducha al muro Notte negro. Diseño contemporáneo, acabado oscuro premium. Despacho a todo Chile."),
    (1494, "Monomando Lavaplatos Vertical Inoxidable", "Griferia", "monomando lavaplatos vertical", "Monomando Lavaplatos Vertical Inox | Victtorino", "Llave monomando lavaplatos vertical inoxidable, ideal para cocinas modernas. Despacho a todo Chile. Victtorino."),
    (1473, "Monomando Lavaplatos Vertical Colomba", "Griferia", "monomando lavaplatos colomba", "Monomando Lavaplatos Colomba Vertical | Victtorino", "Llave monomando lavaplatos vertical modelo Colomba. Diseño elegante. Despacho a todo Chile. Victtorino."),
    (1454, "Llave Lavacopas Inoxidable", "Griferia", "llave lavacopas inoxidable", "Llave Lavacopas Inoxidable | Victtorino", "Llave lavacopas en acero inoxidable, ideal para lavadero auxiliar y cocina. Despacho a todo Chile. Victtorino."),
    (1436, "Lavacopas Bajo Cubierta 44x39", "Griferia", "lavacopas bajo cubierta 44x39", "Lavacopas Bajo Cubierta 44x39 cm | Victtorino", "Lavacopas bajo cubierta 44x39 cm, ideal para mueble auxiliar. Despacho a todo Chile. Calidad Victtorino."),
    (1116, "Monomando Lavamano Alto Inox", "Griferia", "monomando lavamanos alto inox", "Monomando Lavamanos Alto Inox | Victtorino", "Llave monomando lavamanos alto en acero inoxidable, diseño moderno. Despacho a todo Chile. Victtorino."),
    (912, "Monomando Tina Ducha Notte", "Griferia", "monomando tina ducha notte", "Monomando Tina Ducha Notte | Victtorino", "Llave monomando tina ducha Notte, mecanismo cerámico, acabado moderno. Despacho a todo Chile. Victtorino."),
    (873, "Monomando Ducha Notte Täumm", "Griferia", "monomando ducha notte täumm", "Monomando Ducha Notte Täumm | Victtorino", "Llave monomando ducha Notte Täumm, diseño minimalista premium. Despacho a todo Chile. Calidad Victtorino."),
    (834, "Monomando Lavamanos Notte Negro", "Griferia", "monomando lavamanos notte negro", "Llave Monomando Lavamanos Notte Negro | Victtorino", "Llave monomando lavamanos alto Notte negro, acabado oscuro premium. Despacho a todo Chile. Victtorino."),
    (833, "Monomando Lavamanos Modern Negro", "Griferia", "monomando lavamanos modern negro", "Monomando Lavamanos Modern Pituto Negro | Victtorino", "Llave monomando lavamanos alto Modern Pituto negro, contemporáneo. Despacho a todo Chile. Victtorino."),
    (831, "Monomando Lavaplatos Vertical Profesional", "Griferia", "monomando lavaplatos profesional", "Monomando Lavaplatos Vertical Profesional | Victtorino", "Llave monomando lavaplatos vertical profesional para cocinas exigentes. Despacho a todo Chile. Victtorino."),
    (830, "Monomando Lavaplatos Vertical Notte Negro", "Griferia", "monomando lavaplatos notte negro", "Monomando Lavaplatos Notte Negro | Victtorino", "Llave monomando lavaplatos vertical Notte negro, acabado oscuro premium. Despacho a todo Chile. Victtorino."),

    # Lavaplatos (4)
    (1260, "Lavaplatos Empotrado 100x44 Derecho", "Lavaplatos", "lavaplatos empotrado 100x44 derecho", "Lavaplatos Empotrado 100x44 Derecho | Victtorino", "Lavaplatos empotrado simple 100x44 cm secador derecho, acero inoxidable. Despacho a todo Chile. Victtorino."),
    (1259, "Lavaplatos Empotrado 80x44 Izquierdo", "Lavaplatos", "lavaplatos empotrado 80x44 izquierdo", "Lavaplatos Empotrado 80x44 Izquierdo | Victtorino", "Lavaplatos empotrado 80x44 cm secador izquierdo, acero inoxidable. Despacho a todo Chile. Calidad Victtorino."),
    (1236, "Lavaplatos Empotrado 80x44 Derecho", "Lavaplatos", "lavaplatos empotrado 80x44 derecho", "Lavaplatos Empotrado 80x44 Derecho | Victtorino", "Lavaplatos empotrado 80x44 cm secador derecho, acero inoxidable. Despacho a todo Chile. Calidad Victtorino."),
    (769, "Lavaplatos Sobrepuesto 100x50 Izquierdo", "Lavaplatos", "lavaplatos sobrepuesto 100x50 izquierdo", "Lavaplatos Sobrepuesto 100x50 Izquierdo | Victtorino", "Lavaplatos sobrepuesto 100x50 cm secador izquierdo, instalación rápida. Despacho a todo Chile. Victtorino."),

    # Lavamanos (1)
    (770, "Lavamanos Cristal Sobrepuesto 415x415", "Lavamanos", "lavamanos cristal sobrepuesto 415", "Lavamanos Cristal Sobrepuesto 415x415 mm | Victtorino", "Lavamanos cristal sobrepuesto 415x415 mm, diseño contemporáneo. Despacho a todo Chile. Calidad Victtorino."),

    # Espejos (4)
    (667, "Espejo Rectangular 60x90 Ondas", "Espejos", "espejo rectangular 60x90", "Espejo Rectangular 60x90 cm Ondas | Victtorino", "Espejo rectangular 60x90 cm con diseño de ondas, decorativo y moderno. Despacho a todo Chile. Victtorino."),
    (639, "Espejo Rectangular 50x70 LED", "Espejos", "espejo led 50x70", "Espejo Led 50x70 cm para Baño | Victtorino", "Espejo rectangular 50x70 cm con luz led integrada, ideal para baño moderno. Despacho a todo Chile. Victtorino."),
    (638, "Espejo Rectangular 45x60 LED", "Espejos", "espejo led 45x60", "Espejo Led 45x60 cm para Baño | Victtorino", "Espejo rectangular 45x60 cm con luz led integrada, ideal para baños compactos. Despacho a todo Chile. Victtorino."),
    (634, "Espejo 70x100 3 Luces Led", "Espejos", "espejo led 70x100", "Espejo Led 70x100 cm 3 Luces | Victtorino", "Espejo 70x100 cm con 3 luces led de distintas tonalidades. Diseño premium. Despacho a todo Chile. Victtorino."),

    # Dispensador (2)
    (1403, "Dispensador Jabón Blanco 500cc", "Dispensador", "dispensador jabón blanco 500cc", "Dispensador Jabón Blanco 500cc | Victtorino", "Dispensador de jabón blanco simple 500 cc, ideal para baño y cocina. Despacho a todo Chile. Calidad Victtorino."),
    (1402, "Dispensador Jabón Lavaplatos Metálico", "Dispensador", "dispensador jabón lavaplatos", "Dispensador Jabón Lavaplatos Metálico | Victtorino", "Dispensador de jabón para lavaplatos con cuerpo metálico. Instalación bajo encimera. Despacho a todo Chile."),

    # Accesorios (40)
    (1581, "Fluxómetro con Palanca", "Accesorios", "fluxómetro palanca wc", "Fluxómetro con Palanca para WC | Victtorino", "Fluxómetro con palanca para WC institucional, descarga eficiente. Despacho a todo Chile. Calidad Victtorino."),
    (1560, "Flexible Agua HI HI 1/2 100cm", "Accesorios", "flexible agua 1/2 100cm", "Flexible Agua HI HI 1/2 100 cm | Victtorino", "Flexible de agua HI HI 1/2 pulgada por 100 cm, ideal para instalaciones extendidas. Despacho a todo Chile."),
    (1557, "Sifón Lavaplatos 1 1/2 Salida Recta", "Accesorios", "sifón lavaplatos 1 1/2 recto", "Sifón Lavaplatos 1 1/2 Salida Recta | Victtorino", "Sifón lavaplatos 1 1/2 con salida recta, plástico durable. Despacho a todo Chile. Calidad Victtorino."),
    (1556, "Sifón Lavaplatos 1 1/2 Botella Metálico", "Accesorios", "sifón lavaplatos botella metálico", "Sifón Botella Lavaplatos 1 1/2 Metálico | Victtorino", "Sifón botella metálico para lavaplatos 1 1/2 pulgadas, durabilidad superior. Despacho a todo Chile. Victtorino."),
    (1501, "Urinario sin Válvula", "Accesorios", "urinario sin válvula", "Urinario sin Válvula de Corte | Victtorino", "Urinario sin válvula de corte, instalación profesional para baños institucionales. Despacho a todo Chile. Victtorino."),
    (1472, "Teflón 1/2", "Accesorios", "teflón 1/2", "Teflón 1/2 para Gasfitería | Victtorino", "Teflón 1/2 pulgada para sellado de conexiones de agua y gas. Despacho a todo Chile. Calidad Victtorino."),
    (1471, "Teflón 1", "Accesorios", "teflón 1 pulgada", "Teflón 1 Pulgada Gasfitería | Victtorino", "Teflón 1 pulgada para sellado de conexiones de agua y gas residencial. Despacho a todo Chile. Victtorino."),
    (1470, "Flexible Gas HI HI 3/8 1/2 100cm", "Accesorios", "flexible gas 3/8 1/2", "Flexible Gas 3/8 x 1/2 - 100 cm | Victtorino", "Flexible gas HI HI 3/8 x 1/2 pulgada por 100 cm, certificado y seguro. Despacho a todo Chile. Victtorino."),
    (1458, "Flexible Agua HE HI 1/2 35cm", "Accesorios", "flexible agua 1/2 35cm", "Flexible Agua HE HI 1/2 - 35 cm | Victtorino", "Flexible de agua HE HI 1/2 pulgada por 35 cm, conexión estándar chilena. Despacho a todo Chile. Victtorino."),
    (1457, "Manguera Entrada Agua Lavadora", "Accesorios", "manguera entrada agua lavadora", "Manguera Entrada Agua Lavadora | Victtorino", "Manguera entrada de agua para lavadora, conexión universal estándar. Despacho a todo Chile. Calidad Victtorino."),
    (1456, "Flexible Agua HI HE 1/2 50cm", "Accesorios", "flexible agua 1/2 50cm", "Flexible Agua HI HE 1/2 - 50 cm | Victtorino", "Flexible de agua HI HE 1/2 pulgada por 50 cm, ideal para lavamanos y WC. Despacho a todo Chile. Victtorino."),
    (1455, "Sifón Lavaplatos 1 1/2 Tubo Plástico", "Accesorios", "sifón lavaplatos tubo plástico", "Sifón Lavaplatos Tubo Plástico 1 1/2 | Victtorino", "Sifón tubo plástico para lavaplatos 1 1/2 pulgadas, ligero y económico. Despacho a todo Chile. Victtorino."),
    (1433, "Portavaso Colomba", "Accesorios", "portavaso colomba", "Portavaso Baño Colomba | Victtorino", "Portavaso para baño modelo Colomba, combina con set Colomba. Despacho a todo Chile. Calidad Victtorino."),
    (1432, "Perchero Inoxidable", "Accesorios", "perchero inoxidable baño", "Perchero Inoxidable para Baño | Victtorino", "Perchero acero inoxidable para baño, resistente a la humedad. Despacho a todo Chile. Calidad Victtorino."),
    (1406, "Jabonera Colomba", "Accesorios", "jabonera colomba", "Jabonera Baño Colomba | Victtorino", "Jabonera para baño modelo Colomba, combina con accesorios línea Colomba. Despacho a todo Chile. Victtorino."),
    (1404, "Porta Toallas Cromado", "Accesorios", "porta toallas cromado", "Porta Toallas de Manos Cromado | Victtorino", "Porta toallas de manos cromado, instalación al muro. Despacho a todo Chile. Calidad Victtorino."),
    (1373, "Sifón Urinario 1 1/4 Metálico", "Accesorios", "sifón urinario 1 1/4", "Sifón Urinario 1 1/4 Metálico | Victtorino", "Sifón urinario 1 1/4 metálico para baños comerciales. Despacho a todo Chile. Calidad Victtorino."),
    (1349, "Válvula Descarga Dual Flush One Piece", "Accesorios", "válvula dual flush one piece", "Válvula Dual Flush WC One Piece | Victtorino", "Válvula descarga dual flush para WC monobloque (one piece). Ahorro de agua. Despacho a todo Chile. Victtorino."),
    (1348, "Válvula Descarga Dual Flush Estanque", "Accesorios", "válvula dual flush estanque", "Válvula Dual Flush Estanque Tradicional | Victtorino", "Válvula descarga dual flush para estanque tradicional. Ahorro automático. Despacho a todo Chile. Victtorino."),
    (1318, "Kit Anclaje WC", "Accesorios", "kit anclaje wc", "Kit Anclaje para WC | Victtorino", "Kit completo de anclaje para WC al piso, tornillería y sellos incluidos. Despacho a todo Chile. Victtorino."),
    (1317, "Sifón Botella Metálico Lavaplatos 1 1/4", "Accesorios", "sifón botella metálico lavaplatos", "Sifón Botella Metálico Lavaplatos 1 1/4 | Victtorino", "Sifón botella metálico lavaplatos 1 1/4 pulgadas, durable. Despacho a todo Chile. Calidad Victtorino."),
    (1316, "Plato Ducha Redondo 20cm", "Accesorios", "plato ducha redondo 20cm", "Plato Ducha Redondo 20 cm | Victtorino", "Plato de ducha redondo 20 cm, lluvia uniforme. Despacho a todo Chile. Calidad Victtorino."),
    (1289, "Barra 3 Apoyos 90° Esquinera", "Accesorios", "barra esquinera 3 apoyos", "Barra Esquinera 3 Apoyos 90° | Victtorino", "Barra de seguridad esquinera con 3 apoyos a 90°, máxima estabilidad. Despacho a todo Chile. Victtorino."),
    (1288, "Barra Seguridad Recta 60cm", "Accesorios", "barra seguridad recta 60 cm", "Barra Seguridad Recta 60 cm | Victtorino", "Barra de seguridad recta 60 cm para ducha y baño. Despacho a todo Chile. Calidad Victtorino."),
    (1234, "Flexible Agua Trenzada HE/HI 35cm", "Accesorios", "flexible agua trenzada 35cm", "Flexible Agua Trenzada HE/HI 35 cm | Victtorino", "Flexible de agua trenzada HE/HI 35 cm, mayor resistencia. Despacho a todo Chile. Calidad Victtorino."),
    (1233, "Flexible Agua HE HI 1/2 50cm", "Accesorios", "flexible agua he hi 1/2 50cm", "Flexible Agua HE HI 1/2 - 50 cm | Victtorino", "Flexible de agua HE HI 1/2 pulgada por 50 cm, conexión estándar chilena. Despacho a todo Chile. Victtorino."),
    (1231, "Desagüe Rebalse Flexible 3 1/2", "Accesorios", "desagüe rebalse flexible 3 1/2", "Desagüe Rebalse Flexible 3 1/2 | Victtorino", "Desagüe de rebalse flexible 3 1/2 pulgadas para lavaplatos. Despacho a todo Chile. Calidad Victtorino."),
    (1230, "Barra Apoyo 30cm Baño", "Accesorios", "barra apoyo 30 cm baño", "Barra de Apoyo 30 cm para Baño | Victtorino", "Barra de apoyo 30 cm para baño, seguridad para abuelos y niños. Despacho a todo Chile. Victtorino."),
    (1186, "WC Victtorino Compacto", "Accesorios", "wc victtorino compacto", "WC Victtorino Compacto y Eficiente | Victtorino", "WC Victtorino diseño compacto y eficiente, ideal para baños chicos. Despacho a todo Chile. Calidad Victtorino."),
    (1117, "Manguera Ducha Cromada 185cm", "Accesorios", "manguera ducha 185cm", "Manguera de Ducha 185 cm Inox | Victtorino", "Manguera de ducha flexible cromada 185 cm acero inoxidable. Despacho a todo Chile. Calidad Victtorino."),
    (1115, "Bandeja Lavaplatos Hannover", "Accesorios", "bandeja lavaplatos hannover", "Bandeja Escurridora Hannover | Victtorino", "Bandeja escurridora Hannover para lavaplatos. Acero inoxidable. Despacho a todo Chile. Calidad Victtorino."),
    (1087, "Válvula Carga Lateral WC", "Accesorios", "válvula carga lateral wc", "Válvula Carga Lateral WC | Victtorino", "Válvula de carga lateral para estanque WC, mecanismo silencioso. Despacho a todo Chile. Calidad Victtorino."),
    (1084, "Toallero Schwarz 61cm", "Accesorios", "toallero schwarz 61 cm", "Toallero Línea Schwarz 61 cm | Victtorino", "Toallero baño barra de 61 cm línea Schwarz, acabado oscuro premium. Despacho a todo Chile. Victtorino."),
    (1080, "Sifón Económico Lavamanos 1 1/4", "Accesorios", "sifón lavamanos económico 1 1/4", "Sifón Económico Lavamanos 1 1/4 | Victtorino", "Sifón económico para lavamanos 1 1/4 pulgada. Reemplazo simple. Despacho a todo Chile. Calidad Victtorino."),
    (1025, "Set Accesorios Baño Colomba 6 Piezas", "Accesorios", "set accesorios baño colomba 6 piezas", "Set Accesorios Baño Colomba 6 Piezas | Victtorino", "Set completo de 6 accesorios para baño línea Colomba. Despacho a todo Chile. Calidad Victtorino."),
    (1024, "Set Accesorios Baño Colomba 5 Piezas", "Accesorios", "set accesorios baño colomba 5 piezas", "Set Accesorios Baño Colomba 5 Piezas | Victtorino", "Set completo de 5 accesorios para baño línea Colomba. Despacho a todo Chile. Calidad Victtorino."),
    (1021, "Toalla Papel Rollo 250mt", "Accesorios", "toalla papel rollo industrial", "Toalla Papel Rollo 250 metros Pack 2 | Victtorino", "Pack 2 rollos de toalla de papel industrial 250 metros para baño público. Despacho a todo Chile."),
    (986, "Portapapel Higiénico Industrial Acero", "Accesorios", "portapapel higiénico industrial", "Portapapel Higiénico Industrial Acero | Victtorino", "Portapapel higiénico industrial acero inoxidable, rollo grande. Despacho a todo Chile. Victtorino."),
    (559, "Brazo Ducha 40cm Schwartz Negro", "Accesorios", "brazo ducha 40 cm negro", "Brazo Ducha al Muro 40 cm Negro | Victtorino", "Brazo de ducha al muro 40 cm acabado Schwartz negro. Diseño contemporáneo. Despacho a todo Chile."),
    (548, "Bandeja Canasto Esquinero Baño", "Accesorios", "bandeja esquinera baño", "Bandeja Canasto Esquinero Baño | Victtorino", "Bandeja canasto esquinero para baño, aprovecha esquinas de la ducha. Despacho a todo Chile. Victtorino."),
]

assert len(DATOS) == 77, f"Esperaba 77, son {len(DATOS)}"

wb = Workbook()
ws = wb.active
ws.title = "Lote A — 77 productos"

HEADERS = [
    ("Decisión", 12),
    ("Woo ID", 8),
    ("Categoría", 22),
    ("Producto", 42),
    ("Focus keyword", 32),
    ("Meta title (≤60)", 50),
    ("Meta description (≤155)", 70),
    ("Notas / cambios", 30),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "B2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

# Color por categoría para escaneo visual
cat_color = {
    "Shower/Mamparas/Receptaculos": "E7F0FA",
    "Griferia": "FCE7EF",
    "Lavaplatos": "FFF5E0",
    "Lavamanos": "E5F4E5",
    "Espejos": "F0E5F4",
    "Dispensador": "FFFBE5",
    "Accesorios": "F4F4F4",
}

for r_idx, (pid, name, cat, focus, mt, md) in enumerate(DATOS, start=2):
    fila = ["", pid, cat, name, focus, mt, md, ""]
    fill = PatternFill("solid", fgColor=cat_color.get(cat, "FFFFFF"))
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 1:
            c.fill = PatternFill("solid", fgColor="FFFFFF")
        elif c_idx in (3, 4):
            c.fill = fill
    # Validar longitudes y marcar si exceden
    if len(mt) > 60:
        ws.cell(row=r_idx, column=6).fill = PatternFill("solid", fgColor="FFCCCC")
    if len(md) > 155:
        ws.cell(row=r_idx, column=7).fill = PatternFill("solid", fgColor="FFCCCC")
    ws.row_dimensions[r_idx].height = 55

# Dropdown en columna A
dv = DataValidation(type="list", formula1='"OK,Editar,Saltar"', allow_blank=True)
dv.add(f"A2:A{len(DATOS) + 1}")
ws.add_data_validation(dv)

# Hoja de instrucciones
ws2 = wb.create_sheet("Instrucciones", 0)
inst = [
    ("Cómo revisar la propuesta SEO Lote A", True),
    ("", False),
    ("La hoja 'Lote A — 77 productos' tiene los 77 productos del catálogo que están sin focus", False),
    ("keyword pero ya tienen descripción decente (≥200 palabras).", False),
    ("", False),
    ("Para cada producto te propongo:", True),
    ("  - Focus keyword (lo que el cliente buscaría en Google Chile)", False),
    ("  - Meta title (≤60 caracteres, termina en | Victtorino)", False),
    ("  - Meta description (≤155 caracteres, con CTA Despacho a todo Chile)", False),
    ("", False),
    ("Tu trabajo:", True),
    ("  Por cada fila, en la columna A 'Decisión' eliges del dropdown:", False),
    ("    OK     = aplicar tal como está propuesto", False),
    ("    Editar = quiero cambiar algo (escríbelo en la columna 'Notas / cambios')", False),
    ("    Saltar = no aplicar este producto", False),
    ("", False),
    ("Las filas que dejes vacías en la columna A se interpretan como OK por defecto.", False),
    ("", False),
    ("Cuando termines:", True),
    ("  Guarda el Excel (Ctrl+S) y dile a Claude: 'aplica el Lote A'.", False),
    ("  Claude leerá tus decisiones, ajustará lo marcado como Editar, saltará lo marcado", False),
    ("  como Saltar, y aplicará todo el resto vía API.", False),
    ("", False),
    ("Pistas visuales:", True),
    ("  - Colores de categoría: azul=Shower, rosa=Grifería, ámbar=Lavaplatos, etc.", False),
    ("  - Si un meta title o description excede el límite, se pinta en rojo.", False),
    ("  - La columna B (Woo ID) tiene un freeze para que no se mueva al hacer scroll.", False),
]
for r, (txt, bold) in enumerate(inst, start=1):
    c = ws2.cell(row=r, column=1, value=txt)
    if bold:
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 110

out = r"C:\Users\dell\victtorino\propuesta_lote_a.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"Filas: {len(DATOS)}")

# Estadísticas
mt_over = sum(1 for d in DATOS if len(d[4]) > 60)
md_over = sum(1 for d in DATOS if len(d[5]) > 155)
print(f"\nValidación:")
print(f"  Meta titles que exceden 60 chars: {mt_over}")
print(f"  Meta descriptions que exceden 155 chars: {md_over}")
