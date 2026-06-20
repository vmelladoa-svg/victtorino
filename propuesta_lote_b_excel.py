"""Excel Lote B - 28 productos físicos con descripción 100-200 palabras sin focus."""
import sys, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATOS = [
    # Shower/Mamparas/Receptaculos (1)
    (1829, "SHOWER DOOR 90X90X195 CUADRADO DISCRETO", "Shower/Mamparas/Receptaculos",
     "shower door 90x90x195 discreto",
     "Shower Door 90x90x195 Cuadrado Discreto | Victtorino",
     "Shower door 90x90x195 cuadrado discreto, vidrio templado de seguridad. Despacho a todo Chile. Calidad Victtorino."),

    # Lavaplatos (9)
    (1825, "Pack Lavaplatos 80x44 Izquierdo + Llave Monomando", "Lavaplatos",
     "pack lavaplatos 80x44 izquierdo",
     "Pack Lavaplatos 80x44 Izquierdo + Monomando | Victtorino",
     "Pack lavaplatos 80x44 secador izquierdo con llave monomando. Cocina completa. Despacho a todo Chile. Victtorino."),
    (1824, "Pack Lavaplatos 80x44 Derecho + Llave Monomando", "Lavaplatos",
     "pack lavaplatos 80x44 derecho",
     "Pack Lavaplatos 80x44 Derecho + Monomando | Victtorino",
     "Pack lavaplatos 80x44 secador derecho con llave monomando. Cocina completa. Despacho a todo Chile. Victtorino."),
    (1823, "Pack Lavaplatos 37x32 + Llave Lavacopa Inox", "Lavaplatos",
     "pack lavaplatos 37x32",
     "Pack Lavaplatos 37x32 + Llave Lavacopa | Victtorino",
     "Pack lavaplatos compacto 37x32 con llave lavacopa inoxidable. Ideal para isla. Despacho a todo Chile. Victtorino."),
    (1822, "Pack Lavaplatos 100x44 Izquierdo + Llave Monomando", "Lavaplatos",
     "pack lavaplatos 100x44 izquierdo",
     "Pack Lavaplatos 100x44 Izquierdo + Monomando | Victtorino",
     "Pack lavaplatos 100x44 secador izquierdo con llave monomando. Para cocinas amplias. Despacho a todo Chile."),
    (1821, "Pack Lavaplatos 100x44 Derecho + Llave Monomando", "Lavaplatos",
     "pack lavaplatos 100x44 derecho",
     "Pack Lavaplatos 100x44 Derecho + Monomando | Victtorino",
     "Pack lavaplatos 100x44 secador derecho con llave monomando. Cocina familiar. Despacho a todo Chile. Victtorino."),
    (1819, "Lavaplatos Simple Empotrado 686x456 Hamburg", "Lavaplatos",
     "lavaplatos hamburg 686x456",
     "Lavaplatos Empotrado Hamburg 686x456 mm | Victtorino",
     "Lavaplatos empotrado modelo Hamburg 686x456 mm acero inoxidable. Despacho a todo Chile. Calidad Victtorino."),
    (1818, "Lavaplatos Sobreponer 180x50 2 Bachas 2 Secadores", "Lavaplatos",
     "lavaplatos sobreponer 180x50 doble",
     "Lavaplatos Sobreponer 180x50 2 Bachas | Victtorino",
     "Lavaplatos para sobreponer 180x50 cm con 2 bachas y 2 secadores. Para cocinas grandes. Despacho a todo Chile."),
    (1817, "Lavaplatos Sobreponer 150x50x16 2 Bachas 2 Secadores", "Lavaplatos",
     "lavaplatos sobreponer 150x50 doble",
     "Lavaplatos Sobreponer 150x50 2 Bachas | Victtorino",
     "Lavaplatos para sobreponer 150x50x16 cm con 2 bachas y 2 secadores. Cocina amplia. Despacho a todo Chile."),
    (1585, "Lavaplatos empotrado 37x32 chocolate", "Lavaplatos",
     "lavaplatos 37x32 chocolate",
     "Lavaplatos Empotrado 37x32 Chocolate | Victtorino",
     "Lavaplatos empotrado 37x32 cm en color chocolate, ideal para isla. Despacho a todo Chile. Calidad Victtorino."),

    # Griferia (5)
    (1816, "LAVACOPA NOTTE", "Griferia",
     "lavacopa notte",
     "Lavacopa Notte | Victtorino",
     "Llave lavacopa modelo Notte, diseño contemporáneo en negro premium. Despacho a todo Chile. Calidad Victtorino."),
    (875, "Llave Monomando Lavaplatos Vertical Domenica", "Griferia",
     "monomando lavaplatos domenica vertical",
     "Monomando Lavaplatos Vertical Doménica | Victtorino",
     "Llave monomando lavaplatos vertical modelo Doménica, acabado clásico premium. Despacho a todo Chile. Victtorino."),
    (801, "Llave Monomando Ducha Empotrada Dusseldorf", "Griferia",
     "monomando ducha empotrado dusseldorf",
     "Monomando Ducha Empotrado Dusseldorf | Victtorino",
     "Llave monomando ducha empotrada modelo Dusseldorf, instalación al muro. Despacho a todo Chile. Victtorino."),
    (799, "Llave Lavadero Lavadora Doble 3/4", "Griferia",
     "llave lavadero lavadora doble 3/4",
     "Llave Lavadero Lavadora Doble 3/4 | Victtorino",
     "Llave lavadero para lavadora con doble salida 3/4 pulgada. Ideal para patio. Despacho a todo Chile. Victtorino."),
    (772, "Llave Combinación Lavaplatos Antonella", "Griferia",
     "llave combinación antonella",
     "Llave Combinación Lavaplatos Antonella | Victtorino",
     "Llave combinación lavaplatos modelo Antonella, doble salida agua fría/caliente. Despacho a todo Chile. Victtorino."),

    # Accesorios (11)
    (1815, "Escobilla Limpia WC Acero Inoxidable", "Accesorios",
     "escobilla wc acero inoxidable",
     "Escobilla Limpia WC Acero Inoxidable | Victtorino",
     "Escobilla limpia WC en acero inoxidable, resistente y elegante. Despacho a todo Chile. Calidad Victtorino."),
    (1812, "Asiento Y Tapa WC Cierre Lento Modern", "Accesorios",
     "asiento wc cierre lento modern",
     "Asiento Tapa WC Cierre Lento Modern | Victtorino",
     "Asiento y tapa WC cierre lento blanco modelo Modern. Silencioso y duradero. Despacho a todo Chile. Victtorino."),
    (1811, "Asiento Y Tapa WC Fuscher Cierre Lento", "Accesorios",
     "asiento wc cierre lento fuscher",
     "Asiento Tapa WC Cierre Lento Fuscher | Victtorino",
     "Asiento y tapa WC cierre lento blanco modelo Fuscher. Sin golpes. Despacho a todo Chile. Calidad Victtorino."),
    (1809, "Asiento Con Tapa WC Eco Ovalada Blanco", "Accesorios",
     "asiento wc eco ovalado",
     "Asiento Tapa WC Eco Ovalada Blanco | Victtorino",
     "Asiento con tapa WC modelo Eco ovalada blanca, económico y resistente. Despacho a todo Chile. Victtorino."),
    (1315, "Brazo Ducha Entrada Flexible 30 cm", "Accesorios",
     "brazo ducha flexible 30 cm",
     "Brazo Ducha Entrada Flexible 30 cm | Victtorino",
     "Brazo de ducha entrada flexible 30 cm, ideal para ajustar ángulo del plato. Despacho a todo Chile. Victtorino."),
    (1119, "Barra Seguridad Acero Inox 3 Puntos", "Accesorios",
     "barra seguridad 3 puntos baño",
     "Barra Seguridad Inox 3 Puntos para Baño | Victtorino",
     "Barra de seguridad acero inoxidable con 3 puntos de apoyo. Máxima estabilidad en baño. Despacho a todo Chile."),
    (1114, "Válvula Desvío Diverter 3 Vías 1/2 x 3/4", "Accesorios",
     "válvula desvío diverter 3 vías",
     "Válvula Diverter 3 Vías 1/2 x 3/4 | Victtorino",
     "Válvula desvío diverter 3 vías 1/2 x 3/4, control de flujo agua caliente y fría. Despacho a todo Chile."),
    (1083, "Toallas Interfoliadas Paquete 200 Unidades", "Accesorios",
     "toallas interfoliadas 200 unidades",
     "Toallas Interfoliadas Pack 200 Unidades | Victtorino",
     "Pack 200 toallas interfoliadas para dispensador, papel absorbente premium. Despacho a todo Chile. Victtorino."),
    (914, "Papel Higiénico Industrial 216 mt", "Accesorios",
     "papel higiénico industrial 216 metros",
     "Papel Higiénico Industrial 216 metros | Victtorino",
     "Rollo papel higiénico industrial 216 metros para baño público comercial. Despacho a todo Chile. Victtorino."),
    (872, "Mango Ducha Premium 3 Funciones Conster", "Accesorios",
     "mango ducha 3 funciones conster",
     "Mango Ducha Premium 3 Funciones Conster | Victtorino",
     "Mango de ducha premium con 3 funciones de lluvia modelo Conster. Despacho a todo Chile. Calidad Victtorino."),
    (549, "Barra Abatible para WC 78 cm", "Accesorios",
     "barra abatible wc 78 cm",
     "Barra Abatible para WC 78 cm | Victtorino",
     "Barra abatible para WC de 78 cm, seguridad para adultos mayores. Despacho a todo Chile. Calidad Victtorino."),

    # Espejos (1)
    (666, "Espejo Rectangular 60x90 Borde Azul", "Espejos",
     "espejo rectangular 60x90 azul",
     "Espejo Rectangular 60x90 Borde Azul | Victtorino",
     "Espejo rectangular 60x90 cm con borde azul y líneas celestes, decorativo. Despacho a todo Chile. Victtorino."),

    # Dispensador (1)
    (633, "Dispensador de Jabón 500 ml", "Dispensador",
     "dispensador jabón 500 ml",
     "Dispensador de Jabón 500 ml | Victtorino",
     "Dispensador de jabón líquido 500 ml, recargable para baño y cocina. Despacho a todo Chile. Calidad Victtorino."),
]

assert len(DATOS) == 28, f"Esperaba 28, son {len(DATOS)}"

wb = Workbook()
ws = wb.active
ws.title = "Lote B — 28 productos"

HEADERS = [
    ("Decisión", 12), ("Woo ID", 8), ("Categoría", 22), ("Producto", 42),
    ("Focus keyword", 32), ("Meta title (≤60)", 50), ("Meta description (≤155)", 70),
    ("Notas / cambios", 30),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "B2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

cat_color = {
    "Shower/Mamparas/Receptaculos": "E7F0FA",
    "Griferia": "FCE7EF",
    "Lavaplatos": "FFF5E0",
    "Lavamanos": "E5F4E5",
    "Espejos": "F0E5F4",
    "Dispensador": "FFFBE5",
    "Accesorios": "F4F4F4",
}

for r_idx, (pid, name, cat, focus, mt, md) in enumerate(DATOS, start=2):
    fila = ["", pid, cat, name, focus, mt, md, ""]
    fill = PatternFill("solid", fgColor=cat_color.get(cat, "FFFFFF"))
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 1:
            c.fill = PatternFill("solid", fgColor="FFFFFF")
        elif c_idx in (3, 4):
            c.fill = fill
    if len(mt) > 60:
        ws.cell(row=r_idx, column=6).fill = PatternFill("solid", fgColor="FFCCCC")
    if len(md) > 155:
        ws.cell(row=r_idx, column=7).fill = PatternFill("solid", fgColor="FFCCCC")
    ws.row_dimensions[r_idx].height = 55

dv = DataValidation(type="list", formula1='"OK,Editar,Saltar"', allow_blank=True)
dv.add(f"A2:A{len(DATOS) + 1}")
ws.add_data_validation(dv)

out = r"C:\Users\dell\victtorino\propuesta_lote_b.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"Filas: {len(DATOS)}")
mt_over = sum(1 for d in DATOS if len(d[4]) > 60)
md_over = sum(1 for d in DATOS if len(d[5]) > 155)
print(f"Meta titles >60: {mt_over}, Meta descriptions >155: {md_over}")
