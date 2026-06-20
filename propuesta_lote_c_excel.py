"""Excel Lote C - 53 productos físicos con descripción <100 palabras sin focus.
Cada uno tendrá descripción HTML premium generada al aplicar (plantilla por categoría)."""
import sys, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# (id, name, cat_actual, focus, meta_title, meta_description, cat_destino_para_plantilla)
DATOS = [
    # Accesorios -> mapeo a la categoría correcta para plantilla SEO premium
    (1810, "Asiento Y Tapa De Wc Blanca Colomba", "Accesorios", "asiento wc colomba", "Asiento y Tapa WC Colomba Blanca | Victtorino", "Asiento y tapa WC blanca modelo Colomba, instalación universal. Despacho a todo Chile. Calidad Victtorino.", "WC e Inodoros"),
    (1089, "Valvula Descarga Wc Flapper", "Accesorios", "válvula descarga wc flapper", "Válvula Descarga WC Flapper | Victtorino", "Válvula descarga WC tipo flapper, mecanismo simple y duradero. Despacho a todo Chile. Calidad Victtorino.", "WC e Inodoros"),
    (1088, "Válvula De Carga Wc Con Flotador", "Accesorios", "válvula carga wc flotador", "Válvula Carga WC con Flotador | Victtorino", "Válvula de carga WC con flotador, reemplazo simple para estanques. Despacho a todo Chile. Calidad Victtorino.", "WC e Inodoros"),
    (1086, "Válvula Carga Silenciosa Simple Wc", "Accesorios", "válvula carga wc silenciosa", "Válvula Carga WC Silenciosa Simple | Victtorino", "Válvula carga WC silenciosa, evita ruidos durante el llenado. Despacho a todo Chile. Calidad Victtorino.", "WC e Inodoros"),
    (1081, "Sifón Lavamanos 1 1/4 Codo Plástico", "Accesorios", "sifón lavamanos 1 1/4 codo", "Sifón Lavamanos 1 1/4 Codo Plástico | Victtorino", "Sifón lavamanos 1 1/4 con codo plástico, instalación rápida. Despacho a todo Chile. Calidad Victtorino.", "Sifones y Desagües"),
    (1054, "Sifón De Lavamanos 1 1/4 Con Codo Plástico En Caja", "Accesorios", "sifón lavamanos 1 1/4 caja", "Sifón Lavamanos 1 1/4 con Codo en Caja | Victtorino", "Sifón lavamanos 1 1/4 con codo plástico en caja, listo para instalar. Despacho a todo Chile.", "Sifones y Desagües"),
    (1053, "Sifón De Lavamanos 1 1/2 x 1 1/4 Con Codo Plástico", "Accesorios", "sifón lavamanos 1 1/2 1 1/4", "Sifón Lavamanos 1 1/2 x 1 1/4 | Victtorino", "Sifón lavamanos 1 1/2 x 1 1/4 con codo plástico, adapta medidas. Despacho a todo Chile.", "Sifones y Desagües"),
    (1049, "Set Accesorios De Baños Notte 6 Piezas", "Accesorios", "set accesorios baño notte", "Set Accesorios Baño Notte 6 Piezas | Victtorino", "Set completo 6 accesorios baño modelo Notte, línea negra premium. Despacho a todo Chile.", "Accesorios"),
    (1022, "Sello Antifuga Mirage", "Accesorios", "sello antifuga mirage wc", "Sello Antifuga WC Mirage | Victtorino", "Sello antifuga Mirage para WC, evita filtraciones en la base. Despacho a todo Chile. Victtorino.", "WC e Inodoros"),
    (1019, "Portavaso Inoxidable", "Accesorios", "portavaso inoxidable baño", "Portavaso Inoxidable para Baño | Victtorino", "Portavaso baño acero inoxidable, resistente a humedad. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (985, "Porta Toalla Repisa 60 cm Al Muro", "Accesorios", "porta toalla repisa 60 cm", "Porta Toalla Repisa 60 cm al Muro | Victtorino", "Porta toalla con repisa 60 cm al muro, dos en uno. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (984, "Porta Rollo Linea Schwarz", "Accesorios", "porta rollo schwarz negro", "Porta Rollo Línea Schwarz Negro | Victtorino", "Porta rollo papel higiénico línea Schwarz negro, premium. Despacho a todo Chile. Victtorino.", "Accesorios"),
    (983, "Porta Rollo Inoxidable", "Accesorios", "porta rollo inoxidable", "Porta Rollo Acero Inoxidable | Victtorino", "Porta rollo papel higiénico acero inoxidable, resistente a humedad. Despacho a todo Chile.", "Accesorios"),
    (982, "Plato Ducha Redondo Abs 20 cm", "Accesorios", "plato ducha abs 20 cm", "Plato Ducha Redondo ABS 20 cm | Victtorino", "Plato ducha redondo ABS cromado 20 cm, ligero y económico. Despacho a todo Chile. Victtorino.", "Accesorios"),
    (944, "Papelero Pedal 8 L", "Accesorios", "papelero pedal 8 litros", "Papelero Pedal 8 Litros | Victtorino", "Papelero con pedal 8 litros para baño u oficina, acero inoxidable. Despacho a todo Chile.", "Accesorios"),
    (918, "Papelero Pedal 5 L", "Accesorios", "papelero pedal 5 litros", "Papelero Pedal 5 Litros | Victtorino", "Papelero con pedal 5 litros para baño, acero inoxidable plateado. Despacho a todo Chile.", "Accesorios"),
    (917, "Papelero Pedal 3 L", "Accesorios", "papelero pedal 3 litros", "Papelero Pedal 3 Litros | Victtorino", "Papelero con pedal 3 litros compacto para baño chico. Despacho a todo Chile. Victtorino.", "Accesorios"),
    (915, "Papelero Pedal 12 L", "Accesorios", "papelero pedal 12 litros", "Papelero Pedal 12 Litros | Victtorino", "Papelero con pedal 12 litros amplio, ideal para oficina. Despacho a todo Chile. Victtorino.", "Accesorios"),
    (911, "Monomando Lavamanos Domenica", "Griferia", "monomando lavamanos domenica", "Monomando Lavamanos Doménica | Victtorino", "Llave monomando lavamanos modelo Doménica, diseño clásico. Despacho a todo Chile. Victtorino.", "Griferia"),
    (910, "Monomando Lavamanos Alto Modern Pituto Cromado", "Griferia", "monomando lavamanos modern pituto", "Monomando Lavamanos Alto Modern Pituto | Victtorino", "Llave monomando lavamanos alto Modern Pituto cromado. Despacho a todo Chile. Victtorino.", "Griferia"),
    (871, "Llave Monomando Lavamanos Notte Negro", "Griferia", "llave monomando lavamanos notte negro", "Monomando Lavamanos Notte Negro | Victtorino", "Llave monomando lavamanos Notte negro premium. Despacho a todo Chile. Calidad Victtorino.", "Griferia"),
    (870, "Llave Monomando Lavamanos Modern Pituto Cromado", "Griferia", "monomando lavamanos modern cromado", "Monomando Lavamanos Modern Cromado | Victtorino", "Llave monomando lavamanos Modern Pituto cromado, contemporáneo. Despacho a todo Chile.", "Griferia"),
    (869, "Llave Monomando Lavamanos Modern Paleta Inoxidable", "Griferia", "monomando lavamanos modern paleta", "Monomando Lavamanos Modern Paleta Inox | Victtorino", "Llave monomando lavamanos Modern con paleta inoxidable. Despacho a todo Chile. Victtorino.", "Griferia"),
    (868, "Llave Monomando Lavamanos Colomba", "Griferia", "monomando lavamanos colomba", "Monomando Lavamanos Colomba | Victtorino", "Llave monomando lavamanos modelo Colomba, elegante y duradero. Despacho a todo Chile.", "Griferia"),
    (798, "Llave Lavadero Con Flange He He 1/2 3/4", "Griferia", "llave lavadero flange 1/2 3/4", "Llave Lavadero Flange 1/2 3/4 | Victtorino", "Llave lavadero con flange HE HE 1/2 3/4, instalación al muro. Despacho a todo Chile.", "Griferia"),
    (797, "Llave De Paso Para Gas Hi-He 1/2", "Griferia", "llave paso gas 1/2", "Llave de Paso para Gas 1/2 | Victtorino", "Llave de paso para gas HI-HE 1/2 pulgada, certificada. Despacho a todo Chile. Calidad Victtorino.", "Griferia"),
    (708, "Kit Wc De Anclaje Antifuga Con Admisión", "Accesorios", "kit anclaje wc antifuga", "Kit WC Anclaje Antifuga con Admisión | Victtorino", "Kit anclaje WC antifuga con admisión, instalación completa. Despacho a todo Chile. Victtorino.", "WC e Inodoros"),
    (706, "Kit Ducha Bidet Árabe", "Accesorios", "kit ducha bidet árabe", "Kit Ducha Bidet Árabe | Victtorino", "Kit ducha bidet árabe para higiene íntima, instalación simple. Despacho a todo Chile.", "Accesorios"),
    (705, "Jabonera Inoxidable", "Accesorios", "jabonera inoxidable baño", "Jabonera Acero Inoxidable | Victtorino", "Jabonera baño acero inoxidable resistente a humedad. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (704, "Jabonera Botella Linea Schwarz", "Accesorios", "jabonera botella schwarz", "Jabonera Botella Línea Schwarz | Victtorino", "Jabonera botella línea Schwarz negro premium. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (703, "Fluxómetro De Palanca Con Llave De Corte", "Accesorios", "fluxómetro palanca corte", "Fluxómetro Palanca con Llave de Corte | Victtorino", "Fluxómetro palanca con llave de corte para WC institucional. Despacho a todo Chile.", "WC e Inodoros"),
    (701, "Flexible Para Llave Profesional", "Accesorios", "flexible llave profesional", "Flexible Llave Profesional | Victtorino", "Flexible para llave profesional, conexión reforzada. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (700, "Flexible Ducha Swertt 160 cm", "Accesorios", "flexible ducha swertt 160 cm", "Flexible Ducha Swertt 160 cm | Victtorino", "Flexible ducha Swertt 160 cm, manguera reforzada cromada. Despacho a todo Chile.", "Accesorios"),
    (672, "Flexible De Gas Hi-Hi 3/8 x 1/2 De 60 cm", "Accesorios", "flexible gas 3/8 1/2 60 cm", "Flexible Gas 3/8 x 1/2 - 60 cm | Victtorino", "Flexible gas HI-HI 3/8 x 1/2 por 60 cm, certificado y seguro. Despacho a todo Chile.", "Accesorios"),
    (670, "Flexible De Ducha Extensible", "Accesorios", "flexible ducha extensible", "Flexible de Ducha Extensible | Victtorino", "Flexible de ducha extensible para alcance ajustable. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (669, "Flexible De Agua He-Hi 1/2 Con Llave Angular 35 cm", "Accesorios", "flexible agua llave angular 35 cm", "Flexible Agua HE-HI con Llave Angular 35 cm | Victtorino", "Flexible agua HE-HI 1/2 con llave angular 35 cm. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (635, "Espejo Circular 60 cm Led Triple Borde Metal", "Espejos", "espejo led circular 60 cm", "Espejo Circular LED 60 cm Borde Metal | Victtorino", "Espejo circular 60 cm LED triple con borde metal, baño moderno. Despacho a todo Chile.", "Espejos"),
    (599, "Dispensador De Jabón Simple Mejorado 360 ml", "Dispensador", "dispensador jabón 360 ml", "Dispensador Jabón Simple 360 ml | Victtorino", "Dispensador jabón simple mejorado 360 ml, instalación al muro. Despacho a todo Chile.", "Dispensador"),
    (576, "Desagüe Para Tina Con Tubo Y Campana De Rebalse 1 1/2", "Accesorios", "desagüe tina rebalse 1 1/2", "Desagüe Tina con Campana Rebalse 1 1/2 | Victtorino", "Desagüe tina con tubo y campana de rebalse 1 1/2. Despacho a todo Chile. Calidad Victtorino.", "Sifones y Desagües"),
    (575, "Desague Tina 1 1/2 Metalico", "Accesorios", "desagüe tina 1 1/2 metálico", "Desagüe Tina 1 1/2 Metálico | Victtorino", "Desagüe tina 1 1/2 metálico, conexión estándar. Despacho a todo Chile. Calidad Victtorino.", "Sifones y Desagües"),
    (574, "Desagüe Para Receptáculo Con Tapón De Goma 1 1/2", "Accesorios", "desagüe receptáculo tapón goma", "Desagüe Receptáculo Tapón Goma 1 1/2 | Victtorino", "Desagüe receptáculo con tapón de goma 1 1/2, sello hermético. Despacho a todo Chile.", "Sifones y Desagües"),
    (571, "Desagüe Click Clack Metal Con Rebalse 1 1/4 Para Lavamanos", "Accesorios", "desagüe click clack lavamanos", "Desagüe Click Clack Metal con Rebalse | Victtorino", "Desagüe click clack metal con rebalse 1 1/4 lavamanos. Despacho a todo Chile.", "Sifones y Desagües"),
    (570, "Desagüe Click Clack Metal 1 1/4 Cromado Lavamanos", "Accesorios", "desagüe click clack cromado", "Desagüe Click Clack Cromado Lavamanos | Victtorino", "Desagüe click clack metal cromado 1 1/4 lavamanos. Despacho a todo Chile. Victtorino.", "Sifones y Desagües"),
    (567, "Desagüe Con Sifón Corrugado", "Accesorios", "desagüe sifón corrugado", "Desagüe con Sifón Corrugado | Victtorino", "Desagüe con sifón corrugado flexible, adaptable a instalaciones. Despacho a todo Chile.", "Sifones y Desagües"),
    (566, "Combinación Lavaplato En V Vander", "Accesorios", "combinación lavaplato vander", "Combinación Lavaplato en V Vander | Victtorino", "Combinación lavaplato en V modelo Vander, doble salida. Despacho a todo Chile.", "Griferia"),
    (562, "Collarin Goma Wc", "Accesorios", "collarín goma wc", "Collarín de Goma para WC | Victtorino", "Collarín goma para WC, sello entre estanque y taza. Despacho a todo Chile. Calidad Victtorino.", "WC e Inodoros"),
    (561, "Codo Para Lavaplatos Con Salida A Sifón", "Accesorios", "codo lavaplatos salida sifón", "Codo Lavaplatos Salida a Sifón | Victtorino", "Codo lavaplatos con salida a sifón, conexión universal. Despacho a todo Chile. Victtorino.", "Sifones y Desagües"),
    (558, "Brazo Ducha Al Muro 40 cm", "Accesorios", "brazo ducha al muro 40 cm", "Brazo Ducha al Muro 40 cm | Victtorino", "Brazo ducha al muro 40 cm cromado, conexión estándar. Despacho a todo Chile. Calidad Victtorino.", "Accesorios"),
    (557, "Brazo Ducha Al Muro 30 cm", "Accesorios", "brazo ducha al muro 30 cm", "Brazo Ducha al Muro 30 cm | Victtorino", "Brazo ducha al muro 30 cm cromado, ideal para baños compactos. Despacho a todo Chile.", "Accesorios"),
    (555, "Botonera Dual Flush", "Accesorios", "botonera dual flush wc", "Botonera Dual Flush para WC | Victtorino", "Botonera dual flush WC para descarga doble, ahorro de agua. Despacho a todo Chile.", "WC e Inodoros"),
    (553, "Barra De Seguridad Recta 40 cm", "Accesorios", "barra seguridad recta 40 cm", "Barra Seguridad Recta 40 cm | Victtorino", "Barra seguridad recta 40 cm para ducha y baño, acero inox. Despacho a todo Chile.", "Agarraderas y Barras"),
    (550, "Barra De Seguridad 135° Con Jabonera", "Accesorios", "barra seguridad 135 con jabonera", "Barra Seguridad 135° con Jabonera | Victtorino", "Barra seguridad 135° con jabonera integrada, doble función. Despacho a todo Chile.", "Agarraderas y Barras"),
    (547, "Acople Lavaplatos Dobles", "Accesorios", "acople lavaplatos dobles", "Acople Lavaplatos Dobles | Victtorino", "Acople para lavaplatos dobles, conexión a sifón compartido. Despacho a todo Chile.", "Sifones y Desagües"),
]

assert len(DATOS) == 53, f"Esperaba 53, son {len(DATOS)}"

wb = Workbook()
ws = wb.active
ws.title = "Lote C — 53 productos"

HEADERS = [
    ("Decisión", 12), ("Woo ID", 8), ("Cat actual", 18), ("Producto", 42),
    ("Focus keyword", 30), ("Meta title (≤60)", 46),
    ("Meta description (≤155)", 65), ("Cat destino plantilla", 25),
    ("Notas / cambios", 25),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "B2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

cat_color = {
    "Shower/Mamparas/Receptaculos": "E7F0FA", "Griferia": "FCE7EF",
    "Lavaplatos": "FFF5E0", "Lavamanos": "E5F4E5", "Espejos": "F0E5F4",
    "Dispensador": "FFFBE5", "Accesorios": "F4F4F4",
    "Agarraderas y Barras": "FBE5E5", "Sifones y Desagües": "E5EBFB",
    "WC e Inodoros": "FBF1E5",
}

for r_idx, (pid, name, cat_actual, focus, mt, md, cat_destino) in enumerate(DATOS, start=2):
    fila = ["", pid, cat_actual, name, focus, mt, md, cat_destino, ""]
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 1:
            c.fill = PatternFill("solid", fgColor="FFFFFF")
        elif c_idx == 3:
            c.fill = PatternFill("solid", fgColor=cat_color.get(cat_actual, "FFFFFF"))
        elif c_idx == 8:
            c.fill = PatternFill("solid", fgColor=cat_color.get(cat_destino, "FFFFFF"))
    if len(mt) > 60:
        ws.cell(row=r_idx, column=6).fill = PatternFill("solid", fgColor="FFCCCC")
    if len(md) > 155:
        ws.cell(row=r_idx, column=7).fill = PatternFill("solid", fgColor="FFCCCC")
    ws.row_dimensions[r_idx].height = 50

dv = DataValidation(type="list", formula1='"OK,Editar,Saltar"', allow_blank=True)
dv.add(f"A2:A{len(DATOS) + 1}")
ws.add_data_validation(dv)

# Instrucciones diferenciadas
ws2 = wb.create_sheet("Instrucciones", 0)
inst = [
    ("Lote C — 53 productos con descripción muy corta (<100 palabras)", True),
    ("", False),
    ("Diferencia clave con A y B: además de focus + meta tags, voy a", False),
    ("REESCRIBIR la descripción HTML completa con plantilla premium", False),
    ("(~400 palabras, H2 con focus, bullets, enlaces internos).", False),
    ("", False),
    ("La plantilla se aplica según la columna 'Cat destino plantilla'.", True),
    ("", False),
    ("Ejemplo: el producto 1089 'Válvula Descarga WC Flapper' tiene", False),
    ("categoría actual 'Accesorios' pero su plantilla destino es 'WC e", False),
    ("Inodoros' — eso le da una descripción más específica al rubro.", False),
    ("", False),
    ("Esto NO cambia la categoría del producto, solo aplica el texto", False),
    ("de la plantilla correspondiente. Si quieres también mover la", False),
    ("categoría, escríbelo en Notas.", False),
    ("", False),
    ("Decisiones en columna A:", True),
    ("  OK     = aplicar tal como está", False),
    ("  Editar = cambiar focus/meta o cat destino (escribir en Notas)", False),
    ("  Saltar = no tocar este producto", False),
    ("  vacío  = OK por defecto", False),
    ("", False),
    ("Al terminar: guarda el Excel y dice 'aplica el Lote C'.", False),
]
for r, (t, b) in enumerate(inst, start=1):
    c = ws2.cell(row=r, column=1, value=t)
    if b:
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 110

out = r"C:\Users\dell\victtorino\propuesta_lote_c.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"Filas: {len(DATOS)}")
mt_over = sum(1 for d in DATOS if len(d[4]) > 60)
md_over = sum(1 for d in DATOS if len(d[5]) > 155)
print(f"Meta titles >60: {mt_over}, Meta descriptions >155: {md_over}")
