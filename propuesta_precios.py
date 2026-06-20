# -*- coding: utf-8 -*-
"""Propuesta de precios Falabella = Costo x 3.14 (margen 68%, comision 23%).
Recalcula los 157 costos Defontana, y para los que tienen precio actual
(pool envio gratis) marca el desvio vs la formula. Solo lectura."""
import re, unicodedata, difflib, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

SNAP=r"C:\Users\dell\victtorino\fala_editor_save.md"
COSTS=r"C:\Users\dell\Downloads\costos_defontana_limpio.xlsx"
OUT=r"C:\Users\dell\victtorino\propuesta_precios_falabella.xlsx"
MULT=3.14   # Falabella = Costo x 3.14

def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().upper()
    return re.sub(r'\s+',' ',re.sub(r'[^A-Z0-9 ]',' ',s)).strip()
def r990(x):
    x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

# precios actuales (pool)
price_by_norm={}
for ln in open(SNAP,encoding='utf-8').read().splitlines():
    m=re.search(r"row \"(.+?) SKU: (\d+) - EAS: - ([\d.]+) CLP\"",ln)
    if m: price_by_norm[norm(m.group(1))]={'sku':m.group(2),'price':int(m.group(3).replace('.',''))}
pn=list(price_by_norm.keys())

wb=openpyxl.load_workbook(COSTS,read_only=True); ws=wb.active
rows=[]
for r in ws.iter_rows(values_only=True):
    if not r or r[0] in (None,'Codigo') or not isinstance(r[2],(int,float)): continue
    cod,desc,cost=r[0],r[1],int(r[2])
    sug=r990(cost*MULT) if cost>0 else None
    # match current price by name
    n=norm(desc); cur=None;sku=''
    if n in price_by_norm: cur=price_by_norm[n]
    else:
        cand=difflib.get_close_matches(n,pn,n=1,cutoff=0.62)
        if cand: cur=price_by_norm[cand[0]]
    dev=None;flag=''
    if cur and sug:
        sku=cur['sku']; dev=(cur['price']-sug)/sug
        if cost==0: flag='COSTO 0'
        elif cur['price']<=cost: flag='VENTA A PERDIDA'
        elif dev>=0.5: flag='SOBREPRECIO (>50%)'
        elif dev>=0.15: flag='sobre formula'
        elif dev<=-0.15: flag='BAJO FORMULA (margen perdido)'
        else: flag='OK'
    rows.append({'cod':cod,'desc':desc,'cost':cost,'sug':sug,
                 'sku':sku,'cur':cur['price'] if cur else None,'dev':dev,'flag':flag})

withcur=[r for r in rows if r['cur'] is not None]
print(f"Costos: {len(rows)} | con precio actual cruzado: {len(withcur)}")
print('Diagnostico (de los cruzados):', Counter(r['flag'] for r in withcur))
print('\n--- DESVIOS FUERTES (precio actual vs formula Costo x3.14) ---')
for r in sorted(withcur,key=lambda x:-(abs(x['dev']) if x['dev'] is not None else 0)):
    if r['flag'] in('OK','sobre formula',''): continue
    print(f"  [{r['flag']:>28}] actual ${r['cur']:>8,}  formula ${r['sug']:>8,}  ({r['dev']*100:+5.0f}%)  costo ${r['cost']:>6,}  {r['desc'][:34]}")

wbo=openpyxl.Workbook(); w=wbo.active; w.title='Precios Falabella x3.14'
w.append(['Codigo','SKU Fala','Descripcion','Costo','Precio actual','Precio FORMULA (x3.14)','Desvio %','Diagnostico'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center')
fills={'VENTA A PERDIDA':'FFC7CE','SOBREPRECIO (>50%)':'FFC7CE','BAJO FORMULA (margen perdido)':'FFEB9C','COSTO 0':'D9D9D9','sobre formula':'FFF2CC','OK':'C6EFCE','':'FFFFFF'}
for r in sorted(rows,key=lambda x:(x['cur'] is None, -(abs(x['dev']) if x['dev'] is not None else -1))):
    w.append([r['cod'],r['sku'],r['desc'],r['cost'],r['cur'],r['sug'],
              round(r['dev'],3) if r['dev'] is not None else None,r['flag']])
    for c in w[w.max_row]: c.fill=PatternFill('solid',fgColor=fills.get(r['flag'],'FFFFFF'))
for col,wd in zip('ABCDEFGH',[12,11,44,9,13,18,9,30]): w.column_dimensions[col].width=wd
for rr in range(2,w.max_row+1): w.cell(rr,7).number_format='0.0%'
wbo.save(OUT); print('\nGuardado:',OUT)
