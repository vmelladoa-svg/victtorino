"""
Excel de reclasificacion para los 9 productos importados desde duda_revision.
Mismo formato que reclasificar_borradores.xlsx pero para los IDs Woo: 3084, 3088,
3095, 3099, 3111, 3117, 3123, 3126, 3129.
"""
import sys
import io
import warnings
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

IDS = [3084, 3088, 3095, 3099, 3111, 3117, 3123, 3126, 3129]
CATEGORIAS_VALIDAS = ["Accesorios", "Agarraderas y Barras", "Dispensador", "Espejos",
                     "Griferia", "Lavamanos", "Lavaplatos", "Shower/Mamparas/Receptaculos",
                     "Sifones y Desagües", "WC e Inodoros"]

print("Consultando productos...")
productos = []
for pid in IDS:
    r = requests.get(f"{WC}/wp-json/wc/v3/products/{pid}", params=P, timeout=30)
    if r.status_code != 200:
        continue
    p = r.json()
    productos.append({
        "woo_id": p["id"],
        "name": p["name"],
        "cat_actual": ", ".join(c["name"] for c in p.get("categories", [])),
        "permalink": p.get("permalink"),
        "edit_url": f"{WC}/wp-admin/post.php?post={p['id']}&action=edit",
        "thumbnail": p["images"][0]["src"] if p.get("images") else "",
        "precio": p.get("regular_price"),
        "stock": p.get("stock_quantity"),
    })

# Ordenar por cat actual + nombre
productos.sort(key=lambda x: (x["cat_actual"], x["name"]))

wb = Workbook()
ws = wb.active
ws.title = "Reclasificar 9"

HEADERS = [
    ("Categoría nueva", 30),
    ("Woo ID", 9),
    ("Título", 60),
    ("Categoría actual", 22),
    ("Precio", 10),
    ("Stock", 7),
    ("Foto", 8),
    ("Ver", 8),
    ("Editar", 9),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

for r_idx, p in enumerate(productos, start=2):
    fila = ["", p["woo_id"], p["name"], p["cat_actual"], p["precio"], p["stock"], "Foto", "Ver", "Editar"]
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if c_idx == 5:
            c.number_format = "#,##0"
    if p["thumbnail"]:
        ws.cell(row=r_idx, column=7).hyperlink = p["thumbnail"]
        ws.cell(row=r_idx, column=7).font = Font(color="0563C1", underline="single")
    if p["permalink"]:
        ws.cell(row=r_idx, column=8).hyperlink = p["permalink"]
        ws.cell(row=r_idx, column=8).font = Font(color="0563C1", underline="single")
    ws.cell(row=r_idx, column=9).hyperlink = p["edit_url"]
    ws.cell(row=r_idx, column=9).font = Font(color="0563C1", underline="single")
    ws.row_dimensions[r_idx].height = 35

dv = DataValidation(type="list", formula1='"' + ",".join(CATEGORIAS_VALIDAS) + '"', allow_blank=True)
dv.add(f"A2:A{len(productos) + 1}")
ws.add_data_validation(dv)

# Instrucciones simples
ws2 = wb.create_sheet("Instrucciones", 0)
inst = [
    "Cómo reclasificar:",
    "",
    "- Si la categoría actual está bien -> deja la columna A en blanco.",
    "- Si quieres cambiarla -> elige del dropdown en columna A.",
    "- Si quieres MÚLTIPLES categorías para un producto -> escribe en la columna D",
    "  separadas por '-' (ej: 'Griferia - Accesorios'). Yo las leo igual.",
    "",
    "Cuando termines: guarda y dime 'ya'.",
]
for r, t in enumerate(inst, start=1):
    c = ws2.cell(row=r, column=1, value=t)
    if r == 1 or t.endswith(":"):
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 100

out = r"C:\Users\dell\victtorino\reclasificar_9_duda.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"\nDistribucion actual:")
from collections import Counter
for cat, n in Counter(p["cat_actual"] for p in productos).most_common():
    print(f"  {n}  {cat}")
