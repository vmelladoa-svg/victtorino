"""
Genera Excel para reclasificar los 44 productos draft del lote NO.
Una fila por producto, dropdown con las 10 categorias validas.
"""
import json
import sys
import io
import warnings
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

# 1) Leer Excel para tener la lista de los 44 NO
print("[1/3] Leyendo lista NO del Excel revisado...")
wb_in = load_workbook(r"C:\Users\dell\victtorino\dudosos_revision.xlsx", data_only=True)
ws_in = wb_in["Dudosos ML vs Web"]
ml_ids_no = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    dec = (row[0] or "").strip().upper() if row[0] else ""
    if dec == "NO":
        ml_ids_no.append(row[2])
print(f"      {len(ml_ids_no)} ml_ids NO")

# 2) Para cada uno, ubicar el producto en Woo por SKU
print("[2/3] Consultando Woo por SKU...")
productos = []
for mid in ml_ids_no:
    sku = f"ML-{mid}"
    r = requests.get(f"{WC}/wp-json/wc/v3/products",
                     params={**P, "sku": sku}, timeout=30)
    d = r.json()
    if not d:
        print(f"      {mid} -> no encontrado en Woo")
        continue
    p = d[0]
    productos.append({
        "ml_id": mid,
        "woo_id": p["id"],
        "name": p["name"],
        "cat_actual": ", ".join(c["name"] for c in p.get("categories", [])),
        "permalink": p.get("permalink"),
        "edit_url": f"{WC}/wp-admin/post.php?post={p['id']}&action=edit",
        "thumbnail": p["images"][0]["src"] if p.get("images") else "",
        "status": p.get("status"),
        "stock": p.get("stock_quantity"),
        "precio": p.get("regular_price"),
    })
print(f"      OK, {len(productos)} productos cargados")

# 3) Generar Excel
print("[3/3] Generando Excel...")
CATEGORIAS_VALIDAS = [
    "Accesorios",
    "Agarraderas y Barras",
    "Dispensador",
    "Espejos",
    "Griferia",
    "Lavamanos",
    "Lavaplatos",
    "Shower/Mamparas/Receptaculos",
    "Sifones y Desagües",
    "WC e Inodoros",
]

wb = Workbook()
ws = wb.active
ws.title = "Reclasificar borradores"

HEADERS = [
    ("Categoría nueva", 28),  # dropdown
    ("Woo ID", 9),
    ("Título", 60),
    ("Categoría actual", 22),
    ("Precio", 10),
    ("Stock", 7),
    ("Foto", 8),
    ("Ver producto", 14),
    ("Editar", 9),
]
for col, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

borde = Border(left=Side(style="thin", color="DDDDDD"),
               right=Side(style="thin", color="DDDDDD"),
               top=Side(style="thin", color="DDDDDD"),
               bottom=Side(style="thin", color="DDDDDD"))

# Ordenar por categoría actual para que sea más fácil revisar en grupos
productos.sort(key=lambda x: (x["cat_actual"], x["name"]))

for r_idx, p in enumerate(productos, start=2):
    fila = [
        "",  # Categoría nueva (dropdown vacío)
        p["woo_id"],
        p["name"],
        p["cat_actual"],
        p["precio"],
        p["stock"],
        "Foto",
        "Ver",
        "Editar",
    ]
    for c_idx, val in enumerate(fila, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = borde
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if c_idx == 5:
            c.number_format = "#,##0"
    if p["thumbnail"]:
        ws.cell(row=r_idx, column=7).hyperlink = p["thumbnail"]
        ws.cell(row=r_idx, column=7).font = Font(color="0563C1", underline="single")
    if p["permalink"]:
        ws.cell(row=r_idx, column=8).hyperlink = p["permalink"]
        ws.cell(row=r_idx, column=8).font = Font(color="0563C1", underline="single")
    ws.cell(row=r_idx, column=9).hyperlink = p["edit_url"]
    ws.cell(row=r_idx, column=9).font = Font(color="0563C1", underline="single")
    ws.row_dimensions[r_idx].height = 35

# Dropdown en columna A con las 10 categorías válidas
formula = '"' + ",".join(CATEGORIAS_VALIDAS) + '"'
dv = DataValidation(type="list", formula1=formula, allow_blank=True)
dv.add(f"A2:A{len(productos) + 1}")
ws.add_data_validation(dv)

# Hoja de instrucciones
ws2 = wb.create_sheet("Instrucciones", 0)
instrucciones = [
    ("Cómo reclasificar los borradores", True),
    ("", False),
    ("Mira cada producto, decide a qué categoría debería pertenecer, y elige del dropdown en la columna A.", False),
    ("", False),
    ("Categorías disponibles:", True),
    *[(f"  - {c}", False) for c in CATEGORIAS_VALIDAS],
    ("", False),
    ("Tips:", True),
    ("", False),
    ("  - Si la categoría actual está bien, déjala en blanco. Solo marca las que quieras cambiar.", False),
    ("  - Click en 'Foto' abre la imagen del producto.", False),
    ("  - Click en 'Ver' abre la ficha del producto en victtorino.cl (incluso siendo borrador).", False),
    ("  - Click en 'Editar' abre el producto en el admin de WordPress.", False),
    ("", False),
    ("Cuando termines:", True),
    ("", False),
    ("  Guarda el archivo (Ctrl+S) y dile a Claude: \"ya reclasifiqué\".", False),
    ("  Claude actualizará la categoría solo en los productos que marcaste.", False),
    ("", False),
    (f"Total filas: {len(productos)}", True),
]
for r, (txt, bold) in enumerate(instrucciones, start=1):
    c = ws2.cell(row=r, column=1, value=txt)
    if bold:
        c.font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 110

out_path = r"C:\Users\dell\victtorino\reclasificar_borradores.xlsx"
wb.save(out_path)
print(f"\nGuardado: {out_path}")
print(f"Filas a revisar: {len(productos)}")

# Distribución actual
from collections import Counter
print("\nDistribución actual por categoría:")
for cat, n in Counter(p["cat_actual"] for p in productos).most_common():
    print(f"  {n:3}  {cat}")
