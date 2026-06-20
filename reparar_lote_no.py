"""
Recupera los 8 productos que fallaron en importar_dudosos_no.py:
- 6 estan creados pero algunos sin categoria/imagenes (cortes de Apache 503).
- 2 faltan totalmente.

Estrategia:
- Lee Excel, ubica los 8 ml_ids.
- Para cada uno, hace un POST si no existe, o un PUT si existe pero le falta categoria/imagenes.
- Espera 5s entre requests para evitar 503.
"""
import json
import sys
import io
import time
import re
import unicodedata
import requests
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore", category=UserWarning)
from seo_premium_lote import plantilla_premium, meta_desc_premium, meta_title_premium
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WC = "https://victtorino.cl"
P = {"consumer_key": "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15",
     "consumer_secret": "cs_3604e7ebdb8ff78442731344cc95af50516188a5"}

with open(r"C:\Users\dell\victtorino\tokens_cuenta3.json", encoding="utf-8") as f:
    tk = json.load(f)
ML_HEAD = {"Authorization": f"Bearer {tk['access_token']}"}

PENDIENTES = [
    "MLC1293267551", "MLC1404800753", "MLC1293224426", "MLC3724919200",
    "MLC1293249558", "MLC3994366480", "MLC3955887032", "MLC3969900116",
]

MAPEO = {
    "Lavaplatos de Cocina":          ("Lavaplatos",                    "lavaplatos cocina"),
    "Griferías Convencionales":      ("Griferia",                      "llave grifería"),
    "Sifones de Desague":            ("Sifones y Desagües",            "sifón desagüe"),
    "Agarraderas":                   ("Agarraderas y Barras",          "agarradera baño"),
    "De Papel":                      ("Accesorios",                    "dispensador papel higiénico"),
    "Otros":                         ("Accesorios",                    "accesorios baño"),
    "De Jabón y Alcohol en Gel":     ("Dispensador",                   "dispensador de jabón"),
    "Sets de Baño":                  ("Accesorios",                    "set accesorios baño"),
    "Mamparas y Cabinas":            ("Shower/Mamparas/Receptaculos",  "mampara ducha"),
    "Duchas Higiénicas":             ("Griferia",                      "ducha higiénica"),
    "Mangos para Ducha":             ("Griferia",                      "ducha teléfono"),
    "Válvulas de Descarga para WC":  ("WC e Inodoros",                 "válvula descarga WC"),
    "Mangueras":                     ("Accesorios",                    "manguera baño"),
    "Papeles Higiénicos":            ("Accesorios",                    "dispensador papel higiénico"),
}


def slugify(s):
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:80]


def woo_request(method, path, **kwargs):
    url = f"{WC}/wp-json/wc/v3{path}"
    kwargs.setdefault("timeout", 180)
    kwargs.setdefault("params", {}).update(P)
    for n in range(1, 5):
        try:
            r = requests.request(method, url, **kwargs)
            if r.status_code >= 400:
                print(f"    HTTP {r.status_code}: {r.text[:200]}")
                if r.status_code == 503:
                    espera = 10 * n
                    print(f"    503 -> espero {espera}s y reintento")
                    time.sleep(espera)
                    continue
                return None
            return r.json()
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            espera = 5 * n
            print(f"    RED {type(e).__name__}; espero {espera}s")
            time.sleep(espera)
    return None


# Leer Excel para obtener ml_cat de cada pendiente
print("[1/3] Leyendo Excel para obtener cat_ml de los 8 pendientes...")
wb = load_workbook(r"C:\Users\dell\victtorino\dudosos_revision.xlsx", data_only=True)
ws = wb["Dudosos ML vs Web"]
info_excel = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2] in PENDIENTES:
        info_excel[row[2]] = {"titulo": row[3], "ml_cat": row[6]}
print(f"      OK, encontrados {len(info_excel)} en Excel")

# Cargar IDs cats Woo
cats_woo = woo_request("GET", "/products/categories", params={"per_page": 100})
cat_id_por_nombre = {c["name"]: c["id"] for c in cats_woo}

# Procesar cada pendiente
print(f"\n[2/3] Procesando 8 pendientes con pausa de 5s entre cada uno...\n")
resultados = []

for idx, mid in enumerate(PENDIENTES, start=1):
    print(f"({idx}/8) {mid}")
    sku = f"ML-{mid}"
    info = info_excel.get(mid, {})
    ml_cat = info.get("ml_cat", "Otros")
    mapeo = MAPEO.get(ml_cat, ("Accesorios", "accesorios baño"))
    cat_destino, focus = mapeo

    # Trae detalle ML
    item = requests.get(f"https://api.mercadolibre.com/items/{mid}", headers=ML_HEAD, timeout=30).json()
    titulo = item.get("title", info.get("titulo", "")) or info.get("titulo", "")
    precio = item.get("price")
    stock = int(item.get("available_quantity") or 0)
    pictures = [p["url"] for p in item.get("pictures", []) if p.get("url")]

    # Construir SEO premium
    desc_html = plantilla_premium(cat_destino, titulo, focus)
    meta_t = meta_title_premium(titulo, focus)
    meta_d = meta_desc_premium(titulo, cat_destino, focus)
    short_d = (f"<p>{titulo}. <strong>{focus.capitalize()}</strong> con diseño "
               "contemporáneo, materiales resistentes y despacho a todo Chile. "
               "Renueva tu baño con piezas que duran y se ven bien.</p>")
    images_payload = [{"src": u, "alt": (titulo if i == 0 else f"{titulo} - foto {i+1}")[:120]}
                      for i, u in enumerate(pictures)]

    body_completo = {
        "name": titulo,
        "slug": slugify(titulo),
        "type": "simple",
        "status": "draft",
        "regular_price": str(int(precio)) if precio is not None else "",
        "manage_stock": True,
        "stock_quantity": stock,
        "description": desc_html,
        "short_description": short_d,
        "sku": sku,
        "categories": [{"id": cat_id_por_nombre[cat_destino]}],
        "images": images_payload,
        "meta_data": [
            {"key": "rank_math_title", "value": meta_t},
            {"key": "rank_math_description", "value": meta_d},
            {"key": "rank_math_focus_keyword", "value": focus},
        ],
    }

    # Existe?
    existente = woo_request("GET", "/products", params={"sku": sku})
    if existente:
        pid = existente[0]["id"]
        sin_cat = not existente[0].get("categories")
        sin_img = not existente[0].get("images")
        print(f"      Ya existe (Woo {pid}, sin_cat={sin_cat}, sin_img={sin_img})")
        if sin_cat or sin_img:
            print(f"      PUT para completar categoria/imagenes/SEO...")
            # PUT con todo (no incluyo slug en el PUT para evitar conflicto si ya existe)
            body_put = {k: v for k, v in body_completo.items() if k != "slug"}
            res = woo_request("PUT", f"/products/{pid}", json=body_put)
            if res:
                print(f"      OK Woo {pid} categorias={[c['name'] for c in res.get('categories',[])]} images={len(res.get('images',[]))}")
                resultados.append({"ml_id": mid, "woo_id": pid, "accion": "completado", "permalink": res.get("permalink")})
            else:
                resultados.append({"ml_id": mid, "woo_id": pid, "accion": "completado_fallo"})
        else:
            print(f"      Ya estaba completo. Skip.")
            resultados.append({"ml_id": mid, "woo_id": pid, "accion": "ya_completo"})
    else:
        # Crear desde cero
        print(f"      No existe -> POST...")
        nuevo = woo_request("POST", "/products", json=body_completo)
        if nuevo:
            print(f"      OK Woo {nuevo['id']} cats={[c['name'] for c in nuevo.get('categories',[])]} imgs={len(nuevo.get('images',[]))}")
            resultados.append({"ml_id": mid, "woo_id": nuevo["id"], "accion": "creado",
                               "permalink": nuevo.get("permalink")})
        else:
            print(f"      FALLO POST")
            resultados.append({"ml_id": mid, "accion": "fallo"})

    # Pausa entre items para evitar saturar Apache
    if idx < len(PENDIENTES):
        time.sleep(6)

print("\n[3/3] RESUMEN:")
from collections import Counter
for accion, n in Counter(r["accion"] for r in resultados).most_common():
    print(f"  {n:2}  {accion}")

with open(r"C:\Users\dell\victtorino\reparar_lote_resultado.json", "w", encoding="utf-8") as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)
print("\nDetalle -> reparar_lote_resultado.json")
