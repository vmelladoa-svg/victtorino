# -*- coding: utf-8 -*-
"""RECUPERACION: lee la base de monitoreo y genera el archivo de carga masiva
Falabella con TODOS los productos que tienen oferta. Si Falabella apaga/cambia
los specials, se corre esto y se sube el archivo para restaurar todo de una.
Uso: editar precios en la base si hace falta, correr, y subir SellerPriceTemplate_UPLOAD_nuevo.xlsx
"""
import openpyxl

BASE=r'C:\Users\dell\Downloads\Falabella_BASE_monitoreo.xlsx'
SALE_START='2026-06-16 00:00:01'; SALE_END='2026-07-16 23:59:59'
def r990(x): x=int(round(x)); return (x//1000)*1000+990 if x>=1000 else max(0,x)

ws=openpyxl.load_workbook(BASE,data_only=True).active
rows=list(ws.iter_rows(values_only=True)); H=rows[0]
iS=H.index('SKU Seller'); iF=H.index('SKU Falabella'); iN=H.index('Producto')
iNorm=H.index('Precio NORMAL'); iOf=H.index('PRECIO OFERTA')

out=openpyxl.Workbook(); w=out.active; w.title='Sheet'
w.append(['SellerSku','ShopSku','PriceFalabella','SalePriceFalabella','SaleStartDateFalabella','SaleEndDateFalabella','Name'])
n=0
for r in rows[1:]:
    sku=r[iS]; shop=r[iF]; venta=r[iOf]; normal=r[iNorm]
    if not sku or not shop or not isinstance(venta,(int,float)): continue
    normal = normal if isinstance(normal,(int,float)) else r990(venta/0.5)
    w.append([sku, shop, float(normal), float(venta), SALE_START, SALE_END, r[iN]])
    n+=1
out.save('SellerPriceTemplate_UPLOAD_nuevo.xlsx')
print('Archivo de recuperacion generado con',n,'productos -> SellerPriceTemplate_UPLOAD_nuevo.xlsx')
print('Subir en: Seller Center > Carga Masiva > paso 2 > Actualizar productos')
