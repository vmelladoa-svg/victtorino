"""
Compara baseline vs snapshot fresco. Genera tabla de Δ por item y por cohorte.

Uso:
  1. Re-correr: python auditoria_ml.py + python fix_visitas_snapshots.py + python analisis_auditoria_ml.py
  2. Después: python seguimiento_comparar.py
       → genera seguimiento_diff_<fecha>.xlsx

Si baseline_path o analisis_path no se pasan, usa los defaults.
"""
import json
import pickle
import sys
from datetime import datetime, date
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).parent
BASELINE = ROOT / "data" / "auditoria" / "seguimiento_baseline_2026-05-23.json"
ANALISIS = ROOT / "data" / "auditoria" / "analisis.pkl"


def cohort_summary(baseline_items, current_df, ids, cohorte_name):
    """Calcula deltas agregados por cohorte."""
    rows = []
    for iid in ids:
        b = baseline_items.get(iid)
        if not b: continue
        cur_row = current_df[current_df["ItemID"] == iid]
        if cur_row.empty:
            rows.append({**b, "estado_actual": "NOT_FOUND"})
            continue
        c = cur_row.iloc[0]
        rows.append({
            "item_id": iid,
            "cuenta": b["cuenta"],
            "titulo": b["titulo"][:50] if b["titulo"] else "",
            "listing_pre": b["listing_baseline"],
            "listing_post": c["ListingType"],
            "precio_pre": b["precio_baseline"],
            "precio_post": int(c["Precio"]),
            "stock_pre": b["stock_baseline"],
            "stock_post": int(c["Stock"]),
            "visitas_pre": b["visitas_30d_baseline"],
            "visitas_post": int(c["Visitas30d"]),
            "Δ_visitas": int(c["Visitas30d"]) - b["visitas_30d_baseline"],
            "vendidos_pre": b["vendidos_180d_baseline"],
            "vendidos_post": int(c["Vendidos180d"]),
            "Δ_vendidos": int(c["Vendidos180d"]) - b["vendidos_180d_baseline"],
            "revenue_pre": b["revenue_180d_baseline"],
            "revenue_post": int(c["Revenue180d"]),
            "Δ_revenue": int(c["Revenue180d"]) - b["revenue_180d_baseline"],
            "health_pre": b["health_calc_baseline"],
            "health_post": int(c["HealthCalc"]),
            "Δ_health": int(c["HealthCalc"]) - b["health_calc_baseline"],
            "cohorte": cohorte_name,
        })
    return rows


def main():
    if not BASELINE.exists():
        print(f"ERROR: no se encontró baseline en {BASELINE}")
        print("Ejecutá primero: python seguimiento_baseline.py")
        return
    baseline = json.loads(BASELINE.read_text(encoding="utf-8"))
    if not ANALISIS.exists():
        print(f"ERROR: no se encontró análisis en {ANALISIS}")
        print("Ejecutá primero: python auditoria_ml.py && python fix_visitas_snapshots.py && python analisis_auditoria_ml.py")
        return

    data = pickle.loads(ANALISIS.read_bytes())
    df_current = data["df_publicaciones"]
    baseline_items = baseline["items"]
    cohortes = baseline["cohortes"]

    print(f"=== Comparación pre vs post ===")
    print(f"Baseline: {baseline['fecha_baseline']}")
    print(f"Hoy:      {date.today().isoformat()}")
    days = (date.today() - date.fromisoformat(baseline['fecha_baseline'])).days
    print(f"Días transcurridos: {days}\n")

    # Por cohorte
    all_diff = []
    cohort_aggs = []
    for cohorte, ids in cohortes.items():
        rows = cohort_summary(baseline_items, df_current, ids, cohorte)
        if not rows: continue
        all_diff.extend(rows)
        # Agregados
        valid = [r for r in rows if isinstance(r.get("Δ_visitas"), int)]
        if not valid: continue
        total_visitas_pre = sum(r["visitas_pre"] for r in valid)
        total_visitas_post = sum(r["visitas_post"] for r in valid)
        total_vendidos_pre = sum(r["vendidos_pre"] for r in valid)
        total_vendidos_post = sum(r["vendidos_post"] for r in valid)
        total_revenue_pre = sum(r["revenue_pre"] for r in valid)
        total_revenue_post = sum(r["revenue_post"] for r in valid)
        avg_health_pre = sum(r["health_pre"] for r in valid) / len(valid)
        avg_health_post = sum(r["health_post"] for r in valid) / len(valid)
        cohort_aggs.append({
            "Cohorte": cohorte,
            "Items": len(valid),
            "Visitas 30d pre": total_visitas_pre,
            "Visitas 30d post": total_visitas_post,
            "Δ Visitas": total_visitas_post - total_visitas_pre,
            "Δ Visitas %": round((total_visitas_post - total_visitas_pre) / max(total_visitas_pre, 1) * 100, 1),
            "Vendidos 180d pre": total_vendidos_pre,
            "Vendidos 180d post": total_vendidos_post,
            "Δ Vendidos": total_vendidos_post - total_vendidos_pre,
            "Revenue 180d pre": total_revenue_pre,
            "Revenue 180d post": total_revenue_post,
            "Δ Revenue": total_revenue_post - total_revenue_pre,
            "Health avg pre": round(avg_health_pre, 1),
            "Health avg post": round(avg_health_post, 1),
            "Δ Health": round(avg_health_post - avg_health_pre, 1),
        })

    print("=== Agregados por cohorte ===")
    df_agg = pd.DataFrame(cohort_aggs)
    if not df_agg.empty:
        print(df_agg.to_string(index=False))

    # Output Excel
    out_xlsx = ROOT / f"seguimiento_diff_{date.today().isoformat()}.xlsx"
    df_detalle = pd.DataFrame(all_diff)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_agg.to_excel(writer, sheet_name="Resumen por cohorte", index=False)
        df_detalle.to_excel(writer, sheet_name="Detalle por item", index=False)
        # Hojas por cohorte para análisis
        for cohorte in cohortes:
            sub = df_detalle[df_detalle["cohorte"] == cohorte]
            if not sub.empty:
                sub.to_excel(writer, sheet_name=cohorte[:30], index=False)

    # Formato
    wb = load_workbook(out_xlsx)
    HDR = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = HDR; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        widths = {}
        for row in ws.iter_rows():
            for c in row:
                if hasattr(c, "column_letter") and c.value is not None:
                    widths[c.column_letter] = max(widths.get(c.column_letter, 0), min(len(str(c.value)), 50))
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = min(w + 2, 50)
    wb.save(out_xlsx)
    print(f"\nOK Excel: {out_xlsx}")


if __name__ == "__main__":
    main()
