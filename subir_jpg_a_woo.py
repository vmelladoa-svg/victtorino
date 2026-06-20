"""
subir_jpg_a_woo.py
------------------
Sube las 5 JPG de paris_imagenes/MLC32065874/ al producto WooCommerce
#799 (Llave Lavadero Lavadora Doble 3/4) en victtorino.cl, sin perder las
imagenes existentes, y actualiza el Excel paris_productos_carga_masiva.xlsx
con las URLs finales (https://victtorino.cl/wp-content/uploads/...) en la
fila del SKU MLC32065874.

Flujo:
  1. Subir cada JPG a catbox.moe (host efimero publico) para obtener URL.
  2. GET /wp-json/wc/v3/products/799 para listar imagenes actuales.
  3. PUT al mismo producto con el array `images` ampliado:
       - imagenes existentes (referenciadas por id para no re-descargarlas)
       - nuevas imagenes con src=URL catbox
     WC descarga cada src y crea attachments en /wp-content/uploads/.
  4. Leer la respuesta de WC, identificar los nuevos attachments (los que
     no tenian id previamente) y obtener su src final en victtorino.cl.
  5. Actualizar el Excel: agregar esas URLs a la celda "Imagenes" del SKU.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "paris_productos_carga_masiva.xlsx"
IMG_DIR = BASE_DIR / "paris_imagenes" / "MLC32065874"
SKU = "MLC32065874"
WOO_PRODUCT_ID = 799
DATA_SHEET = "herramientas"
COL_SKU = 2
COL_IMAGENES = 13

WC_BASE = "https://victtorino.cl"
WC_KEY = "ck_0c13f81e932be64b0a2c8ba340db4717ef6b5c15"
WC_SEC = "cs_3604e7ebdb8ff78442731344cc95af50516188a5"


def upload_to_catbox(path: Path) -> str:
    with open(path, "rb") as f:
        data = f.read()
    r = requests.post(
        "https://catbox.moe/user/api.php",
        files={"fileToUpload": (path.name, data, "image/jpeg")},
        data={"reqtype": "fileupload"},
        timeout=120,
    )
    r.raise_for_status()
    url = r.text.strip()
    if not url.startswith("https://"):
        raise RuntimeError(f"Catbox respondio: {url[:200]}")
    return url


def find_row_by_sku(ws, sku: str) -> int | None:
    for row in range(7, ws.max_row + 2):
        if ws.cell(row=row, column=COL_SKU).value == sku:
            return row
    return None


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

    jpgs = sorted(IMG_DIR.glob("*.jpg"))
    if not jpgs:
        print(f"[ERROR] No hay JPGs en {IMG_DIR}")
        return 1

    print(f"[1/5] Subiendo {len(jpgs)} JPGs a catbox.moe...")
    catbox_urls: list[str] = []
    for p in jpgs:
        url = upload_to_catbox(p)
        catbox_urls.append(url)
        print(f"      {p.name} -> {url}")

    print(f"[2/5] Leyendo estado actual de producto Woo #{WOO_PRODUCT_ID}")
    r = requests.get(
        f"{WC_BASE}/wp-json/wc/v3/products/{WOO_PRODUCT_ID}",
        auth=(WC_KEY, WC_SEC),
        timeout=30,
    )
    r.raise_for_status()
    product = r.json()
    existing = product.get("images") or []
    existing_ids = {img.get("id") for img in existing}
    print(f"      Imagenes existentes: {len(existing)} (IDs: {sorted(existing_ids)})")

    # Construir nuevo array preservando existentes y agregando catbox
    new_images = []
    seen = set()
    for img in existing:
        iid = img.get("id")
        if iid in seen:
            continue
        seen.add(iid)
        new_images.append({"id": iid})
    for url in catbox_urls:
        new_images.append({"src": url})

    print(f"[3/5] PUT producto con {len(new_images)} imagenes "
          f"({len(existing_ids)} existentes + {len(catbox_urls)} nuevas)")
    r2 = requests.put(
        f"{WC_BASE}/wp-json/wc/v3/products/{WOO_PRODUCT_ID}",
        auth=(WC_KEY, WC_SEC),
        json={"images": new_images},
        timeout=120,
    )
    if r2.status_code >= 400:
        print(f"[ERROR] WC PUT fallo: {r2.status_code} {r2.text[:400]}")
        return 2
    updated = r2.json()
    final_images = updated.get("images") or []
    print(f"      Producto actualizado, ahora tiene {len(final_images)} imagenes:")
    for img in final_images:
        print(f"        id={img.get('id')} src={img.get('src')}")

    # Identificar las nuevas (id no estaba antes)
    nuevos_attachments = [img for img in final_images if img.get("id") not in existing_ids]
    nuevos_urls = [img.get("src") for img in nuevos_attachments if img.get("src")]
    print(f"[4/5] Nuevos attachments creados: {len(nuevos_urls)}")
    for u in nuevos_urls:
        print(f"        {u}")

    if not nuevos_urls:
        print("[WARN] WC no creo attachments nuevos. No actualizo Excel.")
        return 3

    print(f"[5/5] Actualizando Excel con las URLs nuevas en fila de SKU={SKU}")
    wb = load_workbook(EXCEL_FILE)
    ws = wb[DATA_SHEET]
    row = find_row_by_sku(ws, SKU)
    if row is None:
        print(f"[ERROR] No se encontro fila para SKU {SKU}")
        return 4
    current = (ws.cell(row=row, column=COL_IMAGENES).value or "").strip()
    existing_in_cell = [u.strip() for u in current.split(",") if u.strip()] if current else []
    merged = list(existing_in_cell)
    for u in nuevos_urls:
        if u not in merged:
            merged.append(u)
    ws.cell(row=row, column=COL_IMAGENES, value=", ".join(merged))
    wb.save(EXCEL_FILE)
    print(f"      Fila {row}: {len(merged)} imagenes")

    print()
    print("=" * 60)
    print("OK")
    print("=" * 60)
    print(f"Producto Woo #{WOO_PRODUCT_ID}: {len(existing)} -> {len(final_images)} imagenes")
    print(f"Excel SKU {SKU}: {len(existing_in_cell)} -> {len(merged)} imagenes")
    print(f"Archivo: {EXCEL_FILE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
