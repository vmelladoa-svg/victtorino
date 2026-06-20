# -*- coding: utf-8 -*-
"""Tablero de rentabilidad Falabella: cruza precios del catalogo (snapshot picker
Envio gratis) con costos Defontana por NOMBRE, calcula margen y recomienda
palancas (Ads / Envio gratis / CMR) segun la regla del piso de margen."""
import re, unicodedata, difflib, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SNAP = r"C:\Users\dell\victtorino\fala_editor_save.md"
COSTS = r"C:\Users\dell\Downloads\costos_defontana_limpio.xlsx"
OUT  = r"C:\Users\dell\victtorino\tablero_rentabilidad.xlsx"
FLOOR = 0.15  # piso de margen neto objetivo

def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii','ignore').decode()
    s = s.upper()
    s = re.sub(r'[^A-Z0-9 ]',' ', s)
    s = re.sub(r'\s+',' ', s).strip()
    return s

# --- 1) parse catalog rows from snapshot ---
prods = []
lines = open(SNAP, encoding='utf-8').read().splitlines()
for i, ln in enumerate(lines):
    m = re.search(r"row \"(.+?) SKU: (\d+) - EAS: - ([\d.]+) CLP\"", ln)
    if not m:
        continue
    name, sku, price = m.group(1), m.group(2), int(m.group(3).replace('.',''))
    checked = False
    for j in range(i, min(i+8, len(lines))):
        if 'checkbox [checked]' in lines[j]:
            checked = True; break
        if 'row "' in lines[j] and j > i:
            break
    prods.append({'name':name,'sku':sku,'price':price,'envio':checked})
# dedupe by sku
seen={};
for p in prods: seen[p['sku']]=p
prods=list(seen.values())

# --- 2) load costs ---
wb=openpyxl.load_workbook(COSTS, read_only=True); ws=wb.active
costs=[]
for r in ws.iter_rows(values_only=True):
    if not r or r[0] in (None,'Codigo'): continue
    if not isinstance(r[2],(int,float)): continue
    costs.append({'cod':r[0],'desc':r[1],'cost':int(r[2]),'ndesc':norm(r[1])})
cmap={c['ndesc']:c for c in costs}
cnames=list(cmap.keys())

# --- 3) match by name ---
def match_cost(name):
    n=norm(name)
    if n in cmap: return cmap[n],1.0
    cand=difflib.get_close_matches(n, cnames, n=1, cutoff=0.6)
    if cand:
        return cmap[cand[0]], difflib.SequenceMatcher(None,n,cand[0]).ratio()
    # token overlap fallback
    best=None;bests=0
    nt=set(n.split())
    for c in costs:
        ct=set(c['ndesc'].split())
        if not ct: continue
        j=len(nt&ct)/len(nt|ct)
        if j>bests: bests=j;best=c
    if best and bests>=0.45: return best,bests
    return None,0

for p in prods:
    c,score=match_cost(p['name'])
    if c and c['cost']>0:
        p['cost']=c['cost']; p['cod']=c['cod']; p['mdesc']=c['desc']; p['score']=round(score,2)
        p['marg$']=p['price']-c['cost']
        p['marg%']=p['marg$']/p['price']
        # max descuento CMR manteniendo piso 15% margen neto (sin ads/envio)
        dmax=1-(c['cost']/(1-FLOOR))/p['price']
        p['cmr_max']=max(0,dmax)
    else:
        p['cost']=None;p['cod']='';p['mdesc']='';p['score']=0;p['marg$']=None;p['marg%']=None;p['cmr_max']=None

# --- 4) report ---
matched=[p for p in prods if p['cost']]
print(f"Catalogo parseado: {len(prods)} | con costo cruzado: {len(matched)} | sin costo: {len(prods)-len(matched)}")
def band(m):
    if m is None: return 'sin costo'
    if m>=0.40: return 'ALTO'
    if m>=0.25: return 'MEDIO'
    return 'BAJO'
from collections import Counter
print('Bandas:', Counter(band(p['marg%']) for p in prods))
print('\nTOP candidatos CMR (alto margen, ordenado por margen%):')
for p in sorted(matched,key=lambda x:-x['marg%'])[:12]:
    print(f"  {p['marg%']*100:5.1f}%  ${p['price']:>7,}  cmrmax {p['cmr_max']*100:4.1f}%  {p['name'][:45]}")
print('\nRIESGO (margen BAJO <25% que ya estan en Envio gratis):')
for p in matched:
    if p['marg%']<0.25 and p['envio']:
        print(f"  {p['marg%']*100:5.1f}%  ${p['price']:>7,}  {p['name'][:50]}")

# --- 5) excel ---
wbo=openpyxl.Workbook(); wso=wbo.active; wso.title='Tablero'
hdr=['SKU Fala','Codigo','Producto','Precio','Costo','Margen $','Margen %','Banda','En Envio gratis','CMR max (piso 15%)','Match %']
wso.append(hdr)
for c in wso[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center')
fills={'ALTO':'C6EFCE','MEDIO':'FFEB9C','BAJO':'FFC7CE','sin costo':'D9D9D9'}
for p in sorted(prods,key=lambda x:-(x['marg%'] if x['marg%'] is not None else -9)):
    b=band(p['marg%'])
    wso.append([p['sku'],p.get('cod',''),p['name'],p['price'],p['cost'],p['marg$'],
                round(p['marg%'],3) if p['marg%'] is not None else None, b,
                'SI' if p['envio'] else '', round(p['cmr_max'],3) if p['cmr_max'] is not None else None, p['score']])
    for c in wso[wso.max_row]: c.fill=PatternFill('solid',fgColor=fills[b])
for col,w in zip('ABCDEFGHIJK',[12,12,46,10,9,9,9,8,14,16,8]):
    wso.column_dimensions[col].width=w
wso['G1'].number_format='0.0%'
for r in range(2,wso.max_row+1):
    wso.cell(r,7).number_format='0.0%'; wso.cell(r,10).number_format='0.0%'
wbo.save(OUT)
print(f"\nTablero guardado: {OUT}")
