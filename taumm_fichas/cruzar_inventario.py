# -*- coding: utf-8 -*-
"""Cruza el inventario de Victor (Salud_Inventario) con las paginas-ficha del
catalogo Taumm. Salida: que ficha tecnica (PDF + pagina) corresponde a cada
producto en stock, con nivel de confianza."""
import os, re, json, unicodedata, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

def norm(s):
    if s is None: return ""
    s = str(s)
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

STOP = set("DE LA EL LOS LAS Y O CON SIN PARA POR EN AL A UN UNA MM CM CMS MT M T C K W".split())
# nombres de modelo Taumm (boost fuerte)
MODELOS = set("NOTTE COLOMBA DOMENICA MODERN CONSTER NEUE FREIBURG DUSSELDORF HANNOVER "
              "HAMBURG HAMBURGO SWERTT SCHWARZ SPIEGEL MIRAGE VANDER ANTONELLA AGUSTINA "
              "GIACOMMO PASCAL FABRICIO MICHELE NEUE".split())

def tokens(s):
    return [t for t in norm(s).split() if t not in STOP and len(t) > 1]

# --- inventario ---
wb = openpyxl.load_workbook(os.path.join(ROOT, 'Salud_Inventario_26052026.xlsx'),
                            read_only=True, data_only=True)
ws = wb['Detalle']
inv = []
for r in list(ws.iter_rows(values_only=True))[1:]:
    if not r[0]:
        continue
    cod = str(r[0]).strip()
    desc = str(r[1]) if r[1] else ""
    marca = ""
    m = re.match(r'^([A-Z])-', desc.strip())
    if m:
        marca = {'T': 'Taumm', 'C': 'Otro (C)', 'K': 'Insumo (K)', 'W': 'Otro (W)'}.get(m.group(1), m.group(1))
        desc_clean = desc.strip()[2:].strip()
    else:
        desc_clean = desc.strip()
    inv.append({'codigo': cod, 'desc': desc, 'desc_clean': desc_clean,
                'marca': marca or 'Sin marca', 'familia_inv': r[2], 'stock': r[3]})

# --- paginas catalogo ---
pages = json.load(open(os.path.join(HERE, 'paginas_catalogo.json'), encoding='utf-8'))
for p in pages:
    p['norm'] = norm(p['texto'])
    p['tokset'] = set(tokens(p['texto']))
    p['codeset'] = set(p['codigos'])

def best_page(item):
    # 1) match por codigo exacto
    for p in pages:
        if item['codigo'] in p['codeset']:
            return p, 100, 'ALTA (codigo exacto)'
    it = tokens(item['desc_clean'])
    itset = set(it)
    if not itset:
        return None, 0, 'sin datos'
    model_in_item = itset & MODELOS
    best = None; best_score = 0
    for p in pages:
        common = itset & p['tokset']
        if not common:
            continue
        score = 0
        for t in common:
            score += 5 if t in MODELOS else 1
        # cobertura relativa
        cov = len(common) / len(itset)
        score += cov * 3
        # penaliza si item tiene modelo y la pagina no lo contiene
        if model_in_item and not (model_in_item & p['tokset']):
            score -= 4
        if score > best_score:
            best_score = score; best = p
    if not best:
        return None, 0, 'sin coincidencia'
    # confianza
    common = itset & best['tokset']
    model_match = bool(model_in_item & best['tokset'])
    cov = len(common) / len(itset)
    if (model_match and cov >= 0.5) or cov >= 0.7:
        conf = 'MEDIA-ALTA'
    elif model_match or cov >= 0.45:
        conf = 'MEDIA'
    else:
        conf = 'BAJA (revisar)'
    return best, round(best_score, 1), conf

results = []
for item in inv:
    p, score, conf = best_page(item)
    results.append({**item, 'page': p, 'score': score, 'conf': conf})

# --- escribir Excel ---
out = openpyxl.Workbook()
sh = out.active
sh.title = 'Fichas por inventario'
headers = ['Codigo inv.', 'Descripcion', 'Marca', 'Stock',
           'Ficha PDF', 'Pagina', 'Familia catalogo', 'Codigos en pagina',
           'Confianza', 'Link ficha (PDF)']
sh.append(headers)
hdr_fill = PatternFill('solid', fgColor='1F4E78'); hdr_font = Font(bold=True, color='FFFFFF')
for c in sh[1]:
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(vertical='center')
conf_fill = {'ALTA (codigo exacto)': 'C6EFCE', 'MEDIA-ALTA': 'D9EAD3',
             'MEDIA': 'FFF2CC', 'BAJA (revisar)': 'FCE4D6',
             'sin coincidencia': 'F2F2F2', 'sin datos': 'F2F2F2'}
BASE_URL = 'https://www.taumm.cl/fotosotrastaumm/Catalogo/'
results.sort(key=lambda r: (r['marca'] != 'Taumm', -r['stock'] if isinstance(r['stock'], int) else 0))
for r in results:
    p = r['page']
    pdf = p['pdf'] if p else ''
    pag = p['pagina'] if p else ''
    fam = p['familia'] if p else ''
    codes = ', '.join(p['codigos'][:8]) if p else ''
    npdf = re.match(r'^(\d+(?:_\d+)?)_', pdf).group(1) + '.pdf' if pdf else ''
    link = BASE_URL + npdf if npdf else ''
    row = [r['codigo'], r['desc'], r['marca'], r['stock'], pdf, pag, fam, codes, r['conf'], link]
    sh.append(row)
    fill = conf_fill.get(r['conf'])
    if fill:
        sh.cell(sh.max_row, 9).fill = PatternFill('solid', fgColor=fill)

widths = [13, 52, 11, 7, 34, 8, 24, 32, 18, 46]
for i, w in enumerate(widths, 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
sh.freeze_panes = 'A2'
sh.auto_filter.ref = f"A1:J{sh.max_row}"

out_path = os.path.join(ROOT, 'fichas_tecnicas_inventario_taumm.xlsx')
out.save(out_path)

# resumen consola
from collections import Counter
conf_cnt = Counter(r['conf'] for r in results)
marca_cnt = Counter(r['marca'] for r in results)
print('Inventario total:', len(results))
print('Por marca:', dict(marca_cnt))
print('Por confianza:')
for k, v in conf_cnt.most_common():
    print(f'  {v:3d}  {k}')
sinm = [r for r in results if not r['page']]
print(f'\nSin ficha encontrada ({len(sinm)}):')
for r in sinm:
    print('  ', r['codigo'], '|', r['desc'][:50])
print('\nExcel:', out_path)
