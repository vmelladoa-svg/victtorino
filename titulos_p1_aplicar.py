"""
Excel solo con los 10 P1 (top mejora). Formato lado a lado antes/después,
agrupado por cuenta para que sea fácil ejecutar el batch en Seller Center.
"""
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from titulos_top30_curated import CURATED, TOP30, title_score

import sys
ROOT = Path(__file__).parent
# Modo: --p1 (default), --p2, --all
MODE = "p1"
for a in sys.argv:
    if a == "--p2": MODE = "p2"
    elif a == "--p3": MODE = "p3"
    elif a == "--all": MODE = "all"
OUT = ROOT / f"titulos_{MODE.upper()}_solo.xlsx"


def main():
    by_id = {it["iid"]: it for it in TOP30}
    rows = []
    for iid, new_title in CURATED.items():
        it = by_id.get(iid)
        if not it: continue
        current = it["title"]
        rows.append({
            "iid": iid, "cuenta": it["cuenta"], "stock": it["stock"], "precio": it["precio"],
            "urgencia": it["urgencia"],
            "actual": current, "nuevo": new_title,
            "delta": title_score(new_title) - title_score(current),
            "score_actual": title_score(current),
            "score_nuevo": title_score(new_title),
        })
    df = pd.DataFrame(rows).sort_values(["delta", "urgencia"], ascending=[False, False]).reset_index(drop=True)
    if MODE == "p1":
        slice_df = df.iloc[:10].copy()
    elif MODE == "p2":
        slice_df = df.iloc[10:20].copy()
    elif MODE == "p3":
        slice_df = df.iloc[20:30].copy()
    else:
        slice_df = df.copy()
    # Ordenar por cuenta para facilitar ejecución secuencial
    p1 = slice_df.sort_values(["cuenta", "delta"], ascending=[True, False])

    # Construir DataFrame final
    out_rows = []
    for i, r in enumerate(p1.itertuples(index=False), 1):
        cambio = []
        # Detectar tipo de cambio
        if "  " in r.actual:
            cambio.append("doble espacio")
        if r.actual.lower() != r.nuevo.lower():
            # Diferencias palabra por palabra (simplificado)
            act_words = set(r.actual.lower().split())
            new_words = set(r.nuevo.lower().split())
            added = new_words - act_words
            removed = act_words - new_words
            if added:
                cambio.append(f"+{','.join(sorted(added))[:80]}")
            if removed:
                cambio.append(f"-{','.join(sorted(removed))[:60]}")
        out_rows.append({
            "#": i,
            "Cuenta": r.cuenta,
            "ItemID": r.iid,
            "Stock": r.stock,
            "Precio": r.precio,
            "↓ ANTES ↓": r.actual,
            "Long ant": len(r.actual),
            "↑ DESPUÉS ↑": r.nuevo,
            "Long nuevo": len(r.nuevo),
            "Δ Score": f"+{r.delta}",
            "Score": f"{r.score_actual} → {r.score_nuevo}",
            "Qué cambia": " | ".join(cambio) if cambio else "—",
            "URL editar": f"https://www.mercadolibre.cl/publicaciones/{r.iid}/modificar",
            "URL pública": f"https://articulo.mercadolibre.cl/{r.iid[:3]}-{r.iid[3:]}",
        })
    out = pd.DataFrame(out_rows)
    out.to_excel(OUT, index=False)

    # Formato
    wb = load_workbook(OUT)
    ws = wb.active
    HDR = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = HDR
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colores antes/después
    BEFORE = PatternFill("solid", fgColor="FFE5E5")
    AFTER = PatternFill("solid", fgColor="E5F5E5")
    GROUP = {"C1": "FFF2CC", "C2": "DDEBF7", "C3": "E2EFDA"}

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {"A":4,"B":7,"C":15,"D":7,"E":10,"F":55,"G":8,"H":55,"I":9,"J":8,"K":12,"L":40,"M":48,"N":45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(2, ws.max_row+1):
        cuenta = ws.cell(r, 2).value
        for c in range(1, ws.max_column+1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(r, c).border = border
            if c == 6:  # ANTES
                ws.cell(r, c).fill = BEFORE
            elif c == 8:  # DESPUÉS
                ws.cell(r, c).fill = AFTER
        # Tintar cuenta col
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=GROUP.get(cuenta, "FFFFFF"))
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"OK {OUT}")
    print(f"\n=== {len(out)} items {MODE.upper()} a aplicar ===\n")
    for _, r in out.iterrows():
        print(f"#{r['#']}  [{r['Cuenta']}]  {r['ItemID']}  stock={r['Stock']}  Δ={r['Δ Score']}")
        print(f"   ANTES:   {r['↓ ANTES ↓']}")
        print(f"   DESPUÉS: {r['↑ DESPUÉS ↑']}")
        print(f"   Cambio:  {r['Qué cambia']}")
        print(f"   Editar:  {r['URL editar']}")
        print()


if __name__ == "__main__":
    main()
