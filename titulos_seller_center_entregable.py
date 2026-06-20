"""
Genera entregable final para edición manual en Seller Center.
Excel con:
  - URL Seller Center directa al editor del item
  - Permalink público
  - Título actual vs sugerido (copy-paste ready)
  - Score actual vs sugerido
  - Instrucciones por celda
  - Hoja "Instructivo" con paso a paso
"""
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from titulos_top30_curated import CURATED, TOP30, title_score

ROOT = Path(__file__).parent
OUT = ROOT / "titulos_seller_center_2026-05-23.xlsx"

# UIDs por cuenta para construir URL Seller Center
UIDS = {"C1": 483903060, "C2": 483904870, "C3": 1194418785}


def seller_center_edit_url(cuenta, iid):
    """URL directa al editor de la publicación en Seller Center."""
    return f"https://www.mercadolibre.cl/publicaciones/{iid}/modificar"


def main():
    items_by_id = {it["iid"]: it for it in TOP30}
    rows = []
    for iid, new_title in CURATED.items():
        it = items_by_id.get(iid)
        if not it:
            continue
        current = it["title"]
        rows.append({
            "Prioridad": "",
            "Cuenta": it["cuenta"],
            "ItemID": iid,
            "Stock": it["stock"],
            "Precio CLP": it["precio"],
            "Urgencia 0-100": it["urgencia"],
            "Título ACTUAL": current,
            "Long actual": len(current),
            "Score actual": title_score(current),
            "Título NUEVO (copiar y pegar)": new_title,
            "Long nuevo": len(new_title),
            "Score nuevo": title_score(new_title),
            "Δ Score": title_score(new_title) - title_score(current),
            "Editar en Seller Center": seller_center_edit_url(it["cuenta"], iid),
            "Ver publicación pública": f"https://articulo.mercadolibre.cl/{iid[:3]}-{iid[3:]}",
            "Notas": "",
        })
    df = pd.DataFrame(rows).sort_values("Δ Score", ascending=False)
    # Asignar prioridad
    for i, idx in enumerate(df.index, 1):
        if i <= 10:
            df.at[idx, "Prioridad"] = f"P1 ({i})"
        elif i <= 20:
            df.at[idx, "Prioridad"] = f"P2 ({i})"
        else:
            df.at[idx, "Prioridad"] = f"P3 ({i})"

    # Hoja Instructivo
    instructivo = pd.DataFrame([
        {"Paso": 1, "Acción": "Abrir Seller Center", "Detalle": "Login en https://www.mercadolibre.cl/ con la cuenta correspondiente (ver columna 'Cuenta' de la otra hoja)"},
        {"Paso": 2, "Acción": "Ir al editor del item", "Detalle": "Click en columna 'Editar en Seller Center' (URL pre-construida). Te abre directo el editor."},
        {"Paso": 3, "Acción": "Cambiar título", "Detalle": "Copiar el texto de columna 'Título NUEVO' y pegar en el campo título. ML te preguntará si querés cambiar la familia — confirmar."},
        {"Paso": 4, "Acción": "Guardar", "Detalle": "Click 'Guardar cambios'. El item queda activo con el nuevo título."},
        {"Paso": 5, "Acción": "Verificar", "Detalle": "Visitar el permalink público (columna 'Ver publicación') para confirmar que el cambio se reflejó."},
        {"Paso": 6, "Acción": "Medir efecto", "Detalle": "En 7-14 días, re-ejecutar auditoria_ml.py + fix_visitas_snapshots.py para ver impacto en visitas y conversión."},
        {"Paso": "—", "Acción": "Limitación ML", "Detalle": "ML asocia automáticamente 'family_name' a cada item, lo cual bloquea cambio vía API. Solo Seller Center permite forzar el cambio con confirmación explícita."},
        {"Paso": "—", "Acción": "Tiempo estimado", "Detalle": "3-5 minutos por item × 30 items = 1.5 a 2.5 horas total. Priorizar P1 (mayor mejora) primero."},
    ])

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Títulos a aplicar", index=False)
        instructivo.to_excel(writer, sheet_name="Instructivo", index=False)

    # Formato
    wb = load_workbook(OUT)
    ws = wb["Títulos a aplicar"]
    # Header
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Anchos
    widths = {
        "A": 8, "B": 8, "C": 16, "D": 8, "E": 12, "F": 11,
        "G": 50, "H": 11, "I": 12, "J": 50, "K": 11, "L": 11, "M": 9,
        "N": 50, "O": 50, "P": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # Color por prioridad
    p1_fill = PatternFill("solid", fgColor="C6EFCE")
    p2_fill = PatternFill("solid", fgColor="FFEB9C")
    p3_fill = PatternFill("solid", fgColor="FFD68A")
    for row in ws.iter_rows(min_row=2):
        p = row[0].value or ""
        if p.startswith("P1"):
            row[0].fill = p1_fill
        elif p.startswith("P2"):
            row[0].fill = p2_fill
        else:
            row[0].fill = p3_fill
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # Hoja instructivo
    wi = wb["Instructivo"]
    for cell in wi[1]:
        cell.fill = header_fill
        cell.font = header_font
    wi.column_dimensions["A"].width = 10
    wi.column_dimensions["B"].width = 30
    wi.column_dimensions["C"].width = 80
    for row in wi.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"OK {OUT}")
    print(f"\nPriorización (Δ Score desc):")
    print(df[["Prioridad", "Cuenta", "ItemID", "Δ Score", "Long actual", "Long nuevo"]].to_string(index=False))


if __name__ == "__main__":
    main()
