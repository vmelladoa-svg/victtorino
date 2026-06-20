# -*- coding: utf-8 -*-
"""Planilla de validacion de costos: Defontana vs mejor match Taumm, con score y
confianza por color. Sugiere 'Costo a usar'. Para que Victor valide los dudosos."""
import json,re,unicodedata,difflib,pandas as pd,openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
def norm(s):
    s=str(s).replace('�',' ').replace('�',' ')
    s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode().upper()
    return re.sub(r'\s+',' ',re.sub(r'[^A-Z0-9 ]',' ',s)).strip()
def n(x):
    s=re.sub(r'[^0-9]','',str(x)); return int(s) if s else 0

# Taumm
d=json.load(open('taumm_full.json',encoding='utf-8'))
if isinstance(d,str): d=json.loads(d)
twd={r['cod']:r for r in d['all']}
tw=[{'cod':r['cod'],'desc':r['desc'].replace('�','n'),'cost':n(r['precio']),'stock':n(r['stock']),'nd':norm(r['desc'])} for r in twd.values() if n(r['precio'])>0]
twn=[x['nd'] for x in tw]

# Defontana
t=pd.read_html(r'C:\Users\dell\Downloads\Informe de artículos 20250627210343860887 14062026 0314.xls')[0].iloc[3:]
dfn=[]
for _,r in t.iterrows():
    m=re.match(r'^(\d+-[A-Z0-9]+)-(.+)$',str(r.iloc[0]))
    if not m: continue
    dfn.append({'cod':m.group(1),'desc':m.group(2).strip(),'cost':n(r.iloc[2]),'nd':norm(m.group(2))})

def best(ndesc):
    cand=difflib.get_close_matches(ndesc,twn,n=1,cutoff=0.5)
    if cand:
        i=twn.index(cand[0]); sc=difflib.SequenceMatcher(None,ndesc,cand[0]).ratio()
        return tw[i],sc
    # token overlap fallback
    nt=set(ndesc.split()); bx=None;bs=0
    for x in tw:
        ct=set(x['nd'].split())
        if ct:
            j=len(nt&ct)/len(nt|ct)
            if j>bs: bs=j;bx=x
    return (bx,bs) if bx and bs>=0.35 else (None,0)

rows=[]
for x in dfn:
    tm,sc=best(x['nd'])
    err_def = x['cost']>0 and x['cost']<500   # costo defontana absurdo
    if sc>=0.85: conf='ALTA'
    elif sc>=0.70: conf='MEDIA'
    elif sc>=0.50: conf='BAJA'
    else: conf='SIN MATCH'
    # costo a usar
    if conf in('ALTA',) and tm: usar=tm['cost']; fuente='Taumm'
    elif err_def and tm: usar=tm['cost']; fuente='Taumm (Defontana erroneo)'
    else: usar=x['cost']; fuente='Defontana'
    rows.append({'cod':x['cod'],'desc':x['desc'],'cdef':x['cost'],
                 'ctw':tm['cost'] if tm else None,'tdesc':tm['desc'] if tm else '',
                 'tstock':tm['stock'] if tm else None,'sc':round(sc,2),'conf':conf,
                 'usar':usar,'fuente':fuente,'err':err_def})

from collections import Counter
print('Defontana:',len(dfn))
print('Confianza:',dict(Counter(r['conf'] for r in rows)))
print('Errores Defontana (<$500):',sum(1 for r in rows if r['err']))
rev=[r for r in rows if r['conf'] in('MEDIA','BAJA','SIN MATCH') or r['err']]
print('A REVISAR (media/baja/sin match/erroneos):',len(rev))

wb=openpyxl.Workbook(); w=wb.active; w.title='Validacion costos'
w.append(['Codigo','Descripcion','Costo Defontana','Costo Taumm','Desc Taumm (match)','Stock Taumm','Score','Confianza','COSTO A USAR','Fuente','OK? (validar)'])
for c in w[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E79'); c.alignment=Alignment(horizontal='center',wrap_text=True)
col={'ALTA':'C6EFCE','MEDIA':'FFEB9C','BAJA':'FCE4D6','SIN MATCH':'FFC7CE'}
ordn={'SIN MATCH':0,'BAJA':1,'MEDIA':2,'ALTA':3}
for r in sorted(rows,key=lambda z:(0 if z['err'] else 1, ordn[z['conf']])):
    w.append([r['cod'],r['desc'],r['cdef'],r['ctw'],r['tdesc'],r['tstock'],r['sc'],r['conf'],r['usar'],r['fuente'],''])
    for c in w[w.max_row]: c.fill=PatternFill('solid',fgColor=col[r['conf']])
for col_i in [3,4,6,9]:
    for rr in range(2,w.max_row+1):
        cc=w.cell(rr,col_i)
        if isinstance(cc.value,(int,float)): cc.number_format='#,##0'
for i,wd in enumerate([12,34,14,12,32,11,7,11,14,22,12]): w.column_dimensions[chr(65+i)].width=wd
wb.save('validacion_costos.xlsx')
print('\nGuardado: validacion_costos.xlsx (ordenado: errores y dudosos primero)')
